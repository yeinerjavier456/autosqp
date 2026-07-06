from fastapi import FastAPI, Depends, HTTPException, status, Query, Query, Body, UploadFile, File, Form, Response
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from typing import Optional, List, Dict, Any
from sqlalchemy.orm import Session, joinedload, selectinload, noload
from sqlalchemy import or_, and_, func, text, false
from fastapi.middleware.cors import CORSMiddleware
from database import engine, Base, get_db
import models, schemas, auth_utils
import lead_assignment
from models import LeadNote, LeadFile # Explicitly for create_all to see them
from routers import whatsapp, credits, purchases, notifications, rules, vehicles, meta, tiktok, gmail, appointments # Import the new routers
from jose import JWTError, jwt
import datetime
import os
import traceback
import requests
import time
import re
import json
import random
import base64
import secrets
import shutil
import smtplib
import ssl
import uuid
import unicodedata
from html import escape
from io import BytesIO
from email.message import EmailMessage
from zoneinfo import ZoneInfo
from view_registry import SYSTEM_VIEWS, DEFAULT_ROLE_VIEW_ACCESS, DEFAULT_ROLE_MENU_ORDER, VALID_VIEW_IDS, COMPANY_VIEW_IDS
from fastapi import Request
from fastapi.responses import JSONResponse
from fastapi.responses import FileResponse
from fastapi.responses import StreamingResponse
from schema_ensure import ensure_mysql_schema

# Create tables
models.Base.metadata.create_all(bind=engine)
ensure_mysql_schema(engine)

BOGOTA_TZ = ZoneInfo("America/Bogota")

LEGACY_LEAD_STATUS_MAP = {
    "interested": "in_process",
    "credit_application": "credit_study",
    "qualified": "approvals",
    "ally_managed": "new",
}

LEAD_STATUS_SEQUENCE = [
    "new",
    "contacted",
    "in_process",
    "credit_study",
    "approvals",
    "reserved",
    "preparation",
    "sold",
    "lost",
]

LEAD_CREDIT_FLOW_STATUSES = {"credit_study", "approvals", "reserved", "preparation", "sold"}
LEAD_STATUS_LABELS = {
    "new": "Nuevos",
    "contacted": "Contactados",
    "in_process": "En proceso",
    "credit_study": "Estudio de crédito",
    "approvals": "Aprobaciones",
    "reserved": "Reservas",
    "preparation": "Alistamientos",
    "sold": "Vendidos",
    "lost": "Perdidos",
}

DEFAULT_PUBLIC_COMPANY_CONTEXT = {
    "id": None,
    "name": "AutosQP",
    "public_domain": None,
    "public_domains": [],
    "website_url": None,
    "logo_url": None,
    "contact_address": None,
    "contact_phone": None,
    "social_instagram": None,
    "social_tiktok": None,
    "social_facebook": None,
    "primary_color": "#2563eb",
    "secondary_color": "#0f172a",
    "enabled_modules": [],
    "public_credit_requires_email_validation": True,
    "license_status": "unlimited",
    "license_notice": None,
    "license_days_remaining": None,
}

DEFAULT_COMPANY_ENABLED_MODULES = sorted(COMPANY_VIEW_IDS)
ROLE_REQUIRED_MODULES = {
    "inventario": {"inventory"},
    "compras": {"purchase_board"},
    "gestion_creditos": {"credits"},
    "aliado": {"ally_board"},
}
MODULE_ROLE_FALLBACK_NAME = "user"


def normalize_company_enabled_modules(modules: Optional[List[str]]) -> List[str]:
    if modules is None:
        return DEFAULT_COMPANY_ENABLED_MODULES[:]
    normalized = []
    for module_id in modules:
        candidate = str(module_id or "").strip()
        if not candidate or candidate not in COMPANY_VIEW_IDS or candidate in normalized:
            continue
        normalized.append(candidate)
    return normalized


def get_company_enabled_modules(company: Optional[models.Company]) -> List[str]:
    if not company:
        return DEFAULT_COMPANY_ENABLED_MODULES[:]
    if getattr(company, "enabled_modules_json", None) is None:
        return DEFAULT_COMPANY_ENABLED_MODULES[:]
    return normalize_company_enabled_modules(getattr(company, "enabled_modules", []))


def apply_company_module_limits(view_ids: List[str], company: Optional[models.Company]) -> List[str]:
    enabled_modules = set(get_company_enabled_modules(company))
    filtered = []
    for view_id in view_ids or []:
        if view_id not in COMPANY_VIEW_IDS or view_id in enabled_modules:
            filtered.append(view_id)
    return filtered


def is_role_enabled_for_company(role: Optional[models.Role], company: Optional[models.Company]) -> bool:
    if role is None:
        return False
    if not company:
        return True

    role_name = get_effective_role_key(role)
    required_modules = ROLE_REQUIRED_MODULES.get(role_name)
    if not required_modules:
        return True

    enabled_modules = set(get_company_enabled_modules(company))
    return bool(enabled_modules.intersection(required_modules))


def get_role_disabled_required_modules(role: Optional[models.Role], company: Optional[models.Company]) -> List[str]:
    if role is None or not company:
        return []
    role_name = get_effective_role_key(role)
    required_modules = ROLE_REQUIRED_MODULES.get(role_name, set())
    if not required_modules:
        return []
    enabled_modules = set(get_company_enabled_modules(company))
    return sorted(required_modules - enabled_modules)


def get_company_by_id(db: Session, company_id: Optional[int]) -> Optional[models.Company]:
    if not company_id:
        return None
    return db.query(models.Company).filter(models.Company.id == company_id).first()


def company_has_enabled_module(
    module_id: str,
    *,
    company: Optional[models.Company] = None,
    company_id: Optional[int] = None,
    db: Optional[Session] = None,
) -> bool:
    target_company = company
    if target_company is None and company_id and db is not None:
        target_company = get_company_by_id(db, company_id)
    enabled_modules = set(get_company_enabled_modules(target_company))
    return module_id in enabled_modules


def ensure_public_chat_enabled_for_company(
    db: Session,
    *,
    company_id: Optional[int] = None,
    company: Optional[models.Company] = None,
) -> models.Company:
    target_company = company or get_company_by_id(db, company_id)
    if not company_has_enabled_module("public_sales_chat", company=target_company):
        raise HTTPException(
            status_code=403,
            detail="El chat pUblico no estA habilitado para esta empresa",
        )
    return target_company


def get_company_allowed_lead_statuses(company: Optional[models.Company]) -> List[str]:
    statuses = ["new", "contacted", "in_process"]
    if company_has_enabled_module("credits", company=company):
        statuses.extend(["credit_study", "approvals"])
    if company_has_enabled_module("purchase_board", company=company):
        statuses.extend(["reserved", "preparation"])
    statuses.extend(["sold", "lost"])
    return statuses


def normalize_company_lead_status(
    status: Optional[str],
    company: Optional[models.Company],
    default: str = "new"
) -> str:
    normalized_status = normalize_lead_status_value(status, default)
    allowed_statuses = set(get_company_allowed_lead_statuses(company))

    if normalized_status in allowed_statuses:
        return normalized_status

    if normalized_status in {"credit_study", "approvals"}:
        return models.LeadStatus.IN_PROCESS.value

    if normalized_status in {"reserved", "preparation"}:
        if models.LeadStatus.APPROVALS.value in allowed_statuses:
            return models.LeadStatus.APPROVALS.value
        return models.LeadStatus.IN_PROCESS.value

    if normalized_status in {models.LeadStatus.SOLD.value, models.LeadStatus.LOST.value}:
        return normalized_status if normalized_status in allowed_statuses else default

    return default


def get_bogota_today() -> datetime.date:
    return datetime.datetime.now(BOGOTA_TZ).date()


def get_company_license_state(company: Optional[models.Company]) -> Dict[str, Any]:
    if not company:
        return {"status": "unlimited", "notice": None, "days_remaining": None}

    today = get_bogota_today()
    start_date = getattr(company, "license_start_date", None)
    end_date = getattr(company, "license_end_date", None)

    if start_date and today < start_date:
        days_until_start = (start_date - today).days
        return {
            "status": "pending",
            "notice": f"La licencia de la empresa inicia el {start_date.strftime('%d/%m/%Y')}. Faltan {days_until_start} día(s).",
            "days_remaining": days_until_start,
        }

    if end_date:
        days_remaining = (end_date - today).days
        if days_remaining < 0:
            return {
                "status": "expired",
                "notice": f"La licencia de la empresa terminó el {end_date.strftime('%d/%m/%Y')}.",
                "days_remaining": days_remaining,
            }
        if days_remaining <= 5:
            return {
                "status": "warning",
                "notice": f"La licencia vence el {end_date.strftime('%d/%m/%Y')}. Quedan {days_remaining} día(s).",
                "days_remaining": days_remaining,
            }
        return {"status": "active", "notice": None, "days_remaining": days_remaining}

    return {"status": "unlimited", "notice": None, "days_remaining": None}


def ensure_company_license_allows_login(company: Optional[models.Company]):
    license_state = get_company_license_state(company)
    if license_state["status"] in {"pending", "expired"}:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=license_state["notice"] or "La licencia de la empresa no permite el acceso.",
        )
    return license_state


def count_company_users(db: Session, company_id: Optional[int]) -> int:
    if not company_id:
        return 0
    return db.query(func.count(models.User.id)).filter(models.User.company_id == company_id).scalar() or 0


def count_company_active_users(db: Session, company_id: Optional[int]) -> int:
    if not company_id:
        return 0
    return db.query(func.count(models.User.id)).filter(
        models.User.company_id == company_id,
        models.User.is_active != 0
    ).scalar() or 0


def count_company_leads_in_license_period(db: Session, company: Optional[models.Company]) -> int:
    if not company:
        return 0
    query = db.query(func.count(models.Lead.id)).filter(models.Lead.company_id == company.id)
    if getattr(company, "license_start_date", None):
        query = query.filter(func.date(models.Lead.created_at) >= company.license_start_date)
    if getattr(company, "license_end_date", None):
        query = query.filter(func.date(models.Lead.created_at) <= company.license_end_date)
    return query.scalar() or 0


def ensure_company_user_limit(db: Session, company_id: Optional[int], *, excluding_user_id: Optional[int] = None):
    if not company_id:
        return
    company = db.query(models.Company).filter(models.Company.id == company_id).first()
    if not company or not company.max_users:
        return
    query = db.query(func.count(models.User.id)).filter(models.User.company_id == company_id)
    if excluding_user_id:
        query = query.filter(models.User.id != excluding_user_id)
    total_users = query.scalar() or 0
    if total_users >= company.max_users:
        raise HTTPException(status_code=403, detail=f"La empresa alcanzó el límite de usuarios ({company.max_users}).")


def ensure_company_active_account_limit(db: Session, company_id: Optional[int], *, excluding_user_id: Optional[int] = None):
    if not company_id:
        return
    company = db.query(models.Company).filter(models.Company.id == company_id).first()
    if not company or not company.max_active_accounts:
        return
    query = db.query(func.count(models.User.id)).filter(
        models.User.company_id == company_id,
        models.User.is_active != 0
    )
    if excluding_user_id:
        query = query.filter(models.User.id != excluding_user_id)
    total_active_users = query.scalar() or 0
    if total_active_users >= company.max_active_accounts:
        raise HTTPException(status_code=403, detail=f"La empresa alcanzó el límite de cuentas activas ({company.max_active_accounts}).")


def ensure_company_lead_creation_capacity(db: Session, company_id: Optional[int]):
    if not company_id:
        return
    company = db.query(models.Company).filter(models.Company.id == company_id).first()
    if not company:
        return
    if company.max_leads:
        total_leads = count_company_leads_in_license_period(db, company)
        if total_leads >= company.max_leads:
            raise HTTPException(status_code=403, detail=f"La empresa alcanzó el límite de creación de leads ({company.max_leads}) para la licencia actual.")


def normalize_public_host(raw_host: Optional[str]) -> str:
    host = (raw_host or "").strip().lower()
    if not host:
        return ""
    if ":" in host:
        host = host.split(":", 1)[0]
    if host.startswith("www."):
        host = host[4:]
    return host


def normalize_public_domains(domains: Optional[List[str]], primary_domain: Optional[str] = None) -> List[str]:
    normalized = []
    seed_values = list(domains or [])
    if primary_domain:
        seed_values.insert(0, primary_domain)

    for value in seed_values:
        normalized_value = normalize_public_host(value)
        if normalized_value and normalized_value not in normalized:
            normalized.append(normalized_value)
    return normalized


def company_matches_public_host(company: models.Company, host: str) -> bool:
    if not company or not host:
        return False
    domains = normalize_public_domains(getattr(company, "public_domains", []), getattr(company, "public_domain", None))
    return host in domains


def validate_company_public_domains(db: Session, domains: List[str], exclude_company_id: Optional[int] = None):
    if not domains:
        return

    existing_companies = db.query(models.Company).all()
    for company in existing_companies:
        if exclude_company_id and company.id == exclude_company_id:
            continue
        company_domains = normalize_public_domains(getattr(company, "public_domains", []), getattr(company, "public_domain", None))
        collision = next((domain for domain in domains if domain in company_domains), None)
        if collision:
            raise HTTPException(
                status_code=400,
                detail=f'El dominio "{collision}" ya está configurado en la empresa "{company.name}"'
            )


def resolve_public_company(db: Session, request: Optional[Request]) -> Optional[models.Company]:
    if request is None:
        return None

    host = normalize_public_host(request.headers.get("host"))
    if not host:
        return None

    direct_match = db.query(models.Company).filter(
        func.lower(models.Company.public_domain) == host
    ).first()
    if direct_match:
        return direct_match

    for company in db.query(models.Company).all():
        if company_matches_public_host(company, host):
            return company

    if host in {"autosqp.com", "autosqp.co"}:
        return db.query(models.Company).order_by(models.Company.id.asc()).first()

    return None


def serialize_public_company(company: Optional[models.Company]) -> schemas.PublicCompanyContext:
    if not company:
        return schemas.PublicCompanyContext(**DEFAULT_PUBLIC_COMPANY_CONTEXT)
    license_state = get_company_license_state(company)
    return schemas.PublicCompanyContext(
        id=company.id,
        name=company.name or DEFAULT_PUBLIC_COMPANY_CONTEXT["name"],
        public_domain=company.public_domain,
        public_domains=normalize_public_domains(getattr(company, "public_domains", []), company.public_domain),
        website_url=getattr(company, "website_url", None),
        logo_url=company.logo_url,
        contact_address=getattr(company, "contact_address", None),
        contact_phone=getattr(company, "contact_phone", None),
        social_instagram=getattr(company, "social_instagram", None),
        social_tiktok=getattr(company, "social_tiktok", None),
        social_facebook=getattr(company, "social_facebook", None),
        primary_color=company.primary_color or DEFAULT_PUBLIC_COMPANY_CONTEXT["primary_color"],
        secondary_color=company.secondary_color or DEFAULT_PUBLIC_COMPANY_CONTEXT["secondary_color"],
        enabled_modules=get_company_enabled_modules(company),
        public_credit_requires_email_validation=bool(getattr(company, "public_credit_requires_email_validation", True)),
        license_status=license_state["status"],
        license_notice=license_state["notice"],
        license_days_remaining=license_state["days_remaining"],
    )


def company_requires_public_credit_email_validation(company: Optional[models.Company]) -> bool:
    if not company:
        return True
    return bool(getattr(company, "public_credit_requires_email_validation", True))


PUBLIC_CREDIT_POLICY_TEXT = """
AUTORIZO A AUTOS QP SAS Y A LAS ENTIDADES QUE PERTENEZCAN O LLEGAREN A PERTENECER A SU GRUPO EMPRESARIAL DE ACUERDO CON LA LEY, SUS FILIALES Y/O SUBSIDIARIAS, O A LAS ENTIDADES EN LAS CUALES ÉSTAS, DIRECTA O INDIRECTAMENTE, TENGAN PARTICIPACIÓN ACCIONARIA O SEAN ASOCIADAS, DOMICILIADAS EN COLOMBIA Y/O EN EL EXTERIOR, O A QUIEN REPRESENTE SUS DERECHOS U OSTENTE EN EL FUTURO LA CALIDAD DE ACREEDOR, CESIONARIO O CUALQUIER OTRA CALIDAD FRENTE A MÍ COMO TITULAR DE LA INFORMACIÓN, EN ADELANTE LAS ENTIDADES; Y AUTORIZO A LAS ENTIDADES FINANCIERAS ALIADAS CON LAS QUE LAS ENTIDADES CONSIDEREN Y SOSTENGAN RELACIÓN COMERCIAL, A QUIENES AUTORIZO EN FORMA PERMANENTE PARA QUE: (I) LIBEREN LA INFORMACIÓN NECESARIA QUE LES SOLICITEN SEGÚN MI PERFIL Y SUS POLÍTICAS DE OTORGAMIENTO CREDITICIO, PARA LA BÚSQUEDA DE MI CUPO DE CRÉDITO ANTE LAS ENTIDADES FINANCIERAS ALIADAS, ENTIDADES AVALADORAS U OTRAS, PARA QUE ME SEAN ENVIADAS OFERTAS O AVISOS COMERCIALES RELACIONADOS CON EL TIPO DE CRÉDITO QUE ESTOY SOLICITANDO O CON PRODUCTOS AFINES. ENTIENDO QUE LAS ENTIDADES NO ASUMEN RESPONSABILIDAD ALGUNA POR LA APROBACIÓN O NEGACIÓN DEL CRÉDITO POR PARTE DE LAS ENTIDADES FINANCIERAS ALIADAS, AVALADORAS U OTRAS, NI SE COMPROMETEN A OBTENER SU APROBACIÓN, POR CUANTO SIMPLEMENTE ACTÚAN COMO CANAL DE INFORMACIÓN ENTRE EL SOLICITANTE DEL CRÉDITO Y LA ENTIDAD FINANCIERA, LA ENTIDAD AVALADORA U OTRA. (II) SOLICITEN, CONSULTEN, COMPARTAN, INFORMEN, REPORTEN, PROCESEN, MODIFIQUEN, ACTUALICEN, ACLAREN, RETIREN O DIVULGUEN, ANTE LAS ENTIDADES DE CONSULTA DE BASES DE DATOS U OPERADORES DE INFORMACIÓN Y RIESGO, O ANTE CUALQUIER ENTIDAD QUE MANEJE O ADMINISTRE BASES DE DATOS CON LOS FINES LEGALMENTE DEFINIDOS PARA ESTE TIPO DE ENTIDADES, TODO LO REFERENTE A MI INFORMACIÓN FINANCIERA, COMERCIAL Y CREDITICIA, PRESENTE, PASADA O FUTURA, MI ENDEUDAMIENTO Y EL NACIMIENTO, MODIFICACIÓN Y EXTINCIÓN DE MIS DERECHOS Y OBLIGACIONES ORIGINADOS EN VIRTUD DE CUALQUIER CONTRATO CELEBRADO U OPERACIÓN REALIZADA O QUE LLEGARE A CELEBRAR O REALIZAR CON CUALQUIERA DE LAS ENTIDADES. (III) CONSULTEN, SOLICITEN O VERIFIQUEN INFORMACIÓN SOBRE MIS DATOS DE UBICACIÓN O CONTACTO, LOS BIENES O DERECHOS QUE POSEO O LLEGARE A POSEER Y QUE REPOSEN EN BASES DE DATOS DE ENTIDADES PÚBLICAS O PRIVADAS, O QUE CONOZCAN PERSONAS NATURALES O JURÍDICAS, O SE ENCUENTREN EN BUSCADORES PÚBLICOS, REDES SOCIALES O PUBLICACIONES FÍSICAS O ELECTRÓNICAS, BIEN SEA EN COLOMBIA O EN EL EXTERIOR. (IV) ME CONTACTEN A TRAVÉS DEL ENVÍO DE MENSAJES A MI TERMINAL MÓVIL DE TELECOMUNICACIONES Y/O A TRAVÉS DE CORREO ELECTRÓNICO Y/O REDES SOCIALES EN LAS CUALES ESTÉ INSCRITO. (V) CONSERVEN MI INFORMACIÓN Y DOCUMENTACIÓN AUN CUANDO NO SE HAYA PERFECCIONADO UNA RELACIÓN CONTRACTUAL O DESPUÉS DE FINALIZADA LA MISMA CON CUALQUIERA DE LAS ENTIDADES, IGUALMENTE PARA RECOLECTARLA, ACTUALIZARLA, MODIFICARLA, PROCESARLA Y ELIMINARLA DE CONFORMIDAD CON LA LEY APLICABLE. (VI) LAS ENTIDADES COMPARTAN, REMITAN Y ACCEDAN ENTRE SÍ A MI INFORMACIÓN O DOCUMENTACIÓN CONSIGNADA O ANEXA EN LAS SOLICITUDES DE VINCULACIÓN, ACTUALIZACIONES EN LOS DIFERENTES DOCUMENTOS DE DEPÓSITO Y/O CRÉDITO, OPERACIONES Y/O SISTEMAS DE INFORMACIÓN, ASÍ COMO INFORMACIÓN Y/O DOCUMENTACIÓN RELACIONADA CON LOS PRODUCTOS Y/O SERVICIOS QUE POSEO EN CUALQUIERA DE ELLAS. (VII) ELABOREN ESTADÍSTICAS Y DERIVEN MEDIANTE MODELOS MATEMÁTICOS CONCLUSIONES A PARTIR DE ELLAS. DECLARO HABER LEÍDO CUIDADOSAMENTE EL CONTENIDO DE ESTA CLÁUSULA Y HABERLA COMPRENDIDO A CABALIDAD, RAZÓN POR LA CUAL ENTIENDO SUS ALCANCES E IMPLICACIONES.
""".strip()


def _sanitize_public_credit_filename(file_name: Optional[str], fallback: str) -> str:
    raw_name = str(file_name or fallback).strip() or fallback
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", raw_name)
    return safe_name or fallback


def _save_public_credit_upload(upload, subdirectory: str, fallback_name: str) -> Optional[str]:
    if not upload:
        return None

    os.makedirs(os.path.join("static", "public-credit", subdirectory), exist_ok=True)
    original_name = getattr(upload, "filename", None)
    extension = os.path.splitext(original_name or fallback_name)[1] or ".bin"
    unique_name = f"{uuid.uuid4().hex}{extension}"
    safe_name = _sanitize_public_credit_filename(unique_name, fallback_name)
    target_path = os.path.join("static", "public-credit", subdirectory, safe_name)

    if hasattr(upload, "file"):
        with open(target_path, "wb") as buffer:
            try:
                upload.file.seek(0)
            except Exception:
                pass
            shutil.copyfileobj(upload.file, buffer)
    else:
        with open(target_path, "wb") as buffer:
            buffer.write(upload)

    return f"/api/static/public-credit/{subdirectory}/{safe_name}".replace("\\", "/")


def _save_public_credit_data_url(data_url: Optional[str], subdirectory: str, fallback_name: str) -> Optional[str]:
    raw_value = str(data_url or "").strip()
    if not raw_value or "," not in raw_value:
        return None

    header, encoded = raw_value.split(",", 1)
    mime_match = re.search(r"data:(.*?);base64", header)
    mime_type = mime_match.group(1).lower() if mime_match else "image/png"
    extension = ".png"
    if "jpeg" in mime_type or "jpg" in mime_type:
        extension = ".jpg"
    elif "webp" in mime_type:
        extension = ".webp"

    binary_content = base64.b64decode(encoded)
    return _save_public_credit_upload(binary_content, subdirectory, f"{fallback_name}{extension}")


def _get_public_credit_smtp_settings(db: Session, company: Optional[models.Company] = None) -> Dict[str, Any]:
    integration_settings = None
    if company and getattr(company, "id", None):
        integration_settings = db.query(models.IntegrationSettings).filter(
            models.IntegrationSettings.company_id == company.id
        ).first()

    company_smtp_has_values = bool(
        integration_settings
        and str(getattr(integration_settings, "smtp_host", "") or "").strip()
        and str(getattr(integration_settings, "smtp_from", "") or getattr(integration_settings, "smtp_username", "") or "").strip()
    )
    use_company_smtp = bool(
        integration_settings
        and (getattr(integration_settings, "smtp_enabled", False) or company_smtp_has_values)
    )
    host = ((getattr(integration_settings, "smtp_host", None) if use_company_smtp else None) or os.getenv("SMTP_HOST") or "").strip()
    port = int(
        (
            str(getattr(integration_settings, "smtp_port", "") if use_company_smtp else "").strip()
            or (os.getenv("SMTP_PORT") or "587").strip()
            or 587
        )
    )
    username = ((getattr(integration_settings, "smtp_username", None) if use_company_smtp else None) or os.getenv("SMTP_USERNAME") or "").strip()
    password = ((getattr(integration_settings, "smtp_password", None) if use_company_smtp else None) or os.getenv("SMTP_PASSWORD") or "").strip()
    sender = ((getattr(integration_settings, "smtp_from", None) if use_company_smtp else None) or os.getenv("SMTP_FROM") or username or "").strip()
    use_tls = (
        bool(getattr(integration_settings, "smtp_use_tls", True))
        if use_company_smtp
        else str(os.getenv("SMTP_USE_TLS") or "true").strip().lower() not in {"0", "false", "no"}
    )

    if not host or not sender:
        raise HTTPException(
            status_code=500,
            detail="El servidor no tiene configurado el envío de correos SMTP para validar el formulario de crédito.",
        )

    return {
        "host": host,
        "port": port,
        "username": username,
        "password": password,
        "sender": sender,
        "use_tls": use_tls,
    }


def _build_public_credit_verification_email_html(
    company_name: str,
    verification_code: str,
    primary_color: Optional[str] = None,
    secondary_color: Optional[str] = None,
    logo_url: Optional[str] = None,
) -> str:
    safe_company_name = escape(company_name or "AutosQP")
    safe_code = escape(verification_code)
    brand_primary = (primary_color or "#2563eb").strip() or "#2563eb"
    brand_secondary = (secondary_color or "#0f172a").strip() or "#0f172a"
    logo_block = ""
    if logo_url:
        logo_block = (
            f'<div style="margin-bottom:24px;text-align:center;">'
            f'<img src="{escape(logo_url)}" alt="{safe_company_name}" '
            f'style="max-width:180px;max-height:64px;object-fit:contain;" />'
            f"</div>"
        )

    return f"""
    <html>
      <body style="margin:0;padding:0;background:#f8fafc;font-family:Arial,sans-serif;color:#0f172a;">
        <div style="max-width:640px;margin:0 auto;padding:32px 20px;">
          <div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:24px;overflow:hidden;box-shadow:0 12px 30px rgba(15,23,42,0.08);">
            <div style="padding:28px;background:linear-gradient(135deg,{brand_secondary}, {brand_primary});color:#ffffff;">
              <div style="font-size:12px;letter-spacing:0.16em;text-transform:uppercase;opacity:0.8;">Validación de correo</div>
              <div style="margin-top:10px;font-size:28px;font-weight:800;">Código de verificación</div>
              <div style="margin-top:8px;font-size:14px;opacity:0.9;">Usa este código para continuar tu solicitud de crédito.</div>
            </div>
            <div style="padding:32px;">
              {logo_block}
              <div style="font-size:16px;line-height:1.6;margin-bottom:20px;">
                Recibimos una solicitud de verificación para continuar el formulario público de crédito de <strong>{safe_company_name}</strong>.
              </div>
              <div style="margin:24px 0;padding:22px;border:1px dashed {brand_primary};border-radius:18px;background:#f8fbff;text-align:center;">
                <div style="font-size:12px;letter-spacing:0.14em;text-transform:uppercase;color:#64748b;margin-bottom:10px;">Tu código</div>
                <div style="font-size:38px;font-weight:800;letter-spacing:0.3em;color:{brand_secondary};">{safe_code}</div>
              </div>
              <div style="font-size:14px;line-height:1.7;color:#475569;">
                Este código vence en <strong>10 minutos</strong>.<br/>
                Si no solicitaste este proceso, puedes ignorar este correo.
              </div>
            </div>
          </div>
        </div>
      </body>
    </html>
    """.strip()


def _send_public_credit_verification_email(target_email: str, db: Session, company: Optional[models.Company], verification_code: str):
    smtp_settings = _get_public_credit_smtp_settings(db, company)
    company_name = getattr(company, "name", None) or "AutosQP"

    message = EmailMessage()
    message["Subject"] = f"Código de verificación - {company_name}"
    message["From"] = smtp_settings["sender"]
    message["To"] = target_email
    message.set_content(
        (
            f"Hola.\n\n"
            f"Tu código de verificación para continuar con la solicitud de crédito es: {verification_code}\n\n"
            f"Este código vence en 10 minutos.\n"
            f"Si no solicitaste este proceso, puedes ignorar este correo."
        ),
        charset="utf-8",
    )
    message.add_alternative(
        _build_public_credit_verification_email_html(
            company_name=company_name,
            verification_code=verification_code,
            primary_color=getattr(company, "primary_color", None),
            secondary_color=getattr(company, "secondary_color", None),
            logo_url=getattr(company, "logo_url", None),
        ),
        subtype="html",
    )

    _deliver_smtp_message(message, smtp_settings)


def _deliver_smtp_message(message: EmailMessage, smtp_settings: Dict[str, Any]):
    """Send an already-built message using the company's SMTP connection."""

    if smtp_settings["use_tls"]:
        context = ssl.create_default_context()
        with smtplib.SMTP(smtp_settings["host"], smtp_settings["port"]) as server:
            server.starttls(context=context)
            if smtp_settings["username"]:
                server.login(smtp_settings["username"], smtp_settings["password"])
            server.send_message(message)
    else:
        with smtplib.SMTP_SSL(smtp_settings["host"], smtp_settings["port"]) as server:
            if smtp_settings["username"]:
                server.login(smtp_settings["username"], smtp_settings["password"])
            server.send_message(message)


PUBLIC_CREDIT_SECTION_LABELS = {
    "vehicle": "Datos del vehículo",
    "personal": "Datos personales",
    "employment": "Datos laborales",
    "income": "Ingresos",
    "references": "Referencias",
    "consent": "Consentimiento y firma",
}

PUBLIC_CREDIT_FIELD_LABELS = {
    "vehicleValue": "Valor del vehículo",
    "requestedAmount": "Monto solicitado",
    "requestDate": "Fecha de solicitud",
    "make": "Marca",
    "model": "Modelo",
    "vehicleType": "Tipo de vehículo",
    "label": "Vehículo solicitado",
    "firstName": "Nombres",
    "lastName": "Apellidos",
    "documentType": "Tipo de documento",
    "documentNumber": "Número de documento",
    "issuePlace": "Lugar de expedición",
    "birthDate": "Fecha de nacimiento",
    "gender": "Sexo",
    "profession": "Profesión",
    "birthPlace": "Lugar de nacimiento",
    "maritalStatus": "Estado civil",
    "childrenCount": "Número de hijos",
    "educationLevel": "Nivel de estudio",
    "livesWith": "Con quién vive",
    "housingType": "Tipo de vivienda",
    "mobilePhone": "Teléfono móvil",
    "city": "Ciudad",
    "address": "Dirección",
    "email": "Correo electrónico",
    "activity": "Actividad económica",
    "companyName": "Empresa actual",
    "companyCity": "Ciudad de la empresa",
    "companyAddress": "Dirección de la empresa",
    "jobTitle": "Cargo u ocupación",
    "companyEmail": "Correo de la empresa",
    "startDate": "Fecha de ingreso",
    "salary": "Salario",
    "contractType": "Tipo de contrato",
    "previousCompanyName": "Empresa anterior",
    "previousCompanyActivity": "Actividad de la empresa anterior",
    "previousCompanyRole": "Cargo anterior",
    "previousEmploymentTime": "Tiempo laborado",
    "salaryIncome": "Ingreso por salario",
    "commissionsIncome": "Ingreso por comisiones",
    "otherIncome": "Otros ingresos",
    "otherIncomeDetail": "Detalle de otros ingresos",
    "totalIncome": "Total de ingresos",
    "commercial": "Referencia comercial",
    "personal1": "Primera referencia personal",
    "personal2": "Segunda referencia personal",
    "names": "Nombres",
    "lastNames": "Apellidos",
    "phone": "Teléfono",
    "accepted": "Política de tratamiento aceptada",
    "signatureMode": "Modalidad de firma",
    "signatureName": "Nombre de quien firma",
}

PUBLIC_CREDIT_HIDDEN_FIELDS = {"verificationCode", "signatureDrawnDataUrl", "summary"}
PUBLIC_CREDIT_MONEY_FIELDS = {
    "vehicleValue", "requestedAmount", "salary", "salaryIncome",
    "commissionsIncome", "otherIncome", "totalIncome",
}


def _public_credit_field_label(field_name: str) -> str:
    if field_name in PUBLIC_CREDIT_FIELD_LABELS:
        return PUBLIC_CREDIT_FIELD_LABELS[field_name]
    return re.sub(r"(?<=[a-z])(?=[A-Z])", " ", str(field_name)).strip().capitalize()


def _public_credit_display_value(field_name: str, value: Any) -> str:
    if value is None or value == "":
        return "Sin dato"
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if field_name in PUBLIC_CREDIT_MONEY_FIELDS:
        try:
            amount = int(float(re.sub(r"[^\d.-]", "", str(value)) or 0))
            return f"$ {amount:,}".replace(",", ".")
        except (TypeError, ValueError):
            pass
    if field_name == "signatureMode":
        return "Firma dibujada" if str(value) == "draw" else "Firma adjunta"
    if field_name in {"requestDate", "birthDate", "startDate"} and re.match(r"^\d{4}-\d{2}-\d{2}$", str(value)):
        year, month, day = str(value).split("-")
        return f"{day}/{month}/{year}"
    return str(value)


def _public_credit_attachment_path(file_url: Optional[str]) -> Optional[str]:
    value = str(file_url or "").strip()
    marker = "/api/static/"
    if not value or marker not in value:
        return None
    relative_path = value.split(marker, 1)[1].replace("/", os.sep)
    local_path = os.path.abspath(os.path.join("static", relative_path))
    static_root = os.path.abspath("static")
    if os.path.commonpath([local_path, static_root]) != static_root or not os.path.isfile(local_path):
        return None
    return local_path


def _public_credit_bogota_datetime(value: Optional[datetime.datetime]) -> Optional[datetime.datetime]:
    if not value:
        return None
    aware_value = value if value.tzinfo else value.replace(tzinfo=datetime.timezone.utc)
    return aware_value.astimezone(BOGOTA_TZ)


def _build_public_credit_submission_pdf_three_pages(
    company: Optional[models.Company],
    submission: models.PublicCreditSubmission,
) -> bytes:
    try:
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="La libreria reportlab no esta instalada en el servidor") from exc

    payload = getattr(submission, "form_payload", None) or getattr(submission, "form_data", None) or {}
    attachments = getattr(submission, "attachments", None) or getattr(submission, "files", None) or {}
    company_name = getattr(company, "name", None) or "AutosQP"
    primary_color = getattr(company, "primary_color", None) or "#2563eb"
    secondary_color = getattr(company, "secondary_color", None) or "#0f172a"
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    pdf.setTitle(f"Formulario de credito - {submission.applicant_name}")
    pdf.setAuthor(company_name)

    page_width, page_height = letter
    left_margin = 36
    right_margin = 36
    content_width = page_width - left_margin - right_margin
    label_width = 188
    row_height = 14.2
    section_gap = 9
    primary = HexColor(primary_color)
    secondary = HexColor(secondary_color)
    border = HexColor("#cbd5e1")
    label_fill = HexColor("#f1f5f9")
    text_color = HexColor("#0f172a")

    def safe_text(value: Any) -> str:
        text_value = str(value if value is not None and value != "" else "Sin dato")
        return " ".join(text_value.replace("\n", " ").replace("\r", " ").split())

    def payload_section(section_name: str) -> Dict[str, Any]:
        section_value = payload.get(section_name)
        return section_value if isinstance(section_value, dict) else {}

    vehicle = payload_section("vehicle")
    personal = payload_section("personal")
    employment = payload_section("employment")
    income = payload_section("income")
    references = payload_section("references")
    consent_payload = payload_section("consent")

    def field(section_value: Dict[str, Any], field_name: str) -> str:
        return _public_credit_display_value(field_name, section_value.get(field_name))

    def attachment_filename(key: str) -> str:
        attachment_value = attachments.get(key) if isinstance(attachments, dict) else None
        if isinstance(attachment_value, dict):
            attachment_value = attachment_value.get("original_filename") or attachment_value.get("filename") or attachment_value.get("path")
        value = str(attachment_value or "").strip()
        if not value:
            return "Sin dato"
        return os.path.basename(value.split("?", 1)[0]) or value

    def wrap_text(text: str, max_width: float, font_name: str = "Helvetica", font_size: float = 7) -> List[str]:
        words = safe_text(text).split()
        lines = []
        current = ""
        for word in words:
            candidate = word if not current else f"{current} {word}"
            if pdf.stringWidth(candidate, font_name, font_size) <= max_width:
                current = candidate
                continue
            if current:
                lines.append(current)
            current = word
        if current:
            lines.append(current)
        return lines or ["Sin dato"]

    def clipped_text(text: str, max_width: float, font_name: str = "Helvetica", font_size: float = 7) -> str:
        value = safe_text(text)
        if pdf.stringWidth(value, font_name, font_size) <= max_width:
            return value
        ellipsis = "..."
        while value and pdf.stringWidth(f"{value}{ellipsis}", font_name, font_size) > max_width:
            value = value[:-1]
        return f"{value}{ellipsis}" if value else ellipsis

    def draw_header(page_number: int):
        pdf.setStrokeColor(border)
        pdf.setLineWidth(0.7)
        pdf.rect(left_margin, page_height - 78, content_width, 48, stroke=1, fill=0)
        pdf.setFillColor(secondary)
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawCentredString(page_width / 2, page_height - 46, safe_text(company_name).upper())
        contact_parts = [
            getattr(company, "contact_address", None),
            f"Tel: {getattr(company, 'contact_phone', None)}" if getattr(company, "contact_phone", None) else None,
            getattr(company, "website_url", None) or getattr(company, "public_domain", None),
        ]
        contact_line = "   ".join(safe_text(part) for part in contact_parts if part)
        if contact_line:
            pdf.setFont("Helvetica", 6.4)
            pdf.setFillColor(HexColor("#475569"))
            pdf.drawCentredString(page_width / 2, page_height - 61, clipped_text(contact_line, content_width - 20, font_size=6.4))
        pdf.setFillColor(primary)
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawCentredString(page_width / 2, page_height - 92, "Formulario de Credito")
        pdf.setFillColor(HexColor("#64748b"))
        pdf.setFont("Helvetica", 6.2)
        pdf.drawRightString(page_width - right_margin, 22, f"Pagina {page_number} de 3")

    def draw_section_title(title: str, y: float) -> float:
        pdf.setFillColor(primary)
        pdf.setFont("Helvetica-Bold", 8.5)
        pdf.drawString(left_margin, y, title)
        return y - section_gap

    def draw_rows(rows: List[tuple[str, str]], y: float) -> float:
        value_width = content_width - label_width
        for label, value in rows:
            pdf.setFillColor(label_fill)
            pdf.rect(left_margin, y - row_height + 2, label_width, row_height, fill=1, stroke=0)
            pdf.setStrokeColor(border)
            pdf.rect(left_margin, y - row_height + 2, label_width, row_height, stroke=1, fill=0)
            pdf.rect(left_margin + label_width, y - row_height + 2, value_width, row_height, stroke=1, fill=0)
            pdf.setFillColor(text_color)
            pdf.setFont("Helvetica-Bold", 6.4)
            pdf.drawString(left_margin + 4, y - 8.5, clipped_text(label.upper(), label_width - 8, "Helvetica-Bold", 6.4))
            pdf.setFont("Helvetica", 6.6)
            pdf.drawString(left_margin + label_width + 4, y - 8.5, clipped_text(value, value_width - 8, "Helvetica", 6.6))
            y -= row_height
        return y - 3

    def draw_policy_lines(lines: List[str], y: float, max_lines: int) -> tuple[float, List[str]]:
        pdf.setFillColor(text_color)
        pdf.setFont("Helvetica", 5.8)
        line_height = 7.2
        drawn = 0
        while lines and drawn < max_lines and y > 46:
            pdf.drawString(left_margin, y, lines.pop(0))
            y -= line_height
            drawn += 1
        return y, lines

    def reference_rows(data: Dict[str, Any]) -> List[tuple[str, str]]:
        return [
            ("Nombres", field(data, "names")),
            ("Apellidos", field(data, "lastNames")),
            ("Telefono", field(data, "phone")),
            ("Ciudad", field(data, "city")),
        ]

    pdf.setPageCompression(1)
    draw_header(1)
    y = page_height - 112
    y = draw_section_title("Datos del Vehiculo", y)
    y = draw_rows([
        ("Valor vehiculo $", field(vehicle, "vehicleValue")),
        ("Monto solicitado $", field(vehicle, "requestedAmount")),
        ("Asesor comercial", safe_text(vehicle.get("advisor") or vehicle.get("commercialAdvisor"))),
        ("Fecha de solicitud", field(vehicle, "requestDate")),
        ("Marca", field(vehicle, "make")),
        ("Modelo", field(vehicle, "model")),
        ("Tipo", field(vehicle, "vehicleType")),
    ], y)
    y = draw_section_title("Datos Personales", y)
    y = draw_rows([
        ("Nombres", field(personal, "firstName")),
        ("Apellidos", field(personal, "lastName")),
        ("Documento", field(personal, "documentType")),
        ("N. Documento", field(personal, "documentNumber")),
        ("Lugar expedicion", field(personal, "issuePlace")),
        ("Fecha nacimiento", field(personal, "birthDate")),
        ("Sexo", field(personal, "gender")),
        ("Profesion", field(personal, "profession")),
        ("Lugar nacimiento", field(personal, "birthPlace")),
        ("Estado civil", field(personal, "maritalStatus")),
        ("N. de hijos", field(personal, "childrenCount")),
        ("Nivel de estudio", field(personal, "educationLevel")),
        ("Con quien vive", field(personal, "livesWith")),
        ("Tipo de vivienda", field(personal, "housingType")),
        ("Telefono movil", field(personal, "mobilePhone")),
        ("Ciudad", field(personal, "city")),
        ("Direccion", field(personal, "address")),
        ("Email", field(personal, "email")),
        ("Cedula ciudadania cara frontal", attachment_filename("document_front")),
        ("Cedula ciudadania cara posterior", attachment_filename("document_back")),
    ], y)
    y = draw_section_title("Datos Laborales", y)
    draw_rows([
        ("Actividad economica", field(employment, "activity")),
        ("Nombre empresa", field(employment, "companyName")),
        ("Ciudad", field(employment, "companyCity")),
        ("Direccion", field(employment, "companyAddress")),
        ("Ocupacion o cargo", field(employment, "jobTitle")),
        ("Email de la empresa", field(employment, "companyEmail")),
        ("Fecha ingreso", field(employment, "startDate")),
    ], y)
    pdf.showPage()

    draw_header(2)
    y = page_height - 112
    y = draw_rows([
        ("Salario", field(employment, "salary")),
        ("Tipo de contrato", field(employment, "contractType")),
    ], y)
    y = draw_section_title("Empresa Anterior", y)
    y = draw_rows([
        ("Nombre empresa", field(employment, "previousCompanyName")),
        ("Actividad empresa", field(employment, "previousCompanyActivity")),
        ("Cargo", field(employment, "previousCompanyRole")),
        ("Tiempo laborado", field(employment, "previousEmploymentTime")),
    ], y)
    y = draw_section_title("Ingresos", y)
    y = draw_rows([
        ("Sueldo $", field(income, "salaryIncome")),
        ("Comisiones $", field(income, "commissionsIncome")),
        ("Otros ingresos permanentes", field(income, "otherIncome")),
        ("Total ingresos", field(income, "totalIncome")),
    ], y)
    personal1 = references.get("personal1") if isinstance(references.get("personal1"), dict) else {}
    personal2 = references.get("personal2") if isinstance(references.get("personal2"), dict) else {}
    commercial = references.get("commercial") if isinstance(references.get("commercial"), dict) else {}
    y = draw_section_title("Referencias Personales", y)
    y = draw_rows(reference_rows(personal1), y)
    y = draw_section_title("Referencias Personales", y)
    y = draw_rows(reference_rows(personal2), y)
    if any(commercial.values()):
        y = draw_section_title("Referencia Comercial", y)
        y = draw_rows(reference_rows(commercial), y)
    y = draw_section_title("Consentimiento", y)
    y = draw_rows([
        ("Consentimiento", "Estoy de acuerdo con la politica de privacidad." if consent_payload.get("accepted") else "Sin aceptar"),
    ], y)
    policy_lines = wrap_text((submission.consent_text or PUBLIC_CREDIT_POLICY_TEXT).upper(), content_width, "Helvetica", 5.8)
    page_two_policy_lines = min(12, max(1, len(policy_lines) // 2))
    y, policy_lines = draw_policy_lines(policy_lines, y, page_two_policy_lines)
    pdf.showPage()

    draw_header(3)
    y = page_height - 112
    y, policy_lines = draw_policy_lines(policy_lines, y, 76)
    if policy_lines:
        pdf.setFont("Helvetica", 5.8)
        pdf.setFillColor(text_color)
        pdf.drawString(left_margin, y, clipped_text(" ".join(policy_lines), content_width, "Helvetica", 5.8))
    signer_name = safe_text(consent_payload.get("signatureName") or submission.applicant_name)
    signature_y = 126
    pdf.setStrokeColor(border)
    pdf.line(left_margin, signature_y, left_margin + 230, signature_y)
    pdf.setFillColor(text_color)
    pdf.setFont("Helvetica-Bold", 7.5)
    pdf.drawString(left_margin, signature_y - 14, f"Firmado por: {signer_name}")
    pdf.setFont("Helvetica", 6.5)
    pdf.drawString(left_margin, signature_y - 26, f"Modalidad: {_public_credit_display_value('signatureMode', consent_payload.get('signatureMode'))}")
    signature_file = attachment_filename("signature_upload")
    signature_drawn = attachment_filename("signature_drawn")
    if signature_file != "Sin dato" or signature_drawn != "Sin dato":
        pdf.drawString(left_margin, signature_y - 38, f"Soporte de firma: {signature_file if signature_file != 'Sin dato' else signature_drawn}")
    pdf.showPage()
    pdf.save()
    return buffer.getvalue()


def _build_public_credit_submission_pdf_reference_style(
    company: Optional[models.Company],
    submission: models.PublicCreditSubmission,
) -> bytes:
    try:
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="La libreria reportlab no esta instalada en el servidor") from exc

    payload = getattr(submission, "form_payload", None) or getattr(submission, "form_data", None) or {}
    attachments = getattr(submission, "attachments", None) or getattr(submission, "files", None) or {}
    raw_company_name = getattr(company, "name", None) or "AUTOS QP S.A.S"
    company_name = re.sub(r"\s+admin$", "", str(raw_company_name).strip(), flags=re.IGNORECASE) or str(raw_company_name).strip()
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    pdf.setTitle(f"Formulario de credito - {submission.applicant_name}")
    pdf.setAuthor(company_name)

    page_width, page_height = letter
    left = 40
    right = 40
    usable_width = page_width - left - right
    gap = 8
    col_w = (usable_width - (gap * 3)) / 4
    border = HexColor("#d1d5db")
    text_color = HexColor("#111827")
    muted = HexColor("#4b5563")

    def safe_text(value: Any) -> str:
        text_value = str(value if value is not None and value != "" else "Sin dato")
        return " ".join(text_value.replace("\n", " ").replace("\r", " ").split())

    def section(section_name: str) -> Dict[str, Any]:
        value = payload.get(section_name)
        return value if isinstance(value, dict) else {}

    vehicle = section("vehicle")
    personal = section("personal")
    employment = section("employment")
    income = section("income")
    references = section("references")
    consent_payload = section("consent")

    def field(section_value: Dict[str, Any], field_name: str) -> str:
        return _public_credit_display_value(field_name, section_value.get(field_name))

    def attachment_entry(key: str) -> Any:
        return attachments.get(key) if isinstance(attachments, dict) else None

    def attachment_filename(key: str) -> str:
        entry = attachment_entry(key)
        if isinstance(entry, dict):
            entry = entry.get("original_filename") or entry.get("filename") or entry.get("file_name") or entry.get("path") or entry.get("file_path") or entry.get("url")
        value = str(entry or "").strip()
        if not value:
            return "Sin dato"
        return os.path.basename(value.split("?", 1)[0]) or value

    def attachment_local_path(key: str) -> Optional[str]:
        entry = attachment_entry(key)
        if isinstance(entry, dict):
            entry = entry.get("path") or entry.get("file_path") or entry.get("url") or entry.get("original_filename") or entry.get("filename")
        value = str(entry or "").strip()
        if not value:
            return None
        clean_value = value.split("?", 1)[0].replace("\\", "/")
        candidates = []
        if os.path.isabs(clean_value):
            candidates.append(clean_value)
        for prefix in ("/api/static/", "/static/", "api/static/", "static/"):
            normalized_prefix = prefix.lstrip("/")
            if clean_value.startswith(prefix) or clean_value.startswith(normalized_prefix):
                relative = clean_value[len(prefix):] if clean_value.startswith(prefix) else clean_value[len(normalized_prefix):]
                candidates.append(os.path.join("static", *relative.split("/")))
        candidates.append(clean_value.lstrip("/"))
        for candidate in candidates:
            if candidate and os.path.exists(candidate):
                return candidate
        return None

    def policy_text() -> str:
        value = submission.consent_text or PUBLIC_CREDIT_POLICY_TEXT
        display_name = safe_text(company_name).upper()
        value = re.sub(r"\bAUTOS\s+QP\s+S\.?A\.?S\.?\b", display_name, value, flags=re.IGNORECASE)
        value = re.sub(r"\bAUTOSQP\s+ADMIN\b", display_name, value, flags=re.IGNORECASE)
        value = re.sub(r"\bAUTOSQP\b", display_name, value, flags=re.IGNORECASE)
        return value

    def clipped_text(text: str, max_width: float, font_name: str = "Helvetica", font_size: float = 8.5) -> str:
        value = safe_text(text)
        if pdf.stringWidth(value, font_name, font_size) <= max_width:
            return value
        ellipsis = "..."
        while value and pdf.stringWidth(f"{value}{ellipsis}", font_name, font_size) > max_width:
            value = value[:-1]
        return f"{value}{ellipsis}" if value else ellipsis

    def wrap_text(text: str, max_width: float, font_name: str = "Helvetica", font_size: float = 6.2) -> List[str]:
        words = safe_text(text).split()
        lines = []
        current = ""
        for word in words:
            candidate = word if not current else f"{current} {word}"
            if pdf.stringWidth(candidate, font_name, font_size) <= max_width:
                current = candidate
                continue
            if current:
                lines.append(current)
            current = word
        if current:
            lines.append(current)
        return lines or [""]

    def draw_brand():
        pdf.setFillColor(text_color)
        pdf.setStrokeColor(text_color)
        pdf.setLineWidth(1)
        pdf.setFont("Helvetica", 22)
        pdf.drawString(left + 86, page_height - 58, "Autos")
        pdf.setFont("Helvetica-Bold", 33)
        pdf.drawString(left + 88, page_height - 92, "QP")
        pdf.line(left + 76, page_height - 100, left + 76, page_height - 38)
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawCentredString(left + 38, page_height - 54, "AUTO")
        pdf.drawCentredString(left + 38, page_height - 74, "QP")
        pdf.setFont("Helvetica", 8)
        pdf.drawCentredString(left + 38, page_height - 90, "credito")

    def draw_footer():
        contact_address = safe_text(getattr(company, "contact_address", None) or "Centro Comercial Outlet Factory, Av. de las Americas #62-84 Local 128, Puente Aranda, Bogota")
        contact_phone = safe_text(getattr(company, "contact_phone", None) or "3002523226 - 3218912903")
        contact_email = safe_text(getattr(company, "contact_email", None) or "autosqpc@gmail.com")
        website = safe_text(getattr(company, "website_url", None) or getattr(company, "public_domain", None) or "www.autosqp.com")
        pdf.setFillColor(text_color)
        pdf.setFont("Helvetica-Bold", 9)
        pdf.drawCentredString(page_width / 2, 60, safe_text(company_name).upper())
        pdf.setFont("Helvetica", 8.2)
        pdf.drawCentredString(page_width / 2, 35, clipped_text(f"{contact_address}        Tel: {contact_phone}   {contact_email}   {website}", usable_width, "Helvetica", 8.2))

    def start_page(page_number: int, with_brand: bool = True):
        if with_brand:
            draw_brand()

    def finish_page():
        draw_footer()
        pdf.showPage()

    def draw_title(title: str, y: float) -> float:
        pdf.setFillColor(text_color)
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawString(left, y, title)
        return y - 28

    def draw_section(title: str, y: float) -> float:
        pdf.setFillColor(text_color)
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(left, y, title)
        return y - 25

    def draw_field(label: str, value: str, x: float, y: float, width: float, height: float = 23) -> None:
        pdf.setFillColor(text_color)
        pdf.setFont("Helvetica-Bold", 8.5)
        pdf.drawString(x, y, clipped_text(safe_text(label).upper(), width, "Helvetica-Bold", 8.5))
        pdf.setStrokeColor(border)
        pdf.setLineWidth(0.8)
        pdf.rect(x, y - height - 4, width, height, stroke=1, fill=0)
        pdf.setFillColor(text_color)
        pdf.setFont("Helvetica", 9)
        pdf.drawString(x + 7, y - height + 3, clipped_text(value, width - 14, "Helvetica", 9))

    def draw_image_in_box(path: Optional[str], x: float, y: float, width: float, height: float, label: str) -> None:
        pdf.setFillColor(text_color)
        pdf.setFont("Helvetica-Bold", 8.5)
        pdf.drawString(x, y + height + 8, clipped_text(label.upper(), width, "Helvetica-Bold", 8.5))
        pdf.setStrokeColor(border)
        pdf.setLineWidth(0.8)
        pdf.rect(x, y, width, height, stroke=1, fill=0)
        if not path:
            pdf.setFillColor(muted)
            pdf.setFont("Helvetica", 8)
            pdf.drawCentredString(x + width / 2, y + height / 2 - 3, "Sin imagen adjunta")
            return
        try:
            image = ImageReader(path)
            image_width, image_height = image.getSize()
            ratio = min((width - 10) / image_width, (height - 10) / image_height)
            draw_width = image_width * ratio
            draw_height = image_height * ratio
            pdf.drawImage(
                image,
                x + (width - draw_width) / 2,
                y + (height - draw_height) / 2,
                width=draw_width,
                height=draw_height,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pdf.setFillColor(muted)
            pdf.setFont("Helvetica", 8)
            pdf.drawCentredString(x + width / 2, y + height / 2 - 3, clipped_text(os.path.basename(path), width - 14, "Helvetica", 8))

    def draw_signature_image(path: Optional[str], x: float, y: float, width: float, height: float) -> None:
        if not path:
            return
        try:
            image = ImageReader(path)
            image_width, image_height = image.getSize()
            ratio = min(width / image_width, height / image_height)
            draw_width = image_width * ratio
            draw_height = image_height * ratio
            pdf.drawImage(
                image,
                x,
                y + (height - draw_height) / 2,
                width=draw_width,
                height=draw_height,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            return

    def row_fields(y: float, fields: List[tuple[str, str, int]]) -> float:
        x = left
        for label, value, span in fields:
            width = col_w * span + gap * (span - 1)
            draw_field(label, value, x, y, width)
            x += width + gap
        return y - 45

    def draw_policy(lines: List[str], y: float, max_lines: int, min_y: float = 88) -> tuple[float, List[str]]:
        pdf.setFillColor(text_color)
        pdf.setFont("Helvetica", 6.2)
        line_height = 7.1
        drawn = 0
        while lines and drawn < max_lines and y > min_y:
            pdf.drawString(left, y, lines.pop(0))
            y -= line_height
            drawn += 1
        return y, lines

    def reference_values(data: Dict[str, Any]) -> List[tuple[str, str, int]]:
        return [
            ("Nombres", field(data, "names"), 1),
            ("Apellidos", field(data, "lastNames"), 1),
            ("Telefono", field(data, "phone"), 1),
            ("Ciudad", field(data, "city"), 1),
        ]

    start_page(1)
    y = page_height - 158
    y = draw_title("Formulario de Credito", y)
    y = draw_section("Datos del Vehiculo", y)
    y = row_fields(y, [
        ("Valor vehiculo $", field(vehicle, "vehicleValue"), 1),
        ("Monto solicitado $", field(vehicle, "requestedAmount"), 1),
        ("Asesor comercial", safe_text(vehicle.get("advisor") or vehicle.get("commercialAdvisor")), 1),
        ("Fecha de solicitud", field(vehicle, "requestDate"), 1),
    ])
    y = row_fields(y, [
        ("Marca", field(vehicle, "make"), 1),
        ("Modelo", field(vehicle, "model"), 2),
        ("Tipo", field(vehicle, "vehicleType"), 1),
    ])
    y += 4
    y = draw_section("Datos Personales", y)
    personal_rows = [
        [("Nombres", field(personal, "firstName"), 1), ("Apellidos", field(personal, "lastName"), 1), ("Documento", field(personal, "documentType"), 1), ("N. Documento", field(personal, "documentNumber"), 1)],
        [("Lugar expedicion", field(personal, "issuePlace"), 1), ("Fecha nacimiento", field(personal, "birthDate"), 1), ("Sexo", field(personal, "gender"), 1), ("Profesion", field(personal, "profession"), 1)],
        [("Lugar nacimiento", field(personal, "birthPlace"), 1), ("Estado civil", field(personal, "maritalStatus"), 1), ("N. de hijos", field(personal, "childrenCount"), 1), ("Nivel de estudio", field(personal, "educationLevel"), 1)],
        [("Con quien vive ?", field(personal, "livesWith"), 1), ("Tipo de vivienda ?", field(personal, "housingType"), 1), ("Telefono movil", field(personal, "mobilePhone"), 1), ("Ciudad", field(personal, "city"), 1)],
        [("Direccion", field(personal, "address"), 2), ("Email", field(personal, "email"), 2)],
        [("Cedula ciudadania cara frontal", attachment_filename("document_front"), 2), ("Cedula ciudadania cara posterior", attachment_filename("document_back"), 2)],
    ]
    for items in personal_rows:
        y = row_fields(y, items)
    y += 6
    y = draw_section("Datos Laborales", y)
    y = row_fields(y, [
        ("Actividad economica", field(employment, "activity"), 1),
        ("Nombre empresa", field(employment, "companyName"), 1),
        ("Ciudad", field(employment, "companyCity"), 1),
        ("Direccion", field(employment, "companyAddress"), 1),
    ])
    row_fields(y, [
        ("Ocupacion o cargo", field(employment, "jobTitle"), 1),
        ("Email de la empresa", field(employment, "companyEmail"), 2),
        ("Fecha ingreso", field(employment, "startDate"), 1),
    ])
    finish_page()

    start_page(2)
    y = page_height - 164
    y = row_fields(y, [
        ("Salario", field(employment, "salary"), 1),
        ("Tipo de contrato", field(employment, "contractType"), 1),
    ])
    y += 6
    y = draw_section("Empresa Anterior", y)
    y = row_fields(y, [
        ("Nombre empresa", field(employment, "previousCompanyName"), 1),
        ("Actividad empresa", field(employment, "previousCompanyActivity"), 1),
        ("Cargo", field(employment, "previousCompanyRole"), 1),
        ("Tiempo laborado", field(employment, "previousEmploymentTime"), 1),
    ])
    y += 6
    y = draw_section("Ingresos", y)
    y = row_fields(y, [
        ("Sueldo $", field(income, "salaryIncome"), 1),
        ("Comisiones $", field(income, "commissionsIncome"), 1),
        ("Otros ingresos permanentes", field(income, "otherIncome"), 1),
        ("Total ingresos", field(income, "totalIncome"), 1),
    ])
    personal1 = references.get("personal1") if isinstance(references.get("personal1"), dict) else {}
    personal2 = references.get("personal2") if isinstance(references.get("personal2"), dict) else {}
    commercial = references.get("commercial") if isinstance(references.get("commercial"), dict) else {}
    for title, data in [("Referencias Personales", personal1), ("Referencias Personales", personal2)]:
        y += 4
        y = draw_section(title, y)
        y = row_fields(y, reference_values(data))
    if any(commercial.values()):
        y += 4
        y = draw_section("Referencia Comercial", y)
        y = row_fields(y, reference_values(commercial))
    y += 5
    draw_field("Consentimiento", "Estoy de acuerdo con la politica de privacidad." if consent_payload.get("accepted") else "Sin aceptar", left, y, usable_width)
    y -= 43
    policy_lines = wrap_text(policy_text().upper(), usable_width, "Helvetica", 6.2)
    y, policy_lines = draw_policy(policy_lines, y, 21)
    finish_page()

    start_page(3)
    y = page_height - 164
    document_y = 310
    document_height = 132
    y, policy_lines = draw_policy(policy_lines, y, 78, document_y + document_height + 44)
    if policy_lines:
        pdf.drawString(left, y, clipped_text(" ".join(policy_lines), usable_width, "Helvetica", 6.2))
    document_width = (usable_width - gap) / 2
    draw_image_in_box(attachment_local_path("document_front"), left, document_y, document_width, document_height, "Cedula ciudadania cara frontal")
    draw_image_in_box(attachment_local_path("document_back"), left + document_width + gap, document_y, document_width, document_height, "Cedula ciudadania cara posterior")
    signer_name = safe_text(consent_payload.get("signatureName") or submission.applicant_name)
    signature_file = attachment_filename("signature_upload")
    signature_drawn = attachment_filename("signature_drawn")
    signature_path = attachment_local_path("signature_upload") or attachment_local_path("signature_drawn")
    draw_signature_image(signature_path, left, 146, 240, 72)
    pdf.setStrokeColor(border)
    pdf.line(left, 138, left + 240, 138)
    pdf.setFillColor(text_color)
    pdf.setFont("Helvetica-Bold", 8.5)
    pdf.drawString(left, 122, f"Firmado por: {signer_name}")
    pdf.setFont("Helvetica", 7.2)
    pdf.drawString(left, 108, f"Modalidad: {_public_credit_display_value('signatureMode', consent_payload.get('signatureMode'))}")
    if signature_file != "Sin dato" or signature_drawn != "Sin dato":
        pdf.drawString(left, 94, f"Soporte de firma: {signature_file if signature_file != 'Sin dato' else signature_drawn}")
    finish_page()

    pdf.save()
    return buffer.getvalue()


def _build_public_credit_submission_pdf(
    company: Optional[models.Company],
    submission: models.PublicCreditSubmission,
) -> bytes:
    return _build_public_credit_submission_pdf_reference_style(company, submission)


def _parse_public_credit_recipients(raw_recipients: Optional[str], applicant_email: str) -> List[str]:
    candidates = [applicant_email]
    candidates.extend(re.split(r"[,;\n\r]+", str(raw_recipients or "")))
    result = []
    seen = set()
    for candidate in candidates:
        email_value = str(candidate or "").strip().lower()
        if email_value and "@" in email_value and email_value not in seen:
            seen.add(email_value)
            result.append(email_value)
    return result


def _send_public_credit_submission_email(
    db: Session,
    company: Optional[models.Company],
    submission: models.PublicCreditSubmission,
    pdf_bytes: bytes,
):
    smtp_settings = _get_public_credit_smtp_settings(db, company)
    integration_settings = db.query(models.IntegrationSettings).filter(
        models.IntegrationSettings.company_id == submission.company_id
    ).first()
    recipients = _parse_public_credit_recipients(
        getattr(integration_settings, "smtp_always_recipients", None),
        submission.email,
    )
    if not recipients:
        raise ValueError("No hay destinatarios configurados para el formulario de crédito")

    vehicle = (submission.form_payload or {}).get("vehicle", {}) or {}
    vehicle_value = _public_credit_display_value("vehicleValue", vehicle.get("vehicleValue"))
    requested_amount = _public_credit_display_value("requestedAmount", vehicle.get("requestedAmount"))
    company_name = getattr(company, "name", None) or "AutosQP"
    primary_color = getattr(company, "primary_color", None) or "#2563eb"
    secondary_color = getattr(company, "secondary_color", None) or "#0f172a"

    message = EmailMessage()
    message["Subject"] = f"Formulario de crédito - {submission.applicant_name}"
    message["From"] = smtp_settings["sender"]
    message["To"] = submission.email
    internal_recipients = [email for email in recipients if email != submission.email.lower()]
    if internal_recipients:
        message["Bcc"] = ", ".join(internal_recipients)
    message.set_content(
        "\n".join([
            "Formulario de crédito",
            f"Nombre: {submission.applicant_name}",
            f"Correo: {submission.email}",
            f"Valor del carro: {vehicle_value}",
            f"Valor del crédito solicitado: {requested_amount}",
            "",
            "El formulario completo se encuentra adjunto en PDF.",
        ]),
        charset="utf-8",
    )
    message.add_alternative(
        f"""
        <html><body style="margin:0;background:#f8fafc;font-family:Arial,sans-serif;color:#0f172a;">
          <div style="max-width:640px;margin:0 auto;padding:28px 18px;">
            <div style="overflow:hidden;border:1px solid #e2e8f0;border-radius:22px;background:#fff;">
              <div style="padding:26px;background:linear-gradient(135deg,{secondary_color},{primary_color});color:#fff;">
                <div style="font-size:13px;opacity:.85;">{escape(company_name)}</div>
                <h1 style="margin:8px 0 0;font-size:26px;">Formulario de crédito</h1>
              </div>
              <div style="padding:28px;">
                <p style="margin-top:0;color:#475569;">Se registró correctamente una solicitud de crédito.</p>
                <table style="width:100%;border-collapse:collapse;">
                  <tr><td style="padding:10px;border-bottom:1px solid #e2e8f0;font-weight:bold;">Nombre</td><td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(submission.applicant_name)}</td></tr>
                  <tr><td style="padding:10px;border-bottom:1px solid #e2e8f0;font-weight:bold;">Correo</td><td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(submission.email)}</td></tr>
                  <tr><td style="padding:10px;border-bottom:1px solid #e2e8f0;font-weight:bold;">Valor del carro</td><td style="padding:10px;border-bottom:1px solid #e2e8f0;">{escape(vehicle_value)}</td></tr>
                  <tr><td style="padding:10px;font-weight:bold;">Crédito solicitado</td><td style="padding:10px;">{escape(requested_amount)}</td></tr>
                </table>
                <p style="margin:24px 0 0;color:#475569;">El formulario completo está adjunto en formato PDF.</p>
              </div>
            </div>
          </div>
        </body></html>
        """.strip(),
        subtype="html",
    )
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", submission.applicant_name or "solicitante").strip("_")
    message.add_attachment(
        pdf_bytes,
        maintype="application",
        subtype="pdf",
        filename=f"formulario_credito_{safe_name or submission.id}.pdf",
    )
    _deliver_smtp_message(message, smtp_settings)


def _build_public_credit_lead_message(payload: Dict[str, Any]) -> str:
    vehicle = payload.get("vehicle", {}) or {}
    personal = payload.get("personal", {}) or {}
    employment = payload.get("employment", {}) or {}
    income = payload.get("income", {}) or {}
    references = payload.get("references", {}) or {}

    message_parts = ["Solicitud pública de crédito"]
    desired_vehicle = vehicle.get("label") or payload.get("desired_vehicle") or ""
    if desired_vehicle:
        message_parts.append(f"Vehículo: {desired_vehicle}")
    if vehicle.get("requestedAmount"):
        message_parts.append(f"Monto solicitado: ${int(vehicle['requestedAmount']):,}".replace(",", "."))
    if income.get("totalIncome"):
        message_parts.append(f"Ingresos totales: ${int(income['totalIncome']):,}".replace(",", "."))
    if personal.get("city"):
        message_parts.append(f"Ciudad: {personal['city']}")
    if employment.get("activity"):
        message_parts.append(f"Actividad: {employment['activity']}")
    if references.get("summary"):
        message_parts.append(str(references["summary"]))
    return " | ".join([part for part in message_parts if part])


def serialize_public_credit_submission(submission: models.PublicCreditSubmission) -> dict:
    lead = getattr(submission, "lead", None)
    return {
        "id": submission.id,
        "company_id": submission.company_id,
        "lead_id": submission.lead_id,
        "applicant_name": submission.applicant_name,
        "email": submission.email,
        "document_number": submission.document_number,
        "phone": submission.phone,
        "desired_vehicle": submission.desired_vehicle,
        "status": submission.status,
        "verification_verified_at": submission.verification_verified_at,
        "created_at": submission.created_at,
        "updated_at": submission.updated_at,
        "lead_name": getattr(lead, "name", None),
        "lead_status": getattr(lead, "status", None),
        "form_payload": submission.form_payload or {},
        "attachments": submission.attachments or {},
        "consent_text": submission.consent_text,
    }


def build_lead_summary_query(db: Session):
    return db.query(models.Lead).options(
        joinedload(models.Lead.assigned_to),
        joinedload(models.Lead.deleted_by),
        selectinload(models.Lead.supervisors),
        selectinload(models.Lead.process_detail),
        selectinload(models.Lead.purchase_options),
        noload(models.Lead.history),
        noload(models.Lead.conversation),
        noload(models.Lead.notes),
        noload(models.Lead.files),
    )


def apply_lead_access_filters(
    query,
    db: Session,
    current_user: models.User,
    *,
    source: Optional[str] = None,
    status: Optional[str] = None,
    board_scope: Optional[str] = None,
    q: Optional[str] = None,
    exact_date: Optional[str] = None,
    assigned_mode: Optional[str] = None,
    responsible_user_id: Optional[int] = None,
    only_my_leads: bool = False,
):
    role_name = get_user_role_name(current_user)
    can_view_all_company_leads = role_name in {"admin", "super_admin"}
    active_assignee_filter = and_(
        models.User.role_id.isnot(None),
        or_(models.User.is_active == True, models.User.is_active.is_(None))
    )
    invalid_assignment_filter = or_(
        models.Lead.assigned_to_id.is_(None),
        ~models.Lead.assigned_to.has(active_assignee_filter)
    )

    query = query.filter(models.Lead.deleted_at.is_(None))

    if current_user.company_id:
        query = query.filter(models.Lead.company_id == current_user.company_id)

    aliado_user_ids = get_company_ally_user_ids(db, current_user.company_id)

    if board_scope == "ally":
        if not can_view_all_company_leads:
            query = query.filter(
                or_(
                    models.Lead.assigned_to_id == current_user.id,
                    models.Lead.supervisors.any(models.User.id == current_user.id)
                )
            )
        elif aliado_user_ids:
            query = query.filter(
                or_(
                    models.Lead.assigned_to_id.in_(aliado_user_ids),
                    models.Lead.supervisors.any(models.User.id.in_(aliado_user_ids))
                )
            )
        else:
            query = query.filter(false())
    elif aliado_user_ids and not can_view_all_company_leads:
        query = query.filter(
            or_(
                models.Lead.assigned_to_id.is_(None),
                ~models.Lead.assigned_to_id.in_(aliado_user_ids)
            )
        )

    if not can_view_all_company_leads:
        tracked_advisor_ids = get_user_tracked_advisor_ids(current_user)
        access_conditions = [
            models.Lead.assigned_to_id == current_user.id,
            models.Lead.supervisors.any(models.User.id == current_user.id),
            and_(
                invalid_assignment_filter,
                or_(
                    models.Lead.created_by_id == current_user.id,
                    models.Lead.history.any(models.LeadHistory.user_id == current_user.id)
                )
            )
        ]
        if tracked_advisor_ids:
            access_conditions.extend([
                models.Lead.assigned_to_id.in_(tracked_advisor_ids),
                models.Lead.supervisors.any(models.User.id.in_(tracked_advisor_ids))
            ])
        query = query.filter(or_(*access_conditions))

    if source:
        query = query.filter(models.Lead.source == source)

    if status:
        query = query.filter(models.Lead.status == normalize_lead_status_value(status))

    if q:
        search = f"%{q}%"
        query = query.filter(
            or_(
                models.Lead.name.ilike(search),
                models.Lead.email.ilike(search),
                models.Lead.phone.ilike(search)
            )
        )

    if exact_date:
        try:
            parsed_date = datetime.date.fromisoformat(exact_date)
            query = query.filter(func.date(models.Lead.created_at) == parsed_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha inválido")

    if only_my_leads:
        query = query.filter(
            or_(
                models.Lead.assigned_to_id == current_user.id,
                models.Lead.supervisors.any(models.User.id == current_user.id)
            )
        )

    if responsible_user_id:
        query = query.filter(
            or_(
                models.Lead.assigned_to_id == responsible_user_id,
                models.Lead.supervisors.any(models.User.id == responsible_user_id)
            )
        )

    if assigned_mode == "assigned":
        query = query.filter(
            or_(
                models.Lead.assigned_to_id.isnot(None),
                models.Lead.supervisors.any()
            )
        )
    elif assigned_mode == "unassigned":
        query = query.filter(
            and_(
                models.Lead.assigned_to_id.is_(None),
                ~models.Lead.supervisors.any()
            )
        )

    return query


def hydrate_lead_summary_fields(db: Session, leads: List[models.Lead]):
    if not leads:
        return

    lead_ids = [lead.id for lead in leads]
    related_credits = db.query(models.CreditApplication).filter(
        models.CreditApplication.lead_id.in_(lead_ids)
    ).order_by(
        models.CreditApplication.updated_at.desc(),
        models.CreditApplication.created_at.desc()
    ).all()

    records_by_lead_id: Dict[int, List[models.CreditApplication]] = {}
    for record in related_credits:
        if not record.lead_id:
            continue
        records_by_lead_id.setdefault(record.lead_id, []).append(record)

    for lead in leads:
        lead.status = normalize_lead_status_value(lead.status)
        related_records = records_by_lead_id.get(lead.id, [])
        related_credit = pick_related_credit_record(related_records, lead.status)
        related_purchase = pick_related_purchase_record(related_records)
        if related_credit and is_credit_stage_status(lead.status):
            lead.credit_application_id = related_credit.id
            lead.credit_application_status = related_credit.status
            lead.credit_application_updated_at = related_credit.updated_at
        if related_purchase and should_attach_purchase_request(lead, related_purchase):
            lead.purchase_request_id = related_purchase.id
            lead.purchase_request_status = related_purchase.status
            lead.purchase_request_updated_at = related_purchase.updated_at
            lead.purchase_request_notes = related_purchase.notes
        sanitize_lead_assignment_for_response(lead)


def normalize_lead_status_value(status: Optional[str], default: str = "new") -> str:
    normalized = (status or "").strip().lower()
    if not normalized:
        return default
    return LEGACY_LEAD_STATUS_MAP.get(normalized, normalized)


def is_credit_stage_status(status: Optional[str]) -> bool:
    return normalize_lead_status_value(status) in LEAD_CREDIT_FLOW_STATUSES


def is_purchase_stage_status(status: Optional[str]) -> bool:
    return normalize_lead_status_value(status) in {
        models.LeadStatus.RESERVED.value,
        models.LeadStatus.PREPARATION.value,
        models.LeadStatus.SOLD.value,
    }


def should_attach_purchase_request(lead: Optional[models.Lead], purchase: Optional[models.CreditApplication]) -> bool:
    """
    Expose purchase request snapshot fields in lead responses.

    We attach it when the lead is in the purchase flow OR when the purchase record
    itself already progressed (e.g. car_purchased) so the UI can show "Carro comprado"
    even after `has_vehicle=True`.
    """
    if not purchase:
        return False
    if lead and is_purchase_stage_status(getattr(lead, "status", None)):
        return True
    purchase_status = (getattr(purchase, "status", None) or "").strip().lower()
    return purchase_status in {
        "purchase_process",
        "car_purchased",
        models.CreditStatus.COMPLETED.value,
        models.CreditStatus.REJECTED.value,
        models.CreditStatus.IN_REVIEW.value,
        models.CreditStatus.APPROVED.value,
        models.CreditStatus.PENDING.value,
    }


def is_purchase_request_record(record: Optional[models.CreditApplication]) -> bool:
    if not record:
        return False
    notes_text = (getattr(record, "notes", None) or "").strip().lower()
    purchase_markers = (
        "[purchase_request]",
        "solicitud de compra",
        "compra",
        "busqueda de vehiculo",
        "búsqueda de vehiculo",
        "busqueda del vehiculo",
        "búsqueda del vehículo",
    )
    return any(marker in notes_text for marker in purchase_markers)


def pick_related_credit_record(
    records: list[models.CreditApplication],
    lead_status: Optional[str] = None,
) -> Optional[models.CreditApplication]:
    credit_records = [record for record in (records or []) if not is_purchase_request_record(record)]
    if not credit_records:
        return None

    normalized_lead_status = normalize_lead_status_value(lead_status)
    advanced_credit_stages = {
        models.LeadStatus.APPROVALS.value,
        models.LeadStatus.RESERVED.value,
        models.LeadStatus.PREPARATION.value,
        models.LeadStatus.SOLD.value,
    }
    status_priority = {
        models.CreditStatus.COMPLETED.value: 5,
        models.CreditStatus.APPROVED.value: 4,
        models.CreditStatus.REJECTED.value: 3,
        models.CreditStatus.IN_REVIEW.value: 2,
        models.CreditStatus.PENDING.value: 1,
    }

    def _score(record: models.CreditApplication):
        record_status = (getattr(record, "status", None) or "").strip().lower()
        priority = status_priority.get(record_status, 0)
        notes_text = (getattr(record, "notes", None) or "").strip().lower()
        if normalized_lead_status in advanced_credit_stages and record_status == models.CreditStatus.PENDING.value:
            if notes_text.startswith("generado automaticamente desde lead") or notes_text.startswith("[credit_request] generado automaticamente desde lead"):
                priority = -1
        sort_date = getattr(record, "updated_at", None) or getattr(record, "created_at", None) or datetime.datetime.min
        return (priority, sort_date)

    return max(credit_records, key=_score)


def pick_related_purchase_record(records: list[models.CreditApplication]) -> Optional[models.CreditApplication]:
    purchase_records = [record for record in (records or []) if is_purchase_request_record(record)]
    if not purchase_records:
        return None
    return max(
        purchase_records,
        key=lambda record: getattr(record, "updated_at", None) or getattr(record, "created_at", None) or datetime.datetime.min
    )


def normalize_credit_desired_vehicle(value: Optional[str]) -> str:
    normalized = " ".join(str(value or "").split()).strip()
    return normalized or "Por definir"


def get_lead_credit_vehicle_context(
    db: Session,
    lead: Optional[models.Lead],
    process_detail: Optional[models.LeadProcessDetail] = None
) -> Dict[str, str]:
    detail = process_detail
    if detail is None and lead:
        detail = db.query(models.LeadProcessDetail).filter(
            models.LeadProcessDetail.lead_id == lead.id
        ).first()

    vehicle_label = ""
    if detail:
        if detail.vehicle_id:
            vehicle = db.query(models.Vehicle).filter(models.Vehicle.id == detail.vehicle_id).first()
            if vehicle:
                vehicle_label = " ".join(
                    part for part in [vehicle.make, vehicle.model, str(vehicle.year or ""), getattr(vehicle, "plate", None)] if part
                ).strip()
        if not vehicle_label:
            vehicle_label = (detail.desired_vehicle or "").strip()

    if not vehicle_label and lead:
        vehicle_label = (lead.message or "").strip()

    vehicle_label = normalize_credit_desired_vehicle(vehicle_label)
    mode_label = "vehiculo de inventario" if detail and detail.has_vehicle else "vehiculo por buscar"
    return {
        "vehicle": vehicle_label,
        "mode": mode_label
    }

def ensure_role_configuration_columns():
    with engine.begin() as conn:
        role_columns = {
            "company_id": "ALTER TABLE roles ADD COLUMN company_id INTEGER NULL",
            "permissions_json": "ALTER TABLE roles ADD COLUMN permissions_json TEXT NULL",
            "menu_order_json": "ALTER TABLE roles ADD COLUMN menu_order_json TEXT NULL",
            "is_system": "ALTER TABLE roles ADD COLUMN is_system BOOLEAN NOT NULL DEFAULT 0",
            "base_role_name": "ALTER TABLE roles ADD COLUMN base_role_name VARCHAR(50) NULL",
            "auto_assign_leads": "ALTER TABLE roles ADD COLUMN auto_assign_leads BOOLEAN NOT NULL DEFAULT 0",
            "assignable_role_ids_json": "ALTER TABLE roles ADD COLUMN assignable_role_ids_json TEXT NULL",
            "advisor_tracking_enabled": "ALTER TABLE roles ADD COLUMN advisor_tracking_enabled BOOLEAN NOT NULL DEFAULT 0",
            "tracked_advisor_ids_json": "ALTER TABLE roles ADD COLUMN tracked_advisor_ids_json TEXT NULL",
        }

        for column_name, ddl in role_columns.items():
            exists = conn.execute(text(
                "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'roles' "
                "AND COLUMN_NAME = :column_name"
            ), {"column_name": column_name}).scalar()
            if not exists:
                conn.execute(text(ddl))


def ensure_payment_receipts_columns():
    with engine.begin() as conn:
        table_exists = conn.execute(text(
            "SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'payment_receipts'"
        )).scalar()
        if not table_exists:
            return

        receipt_columns = {
            "display_name": "ALTER TABLE payment_receipts ADD COLUMN display_name VARCHAR(180) NULL",
            "customer_name": "ALTER TABLE payment_receipts ADD COLUMN customer_name VARCHAR(180) NULL",
            "customer_document": "ALTER TABLE payment_receipts ADD COLUMN customer_document VARCHAR(80) NULL",
            "concept": "ALTER TABLE payment_receipts ADD COLUMN concept VARCHAR(200) NULL",
            "movement_type": "ALTER TABLE payment_receipts ADD COLUMN movement_type VARCHAR(20) NOT NULL DEFAULT 'income'",
            "payment_method": "ALTER TABLE payment_receipts ADD COLUMN payment_method VARCHAR(30) NULL",
            "bank": "ALTER TABLE payment_receipts ADD COLUMN bank VARCHAR(80) NULL"
        }
        for column_name, ddl in receipt_columns.items():
            exists = conn.execute(text(
                "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'payment_receipts' "
                "AND COLUMN_NAME = :column_name"
            ), {"column_name": column_name}).scalar()
            if not exists:
                conn.execute(text(ddl))

        sale_id_nullable = conn.execute(text(
            "SELECT IS_NULLABLE FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'payment_receipts' "
            "AND COLUMN_NAME = 'sale_id'"
        )).scalar()
        if sale_id_nullable == "NO":
            conn.execute(text("ALTER TABLE payment_receipts MODIFY COLUMN sale_id INTEGER NULL"))

        index_exists = conn.execute(text(
            "SELECT COUNT(*) FROM INFORMATION_SCHEMA.STATISTICS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'roles' "
            "AND INDEX_NAME = 'ix_roles_company_id'"
        )).scalar()
        if not index_exists:
            conn.execute(text("CREATE INDEX ix_roles_company_id ON roles (company_id)"))

        conn.execute(text(
            "UPDATE roles SET is_system = 1 WHERE name IN "
            "('super_admin', 'admin', 'asesor', 'gestion_creditos', 'aliado', 'inventario', 'compras', 'user')"
        ))

ensure_role_configuration_columns()
ensure_payment_receipts_columns()

def ensure_company_public_domain_columns():
    with engine.begin() as conn:
        table_exists = conn.execute(text(
            "SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'companies'"
        )).scalar()
        if not table_exists:
            return

        column_exists = conn.execute(text(
            "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'companies' "
            "AND COLUMN_NAME = 'public_domain'"
        )).scalar()
        if not column_exists:
            conn.execute(text("ALTER TABLE companies ADD COLUMN public_domain VARCHAR(255) NULL"))

        domains_json_exists = conn.execute(text(
            "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'companies' "
            "AND COLUMN_NAME = 'public_domains_json'"
        )).scalar()
        if not domains_json_exists:
            conn.execute(text("ALTER TABLE companies ADD COLUMN public_domains_json TEXT NULL"))

        index_exists = conn.execute(text(
            "SELECT COUNT(*) FROM INFORMATION_SCHEMA.STATISTICS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'companies' "
            "AND INDEX_NAME = 'ix_companies_public_domain'"
        )).scalar()
        if not index_exists:
            conn.execute(text("CREATE INDEX ix_companies_public_domain ON companies (public_domain)"))

        company_columns = {
            "max_users": "ALTER TABLE companies ADD COLUMN max_users INTEGER NULL",
            "max_leads": "ALTER TABLE companies ADD COLUMN max_leads INTEGER NULL",
            "max_active_accounts": "ALTER TABLE companies ADD COLUMN max_active_accounts INTEGER NULL",
            "license_start_date": "ALTER TABLE companies ADD COLUMN license_start_date DATE NULL",
            "license_end_date": "ALTER TABLE companies ADD COLUMN license_end_date DATE NULL",
            "enabled_modules_json": "ALTER TABLE companies ADD COLUMN enabled_modules_json TEXT NULL",
            "website_url": "ALTER TABLE companies ADD COLUMN website_url VARCHAR(500) NULL",
            "contact_address": "ALTER TABLE companies ADD COLUMN contact_address VARCHAR(500) NULL",
            "contact_phone": "ALTER TABLE companies ADD COLUMN contact_phone VARCHAR(80) NULL",
            "social_instagram": "ALTER TABLE companies ADD COLUMN social_instagram VARCHAR(255) NULL",
            "social_tiktok": "ALTER TABLE companies ADD COLUMN social_tiktok VARCHAR(255) NULL",
            "social_facebook": "ALTER TABLE companies ADD COLUMN social_facebook VARCHAR(255) NULL",
            "public_credit_requires_email_validation": "ALTER TABLE companies ADD COLUMN public_credit_requires_email_validation BOOLEAN NOT NULL DEFAULT 1",
        }
        for column_name, ddl in company_columns.items():
            exists = conn.execute(text(
                "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'companies' "
                "AND COLUMN_NAME = :column_name"
            ), {"column_name": column_name}).scalar()
            if not exists:
                conn.execute(text(ddl))

ensure_company_public_domain_columns()

def ensure_user_status_columns():
    try:
        with engine.begin() as conn:
            user_columns = {
                "is_active": "ALTER TABLE users ADD COLUMN is_active BOOLEAN NOT NULL DEFAULT 1",
                "auto_assign_leads": "ALTER TABLE users ADD COLUMN auto_assign_leads BOOLEAN NOT NULL DEFAULT 0",
                "tracked_advisor_ids_json": "ALTER TABLE users ADD COLUMN tracked_advisor_ids_json TEXT NULL",
            }
            created_columns = set()
            for column_name, ddl in user_columns.items():
                exists = conn.execute(text(
                    "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                    "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'users' "
                    "AND COLUMN_NAME = :column_name"
                ), {"column_name": column_name}).scalar()
                if not exists:
                    conn.execute(text(ddl))
                    created_columns.add(column_name)

            if "auto_assign_leads" in created_columns:
                conn.execute(text("""
                    UPDATE users u
                    LEFT JOIN roles r ON r.id = u.role_id
                    SET u.auto_assign_leads = 1
                    WHERE COALESCE(u.is_active, 1) = 1
                      AND (
                        COALESCE(r.base_role_name, r.name) = 'asesor'
                        OR COALESCE(r.auto_assign_leads, 0) = 1
                      )
                """))
    except Exception as exc:
        print(f"Warning: could not ensure user status columns: {exc}", flush=True)

ensure_user_status_columns()


def ensure_system_roles_exist():
    try:
        system_roles = {
            "super_admin": "Super Admin Global",
            "admin": "Administrador de Empresa",
            "inventario": "Gestor de Inventario",
            "asesor": "Asesor / Vendedor",
            "gestion_creditos": "Gestion de Creditos",
            "aliado": "Aliado Estrategico",
            "compras": "Gestor de Compras",
            "user": "Usuario Basico",
        }

        with Session(engine) as db:
            existing_roles = {
                role.name: role
                for role in db.query(models.Role).filter(models.Role.name.in_(system_roles.keys())).all()
            }
            changed = False

            for role_name, label in system_roles.items():
                role = existing_roles.get(role_name)
                if not role:
                    role = models.Role(
                        name=role_name,
                        label=label,
                        is_system=True,
                        permissions_json=json.dumps(DEFAULT_ROLE_VIEW_ACCESS.get(role_name, [])),
                        menu_order_json=json.dumps(DEFAULT_ROLE_MENU_ORDER.get(role_name, [])),
                        auto_assign_leads=(role_name == "asesor"),
                    )
                    db.add(role)
                    changed = True
                    continue

                if not getattr(role, "label", None):
                    role.label = label
                    changed = True
                if not getattr(role, "is_system", False):
                    role.is_system = True
                    changed = True
                if not getattr(role, "permissions_json", None):
                    role.permissions_json = json.dumps(DEFAULT_ROLE_VIEW_ACCESS.get(role_name, []))
                    changed = True
                if not getattr(role, "menu_order_json", None):
                    role.menu_order_json = json.dumps(DEFAULT_ROLE_MENU_ORDER.get(role_name, []))
                    changed = True

            if changed:
                db.commit()
    except Exception as exc:
        print(f"Warning: could not ensure system roles: {exc}", flush=True)


ensure_system_roles_exist()

def ensure_sales_metadata_columns():
    try:
        with engine.begin() as conn:
            sales_columns = {
                "seller_type": "ALTER TABLE sales ADD COLUMN seller_type VARCHAR(20) NOT NULL DEFAULT 'internal'",
                "external_seller_name": "ALTER TABLE sales ADD COLUMN external_seller_name VARCHAR(150) NULL",
                "tax_transaction_type": "ALTER TABLE sales ADD COLUMN tax_transaction_type VARCHAR(50) NULL DEFAULT 'intermediacion'",
                "tax_transfer_to_cars": "ALTER TABLE sales ADD COLUMN tax_transfer_to_cars VARCHAR(20) NULL",
                "tax_seller_name": "ALTER TABLE sales ADD COLUMN tax_seller_name VARCHAR(180) NULL",
                "tax_seller_document": "ALTER TABLE sales ADD COLUMN tax_seller_document VARCHAR(60) NULL",
                "tax_seller_email": "ALTER TABLE sales ADD COLUMN tax_seller_email VARCHAR(150) NULL",
                "tax_seller_address": "ALTER TABLE sales ADD COLUMN tax_seller_address VARCHAR(250) NULL",
                "tax_seller_phone": "ALTER TABLE sales ADD COLUMN tax_seller_phone VARCHAR(60) NULL",
                "tax_seller_payment_method": "ALTER TABLE sales ADD COLUMN tax_seller_payment_method VARCHAR(120) NULL",
                "tax_buyer_name": "ALTER TABLE sales ADD COLUMN tax_buyer_name VARCHAR(180) NULL",
                "tax_buyer_document": "ALTER TABLE sales ADD COLUMN tax_buyer_document VARCHAR(60) NULL",
                "tax_buyer_email": "ALTER TABLE sales ADD COLUMN tax_buyer_email VARCHAR(150) NULL",
                "tax_buyer_address": "ALTER TABLE sales ADD COLUMN tax_buyer_address VARCHAR(250) NULL",
                "tax_buyer_phone": "ALTER TABLE sales ADD COLUMN tax_buyer_phone VARCHAR(60) NULL",
                "tax_buyer_payment_method": "ALTER TABLE sales ADD COLUMN tax_buyer_payment_method VARCHAR(120) NULL",
                "tax_buyer_financing_entity": "ALTER TABLE sales ADD COLUMN tax_buyer_financing_entity VARCHAR(150) NULL",
            }
            for column_name, ddl in sales_columns.items():
                exists = conn.execute(text(
                    "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
                    "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'sales' "
                    "AND COLUMN_NAME = :column_name"
                ), {"column_name": column_name}).scalar()
                if not exists:
                    conn.execute(text(ddl))

            seller_is_nullable = conn.execute(text(
                "SELECT IS_NULLABLE FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'sales' "
                "AND COLUMN_NAME = 'seller_id'"
            )).scalar()
            if seller_is_nullable == "NO":
                conn.execute(text("ALTER TABLE sales MODIFY COLUMN seller_id INTEGER NULL"))
    except Exception as exc:
        print(f"Warning: could not ensure sales metadata columns: {exc}", flush=True)

ensure_sales_metadata_columns()


def ensure_tax_report_entries_table():
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS tax_report_entries (
                    id INTEGER PRIMARY KEY AUTO_INCREMENT,
                    company_id INTEGER NOT NULL,
                    month VARCHAR(20) NULL,
                    year INTEGER NULL,
                    make VARCHAR(120) NULL,
                    reference VARCHAR(160) NULL,
                    plate VARCHAR(30) NULL,
                    model_year INTEGER NULL,
                    purchase_price INTEGER DEFAULT 0,
                    sale_price INTEGER DEFAULT 0,
                    transaction_type VARCHAR(50) NULL,
                    transfer_to_cars VARCHAR(20) NULL,
                    seller_name VARCHAR(180) NULL,
                    seller_document VARCHAR(60) NULL,
                    seller_email VARCHAR(150) NULL,
                    seller_address VARCHAR(250) NULL,
                    seller_phone VARCHAR(60) NULL,
                    seller_payment_method VARCHAR(120) NULL,
                    buyer_name VARCHAR(180) NULL,
                    buyer_document VARCHAR(60) NULL,
                    buyer_email VARCHAR(150) NULL,
                    buyer_address VARCHAR(250) NULL,
                    buyer_phone VARCHAR(60) NULL,
                    buyer_payment_method VARCHAR(120) NULL,
                    buyer_financing_entity VARCHAR(150) NULL,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX ix_tax_report_entries_company_id (company_id)
                )
            """))
    except Exception as exc:
        print(f"Warning: could not ensure tax report entries table: {exc}", flush=True)


ensure_tax_report_entries_table()

def ensure_gmail_settings_columns():
    """
    Backward-compatible bootstrap for Gmail integration settings.
    """
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'integration_settings'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            gmail_columns = {
                "gmail_enabled": (
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN gmail_enabled BOOLEAN NULL DEFAULT 0"
                ),
                "gmail_client_id": (
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN gmail_client_id TEXT NULL"
                ),
                "gmail_client_secret": (
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN gmail_client_secret TEXT NULL"
                ),
                "gmail_redirect_uri": (
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN gmail_redirect_uri VARCHAR(500) NULL"
                ),
                "gmail_refresh_token": (
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN gmail_refresh_token TEXT NULL"
                ),
                "gmail_monitored_sender": (
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN gmail_monitored_sender VARCHAR(255) NULL"
                ),
                "gmail_label": (
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN gmail_label VARCHAR(120) NULL"
                ),
                "gmail_sync_days": (
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN gmail_sync_days INT NULL DEFAULT 7"
                ),
                "gmail_sync_max_results": (
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN gmail_sync_max_results INT NULL DEFAULT 20"
                ),
            }

            for column_name, ddl in gmail_columns.items():
                if column_name not in existing_cols:
                    conn.execute(text(ddl))

            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure gmail settings columns: {exc}", flush=True)

ensure_gmail_settings_columns()

def ensure_company_smtp_settings_columns():
    try:
        with engine.begin() as connection:
            smtp_columns = {
                "smtp_enabled": "ADD COLUMN smtp_enabled BOOLEAN NULL DEFAULT 0",
                "smtp_host": "ADD COLUMN smtp_host VARCHAR(255) NULL",
                "smtp_port": "ADD COLUMN smtp_port INT NULL DEFAULT 587",
                "smtp_username": "ADD COLUMN smtp_username VARCHAR(255) NULL",
                "smtp_password": "ADD COLUMN smtp_password TEXT NULL",
                "smtp_from": "ADD COLUMN smtp_from VARCHAR(255) NULL",
                "smtp_use_tls": "ADD COLUMN smtp_use_tls BOOLEAN NULL DEFAULT 1",
                "smtp_always_recipients": "ADD COLUMN smtp_always_recipients TEXT NULL",
            }
            for _, ddl in smtp_columns.items():
                try:
                    connection.execute(text(f"ALTER TABLE integration_settings {ddl}"))
                except Exception:
                    pass
    except Exception as exc:
        print(f"Warning: could not ensure smtp settings columns: {exc}", flush=True)

ensure_company_smtp_settings_columns()

def ensure_whatsapp_settings_columns():
    try:
        with engine.begin() as connection:
            whatsapp_columns = {
                "whatsapp_sales_phone_number_id": "ADD COLUMN whatsapp_sales_phone_number_id VARCHAR(100) NULL",
                "whatsapp_purchases_phone_number_id": "ADD COLUMN whatsapp_purchases_phone_number_id VARCHAR(100) NULL",
                "whatsapp_documents_enabled": "ADD COLUMN whatsapp_documents_enabled BOOLEAN NULL DEFAULT 1",
                "whatsapp_calling_enabled": "ADD COLUMN whatsapp_calling_enabled BOOLEAN NULL DEFAULT 0",
                "whatsapp_calling_mode": "ADD COLUMN whatsapp_calling_mode VARCHAR(40) NULL DEFAULT 'whatsapp_link'",
                "whatsapp_calling_provider_url": "ADD COLUMN whatsapp_calling_provider_url VARCHAR(500) NULL",
                "whatsapp_calling_provider_token": "ADD COLUMN whatsapp_calling_provider_token TEXT NULL",
                "whatsapp_sales_agent_name": "ADD COLUMN whatsapp_sales_agent_name VARCHAR(120) NULL",
                "whatsapp_sales_agent_prompt": "ADD COLUMN whatsapp_sales_agent_prompt TEXT NULL",
                "whatsapp_purchases_agent_name": "ADD COLUMN whatsapp_purchases_agent_name VARCHAR(120) NULL",
                "whatsapp_purchases_agent_prompt": "ADD COLUMN whatsapp_purchases_agent_prompt TEXT NULL",
            }
            for _, ddl in whatsapp_columns.items():
                try:
                    connection.execute(text(f"ALTER TABLE integration_settings {ddl}"))
                except Exception:
                    pass
    except Exception as exc:
        print(f"Warning: could not ensure whatsapp settings columns: {exc}", flush=True)

ensure_whatsapp_settings_columns()

def ensure_chatbot_settings_columns():
    """
    Backward-compatible bootstrap for legacy DBs where integration_settings
    still lacks chatbot config columns.
    """
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'integration_settings'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            if "chatbot_bot_name" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN chatbot_bot_name VARCHAR(120) NULL DEFAULT 'Jennifer Quimbayo'"
                ))
            if "chatbot_typing_min_ms" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN chatbot_typing_min_ms INT NULL DEFAULT 7000"
                ))
            if "chatbot_typing_max_ms" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE integration_settings "
                    "ADD COLUMN chatbot_typing_max_ms INT NULL DEFAULT 18000"
                ))
            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure chatbot settings columns: {exc}", flush=True)

ensure_chatbot_settings_columns()


def ensure_lead_reply_columns():
    """
    Backward-compatible bootstrap for unread-reply markers on leads.
    """
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'leads'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            if "has_unread_reply" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE leads "
                    "ADD COLUMN has_unread_reply INT NULL DEFAULT 0"
                ))
            if "last_reply_at" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE leads "
                    "ADD COLUMN last_reply_at DATETIME NULL"
                ))
            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure lead reply columns: {exc}", flush=True)


ensure_lead_reply_columns()


def ensure_lead_soft_delete_columns():
    """
    Backward-compatible bootstrap for soft-delete metadata on leads.
    """
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'leads'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            if "deleted_at" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE leads "
                    "ADD COLUMN deleted_at DATETIME NULL"
                ))
            if "deleted_reason" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE leads "
                    "ADD COLUMN deleted_reason TEXT NULL"
                ))
            if "deleted_by_id" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE leads "
                    "ADD COLUMN deleted_by_id INT NULL"
                ))
            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure lead soft delete columns: {exc}", flush=True)


ensure_lead_soft_delete_columns()


def ensure_credit_application_link_column():
    """
    Backward-compatible bootstrap so credit applications can be tied to a lead.
    """
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'credit_applications'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            if "lead_id" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE credit_applications "
                    "ADD COLUMN lead_id INT NULL"
                ))
                conn.execute(text(
                    "ALTER TABLE credit_applications "
                    "ADD INDEX ix_credit_applications_lead_id (lead_id)"
                ))
            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure credit application lead link column: {exc}", flush=True)


ensure_credit_application_link_column()


def ensure_purchase_option_decision_columns():
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'purchase_options'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            if "decision_status" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE purchase_options "
                    "ADD COLUMN decision_status VARCHAR(30) NULL DEFAULT 'pending'"
                ))
            if "decision_note" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE purchase_options "
                    "ADD COLUMN decision_note TEXT NULL"
                ))
            if "decision_user_id" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE purchase_options "
                    "ADD COLUMN decision_user_id INT NULL"
                ))
            if "decision_at" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE purchase_options "
                    "ADD COLUMN decision_at DATETIME NULL"
                ))
            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure purchase option decision columns: {exc}", flush=True)


ensure_purchase_option_decision_columns()


def ensure_credit_notes_text_column():
    try:
        with engine.connect() as conn:
            notes_column_type = conn.execute(text(
                "SELECT DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'credit_applications' "
                "AND COLUMN_NAME = 'notes'"
            )).scalar()

            if notes_column_type and notes_column_type.lower() != "text":
                conn.execute(text(
                    "ALTER TABLE credit_applications MODIFY COLUMN notes TEXT NULL"
                ))
                conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure credit notes text column: {exc}", flush=True)


ensure_credit_notes_text_column()


def ensure_credit_desired_vehicle_text_column():
    try:
        with engine.connect() as conn:
            desired_vehicle_column_type = conn.execute(text(
                "SELECT DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'credit_applications' "
                "AND COLUMN_NAME = 'desired_vehicle'"
            )).scalar()

            if desired_vehicle_column_type and desired_vehicle_column_type.lower() != "text":
                conn.execute(text(
                    "ALTER TABLE credit_applications MODIFY COLUMN desired_vehicle TEXT NULL"
                ))
                conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure credit desired_vehicle text column: {exc}", flush=True)


ensure_credit_desired_vehicle_text_column()


def ensure_credit_approval_columns():
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'credit_applications'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            if "approved_amount" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE credit_applications "
                    "ADD COLUMN approved_amount INTEGER NULL"
                ))
            if "approval_percentage" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE credit_applications "
                    "ADD COLUMN approval_percentage INTEGER NULL"
                ))
            if "approved_down_payment" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE credit_applications "
                    "ADD COLUMN approved_down_payment INTEGER NULL"
                ))
            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure credit approval columns: {exc}", flush=True)


ensure_credit_approval_columns()


def ensure_lead_process_detail_reservation_columns():
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'lead_process_details'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            if "reservation_amount" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE lead_process_details "
                    "ADD COLUMN reservation_amount INTEGER NULL"
                ))
            if "credit_used_amount" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE lead_process_details "
                    "ADD COLUMN credit_used_amount INTEGER NULL"
                ))
            if "reservation_payment_method" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE lead_process_details "
                    "ADD COLUMN reservation_payment_method VARCHAR(50) NULL"
                ))
            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure lead process detail reservation columns: {exc}", flush=True)


ensure_lead_process_detail_reservation_columns()

def ensure_automation_rule_reassignment_columns():
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'automation_rules'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            if "reassign_after_alerts_enabled" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE automation_rules "
                    "ADD COLUMN reassign_after_alerts_enabled BOOLEAN NOT NULL DEFAULT 0"
                ))
            if "reassign_after_alerts_count" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE automation_rules "
                    "ADD COLUMN reassign_after_alerts_count INTEGER NOT NULL DEFAULT 0"
                ))
            if "reassign_to_user_id" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE automation_rules "
                    "ADD COLUMN reassign_to_user_id INTEGER NULL"
                ))
            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure automation rule reassignment columns: {exc}", flush=True)


ensure_automation_rule_reassignment_columns()

def ensure_sent_alert_logs_indexes():
    try:
        with engine.connect() as conn:
            index_exists = conn.execute(text(
                "SELECT COUNT(*) FROM INFORMATION_SCHEMA.STATISTICS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'sent_alert_logs' "
                "AND INDEX_NAME = 'ix_sent_alert_logs_rule_lead_sent_at'"
            )).scalar()

            if not index_exists:
                conn.execute(text(
                    "CREATE INDEX ix_sent_alert_logs_rule_lead_sent_at "
                    "ON sent_alert_logs (rule_id, lead_id, sent_at)"
                ))
                conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure sent_alert_logs indexes: {exc}", flush=True)


ensure_sent_alert_logs_indexes()

def ensure_lead_query_indexes():
    try:
        with engine.connect() as conn:
            desired_indexes = [
                (
                    "leads",
                    "ix_leads_company_deleted_id",
                    "CREATE INDEX ix_leads_company_deleted_id "
                    "ON leads (company_id, deleted_at, id)"
                ),
                (
                    "leads",
                    "ix_leads_company_deleted_status_id",
                    "CREATE INDEX ix_leads_company_deleted_status_id "
                    "ON leads (company_id, deleted_at, status, id)"
                ),
                (
                    "leads",
                    "ix_leads_company_assigned_deleted_id",
                    "CREATE INDEX ix_leads_company_assigned_deleted_id "
                    "ON leads (company_id, assigned_to_id, deleted_at, id)"
                ),
                (
                    "leads",
                    "ix_leads_company_source_deleted_id",
                    "CREATE INDEX ix_leads_company_source_deleted_id "
                    "ON leads (company_id, source, deleted_at, id)"
                ),
                (
                    "lead_supervisors",
                    "ix_lead_supervisors_user_lead",
                    "CREATE INDEX ix_lead_supervisors_user_lead "
                    "ON lead_supervisors (user_id, lead_id)"
                ),
                (
                    "notifications",
                    "ix_notifications_user_read_created",
                    "CREATE INDEX ix_notifications_user_read_created "
                    "ON notifications (user_id, is_read, created_at)"
                ),
            ]

            for table_name, index_name, ddl in desired_indexes:
                index_exists = conn.execute(text(
                    "SELECT COUNT(*) FROM INFORMATION_SCHEMA.STATISTICS "
                    "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = :table_name "
                    "AND INDEX_NAME = :index_name"
                ), {
                    "table_name": table_name,
                    "index_name": index_name,
                }).scalar()

                if not index_exists:
                    conn.execute(text(ddl))

            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure lead query indexes: {exc}", flush=True)


ensure_lead_query_indexes()

def ensure_user_activity_column():
    try:
        with engine.connect() as conn:
            existing_cols_result = conn.execute(text(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'users'"
            ))
            existing_cols = {row[0] for row in existing_cols_result.fetchall()}

            if "is_active" not in existing_cols:
                conn.execute(text(
                    "ALTER TABLE users "
                    "ADD COLUMN is_active INT NULL DEFAULT 1"
                ))
            conn.execute(text(
                "UPDATE users SET is_active = 1 WHERE is_active IS NULL"
            ))
            conn.commit()
    except Exception as exc:
        print(f"Warning: could not ensure user activity column: {exc}", flush=True)


ensure_user_activity_column()


def ensure_lead_statuses_synced():
    try:
        with engine.begin() as conn:
            for old_status, new_status in LEGACY_LEAD_STATUS_MAP.items():
                conn.execute(
                    text("UPDATE leads SET status = :new_status WHERE status = :old_status"),
                    {"old_status": old_status, "new_status": new_status}
                )
                conn.execute(
                    text(
                        "UPDATE lead_history SET previous_status = :new_status "
                        "WHERE previous_status = :old_status"
                    ),
                    {"old_status": old_status, "new_status": new_status}
                )
                conn.execute(
                    text(
                        "UPDATE lead_history SET new_status = :new_status "
                        "WHERE new_status = :old_status"
                    ),
                    {"old_status": old_status, "new_status": new_status}
                )
    except Exception as exc:
        print(f"Warning: could not sync lead statuses: {exc}", flush=True)


ensure_lead_statuses_synced()

app = FastAPI(title="AutosQP API", description="API para gestión de compra venta de carros")

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    print("CRITICAL ERROR IN REQUEST:", request.url, flush=True)
    traceback.print_exc()
    return JSONResponse(
        status_code=500,
        content={"detail": "Internal Server Error", "error": str(exc)}
    )

app.include_router(whatsapp.router) # Register the router
app.include_router(credits.router) # Register credits router
app.include_router(purchases.router)
app.include_router(notifications.router)
app.include_router(rules.router)
app.include_router(vehicles.router)
app.include_router(meta.router)
app.include_router(tiktok.router)
app.include_router(gmail.router)
app.include_router(appointments.router)


# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:3000",
        "http://34.229.164.25:3000",
        "http://54.226.30.192:3000", # Old IP port
        "http://54.226.30.192", # Old Server IP
        "https://54.226.30.192", 
        "http://3.234.117.124:3000", # New Elastic IP port
        "http://3.234.117.124",      # New Elastic IP
        "https://3.234.117.124",     # New Elastic IP HTTPS
        "http://autosqp.com",        # Canonical Domain HTTP
        "https://autosqp.com",       # Canonical Domain HTTPS
        "http://www.autosqp.com",    # Canonical WWW Domain HTTP
        "https://www.autosqp.com",   # Canonical WWW Domain HTTPS
        "http://autosqp.co",         # Domain HTTP
        "https://autosqp.co",        # Domain HTTPS
        "http://www.autosqp.co",     # WWW Domain HTTP
        "https://www.autosqp.co",    # WWW Domain HTTPS
    ],
    allow_origin_regex=r"https?://([a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}(:\d+)?$",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/ping")
def ping():
    print("DEBUG: Ping called", flush=True)
    return {"message": "pong"}

from dependencies import (
    get_current_user,
    get_user_from_anywhere,
    oauth2_scheme,
    ensure_can_modify_lead,
    ensure_can_manage_lead_supervision,
    is_company_admin,
)

@app.post("/token", response_model=schemas.Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == form_data.username).first()
    if not user or not auth_utils.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    if not getattr(user, "is_active", True):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario inhabilitado",
            headers={"WWW-Authenticate": "Bearer"},
        )
    company = None
    if getattr(user, "company_id", None):
        company = db.query(models.Company).filter(models.Company.id == user.company_id).first()
    license_state = ensure_company_license_allows_login(company)
    access_token = auth_utils.create_access_token(
        data={"sub": str(user.id), "role": user.role.name if user.role else "user"}
    )
    
    # Log the successful login
    log_action_to_db(db, user.id, "LOGIN", "Auth", user.id, "Usuario inició sesión exitosamente")
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "license_status": license_state["status"],
        "license_notice": license_state["notice"],
        "license_days_remaining": license_state["days_remaining"],
    }

def log_action_to_db(db: Session, user_id: int, action: str, entity_type: str, entity_id: int, details: str = None, ip_address: str = None):
    try:
        company_id = None
        if user_id:
            log_user = db.query(models.User.company_id).filter(models.User.id == user_id).first()
            company_id = log_user[0] if log_user else None

        new_log = models.SystemLog(
            company_id=company_id,
            user_id=user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            details=details,
            ip_address=ip_address
        )
        db.add(new_log)
        db.commit()
    except Exception as e:
        print(f"Failed to log action: {e}", flush=True)
        db.rollback()

def is_inventory_editor(current_user: models.User) -> bool:
    role_name = current_user.role.name if current_user.role else ""
    return role_name in ["admin", "inventario"] or (role_name == "super_admin" and bool(current_user.company_id))

def is_company_admin_for_inventory(current_user: models.User) -> bool:
    role_name = current_user.role.name if current_user.role else ""
    return role_name == "admin" or (role_name == "super_admin" and bool(current_user.company_id))

def ensure_inventory_editor(current_user: models.User):
    if not is_inventory_editor(current_user):
        raise HTTPException(status_code=403, detail="Solo usuarios de inventario o administradores pueden crear/editar vehículos")

def ensure_inventory_admin_only(current_user: models.User):
    if not is_inventory_editor(current_user):
        raise HTTPException(status_code=403, detail="Solo usuarios de inventario o administradores pueden desactivar vehículos")

def enforce_ally_managed_status(
    db: Session,
    current_user: models.User,
    base_status: Optional[str],
    assigned_to_id: Optional[int] = None
) -> str:
    return normalize_lead_status_value(base_status, models.LeadStatus.NEW.value)


def can_manually_assign_to_any_role(current_user: models.User) -> bool:
    return bool(get_user_role_name(current_user) in {"admin", "super_admin", "aliado"})


def get_user_role_name(user: Optional[models.User]) -> Optional[str]:
    if not user or not user.role:
        return None
    role_names = {
        normalize_role_text(getattr(user.role, "base_role_name", None)),
        normalize_role_text(getattr(user.role, "name", None)),
        normalize_role_text(getattr(user.role, "label", None)),
    }
    role_names.discard("")

    if (
        "super admin" in role_names
        or "super administrador" in role_names
        or any("super" in role_name and "admin" in role_name for role_name in role_names)
        or any("super" in role_name and "administrador" in role_name for role_name in role_names)
    ):
        return "super_admin"
    if (
        "admin" in role_names
        or "administrador" in role_names
        or "administrador de empresa" in role_names
        or any("admin" in role_name for role_name in role_names)
        or any("administrador" in role_name for role_name in role_names)
    ):
        return "admin"
    if "aliado" in role_names or "aliado estrategico" in role_names:
        return "aliado"
    if "asesor" in role_names or "vendedor" in role_names or "asesor vendedor" in role_names:
        return "asesor"
    if (
        "gestion creditos" in role_names
        or "gestion de creditos" in role_names
        or "gestor de creditos" in role_names
        or "gestor creditos" in role_names
        or any("gestion" in role_name and "credit" in role_name for role_name in role_names)
        or any("gestor" in role_name and "credit" in role_name for role_name in role_names)
    ):
        return "gestion_creditos"
    if any("coordinador" in role_name and "credit" in role_name for role_name in role_names):
        return "gestion_creditos"
    if "compras" in role_names:
        return "compras"
    if "inventario" in role_names:
        return "inventario"

    base_role_name = getattr(user.role, "base_role_name", None)
    if base_role_name:
        return base_role_name
    return getattr(user.role, "name", None)


def is_advisor_role(role: Optional[models.Role]) -> bool:
    if not role:
        return False

    role_names = {
        normalize_role_text(getattr(role, "name", None)),
        normalize_role_text(getattr(role, "base_role_name", None)),
        normalize_role_text(getattr(role, "label", None)),
    }
    role_names.discard("")

    return bool(
        role_names & {"asesor", "vendedor", "asesor vendedor", "asesor_vendedor"}
        or any("asesor" in role_name or "vendedor" in role_name for role_name in role_names)
    )


def is_ally_role(role: Optional[models.Role]) -> bool:
    if not role:
        return False

    role_names = {
        normalize_role_text(getattr(role, "name", None)),
        normalize_role_text(getattr(role, "base_role_name", None)),
        normalize_role_text(getattr(role, "label", None)),
    }
    role_names.discard("")

    if "aliado" in role_names:
        return True

    return any("aliad" in role_name for role_name in role_names)


def get_role_permissions(role: Optional[models.Role]) -> List[str]:
    if not role:
        return []
    fallback = DEFAULT_ROLE_VIEW_ACCESS.get(getattr(role, "base_role_name", None) or role.name, [])
    return sanitize_view_ids(parse_json_list(getattr(role, "permissions_json", None), fallback), getattr(role, "company_id", None))


def parse_json_int_list(value: Optional[str], fallback: Optional[List[int]] = None) -> List[int]:
    if not value:
        return fallback[:] if fallback else []
    try:
        parsed = json.loads(value)
        if isinstance(parsed, list):
            normalized = []
            for item in parsed:
                try:
                    parsed_id = int(item)
                except (TypeError, ValueError):
                    continue
                if parsed_id not in normalized:
                    normalized.append(parsed_id)
            return normalized
    except Exception:
        pass
    return fallback[:] if fallback else []


def sanitize_assignable_role_ids(
    db: Session,
    role_ids: Optional[List[int]],
    company_id: Optional[int]
) -> List[int]:
    if not role_ids:
        return []

    normalized_ids = []
    for role_id in role_ids:
        try:
            parsed_id = int(role_id)
        except (TypeError, ValueError):
            continue
        if parsed_id not in normalized_ids:
            normalized_ids.append(parsed_id)

    if not normalized_ids:
        return []

    allowed_roles = db.query(models.Role).filter(models.Role.id.in_(normalized_ids))
    if company_id:
        allowed_roles = allowed_roles.filter(
            or_(
                models.Role.company_id == company_id,
                and_(models.Role.company_id.is_(None), models.Role.is_system == True)
            )
        )

    company = get_company_by_id(db, company_id)
    allowed_ids = {
        role.id for role in allowed_roles.all()
        if is_role_enabled_for_company(role, company)
    }
    return [role_id for role_id in normalized_ids if role_id in allowed_ids]


def get_role_assignable_role_ids(role: Optional[models.Role]) -> List[int]:
    if not role:
        return []
    return parse_json_int_list(getattr(role, "assignable_role_ids_json", None), [])


def get_user_tracked_advisor_ids(user: Optional[models.User]) -> List[int]:
    if (
        not user
        or not getattr(user, "role", None)
        or not bool(getattr(user.role, "advisor_tracking_enabled", False))
    ):
        return []
    return parse_json_int_list(getattr(user, "tracked_advisor_ids_json", None), [])


def sanitize_tracked_advisor_ids(
    db: Session,
    advisor_ids: Optional[List[int]],
    company_id: Optional[int],
    excluded_user_id: Optional[int] = None
) -> List[int]:
    if not advisor_ids or not company_id:
        return []

    normalized_ids = []
    for advisor_id in advisor_ids:
        try:
            parsed_id = int(advisor_id)
        except (TypeError, ValueError):
            continue
        if parsed_id not in normalized_ids:
            normalized_ids.append(parsed_id)

    if not normalized_ids:
        return []

    users = db.query(models.User).join(models.Role, isouter=True).filter(
        models.User.id.in_(normalized_ids),
        models.User.company_id == company_id,
        or_(models.User.is_active == True, models.User.is_active.is_(None))
    ).all()
    allowed_ids = {
        user.id for user in users
        if is_advisor_role(getattr(user, "role", None))
        and (excluded_user_id is None or user.id != excluded_user_id)
    }
    return [advisor_id for advisor_id in normalized_ids if advisor_id in allowed_ids]


def can_assign_lead_to_user(
    current_user: models.User,
    target_user: Optional[models.User]
) -> bool:
    if not target_user or not target_user.role:
        return False

    configured_role_ids = get_role_assignable_role_ids(getattr(current_user, "role", None))
    if configured_role_ids:
        return target_user.role.id in configured_role_ids

    return can_manually_assign_to_any_role(current_user) or is_advisor_role(target_user.role)


def get_company_ally_user_ids(db: Session, company_id: Optional[int]) -> List[int]:
    if not company_id:
        return []

    company_users = db.query(models.User).join(models.Role, isouter=True).filter(
        models.User.company_id == company_id
    ).all()

    return [
        user.id
        for user in company_users
        if is_ally_role(getattr(user, "role", None))
    ]


def normalize_role_text(value: Optional[str]) -> str:
    if not value:
        return ""
    normalized = str(value).strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "Ã¡": "a",
        "Ã©": "e",
        "Ã­": "i",
        "Ã³": "o",
        "Ãº": "u",
        "_": " ",
        "-": " ",
    }
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    return " ".join(normalized.split())


def get_effective_role_key(role: Optional[models.Role]) -> str:
    if not role:
        return ""

    role_names = {
        normalize_role_text(getattr(role, "base_role_name", None)),
        normalize_role_text(getattr(role, "name", None)),
        normalize_role_text(getattr(role, "label", None)),
    }
    role_names.discard("")

    if any("super" in role_name and ("admin" in role_name or "administrador" in role_name) for role_name in role_names):
        return "super_admin"
    if any("admin" in role_name or "administrador" in role_name for role_name in role_names):
        return "admin"
    if "aliado" in role_names or "aliado estrategico" in role_names:
        return "aliado"
    if "asesor" in role_names or "vendedor" in role_names or "asesor vendedor" in role_names:
        return "asesor"
    if (
        "gestion creditos" in role_names
        or "gestion de creditos" in role_names
        or "gestor de creditos" in role_names
        or "gestor creditos" in role_names
        or any("gestion" in role_name and "credit" in role_name for role_name in role_names)
        or any("gestor" in role_name and "credit" in role_name for role_name in role_names)
        or any("coordinador" in role_name and "credit" in role_name for role_name in role_names)
    ):
        return "gestion_creditos"
    if "compras" in role_names:
        return "compras"
    if "inventario" in role_names:
        return "inventario"

    return getattr(role, "base_role_name", None) or getattr(role, "name", None) or ""


def ensure_role_allowed_for_company(role: models.Role, company: Optional[models.Company]):
    disabled_modules = get_role_disabled_required_modules(role, company)
    if not disabled_modules:
        return
    role_label = getattr(role, "label", None) or getattr(role, "name", None) or "rol"
    raise HTTPException(
        status_code=400,
        detail=(
            f"El rol '{role_label}' depende de módulos deshabilitados para esta empresa: "
            f"{', '.join(disabled_modules)}."
        )
    )


def is_credit_coordinator_role(role: Optional[models.Role]) -> bool:
    if not role:
        return False

    role_names = {
        normalize_role_text(getattr(role, "name", None)),
        normalize_role_text(getattr(role, "base_role_name", None)),
        normalize_role_text(getattr(role, "label", None)),
    }
    role_names.discard("")

    explicit_credit_role_names = {
        "gestion creditos",
        "gestion de creditos",
        "gestor de creditos",
        "gestor creditos",
        "gestion_creditos",
        "coordinador",
        "coordinador credito",
        "coordinador de credito",
        "coordinador creditos",
        "coordinador de creditos",
        "coordinador_credito",
        "coordinador_creditos",
    }

    if role_names & explicit_credit_role_names:
        return True

    return any("coordinador" in role_name and "credit" in role_name for role_name in role_names)


def get_credit_coordinator_users(db: Session, company_id: Optional[int]) -> List[models.User]:
    if not company_id:
        return []

    candidates = db.query(models.User).join(models.Role, isouter=True).filter(
        models.User.company_id == company_id
    ).all()

    coordinators = []
    credit_enabled_users = []
    for user in candidates:
        if not is_active_user(user):
            continue
        if "credits" not in get_role_permissions(user.role):
            continue
        credit_enabled_users.append(user)
        if not is_credit_coordinator_role(user.role):
            continue
        coordinators.append(user)
    return coordinators or credit_enabled_users


def choose_credit_coordinator(db: Session, company_id: Optional[int]) -> Optional[models.User]:
    coordinators = get_credit_coordinator_users(db, company_id)
    if not coordinators:
        return None
    return random.choice(coordinators)


def is_active_user(user: Optional[models.User]) -> bool:
    if not user:
        return False
    return bool(getattr(user, "is_active", 1))


def can_user_receive_auto_assigned_leads(user: Optional[models.User]) -> bool:
    return lead_assignment.can_user_receive_auto_assigned_leads(user)


def is_valid_lead_assignee(user: Optional[models.User], company_id: Optional[int] = None) -> bool:
    if not user or not is_active_user(user):
        return False
    if company_id and getattr(user, "company_id", None) != company_id:
        return False

    role_name = get_user_role_name(user)
    return bool(role_name and role_name != "user")


def is_recoverable_lead_owner(user: Optional[models.User], company_id: Optional[int] = None) -> bool:
    if not is_valid_lead_assignee(user, company_id):
        return False

    role_name = get_user_role_name(user)
    return role_name in {"asesor", "compras", "gestion_creditos", "aliado"}


def find_fallback_lead_assignee(
    db: Session,
    lead: models.Lead,
    preferred_user: Optional[models.User] = None
) -> Optional[models.User]:
    candidate_pool = []

    if preferred_user and is_recoverable_lead_owner(preferred_user, lead.company_id):
        candidate_pool.append(preferred_user)

    created_by_user = getattr(lead, "created_by", None)
    if created_by_user is None and getattr(lead, "created_by_id", None):
        created_by_user = db.query(models.User).join(models.Role, isouter=True).filter(
            models.User.id == lead.created_by_id
        ).first()
    if created_by_user and is_recoverable_lead_owner(created_by_user, lead.company_id):
        candidate_pool.append(created_by_user)

    seen_ids = {user.id for user in candidate_pool if getattr(user, "id", None)}
    if candidate_pool:
        return candidate_pool[0]

    history_users = db.query(models.User).join(
        models.LeadHistory,
        models.LeadHistory.user_id == models.User.id
    ).join(models.Role, isouter=True).filter(
        models.LeadHistory.lead_id == lead.id,
        models.User.company_id == lead.company_id
    ).order_by(models.LeadHistory.created_at.desc()).all()

    for history_user in history_users:
        history_user_id = getattr(history_user, "id", None)
        if history_user_id in seen_ids:
            continue
        if is_recoverable_lead_owner(history_user, lead.company_id):
            return history_user

    return None


def sanitize_lead_assignment_for_response(lead: Optional[models.Lead]) -> Optional[models.Lead]:
    if not lead:
        return lead

    assigned_user = getattr(lead, "assigned_to", None)
    if getattr(lead, "assigned_to_id", None) and not is_valid_lead_assignee(assigned_user, getattr(lead, "company_id", None)):
        lead.assigned_to = None

    valid_supervisors = [
        supervisor
        for supervisor in (getattr(lead, "supervisors", None) or [])
        if is_valid_lead_assignee(supervisor, getattr(lead, "company_id", None))
    ]
    if len(valid_supervisors) != len(getattr(lead, "supervisors", None) or []):
        lead.supervisors = valid_supervisors

    return lead


def is_purchase_manager_role(role: Optional[models.Role]) -> bool:
    if not role:
        return False

    role_names = {
        normalize_role_text(getattr(role, "name", None)),
        normalize_role_text(getattr(role, "base_role_name", None)),
        normalize_role_text(getattr(role, "label", None)),
    }
    role_names.discard("")

    explicit_purchase_role_names = {
        "compras",
        "gestor compras",
        "gestor de compras",
        "gestor compra",
        "gestor de compra",
        "comprador",
    }
    if role_names & explicit_purchase_role_names:
        return True

    return any("compra" in role_name for role_name in role_names)


def get_purchase_manager_users(db: Session, company_id: Optional[int]) -> List[models.User]:
    if not company_id:
        return []

    candidates = db.query(models.User).join(models.Role, isouter=True).filter(
        models.User.company_id == company_id
    ).all()

    purchase_users = []
    for user in candidates:
        if not is_active_user(user):
            continue
        role = getattr(user, "role", None)
        if not role:
            continue
        if is_purchase_manager_role(role):
            purchase_users.append(user)

    return purchase_users


def choose_purchase_manager(db: Session, company_id: Optional[int]) -> Optional[models.User]:
    candidates = get_purchase_manager_users(db, company_id)
    if not candidates:
        return None
    return random.choice(candidates)


def get_auto_assign_candidate_users(db: Session, company_id: Optional[int]) -> List[models.User]:
    return lead_assignment.get_auto_assign_candidate_users(db, company_id)


def choose_auto_assign_user(db: Session, company_id: Optional[int]) -> Optional[models.User]:
    return lead_assignment.choose_auto_assign_user(db, company_id)


def get_active_reassignment_candidates(
    db: Session,
    company_id: Optional[int],
    current_user: models.User,
    exclude_user_id: Optional[int] = None,
    advisor_only: bool = False,
    auto_assign_only: bool = False,
) -> List[models.User]:
    if not company_id:
        return []

    candidates = db.query(models.User).join(models.Role, isouter=True).filter(
        models.User.company_id == company_id
    ).all()

    valid_users = []
    for user in candidates:
        if exclude_user_id and user.id == exclude_user_id:
            continue
        if not is_active_user(user):
            continue
        if not user.role:
            continue
        if advisor_only and get_user_role_name(user) != "asesor":
            continue
        if auto_assign_only and not can_user_receive_auto_assigned_leads(user):
            continue
        if not can_assign_lead_to_user(current_user, user):
            continue
        valid_users.append(user)

    return valid_users


def get_active_advisor_users_for_redistribution(
    db: Session,
    company_id: Optional[int],
    exclude_user_id: Optional[int] = None
) -> List[models.User]:
    if not company_id:
        return []

    candidates = db.query(models.User).join(models.Role, isouter=True).filter(
        models.User.company_id == company_id
    ).all()

    valid_users = []
    for user in candidates:
        if exclude_user_id and user.id == exclude_user_id:
            continue
        if not can_user_receive_auto_assigned_leads(user):
            continue
        valid_users.append(user)

    return valid_users


def should_self_assign_manual_lead(current_user: models.User) -> bool:
    return get_user_role_name(current_user) == "asesor"


def should_auto_assign_manual_lead(current_user: models.User) -> bool:
    return get_user_role_name(current_user) in {"admin", "super_admin"}


def normalize_supervisor_ids(supervisor_ids: Optional[List[int]]) -> List[int]:
    normalized = []
    for raw_id in supervisor_ids or []:
        try:
            user_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if user_id not in normalized:
            normalized.append(user_id)
    return normalized


def validate_supervisors(
    db: Session,
    company_id: Optional[int],
    supervisor_ids: Optional[List[int]]
) -> List[models.User]:
    normalized_ids = normalize_supervisor_ids(supervisor_ids)
    if not normalized_ids:
        return []

    supervisors = db.query(models.User).filter(models.User.id.in_(normalized_ids)).all()
    if len(supervisors) != len(normalized_ids):
        raise HTTPException(status_code=400, detail="Uno o más supervisores no existen")

    for supervisor in supervisors:
        if company_id and supervisor.company_id != company_id:
            raise HTTPException(status_code=400, detail="Todos los supervisores deben pertenecer a la misma empresa")
        if not is_active_user(supervisor):
            raise HTTPException(status_code=400, detail="No puedes agregar supervisores inactivos")

    supervisors_by_id = {user.id: user for user in supervisors}
    return [supervisors_by_id[user_id] for user_id in normalized_ids if user_id in supervisors_by_id]


def sync_lead_supervisors(
    db: Session,
    lead: models.Lead,
    supervisor_ids: Optional[List[int]],
    assigned_by_id: Optional[int] = None
) -> List[models.User]:
    target_supervisors = validate_supervisors(db, lead.company_id, supervisor_ids)
    target_ids = {user.id for user in target_supervisors}

    existing_links = list(lead.supervisor_links or [])
    for link in existing_links:
        if link.user_id not in target_ids:
            db.delete(link)

    existing_ids = {link.user_id for link in existing_links}
    for supervisor in target_supervisors:
        if supervisor.id in existing_ids:
            continue
        db.add(models.LeadSupervisor(
            lead_id=lead.id,
            user_id=supervisor.id,
            assigned_by_id=assigned_by_id
        ))

    db.flush()
    db.refresh(lead)
    return target_supervisors


def ensure_user_in_supervisors(supervisor_ids: Optional[List[int]], user_id: Optional[int]) -> List[int]:
    normalized_ids = normalize_supervisor_ids(supervisor_ids)
    if user_id and user_id not in normalized_ids:
        normalized_ids.append(user_id)
    return normalized_ids


def maybe_assign_credit_coordinator(
    db: Session,
    lead_company_id: Optional[int],
    previous_status: Optional[str],
    target_status: Optional[str],
    assigned_to_id: Optional[int],
    supervisor_ids: Optional[List[int]],
    current_user_id: Optional[int] = None,
    previous_assigned_to_id: Optional[int] = None
):
    next_supervisor_ids = normalize_supervisor_ids(supervisor_ids)
    if not company_has_enabled_module("credits", company_id=lead_company_id, db=db):
        return assigned_to_id, next_supervisor_ids, None

    if normalize_lead_status_value(target_status) != models.LeadStatus.CREDIT_STUDY.value:
        return assigned_to_id, next_supervisor_ids, None

    if normalize_lead_status_value(previous_status) == models.LeadStatus.CREDIT_STUDY.value:
        return assigned_to_id, next_supervisor_ids, None

    # Preserve traceability when the lead enters the credit flow:
    # keep the current responsible user, the previous responsible user,
    # and the user performing the transition inside the lead team.
    next_supervisor_ids = ensure_user_in_supervisors(next_supervisor_ids, assigned_to_id)
    next_supervisor_ids = ensure_user_in_supervisors(next_supervisor_ids, previous_assigned_to_id)
    next_supervisor_ids = ensure_user_in_supervisors(next_supervisor_ids, current_user_id)

    coordinator = choose_credit_coordinator(db, lead_company_id)
    if not coordinator:
        return assigned_to_id, next_supervisor_ids, None

    next_supervisor_ids = ensure_user_in_supervisors(next_supervisor_ids, coordinator.id)
    return assigned_to_id, next_supervisor_ids, coordinator


def get_effective_lead_assignment_for_update(
    db: Session,
    lead: models.Lead,
    lead_update: schemas.LeadUpdate,
    current_user: models.User
):
    lead_company = get_company_by_id(db, lead.company_id)
    previous_assigned_user = lead.assigned_to
    previous_role_name = get_user_role_name(previous_assigned_user)
    target_assigned_user = previous_assigned_user
    current_supervisor_ids = normalize_supervisor_ids(getattr(lead, "supervisor_ids", []))
    target_supervisor_ids = (
        normalize_supervisor_ids(lead_update.supervisor_ids)
        if lead_update.supervisor_ids is not None
        else current_supervisor_ids
    )

    manual_assignment_requested = lead_update.assigned_to_id is not None
    target_assigned_to_id = lead.assigned_to_id

    if not manual_assignment_requested and not is_valid_lead_assignee(previous_assigned_user, lead.company_id):
        recovered_user = find_fallback_lead_assignee(db, lead, preferred_user=current_user)
        target_assigned_user = recovered_user
        target_assigned_to_id = recovered_user.id if recovered_user else None

    if manual_assignment_requested:
        if lead_update.assigned_to_id:
            target_user = db.query(models.User).join(models.Role, isouter=True).filter(
                models.User.id == lead_update.assigned_to_id
            ).first()
            if not target_user or target_user.company_id != lead.company_id:
                raise HTTPException(status_code=400, detail="Invalid assigned user (not in company)")
            if not is_active_user(target_user):
                raise HTTPException(status_code=400, detail="No puedes reasignar el lead a un usuario inactivo")
            if not can_assign_lead_to_user(current_user, target_user):
                raise HTTPException(status_code=400, detail="Tu rol no tiene permitido reasignar leads a este rol")
            target_assigned_user = target_user
            target_assigned_to_id = target_user.id
        else:
            target_assigned_user = None
            target_assigned_to_id = None

    base_status = lead_update.status if lead_update.status is not None else lead.status
    target_role_name = get_user_role_name(target_assigned_user)
    board_transfer = should_reset_status_for_board_transfer(previous_role_name, target_role_name)
    effective_status = enforce_ally_managed_status(
        db=db,
        current_user=current_user,
        base_status=base_status,
        assigned_to_id=target_assigned_to_id
    )
    effective_status = normalize_company_lead_status(
        effective_status,
        lead_company,
        models.LeadStatus.NEW.value
    )
    if board_transfer:
        effective_status = models.LeadStatus.NEW.value

    target_assigned_to_id, target_supervisor_ids, credit_coordinator = maybe_assign_credit_coordinator(
        db=db,
        lead_company_id=lead.company_id,
        previous_status=lead.status,
        target_status=effective_status,
        assigned_to_id=target_assigned_to_id,
        supervisor_ids=target_supervisor_ids,
        current_user_id=current_user.id,
        previous_assigned_to_id=getattr(previous_assigned_user, "id", None)
    )

    if manual_assignment_requested and should_keep_assigner_as_supervisor(current_user.id, target_assigned_to_id):
        target_supervisor_ids = ensure_user_in_supervisors(target_supervisor_ids, current_user.id)

    return {
        "previous_assigned_user": previous_assigned_user,
        "target_assigned_user": target_assigned_user,
        "target_assigned_to_id": target_assigned_to_id,
        "target_supervisor_ids": target_supervisor_ids,
        "effective_status": effective_status,
        "board_transfer": board_transfer,
        "credit_coordinator": credit_coordinator,
        "manual_assignment_requested": manual_assignment_requested,
        "target_role_name": target_role_name,
    }


def should_keep_assigner_as_supervisor(
    current_user_id: Optional[int],
    target_assigned_to_id: Optional[int]
) -> bool:
    return bool(current_user_id and target_assigned_to_id and current_user_id != target_assigned_to_id)


def should_supervise_manual_lead(
    current_user_id: Optional[int],
    target_assigned_to_id: Optional[int]
) -> bool:
    return bool(current_user_id and current_user_id != target_assigned_to_id)


def should_reset_status_for_board_transfer(previous_role_name: Optional[str], target_role_name: Optional[str]) -> bool:
    return (previous_role_name == "aliado") != (target_role_name == "aliado")


def validate_interested_process_detail(
    target_status: Optional[str],
    process_detail_data: Optional[schemas.LeadProcessDetailCreate],
    existing_detail: Optional[models.LeadProcessDetail] = None,
):
    normalized_status = normalize_lead_status_value(target_status)
    if normalized_status == models.LeadStatus.IN_PROCESS.value:
        has_vehicle = process_detail_data.has_vehicle if process_detail_data is not None else getattr(existing_detail, "has_vehicle", None)
        vehicle_id = process_detail_data.vehicle_id if process_detail_data is not None else getattr(existing_detail, "vehicle_id", None)
        desired_vehicle = process_detail_data.desired_vehicle if process_detail_data is not None else getattr(existing_detail, "desired_vehicle", None)
        desired_vehicle = (desired_vehicle or "").strip()

        if has_vehicle is None:
            raise HTTPException(
                status_code=400,
                detail="Debes indicar si el vehículo está disponible en inventario o si toca conseguirlo."
            )

        if has_vehicle and not vehicle_id:
            raise HTTPException(status_code=400, detail="Debes seleccionar un vehículo disponible del inventario.")

        if not has_vehicle and not desired_vehicle:
            raise HTTPException(status_code=400, detail="Debes indicar qué vehículo busca el cliente.")
        return

    if normalized_status == models.LeadStatus.RESERVED.value:
        reservation_amount = (
            process_detail_data.reservation_amount if process_detail_data is not None
            else getattr(existing_detail, "reservation_amount", None)
        )
        payment_method = (
            process_detail_data.reservation_payment_method if process_detail_data is not None
            else getattr(existing_detail, "reservation_payment_method", None)
        )
        try:
            reservation_amount = int(reservation_amount) if reservation_amount is not None else None
        except (TypeError, ValueError):
            reservation_amount = None

        payment_method = (payment_method or "").strip().lower()
        if reservation_amount is None or reservation_amount <= 0:
            raise HTTPException(
                status_code=400,
                detail="Debes indicar el monto de la separación para pasar el lead a Reservas."
            )
        if payment_method not in {"efectivo", "transferencia"}:
            raise HTTPException(
                status_code=400,
                detail="Debes indicar si la separación fue en efectivo o transferencia."
            )


def normalize_interested_process_detail(
    target_status: Optional[str],
    process_detail_data: Optional[schemas.LeadProcessDetailCreate],
    existing_detail: Optional[models.LeadProcessDetail],
    lead: Optional[models.Lead],
) -> Optional[schemas.LeadProcessDetailCreate]:
    normalized_status = normalize_lead_status_value(target_status)
    if process_detail_data is None:
        return process_detail_data

    if normalized_status == models.LeadStatus.RESERVED.value:
        if process_detail_data.reservation_payment_method:
            process_detail_data.reservation_payment_method = process_detail_data.reservation_payment_method.strip().lower()
        return process_detail_data

    if normalized_status != models.LeadStatus.IN_PROCESS.value:
        return process_detail_data

    if process_detail_data.has_vehicle is True and not process_detail_data.vehicle_id:
        fallback_desired_vehicle = (
            (process_detail_data.desired_vehicle or "").strip()
            or (getattr(existing_detail, "desired_vehicle", None) or "").strip()
            or (getattr(lead, "message", None) or "").strip()
            or "Por definir"
        )
        process_detail_data.has_vehicle = False
        process_detail_data.vehicle_id = None
        process_detail_data.desired_vehicle = fallback_desired_vehicle

    return process_detail_data


def build_lead_board_link(target_user: Optional[models.User], lead_id: int) -> str:
    target_role_name = get_user_role_name(target_user)
    base_path = "/aliado/dashboard" if target_role_name == "aliado" else "/admin/leads"
    return f"{base_path}?leadId={lead_id}"


def parse_json_list(value: Optional[str], fallback: Optional[List[str]] = None) -> List[str]:
    if not value:
        return fallback[:] if fallback else []
    try:
        parsed = json.loads(value)
        if isinstance(parsed, list):
            return [str(item) for item in parsed]
    except Exception:
        pass
    return fallback[:] if fallback else []


def sanitize_view_ids(view_ids: Optional[List[str]], company_id: Optional[int] = None) -> List[str]:
    if not view_ids:
        return []
    sanitized = []
    for view_id in view_ids:
        normalized = str(view_id)
        if normalized not in VALID_VIEW_IDS:
            continue
        if company_id and normalized not in COMPANY_VIEW_IDS:
            continue
        if normalized not in sanitized:
            sanitized.append(normalized)
    return sanitized


def ensure_role_view_defaults_synced():
    """
    Keep existing admin/super-admin roles aligned with newly introduced views.
    """
    try:
        with Session(engine) as db:
            roles = db.query(models.Role).all()
            updated = False

            for role in roles:
                effective_role_name = getattr(role, "base_role_name", None) or getattr(role, "name", None)
                if effective_role_name not in {"admin", "super_admin", "asesor", "gestion_creditos", "aliado", "compras", "inventario", "user"}:
                    continue

                default_permissions = DEFAULT_ROLE_VIEW_ACCESS.get(effective_role_name, [])
                permissions = sanitize_view_ids(
                    parse_json_list(getattr(role, "permissions_json", None), default_permissions),
                    getattr(role, "company_id", None)
                )
                menu_order = sanitize_view_ids(
                    parse_json_list(getattr(role, "menu_order_json", None), DEFAULT_ROLE_MENU_ORDER.get(effective_role_name, permissions)),
                    getattr(role, "company_id", None)
                )

                role_changed = False
                for view_id in default_permissions:
                    if view_id not in permissions:
                        permissions.append(view_id)
                        role_changed = True
                    if view_id not in menu_order:
                        menu_order.append(view_id)
                        role_changed = True

                if role_changed:
                    role.permissions_json = json.dumps(permissions)
                    role.menu_order_json = json.dumps(menu_order)
                    updated = True

            if updated:
                db.commit()
    except Exception as exc:
        print(f"Warning: could not sync role view defaults: {exc}", flush=True)


def serialize_role(role: models.Role) -> schemas.Role:
    effective_role_name = getattr(role, "base_role_name", None) or role.name
    default_permissions = DEFAULT_ROLE_VIEW_ACCESS.get(effective_role_name, [])
    permissions = sanitize_view_ids(parse_json_list(getattr(role, "permissions_json", None), default_permissions), getattr(role, "company_id", None))
    menu_order = sanitize_view_ids(parse_json_list(getattr(role, "menu_order_json", None), DEFAULT_ROLE_MENU_ORDER.get(effective_role_name, permissions)), getattr(role, "company_id", None))

    for view_id in permissions:
        if view_id not in menu_order:
            menu_order.append(view_id)

    return schemas.Role(
        id=role.id,
        name=role.name,
        label=role.label,
        permissions=permissions,
        menu_order=menu_order,
        auto_assign_leads=bool(getattr(role, "auto_assign_leads", False) or effective_role_name == "asesor"),
        assignable_role_ids=get_role_assignable_role_ids(role),
        advisor_tracking_enabled=bool(getattr(role, "advisor_tracking_enabled", False)),
        tracked_advisor_ids=[],
        company_id=getattr(role, "company_id", None),
        is_system=bool(getattr(role, "is_system", False)),
        base_role_name=getattr(role, "base_role_name", None)
    )


def ensure_role_management_permissions(current_user: models.User):
    role_name = get_user_role_name(current_user) or ""
    if role_name not in {"admin", "super_admin"}:
        raise HTTPException(status_code=403, detail="Not authorized")


def ensure_public_credit_review_permissions(current_user: models.User):
    role_name = get_user_role_name(current_user) or ""
    if role_name not in {"admin", "super_admin", "gestion_creditos"}:
        raise HTTPException(status_code=403, detail="No autorizado para revisar solicitudes públicas de crédito")


def ensure_public_credit_delete_permissions(current_user: models.User):
    role_name = get_user_role_name(current_user) or ""
    if role_name not in {"admin", "super_admin"}:
        raise HTTPException(status_code=403, detail="Solo un administrador puede eliminar solicitudes públicas de crédito")


ensure_role_view_defaults_synced()


def get_company_role_override(db: Session, company_id: Optional[int], base_role_name: Optional[str]) -> Optional[models.Role]:
    if not company_id or not base_role_name:
        return None
    return db.query(models.Role).filter(
        models.Role.company_id == company_id,
        models.Role.base_role_name == base_role_name
    ).first()


def resolve_assignable_role(db: Session, target_role: models.Role, company_id: Optional[int]) -> models.Role:
    if company_id and target_role and target_role.is_system:
        override = get_company_role_override(db, company_id, target_role.name)
        if override:
            return override
    return target_role


def get_module_role_fallback(db: Session, company_id: Optional[int]) -> models.Role:
    fallback_role = db.query(models.Role).filter(models.Role.name == MODULE_ROLE_FALLBACK_NAME).first()
    if not fallback_role:
        fallback_role = models.Role(
            name=MODULE_ROLE_FALLBACK_NAME,
            label="Usuario Basico",
            is_system=True,
            permissions_json=json.dumps(DEFAULT_ROLE_VIEW_ACCESS.get(MODULE_ROLE_FALLBACK_NAME, [])),
            menu_order_json=json.dumps(DEFAULT_ROLE_MENU_ORDER.get(MODULE_ROLE_FALLBACK_NAME, [])),
            auto_assign_leads=False,
        )
        db.add(fallback_role)
        db.flush()
    return resolve_assignable_role(db, fallback_role, company_id)


def get_users_with_disabled_module_roles(db: Session, company: models.Company) -> List[models.User]:
    if not company:
        return []

    users = (
        db.query(models.User)
        .options(joinedload(models.User.role))
        .filter(models.User.company_id == company.id)
        .all()
    )
    return [
        user for user in users
        if user.role and not is_role_enabled_for_company(user.role, company)
    ]


def remediate_users_with_disabled_module_roles(
    db: Session,
    company: models.Company,
    *,
    actor_user_id: Optional[int] = None,
    source: str = "company_modules_update",
) -> List[dict]:
    affected_users = get_users_with_disabled_module_roles(db, company)
    if not affected_users:
        return []

    fallback_role = get_module_role_fallback(db, company.id)
    enabled_modules = get_company_enabled_modules(company)
    remediated = []

    for affected_user in affected_users:
        old_role = affected_user.role
        old_role_id = affected_user.role_id
        old_role_name = get_effective_role_key(old_role)
        disabled_modules = get_role_disabled_required_modules(old_role, company)

        affected_user.role_id = fallback_role.id
        affected_user.auto_assign_leads = False

        details = {
            "source": source,
            "company_id": company.id,
            "user_id": affected_user.id,
            "email": affected_user.email,
            "old_role_id": old_role_id,
            "old_role_name": old_role_name,
            "old_role_label": getattr(old_role, "label", None),
            "new_role_id": fallback_role.id,
            "new_role_name": get_effective_role_key(fallback_role),
            "disabled_required_modules": disabled_modules,
            "enabled_modules": enabled_modules,
        }
        db.add(models.SystemLog(
            company_id=company.id,
            user_id=actor_user_id,
            action="MODULE_ROLE_FALLBACK",
            entity_type="User",
            entity_id=affected_user.id,
            details=json.dumps(details, ensure_ascii=True),
        ))
        remediated.append(details)

    return remediated


def serialize_user(user: models.User, is_online: Optional[bool] = None) -> dict:
    payload = schemas.User.model_validate(user, from_attributes=True).model_dump()
    payload["role"] = serialize_role(user.role).model_dump() if user.role else None
    company = getattr(user, "company", None)
    if payload.get("role"):
        payload["role"]["permissions"] = apply_company_module_limits(payload["role"].get("permissions", []), company)
        payload["role"]["menu_order"] = apply_company_module_limits(payload["role"].get("menu_order", []), company)
    if payload.get("company"):
        payload["company"]["enabled_modules"] = get_company_enabled_modules(company)
    payload["auto_assign_leads"] = bool(getattr(user, "auto_assign_leads", False) and is_advisor_role(getattr(user, "role", None)))
    payload["tracked_advisor_ids"] = get_user_tracked_advisor_ids(user)
    if is_online is not None:
        payload["is_online"] = is_online
    return payload


def normalize_ecard_slug(value: Optional[str]) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    ascii_value = "".join(char for char in normalized if not unicodedata.combining(char))
    ascii_value = re.sub(r"[^a-z0-9]+", "-", ascii_value).strip("-")
    return ascii_value[:90] or "equipo"


def ensure_unique_ecard_slug(
    db: Session,
    company_id: Optional[int],
    requested_slug: Optional[str],
    fallback_text: Optional[str],
    user_id: Optional[int] = None,
) -> str:
    base_slug = normalize_ecard_slug(requested_slug or fallback_text)
    slug = base_slug
    suffix = 2
    while True:
        query = db.query(models.User).filter(models.User.ecard_slug == slug)
        if company_id is not None:
            query = query.filter(models.User.company_id == company_id)
        else:
            query = query.filter(models.User.company_id.is_(None))
        if user_id is not None:
            query = query.filter(models.User.id != user_id)
        if not query.first():
            return slug
        slug = f"{base_slug[:84]}-{suffix}"
        suffix += 1


def build_ecard_fallback_text(user: models.User | schemas.UserBase) -> str:
    return getattr(user, "full_name", None) or getattr(user, "email", None) or "equipo"


def ensure_user_management_scope(current_user: models.User, target_user: models.User):
    effective_role_name = get_user_role_name(current_user)
    if effective_role_name not in {"super_admin", "admin"}:
        raise HTTPException(status_code=403, detail="No tienes permisos para administrar usuarios")
    if current_user.company_id and effective_role_name != "super_admin" and target_user.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="No tienes permisos para editar usuarios de otra empresa")


def serialize_company(company: models.Company) -> dict:
    payload = schemas.Company.model_validate(company, from_attributes=True).model_dump()
    payload["enabled_modules"] = get_company_enabled_modules(company)
    return payload

def upsert_purchase_request_from_lead(db: Session, lead: models.Lead):
    """
    Create/update a purchase request for purchase managers when a lead reaches
    reservation/purchase flow and is marked as "no vehicle in inventory" with a desired vehicle.
    """
    if not lead or not company_has_enabled_module("purchase_board", company_id=getattr(lead, "company_id", None), db=db):
        return

    normalized_status = normalize_lead_status_value(getattr(lead, "status", None))
    if normalized_status not in {
        models.LeadStatus.RESERVED.value,
        models.LeadStatus.PREPARATION.value,
        models.LeadStatus.SOLD.value,
    }:
        return

    detail = db.query(models.LeadProcessDetail).filter(models.LeadProcessDetail.lead_id == lead.id).first()
    if not detail:
        return
    if detail.has_vehicle:
        return

    desired_vehicle = normalize_credit_desired_vehicle((detail.desired_vehicle or "").strip())
    if not desired_vehicle:
        return

    safe_phone = (lead.phone or "").strip() or f"lead-{lead.id}"
    existing_purchase_candidates = db.query(models.CreditApplication).filter(
        models.CreditApplication.company_id == lead.company_id,
        or_(
            models.CreditApplication.lead_id == lead.id,
            and_(
                models.CreditApplication.phone == safe_phone,
                models.CreditApplication.desired_vehicle == desired_vehicle,
                models.CreditApplication.status.in_([
                    models.CreditStatus.PENDING.value,
                    models.CreditStatus.IN_REVIEW.value,
                    models.CreditStatus.APPROVED.value
                ])
            )
        )
    ).order_by(
        models.CreditApplication.updated_at.desc(),
        models.CreditApplication.created_at.desc()
    ).all()
    existing_purchase = next(
        (record for record in existing_purchase_candidates if is_purchase_request_record(record)),
        None
    )

    note_parts = [
        "[PURCHASE_REQUEST]",
        f"Generado automáticamente desde lead #{lead.id} en estado {LEAD_STATUS_LABELS.get(normalized_status, normalized_status)}."
    ]
    note_parts.append(f"Vehículo por buscar: {desired_vehicle}.")
    if normalized_status == models.LeadStatus.RESERVED.value:
        reservation_amount = getattr(detail, "reservation_amount", None)
        reservation_payment_method = (getattr(detail, "reservation_payment_method", None) or "").strip().lower()
        if reservation_amount:
            note_parts.append(f"Separación registrada: ${reservation_amount:,.0f}".replace(",", "."))
        if reservation_payment_method:
            note_parts.append(f"Medio de pago de separación: {reservation_payment_method.capitalize()}.")
    auto_note = " ".join(part for part in note_parts if part)
    compras_users = get_purchase_manager_users(db, lead.company_id)
    assigned_purchase_user = choose_purchase_manager(db, lead.company_id)
    assigned_compras_id = assigned_purchase_user.id if assigned_purchase_user else None

    if existing_purchase:
        existing_purchase.lead_id = lead.id
        existing_purchase.client_name = lead.name or existing_purchase.client_name
        existing_purchase.email = lead.email or existing_purchase.email
        existing_purchase.company_id = lead.company_id
        existing_purchase.desired_vehicle = desired_vehicle or existing_purchase.desired_vehicle
        if assigned_compras_id and existing_purchase.assigned_to_id != assigned_compras_id:
            existing_purchase.assigned_to_id = assigned_compras_id
        if not existing_purchase.status:
            existing_purchase.status = models.CreditStatus.PENDING.value
        if not (existing_purchase.notes or "").strip():
            existing_purchase.notes = auto_note
        return

    new_credit = models.CreditApplication(
        lead_id=lead.id,
        client_name=lead.name or f"Lead {lead.id}",
        phone=safe_phone,
        email=lead.email,
        desired_vehicle=desired_vehicle,
        monthly_income=0,
        other_income=0,
        occupation="employee",
        application_mode="individual",
        down_payment=0,
        status=models.CreditStatus.PENDING.value,
        notes=auto_note,
        company_id=lead.company_id,
        assigned_to_id=assigned_compras_id
    )
    db.add(new_credit)

    for compras_user in compras_users:
        db.add(models.Notification(
            user_id=compras_user.id,
            title="Nuevo vehículo por buscar",
            message=f"{lead.name or 'Lead'} busca: {desired_vehicle}",
            type="warning",
            link="/admin/purchases"
        ))


def upsert_credit_application_from_lead(
    db: Session,
    lead: models.Lead,
    comment: Optional[str] = None
):
    """
    Create/update a credit application when the lead enters credit request stage.
    This keeps the same lead visible in /creditos without creating duplicates.
    """
    if not lead or not company_has_enabled_module("credits", company_id=getattr(lead, "company_id", None), db=db):
        return
    if normalize_lead_status_value(lead.status) != models.LeadStatus.CREDIT_STUDY.value:
        return

    detail = db.query(models.LeadProcessDetail).filter(
        models.LeadProcessDetail.lead_id == lead.id
    ).first()
    vehicle_context = get_lead_credit_vehicle_context(db, lead, detail)
    desired_vehicle = vehicle_context["vehicle"]

    note_parts = ["[CREDIT_REQUEST]", f"Generado autom?ticamente desde lead #{lead.id}."]
    note_parts.append(f"{vehicle_context['mode'].capitalize()}: {desired_vehicle}.")
    if comment and comment.strip():
        note_parts.append(f"Seguimiento: {comment.strip()}")
    auto_note = " ".join(note_parts)
    credit_assignee_id = lead.assigned_to_id
    if not (lead.assigned_to and is_credit_coordinator_role(getattr(lead.assigned_to, "role", None))):
        for supervisor in getattr(lead, "supervisors", []) or []:
            if is_active_user(supervisor) and is_credit_coordinator_role(getattr(supervisor, "role", None)):
                credit_assignee_id = supervisor.id
                break
        else:
            credit_coordinator = choose_credit_coordinator(db, lead.company_id)
            if credit_coordinator:
                credit_assignee_id = credit_coordinator.id

    existing_credit_candidates = db.query(models.CreditApplication).filter(
        models.CreditApplication.company_id == lead.company_id,
        models.CreditApplication.lead_id == lead.id
    ).order_by(
        models.CreditApplication.updated_at.desc(),
        models.CreditApplication.created_at.desc()
    ).all()
    existing_credit = next(
        (record for record in existing_credit_candidates if not is_purchase_request_record(record)),
        None
    )

    if existing_credit:
        existing_credit.client_name = lead.name or existing_credit.client_name
        existing_credit.phone = (lead.phone or "").strip() or existing_credit.phone
        existing_credit.email = lead.email or existing_credit.email
        existing_credit.desired_vehicle = desired_vehicle or existing_credit.desired_vehicle
        existing_credit.assigned_to_id = credit_assignee_id or existing_credit.assigned_to_id
        existing_credit.notes = auto_note
        if not existing_credit.status:
            existing_credit.status = models.CreditStatus.PENDING.value
        return

    db.add(models.CreditApplication(
        lead_id=lead.id,
        client_name=lead.name or f"Lead {lead.id}",
        phone=(lead.phone or "").strip() or f"lead-{lead.id}",
        email=lead.email,
        desired_vehicle=desired_vehicle,
        monthly_income=0,
        other_income=0,
        occupation="employee",
        application_mode="individual",
        down_payment=0,
        status=models.CreditStatus.PENDING.value,
        notes=auto_note,
        company_id=lead.company_id,
        assigned_to_id=credit_assignee_id
    ))


def upsert_credit_approval_data_for_lead(
    db: Session,
    lead: models.Lead,
    lead_update: schemas.LeadUpdate,
):
    if not company_has_enabled_module("credits", company_id=getattr(lead, "company_id", None), db=db):
        return

    approved_amount = lead_update.approved_amount
    approval_percentage = lead_update.approval_percentage
    approved_down_payment = lead_update.approved_down_payment

    if approved_amount is None and approval_percentage is None and approved_down_payment is None:
        return

    related_credit_records = db.query(models.CreditApplication).filter(
        models.CreditApplication.lead_id == lead.id
    ).order_by(
        models.CreditApplication.updated_at.desc(),
        models.CreditApplication.created_at.desc()
    ).all()
    related_credit = next(
        (record for record in related_credit_records if not is_purchase_request_record(record)),
        None
    )

    detail = db.query(models.LeadProcessDetail).filter(
        models.LeadProcessDetail.lead_id == lead.id
    ).first()
    vehicle_context = get_lead_credit_vehicle_context(db, lead, detail)
    desired_vehicle = vehicle_context["vehicle"]

    if not related_credit:
        credit_assignee_id = lead.assigned_to_id
        if not (lead.assigned_to and is_credit_coordinator_role(getattr(lead.assigned_to, "role", None))):
            for supervisor in getattr(lead, "supervisors", []) or []:
                if is_active_user(supervisor) and is_credit_coordinator_role(getattr(supervisor, "role", None)):
                    credit_assignee_id = supervisor.id
                    break
            else:
                credit_coordinator = choose_credit_coordinator(db, lead.company_id)
                if credit_coordinator:
                    credit_assignee_id = credit_coordinator.id

        related_credit = models.CreditApplication(
            lead_id=lead.id,
            client_name=lead.name or f"Lead {lead.id}",
            phone=(lead.phone or "").strip() or f"lead-{lead.id}",
            email=lead.email,
            desired_vehicle=desired_vehicle,
            monthly_income=0,
            other_income=0,
            occupation="employee",
            application_mode="individual",
            down_payment=0,
            status=models.CreditStatus.APPROVED.value,
            notes=f"[CREDIT_REQUEST] Actualizado desde lead #{lead.id}.",
            company_id=lead.company_id,
            assigned_to_id=credit_assignee_id
        )
        db.add(related_credit)
        db.flush()

    if approved_amount is not None:
        related_credit.approved_amount = approved_amount
    if approval_percentage is not None:
        related_credit.approval_percentage = approval_percentage
    if approved_down_payment is not None:
        related_credit.approved_down_payment = approved_down_payment

    normalized_lead_status = normalize_lead_status_value(lead.status)
    if normalized_lead_status in {
        models.LeadStatus.APPROVALS.value,
        models.LeadStatus.RESERVED.value,
        models.LeadStatus.PREPARATION.value,
    }:
        related_credit.status = models.CreditStatus.APPROVED.value



# --- USER ENDPOINTS ---

@app.get("/users/", response_model=schemas.UserList)
def read_users(
    skip: int = 0, 
    limit: int = 100, 
    q: str = None, 
    role_id: int = None,
    include_inactive: bool = False,
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    effective_role_name = get_user_role_name(current_user)
    today = datetime.datetime.utcnow()
    five_min_ago = today - datetime.timedelta(minutes=5)

    # Update current user last_active
    current_user.last_active = today
    db.commit()

    query = db.query(models.User)
    
    # If user belongs to a company, limit scope to that company specific users.
    # Global super admins without company_id can still inspect all companies.
    if current_user.company_id:
        query = query.filter(models.User.company_id == current_user.company_id)

    if not include_inactive:
        query = query.filter(or_(models.User.is_active == True, models.User.is_active.is_(None)))

    if q:
        query = query.filter(models.User.email.ilike(f"%{q}%"))
    
    if role_id:
        query = query.filter(models.User.role_id == role_id)
    
    total = query.count()
    users = query.offset(skip).limit(limit).all()

    # Compute is_online for each user
    serialized_users = []
    for user in users:
        is_online = bool(user.last_active and user.last_active > five_min_ago)
        serialized_users.append(serialize_user(user, is_online=is_online))

    return {"items": serialized_users, "total": total}

@app.post("/users/", response_model=schemas.User)
def create_user(user: schemas.UserCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    print(f"Creating user with payload: {user.dict()}") 
    effective_role_name = get_user_role_name(current_user)
    acting_company_id = None if effective_role_name == "super_admin" else current_user.company_id
    
    # Verify role exists
    role_obj = db.query(models.Role).filter(models.Role.id == user.role_id).first()
    if not role_obj:
        raise HTTPException(status_code=400, detail="Invalid Role ID")
    role_obj = resolve_assignable_role(db, role_obj, acting_company_id)
    if acting_company_id and role_obj.company_id not in {None, acting_company_id}:
        raise HTTPException(status_code=403, detail="No puedes usar roles de otra empresa")

    # Check permissions and scope
    if acting_company_id:
        user.company_id = acting_company_id
        # Prevent creating Super Admins
        if role_obj.name == "super_admin":
             raise HTTPException(status_code=403, detail="Company admins cannot create Super Users")

    target_company = get_company_by_id(db, user.company_id)
    if user.company_id and not target_company:
        raise HTTPException(status_code=400, detail="Empresa inválida")
    ensure_role_allowed_for_company(role_obj, target_company)
    
    # Check if user exists
    db_user = db.query(models.User).filter(models.User.email == user.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")

    ensure_company_user_limit(db, user.company_id)
    if user.company_id:
        ensure_company_active_account_limit(db, user.company_id)
    
    hashed_password = auth_utils.get_password_hash(user.password)
    auto_assign_leads = bool(user.auto_assign_leads) if is_advisor_role(role_obj) else False
    tracked_advisor_ids = sanitize_tracked_advisor_ids(
        db,
        user.tracked_advisor_ids if bool(getattr(role_obj, "advisor_tracking_enabled", False)) else [],
        user.company_id
    )
    ecard_slug = ensure_unique_ecard_slug(db, user.company_id, user.ecard_slug, build_ecard_fallback_text(user))
    new_user = models.User(
        email=user.email,
        full_name=user.full_name,
        hashed_password=hashed_password,
        role_id=role_obj.id, 
        company_id=user.company_id,
        auto_assign_leads=auto_assign_leads,
        tracked_advisor_ids_json=json.dumps(tracked_advisor_ids),
        ecard_enabled=bool(user.ecard_enabled),
        ecard_slug=ecard_slug,
        ecard_photo_url=user.ecard_photo_url,
        ecard_position=(user.ecard_position or "").strip() or None,
        commission_percentage=user.commission_percentage or 0,
        base_salary=user.base_salary,
        payment_dates=user.payment_dates
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    log_action_to_db(db, current_user.id, "CREATE", "User", new_user.id, f"Usuario creado: {new_user.email} con Rol ID {new_user.role_id}")
    
    return serialize_user(new_user)


# Keep this for backward compatibility or strict "me" endpoint
@app.get("/users/me", response_model=schemas.User)
async def read_users_me(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    ensure_company_license_allows_login(getattr(current_user, "company", None))
    if (
        getattr(current_user, "company", None)
        and getattr(current_user, "role", None)
        and not is_role_enabled_for_company(current_user.role, current_user.company)
    ):
        remediate_users_with_disabled_module_roles(
            db,
            current_user.company,
            actor_user_id=None,
            source="users_me_fallback",
        )
        db.commit()
        db.refresh(current_user)
    return serialize_user(current_user)

@app.get("/users/{user_id}", response_model=schemas.User)
def read_user(user_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    effective_role_name = get_user_role_name(current_user)
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    if current_user.company_id and effective_role_name != "super_admin" and user.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="No tienes permisos para ver usuarios de otra empresa")
    return serialize_user(user)

@app.put("/users/{user_id}", response_model=schemas.User)
def update_user(user_id: int, user_update: schemas.UserUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    print(f"Updating user {user_id} with: {user_update.dict()}") 
    effective_role_name = get_user_role_name(current_user)
    acting_company_id = None if effective_role_name == "super_admin" else current_user.company_id
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    if current_user.company_id and effective_role_name != "super_admin" and db_user.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="No tienes permisos para editar usuarios de otra empresa")
    resolved_role = getattr(db_user, "role", None)

    if user_update.email:
        db_user.email = user_update.email
    if user_update.password:
        db_user.hashed_password = auth_utils.get_password_hash(user_update.password)
    if user_update.role_id is not None:
        target_role = db.query(models.Role).filter(models.Role.id == user_update.role_id).first()
        if not target_role:
            raise HTTPException(status_code=400, detail="Invalid Role ID")
        target_role = resolve_assignable_role(db, target_role, acting_company_id)
        if acting_company_id and target_role.company_id not in {None, acting_company_id}:
            raise HTTPException(status_code=403, detail="No puedes usar roles de otra empresa")
        print(f"Setting role_id to: {target_role.id}") 
        db_user.role_id = target_role.id
        resolved_role = target_role
    if user_update.company_id is not None:
        if acting_company_id and user_update.company_id != acting_company_id:
            raise HTTPException(status_code=403, detail="No puedes mover usuarios a otra empresa")
        if user_update.company_id != db_user.company_id:
            ensure_company_user_limit(db, user_update.company_id, excluding_user_id=db_user.id)
            target_active_state = bool(user_update.is_active) if user_update.is_active is not None else bool(db_user.is_active)
            if target_active_state and user_update.company_id:
                ensure_company_active_account_limit(db, user_update.company_id, excluding_user_id=db_user.id)
        db_user.company_id = user_update.company_id

    target_company = get_company_by_id(db, db_user.company_id)
    if db_user.company_id and not target_company:
        raise HTTPException(status_code=400, detail="Empresa inválida")
    ensure_role_allowed_for_company(resolved_role, target_company)
    
    # Update new fields
    if user_update.full_name is not None:
        db_user.full_name = user_update.full_name
    if user_update.commission_percentage is not None:
        db_user.commission_percentage = user_update.commission_percentage
    if user_update.base_salary is not None:
        db_user.base_salary = user_update.base_salary
    if user_update.payment_dates is not None:
        db_user.payment_dates = user_update.payment_dates
    if user_update.auto_assign_leads is not None:
        db_user.auto_assign_leads = bool(user_update.auto_assign_leads) if is_advisor_role(resolved_role) else False
    elif not is_advisor_role(resolved_role):
        db_user.auto_assign_leads = False
    if user_update.tracked_advisor_ids is not None:
        db_user.tracked_advisor_ids_json = json.dumps(
            sanitize_tracked_advisor_ids(
                db,
                user_update.tracked_advisor_ids if bool(getattr(resolved_role, "advisor_tracking_enabled", False)) else [],
                db_user.company_id,
                excluded_user_id=db_user.id
            )
        )
    elif not bool(getattr(resolved_role, "advisor_tracking_enabled", False)):
        db_user.tracked_advisor_ids_json = json.dumps([])
    if user_update.ecard_enabled is not None:
        db_user.ecard_enabled = bool(user_update.ecard_enabled)
    if user_update.ecard_position is not None:
        db_user.ecard_position = (user_update.ecard_position or "").strip() or None
    if user_update.ecard_photo_url is not None:
        db_user.ecard_photo_url = (user_update.ecard_photo_url or "").strip() or None
    if user_update.ecard_slug is not None:
        db_user.ecard_slug = ensure_unique_ecard_slug(
            db,
            db_user.company_id,
            user_update.ecard_slug,
            build_ecard_fallback_text(db_user),
            user_id=db_user.id,
        )
    elif not db_user.ecard_slug:
        db_user.ecard_slug = ensure_unique_ecard_slug(
            db,
            db_user.company_id,
            None,
            build_ecard_fallback_text(db_user),
            user_id=db_user.id,
        )
    if user_update.is_active is not None:
        next_active = 1 if bool(user_update.is_active) else 0
        if next_active and not bool(db_user.is_active):
            ensure_company_active_account_limit(db, db_user.company_id, excluding_user_id=db_user.id)
        db_user.is_active = next_active
        
    try:
        db.commit()
        db.refresh(db_user)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    log_action_to_db(db, current_user.id, "UPDATE", "User", db_user.id, f"Usuario modificado: {db_user.email}")
        
    return serialize_user(db_user)


@app.post("/users/{user_id}/ecard-photo", response_model=schemas.User)
async def upload_user_ecard_photo(
    user_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    ensure_user_management_scope(current_user, db_user)

    allowed_extensions = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
    original_name = str(file.filename or "").strip()
    extension = os.path.splitext(original_name)[1].lower()
    content_type = str(file.content_type or "").lower()
    if extension not in allowed_extensions or not content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Solo se permiten imágenes JPG, PNG, WEBP o GIF")

    os.makedirs("static/user-ecards", exist_ok=True)
    safe_name = f"{uuid.uuid4().hex}{extension}"
    target_path = os.path.join("static", "user-ecards", safe_name)
    with open(target_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    db_user.ecard_photo_url = f"/api/static/user-ecards/{safe_name}"
    if not db_user.ecard_slug:
        db_user.ecard_slug = ensure_unique_ecard_slug(
            db,
            db_user.company_id,
            None,
            build_ecard_fallback_text(db_user),
            user_id=db_user.id,
        )
    db.commit()
    db.refresh(db_user)

    log_action_to_db(db, current_user.id, "UPLOAD_ECARD_PHOTO", "User", db_user.id, f"Foto de tarjeta actualizada: {db_user.email}")
    return serialize_user(db_user)


@app.delete("/users/{user_id}")
def delete_user(
    user_id: int,
    disable_request: Optional[schemas.UserDisableRequest] = Body(default=None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # 1. Check if the executing user has permission (Admin or Super Admin)
    role_obj = db.query(models.Role).filter(models.Role.id == current_user.role_id).first()
    effective_role_name = getattr(role_obj, "base_role_name", None) or getattr(role_obj, "name", None)
    if not role_obj or effective_role_name not in ["super_admin", "admin"]:
        raise HTTPException(status_code=403, detail="No tienes permisos para eliminar usuarios")
        
    # 2. Prevent self-deletion
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")

    # 3. Find user
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if not getattr(db_user, "is_active", True):
        raise HTTPException(status_code=400, detail="El usuario ya está inhabilitado")
        
    # 4. If current user is just 'admin', they can only delete users in their same company
    if effective_role_name == "admin":
        if db_user.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="No puedes eliminar usuarios de otra empresa")
            
        # Prevent 'admin' from deleting a 'super_admin'
        target_role = db.query(models.Role).filter(models.Role.id == db_user.role_id).first()
        target_role_name = getattr(target_role, "base_role_name", None) or getattr(target_role, "name", None)
        if target_role and target_role_name == "super_admin":
            raise HTTPException(status_code=403, detail="No puedes eliminar a un Súper Administrador")
            
    reassignment_target_id = disable_request.reassign_leads_to_user_id if disable_request else None
    assigned_leads_query = db.query(models.Lead).filter(
        models.Lead.assigned_to_id == user_id,
        models.Lead.deleted_at.is_(None)
    )
    assigned_leads_count = assigned_leads_query.count()

    replacement_user = None
    if reassignment_target_id is not None:
        replacement_user = db.query(models.User).filter(models.User.id == reassignment_target_id).first()
        if not replacement_user:
            raise HTTPException(status_code=404, detail="El usuario destino para reasignación no existe")
        if replacement_user.company_id != db_user.company_id:
            raise HTTPException(status_code=400, detail="Solo puedes reasignar los leads a un usuario de la misma empresa")
        if replacement_user.id == db_user.id:
            raise HTTPException(status_code=400, detail="Debes escoger un usuario diferente para la reasignación")
        if not getattr(replacement_user, "is_active", True):
            raise HTTPException(status_code=400, detail="El usuario destino está inhabilitado")
        if get_user_role_name(replacement_user) != "asesor":
            raise HTTPException(status_code=400, detail="Solo puedes reasignar leads a usuarios con rol Asesor / Vendedor")

    if assigned_leads_count > 0 and replacement_user is None:
        raise HTTPException(
            status_code=400,
            detail=f"Este usuario tiene {assigned_leads_count} lead(s) asignado(s). Debes seleccionar un usuario para reasignarlos antes de inhabilitarlo."
        )

    user_email_snapshot = db_user.email
            
    try:
        if replacement_user is not None:
            assigned_leads_query.update({"assigned_to_id": replacement_user.id}, synchronize_session=False)

        db_user.is_active = False
        db.commit()
        
        action_detail = f"Usuario inhabilitado: {user_email_snapshot}"
        if replacement_user is not None and assigned_leads_count > 0:
            action_detail += f". Leads reasignados a {replacement_user.full_name or replacement_user.email}"
        log_action_to_db(db, current_user.id, "DISABLE", "User", user_id, action_detail)
        
        return {
            "status": "success",
            "message": "Usuario inhabilitado correctamente",
            "reassigned_leads": assigned_leads_count if replacement_user is not None else 0,
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"No se pudo inhabilitar el usuario. {str(e)}")


@app.post("/users/{user_id}/redistribute-leads", response_model=schemas.UserLeadRedistributeResponse)
def redistribute_user_leads(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    role_obj = db.query(models.Role).filter(models.Role.id == current_user.role_id).first()
    effective_role_name = get_user_role_name(current_user)
    if not role_obj or effective_role_name not in ["super_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Solo administradores pueden redistribuir leads")

    source_user = db.query(models.User).join(models.Role, isouter=True).filter(
        models.User.id == user_id
    ).first()
    if not source_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    if current_user.company_id and source_user.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="No puedes redistribuir leads de otra empresa")

    if source_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="No puedes redistribuir tus propios leads desde esta accion")

    recipient_users = get_active_advisor_users_for_redistribution(
        db=db,
        company_id=source_user.company_id,
        exclude_user_id=source_user.id,
    )
    if not recipient_users:
        raise HTTPException(
            status_code=400,
            detail="No hay asesores o vendedores activos con asignación automática habilitada para redistribuir estos leads"
        )

    leads = db.query(models.Lead).options(
        joinedload(models.Lead.supervisors)
    ).filter(
        models.Lead.assigned_to_id == source_user.id,
        models.Lead.deleted_at.is_(None)
    ).all()

    if not leads:
        return {
            "status": "success",
            "message": "El usuario no tiene leads asignados para redistribuir",
            "redistributed_leads": 0,
            "recipient_users": len(recipient_users),
        }

    redistributed = 0
    for lead in leads:
        target_user = random.choice(recipient_users)
        previous_assigned_to_id = lead.assigned_to_id
        lead.assigned_to_id = target_user.id

        supervisor_ids = [
            supervisor.id
            for supervisor in (getattr(lead, "supervisors", None) or [])
            if (
                supervisor
                and supervisor.id != source_user.id
                and is_active_user(supervisor)
            )
        ]
        sync_lead_supervisors(db, lead, supervisor_ids, current_user.id)

        db.add(models.LeadHistory(
            lead_id=lead.id,
            user_id=current_user.id,
            previous_status=lead.status,
            new_status=lead.status,
            comment=(
                f"Lead redistribuido automaticamente desde "
                f"{source_user.full_name or source_user.email} hacia {target_user.full_name or target_user.email}"
            )
        ))
        db.add(models.Notification(
            user_id=target_user.id,
            title="Lead Reasignado",
            message=f"Se te reasigno el lead: {lead.name}. Continúa con la gestión.",
            type="info",
            link=build_lead_board_link(target_user, lead.id)
        ))
        redistributed += 1

    db.commit()
    log_action_to_db(
        db,
        current_user.id,
        "REDISTRIBUTE",
        "User",
        source_user.id,
        f"Redistribuidos {redistributed} lead(s) asignados a {source_user.email}"
    )

    return {
        "status": "success",
        "message": "Leads redistribuidos correctamente",
        "redistributed_leads": redistributed,
        "recipient_users": len(recipient_users),
    }

# --- SYSTEM LOGS ENDPOINTS ---

@app.get("/logs/", response_model=schemas.SystemLogList)
def read_system_logs(
    skip: int = 0, 
    limit: int = 50, 
    user_query: Optional[str] = Query(None),
    module_query: Optional[str] = Query(None),
    log_date: Optional[str] = Query(None),
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    """
    Get all system logs.
    Restricted to Admins and Super Admins.
    """
    role_obj = db.query(models.Role).filter(models.Role.id == current_user.role_id).first()
    effective_role_name = getattr(role_obj, "base_role_name", None) or getattr(role_obj, "name", None)
    if not role_obj or effective_role_name not in ["super_admin", "admin"]:
        raise HTTPException(status_code=403, detail="No tienes permisos para ver auditoría")

    query = db.query(models.SystemLog).outerjoin(models.User, models.SystemLog.user_id == models.User.id)

    if current_user.company_id:
        query = query.filter(
            or_(
                models.SystemLog.company_id == current_user.company_id,
                and_(
                    models.SystemLog.company_id.is_(None),
                    models.User.company_id == current_user.company_id
                )
            )
        )

    if user_query:
        like_value = f"%{user_query.strip()}%"
        query = query.filter(
            or_(
                models.User.full_name.ilike(like_value),
                models.User.email.ilike(like_value)
            )
        )

    if module_query:
        like_value = f"%{module_query.strip()}%"
        query = query.filter(
            or_(
                models.SystemLog.entity_type.ilike(like_value),
                models.SystemLog.action.ilike(like_value)
            )
        )

    if log_date:
        try:
            target_day = datetime.date.fromisoformat(log_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="La fecha debe usar formato YYYY-MM-DD")

        start = datetime.datetime.combine(target_day, datetime.time.min)
        end = datetime.datetime.combine(target_day + datetime.timedelta(days=1), datetime.time.min)
        query = query.filter(
            models.SystemLog.created_at >= start,
            models.SystemLog.created_at < end
        )
    
    total = query.count()
    items = query.options(joinedload(models.SystemLog.user)).order_by(models.SystemLog.created_at.desc()).offset(skip).limit(limit).all()
    
    return {"items": items, "total": total}

@app.get("/dashboard/stats", response_model=schemas.DashboardStats)
def get_dashboard_stats(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    companies_count = db.query(models.Company).count()
    users_count = db.query(models.User).count()
    return {"companies_count": companies_count, "users_count": users_count}

@app.get("/")
def read_root():
    return {"message": "Bienvenido a la API de AutosQP"}

@app.get("/health")
def health_check():
    return {"status": "ok"}

@app.post("/companies/", response_model=schemas.Company)
def create_company(company: schemas.CompanyCreate, db: Session = Depends(get_db)):
    db_company = db.query(models.Company).filter(models.Company.name == company.name).first()
    if db_company:
        raise HTTPException(status_code=400, detail="Company already registered")
    
    public_domains = normalize_public_domains(company.public_domains, company.public_domain)
    validate_company_public_domains(db, public_domains)

    new_company = models.Company(
        name=company.name,
        public_domain=public_domains[0] if public_domains else None,
        public_domains_json=json.dumps(public_domains),
        website_url=company.website_url,
        logo_url=company.logo_url,
        contact_address=company.contact_address,
        contact_phone=company.contact_phone,
        social_instagram=company.social_instagram,
        social_tiktok=company.social_tiktok,
        social_facebook=company.social_facebook,
        primary_color=company.primary_color,
        secondary_color=company.secondary_color,
        max_users=company.max_users,
        max_leads=company.max_leads,
        max_active_accounts=company.max_active_accounts,
        license_start_date=company.license_start_date,
        license_end_date=company.license_end_date,
        enabled_modules_json=json.dumps(normalize_company_enabled_modules(company.enabled_modules)),
        public_credit_requires_email_validation=bool(company.public_credit_requires_email_validation),
    )
    db.add(new_company)
    db.commit()
    db.refresh(new_company)
    return serialize_company(new_company)

@app.get("/companies/", response_model=schemas.CompanyList)
def read_companies(skip: int = 0, limit: int = 10, q: str = None, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    query = db.query(models.Company)
    if q:
        query = query.filter(models.Company.name.ilike(f"%{q}%"))
    
    total = query.count()
    companies = query.offset(skip).limit(limit).all()
    return {"items": [serialize_company(company) for company in companies], "total": total}

@app.get("/companies/{company_id}", response_model=schemas.Company)
def read_company(company_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    company = db.query(models.Company).filter(models.Company.id == company_id).first()
    if company is None:
        raise HTTPException(status_code=404, detail="Company not found")
    return serialize_company(company)

@app.put("/companies/{company_id}", response_model=schemas.Company)
def update_company(company_id: int, company_update: schemas.CompanyCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    db_company = db.query(models.Company).filter(models.Company.id == company_id).first()
    if not db_company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    public_domains = normalize_public_domains(company_update.public_domains, company_update.public_domain)
    validate_company_public_domains(db, public_domains, exclude_company_id=company_id)
    db_company.name = company_update.name
    db_company.public_domain = public_domains[0] if public_domains else None
    db_company.public_domains_json = json.dumps(public_domains)
    db_company.website_url = company_update.website_url
    db_company.logo_url = company_update.logo_url
    db_company.contact_address = company_update.contact_address
    db_company.contact_phone = company_update.contact_phone
    db_company.social_instagram = company_update.social_instagram
    db_company.social_tiktok = company_update.social_tiktok
    db_company.social_facebook = company_update.social_facebook
    db_company.primary_color = company_update.primary_color
    db_company.secondary_color = company_update.secondary_color
    db_company.max_users = company_update.max_users
    db_company.max_leads = company_update.max_leads
    db_company.max_active_accounts = company_update.max_active_accounts
    db_company.license_start_date = company_update.license_start_date
    db_company.license_end_date = company_update.license_end_date
    db_company.enabled_modules_json = json.dumps(normalize_company_enabled_modules(company_update.enabled_modules))
    db_company.public_credit_requires_email_validation = bool(company_update.public_credit_requires_email_validation)
    remediated_users = remediate_users_with_disabled_module_roles(
        db,
        db_company,
        actor_user_id=current_user.id,
        source="company_modules_update",
    )
    
    try:
        db.commit()
        db.refresh(db_company)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    if remediated_users:
        print(
            f"Remediated {len(remediated_users)} users with disabled module roles for company {db_company.id}",
            flush=True,
        )

    return serialize_company(db_company)

@app.get("/public/company-context", response_model=schemas.PublicCompanyContext)
def read_public_company_context(request: Request, db: Session = Depends(get_db)):
    company = resolve_public_company(db, request)
    return serialize_public_company(company)


@app.get("/public/team-cards/{slug}", response_model=schemas.PublicTeamCard)
def read_public_team_card(slug: str, request: Request, db: Session = Depends(get_db)):
    company = resolve_public_company(db, request)
    if not company:
        raise HTTPException(status_code=404, detail="No se encontró una empresa asociada a este dominio")

    normalized_slug = normalize_ecard_slug(slug)
    team_user = db.query(models.User).options(
        joinedload(models.User.role),
        joinedload(models.User.company),
    ).filter(
        models.User.company_id == company.id,
        models.User.ecard_slug == normalized_slug,
        models.User.ecard_enabled == True,
        or_(models.User.is_active == True, models.User.is_active == 1),
    ).first()
    if not team_user:
        raise HTTPException(status_code=404, detail="La tarjeta no está disponible")

    role_label = getattr(getattr(team_user, "role", None), "label", None)
    return schemas.PublicTeamCard(
        full_name=team_user.full_name or team_user.email,
        email=team_user.email,
        position=team_user.ecard_position or role_label,
        photo_url=team_user.ecard_photo_url,
        company=schemas.PublicTeamCardCompany(
            name=company.name or DEFAULT_PUBLIC_COMPANY_CONTEXT["name"],
            logo_url=company.logo_url,
            contact_address=getattr(company, "contact_address", None),
            contact_phone=getattr(company, "contact_phone", None),
            primary_color=company.primary_color or DEFAULT_PUBLIC_COMPANY_CONTEXT["primary_color"],
            secondary_color=company.secondary_color or DEFAULT_PUBLIC_COMPANY_CONTEXT["secondary_color"],
        ),
    )


@app.post("/public/credit-request/send-code", response_model=schemas.PublicCreditVerificationResponse)
def send_public_credit_verification_code(
    payload: schemas.PublicCreditVerificationSendRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    company = resolve_public_company(db, request)
    if not company:
        raise HTTPException(status_code=404, detail="No se encontró una empresa asociada a este dominio")
    if not company_has_enabled_module("public_credit_form", company=company):
        raise HTTPException(status_code=403, detail="El formulario de crédito no está habilitado para esta empresa")
    if not company_requires_public_credit_email_validation(company):
        return schemas.PublicCreditVerificationResponse(
            status="ok",
            message="Esta empresa no requiere código de validación por correo para enviar el formulario.",
            verified=True,
        )

    email_value = str(payload.email).strip().lower()
    verification_code = f"{random.randint(100000, 999999)}"
    now = datetime.datetime.now()

    db.query(models.PublicCreditEmailVerification).filter(
        models.PublicCreditEmailVerification.company_id == company.id,
        func.lower(models.PublicCreditEmailVerification.email) == email_value,
        models.PublicCreditEmailVerification.verified_at.is_(None),
    ).delete(synchronize_session=False)

    verification = models.PublicCreditEmailVerification(
        company_id=company.id,
        email=email_value,
        code=verification_code,
        expires_at=now + datetime.timedelta(minutes=10),
    )
    db.add(verification)
    db.commit()

    try:
        _send_public_credit_verification_email(email_value, db, company, verification_code)
    except HTTPException:
        db.delete(verification)
        db.commit()
        raise
    except Exception as exc:
        db.delete(verification)
        db.commit()
        raise HTTPException(status_code=500, detail=f"No se pudo enviar el correo de validación: {exc}")

    return schemas.PublicCreditVerificationResponse(
        status="ok",
        message="Se envió un código de verificación al correo indicado.",
        verified=False,
    )


@app.post("/public/credit-request/verify-code", response_model=schemas.PublicCreditVerificationResponse)
def verify_public_credit_verification_code(
    payload: schemas.PublicCreditVerificationVerifyRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    company = resolve_public_company(db, request)
    if not company:
        raise HTTPException(status_code=404, detail="No se encontró una empresa asociada a este dominio")
    if not company_requires_public_credit_email_validation(company):
        return schemas.PublicCreditVerificationResponse(
            status="ok",
            message="Esta empresa no requiere código de validación por correo.",
            verified=True,
        )

    email_value = str(payload.email).strip().lower()
    code_value = str(payload.code or "").strip()
    now = datetime.datetime.now()

    verification = db.query(models.PublicCreditEmailVerification).filter(
        models.PublicCreditEmailVerification.company_id == company.id,
        func.lower(models.PublicCreditEmailVerification.email) == email_value,
        models.PublicCreditEmailVerification.code == code_value,
    ).order_by(models.PublicCreditEmailVerification.created_at.desc()).first()

    if not verification:
        raise HTTPException(status_code=400, detail="El código de verificación no es válido.")
    if verification.expires_at < now:
        raise HTTPException(status_code=400, detail="El código de verificación ya expiró.")

    verification.verified_at = now
    db.commit()

    return schemas.PublicCreditVerificationResponse(
        status="ok",
        message="Correo validado correctamente.",
        verified=True,
    )


@app.get("/public/credit-request/access/{token}")
def read_public_credit_lead_access(
    token: str,
    request: Request,
    db: Session = Depends(get_db),
):
    company = resolve_public_company(db, request)
    if not company:
        raise HTTPException(status_code=404, detail="No se encontró una empresa asociada a este dominio")
    if not company_has_enabled_module("public_credit_form", company=company):
        raise HTTPException(status_code=403, detail="El formulario de crédito no está habilitado para esta empresa")

    access = _get_public_credit_lead_access(db, token, company_id=company.id)
    lead = access.lead
    if not lead:
        raise HTTPException(status_code=404, detail="El lead asociado a este acceso ya no existe.")
    submission = db.query(models.PublicCreditSubmission).filter(
        models.PublicCreditSubmission.lead_id == lead.id
    ).order_by(
        models.PublicCreditSubmission.updated_at.desc(),
        models.PublicCreditSubmission.id.desc(),
    ).first()

    return {
        "status": "ok",
        "lead_id": lead.id,
        "email": access.email,
        "applicant_name": lead.name,
        "verified": bool(access.verified_at),
        "requires_email_validation": company_requires_public_credit_email_validation(company),
        "expires_at": access.expires_at,
        "form_payload": _build_lead_credit_access_payload(db, lead),
        "attachments": submission.attachments if submission and isinstance(submission.attachments, dict) else {},
    }


@app.post("/public/credit-request/access/{token}/verify-code", response_model=schemas.PublicCreditVerificationResponse)
def verify_public_credit_lead_access_code(
    token: str,
    payload: schemas.PublicCreditVerificationVerifyRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    company = resolve_public_company(db, request)
    if not company:
        raise HTTPException(status_code=404, detail="No se encontró una empresa asociada a este dominio")
    access = _get_public_credit_lead_access(db, token, company_id=company.id)
    if not company_requires_public_credit_email_validation(company):
        if access.verified_at is None:
            access.verified_at = datetime.datetime.now()
            db.commit()
        return schemas.PublicCreditVerificationResponse(
            status="ok",
            message="Esta empresa no requiere código de validación por correo.",
            verified=True,
        )
    email_value = str(payload.email or "").strip().lower()
    code_value = str(payload.code or "").strip()
    if email_value != str(access.email or "").strip().lower():
        raise HTTPException(status_code=400, detail="El correo no corresponde a este acceso.")
    if code_value != str(access.code or "").strip():
        raise HTTPException(status_code=400, detail="El código no es válido.")

    access.verified_at = datetime.datetime.now()
    db.commit()
    return schemas.PublicCreditVerificationResponse(
        status="ok",
        message="Código validado correctamente.",
        verified=True,
    )


def _serialize_public_credit_capture_session(session: models.PublicCreditCaptureSession) -> schemas.PublicCreditCaptureSessionResponse:
    return schemas.PublicCreditCaptureSessionResponse(
        status="ok",
        token=session.token,
        side=session.side,
        uploaded=bool(session.file_path),
        file_url=session.file_path,
        file_type=session.file_type,
        original_file_name=session.original_file_name,
        expires_at=session.expires_at,
    )


def _get_public_credit_lead_access(
    db: Session,
    token: str,
    company_id: Optional[int] = None,
    require_valid: bool = True,
) -> models.PublicCreditLeadAccess:
    access = db.query(models.PublicCreditLeadAccess).options(
        joinedload(models.PublicCreditLeadAccess.lead)
    ).filter(
        models.PublicCreditLeadAccess.token == str(token or "").strip()
    ).first()
    if not access:
        raise HTTPException(status_code=404, detail="El acceso al formulario no existe o expiró.")
    if company_id and access.company_id != company_id:
        raise HTTPException(status_code=403, detail="El acceso no corresponde a esta empresa.")
    if require_valid and access.expires_at < datetime.datetime.now():
        raise HTTPException(status_code=400, detail="El acceso al formulario expiró. Solicita uno nuevo al asesor.")
    return access


def _build_lead_credit_access_payload(db: Session, lead: models.Lead) -> Dict[str, Any]:
    submission = db.query(models.PublicCreditSubmission).filter(
        models.PublicCreditSubmission.lead_id == lead.id
    ).order_by(
        models.PublicCreditSubmission.updated_at.desc(),
        models.PublicCreditSubmission.id.desc(),
    ).first()
    if submission and isinstance(submission.form_payload, dict):
        return submission.form_payload

    name_parts = str(lead.name or "").strip().split()
    first_name = " ".join(name_parts[:-1]) if len(name_parts) > 1 else (name_parts[0] if name_parts else "")
    last_name = name_parts[-1] if len(name_parts) > 1 else ""
    process_detail = db.query(models.LeadProcessDetail).filter(
        models.LeadProcessDetail.lead_id == lead.id
    ).first()
    desired_vehicle = str(
        (process_detail.desired_vehicle if process_detail else None)
        or lead.message
        or ""
    ).strip()
    return {
        "vehicle": {
            "vehicleValue": "",
            "requestedAmount": "",
            "requestDate": datetime.datetime.now(ZoneInfo("America/Bogota")).date().isoformat(),
            "make": desired_vehicle,
            "model": "",
            "vehicleType": "Automóvil",
            "label": desired_vehicle,
        },
        "personal": {
            "firstName": first_name,
            "lastName": last_name,
            "documentType": "C.C",
            "documentNumber": "",
            "issuePlace": "",
            "birthDate": "",
            "gender": "M",
            "profession": "",
            "birthPlace": "",
            "maritalStatus": "Soltero",
            "childrenCount": "Sin Hijos",
            "educationLevel": "Primaria",
            "livesWith": "Cónyuge",
            "housingType": "Familiar",
            "mobilePhone": lead.phone or "",
            "city": "",
            "address": "",
            "email": lead.email or "",
        },
        "employment": {},
        "income": {},
        "references": {},
        "consent": {
            "accepted": False,
            "signatureMode": "draw",
            "signatureName": lead.name or "",
            "verificationCode": "",
            "signatureDrawnDataUrl": "",
        },
    }


def _get_public_credit_capture_session(
    db: Session,
    token: str,
    company_id: Optional[int] = None,
    require_upload: bool = False,
) -> models.PublicCreditCaptureSession:
    session = db.query(models.PublicCreditCaptureSession).filter(
        models.PublicCreditCaptureSession.token == token
    ).first()

    if not session:
        raise HTTPException(status_code=404, detail="La sesión de captura no existe o expiró.")
    if company_id and session.company_id != company_id:
        raise HTTPException(status_code=403, detail="La captura no pertenece a esta empresa.")
    if session.expires_at < datetime.datetime.now():
        raise HTTPException(status_code=410, detail="La sesión de captura expiró. Genera un nuevo QR.")
    if require_upload and not session.file_path:
        raise HTTPException(status_code=400, detail="Aún no se ha cargado el archivo solicitado.")

    return session


@app.post("/public/credit-request/capture-session", response_model=schemas.PublicCreditCaptureSessionResponse)
def create_public_credit_capture_session(
    payload: schemas.PublicCreditCaptureSessionCreate,
    request: Request,
    db: Session = Depends(get_db),
):
    company = resolve_public_company(db, request)
    if not company:
        raise HTTPException(status_code=404, detail="No se encontró una empresa asociada a este dominio")
    if not company_has_enabled_module("public_credit_form", company=company):
        raise HTTPException(status_code=403, detail="El formulario de crédito no está habilitado para esta empresa")

    side = str(payload.side or "").strip().lower()
    if side not in {"front", "back", "signature"}:
        raise HTTPException(status_code=400, detail="El tipo de captura no es válido.")

    session = models.PublicCreditCaptureSession(
        company_id=company.id,
        token=secrets.token_urlsafe(32),
        side=side,
        expires_at=datetime.datetime.now() + datetime.timedelta(minutes=30),
    )
    db.add(session)
    db.commit()
    db.refresh(session)
    return _serialize_public_credit_capture_session(session)


@app.get("/public/credit-request/capture-session/{token}", response_model=schemas.PublicCreditCaptureSessionResponse)
def read_public_credit_capture_session(
    token: str,
    db: Session = Depends(get_db),
):
    session = _get_public_credit_capture_session(db, token)
    return _serialize_public_credit_capture_session(session)


@app.post("/public/credit-request/capture-session/{token}/upload", response_model=schemas.PublicCreditCaptureSessionResponse)
async def upload_public_credit_capture_session_file(
    token: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    session = _get_public_credit_capture_session(db, token)
    if not file:
        raise HTTPException(status_code=400, detail="Debes adjuntar una imagen.")

    is_signature = session.side == "signature"
    file_url = _save_public_credit_upload(
        file,
        "signatures" if is_signature else "documents",
        "signature_upload" if is_signature else f"document_{session.side}",
    )
    session.file_path = file_url
    session.file_type = getattr(file, "content_type", None)
    session.original_file_name = getattr(file, "filename", None)
    session.uploaded_at = datetime.datetime.now()
    db.commit()
    db.refresh(session)
    return _serialize_public_credit_capture_session(session)


@app.post("/public/credit-request/submit", response_model=schemas.PublicCreditSubmissionResponse)
async def submit_public_credit_request(
    request: Request,
    payload_json: str = Form(...),
    access_token: Optional[str] = Form(None),
    document_front: Optional[UploadFile] = File(None),
    document_back: Optional[UploadFile] = File(None),
    signature_file: Optional[UploadFile] = File(None),
    document_front_capture_token: Optional[str] = Form(None),
    document_back_capture_token: Optional[str] = Form(None),
    signature_capture_token: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    company = resolve_public_company(db, request)
    if not company:
        raise HTTPException(status_code=404, detail="No se encontró una empresa asociada a este dominio")
    if not company_has_enabled_module("public_credit_form", company=company):
        raise HTTPException(status_code=403, detail="El formulario de crédito no está habilitado para esta empresa")

    try:
        payload = json.loads(payload_json)
    except Exception:
        raise HTTPException(status_code=400, detail="La estructura del formulario no es válida.")

    consent = payload.get("consent", {}) or {}
    personal = payload.get("personal", {}) or {}
    vehicle = payload.get("vehicle", {}) or {}
    employment = payload.get("employment", {}) or {}
    income = payload.get("income", {}) or {}
    references = payload.get("references", {}) or {}

    applicant_email = str(personal.get("email") or "").strip().lower()
    applicant_phone = str(personal.get("mobilePhone") or "").strip()
    applicant_name = " ".join(
        [
            str(personal.get("firstName") or "").strip(),
            str(personal.get("lastName") or "").strip(),
        ]
    ).strip()
    desired_vehicle = str(vehicle.get("label") or "").strip()
    document_number = str(personal.get("documentNumber") or "").strip()
    verification_code = str(consent.get("verificationCode") or "").strip()

    if not applicant_name or not applicant_email or not applicant_phone or not desired_vehicle or not document_number:
        raise HTTPException(status_code=400, detail="Faltan datos obligatorios del formulario.")
    if not bool(consent.get("accepted")):
        raise HTTPException(status_code=400, detail="Debes aceptar la política de tratamiento de datos.")

    requires_email_validation = company_requires_public_credit_email_validation(company)
    auto_verified_at = datetime.datetime.now()
    verification = None
    linked_access = None
    linked_lead = None
    if access_token:
        linked_access = _get_public_credit_lead_access(db, str(access_token).strip(), company_id=company.id)
        if applicant_email != str(linked_access.email or "").strip().lower():
            raise HTTPException(status_code=400, detail="El correo no corresponde al acceso enviado por el asesor.")
        if requires_email_validation and verification_code != str(linked_access.code or "").strip():
            raise HTTPException(status_code=400, detail="El código de validación no es correcto.")
        if linked_access.verified_at is None:
            linked_access.verified_at = auto_verified_at
        linked_lead = linked_access.lead
        if not linked_lead:
            raise HTTPException(status_code=404, detail="El lead asociado a este acceso ya no existe.")
    else:
        if requires_email_validation:
            verification = db.query(models.PublicCreditEmailVerification).filter(
                models.PublicCreditEmailVerification.company_id == company.id,
                func.lower(models.PublicCreditEmailVerification.email) == applicant_email,
                models.PublicCreditEmailVerification.code == verification_code,
            ).order_by(models.PublicCreditEmailVerification.created_at.desc()).first()

            if not verification or verification.verified_at is None:
                raise HTTPException(status_code=400, detail="Debes validar el código enviado al correo antes de continuar.")
            if verification.expires_at < datetime.datetime.now():
                raise HTTPException(status_code=400, detail="La validación por correo expiró. Solicita un nuevo código.")

        ensure_company_lead_creation_capacity(db, company.id)

    front_capture = None
    back_capture = None
    signature_capture = None
    if document_front_capture_token:
        front_capture = _get_public_credit_capture_session(
            db,
            str(document_front_capture_token).strip(),
            company_id=company.id,
            require_upload=True,
        )
        if front_capture.side != "front":
            raise HTTPException(status_code=400, detail="La captura frontal no corresponde a la cara frontal.")
    if document_back_capture_token:
        back_capture = _get_public_credit_capture_session(
            db,
            str(document_back_capture_token).strip(),
            company_id=company.id,
            require_upload=True,
        )
        if back_capture.side != "back":
            raise HTTPException(status_code=400, detail="La captura posterior no corresponde a la cara posterior.")
    if signature_capture_token:
        signature_capture = _get_public_credit_capture_session(
            db,
            str(signature_capture_token).strip(),
            company_id=company.id,
            require_upload=True,
        )
        if signature_capture.side != "signature":
            raise HTTPException(status_code=400, detail="La captura de firma no corresponde a una firma.")

    document_front_url = _save_public_credit_upload(document_front, "documents", "document_front") if document_front else (front_capture.file_path if front_capture else None)
    document_back_url = _save_public_credit_upload(document_back, "documents", "document_back") if document_back else (back_capture.file_path if back_capture else None)
    signature_upload_url = _save_public_credit_upload(signature_file, "signatures", "signature_upload") if signature_file else (signature_capture.file_path if signature_capture else None)
    signature_drawn_url = _save_public_credit_data_url(consent.get("signatureDrawnDataUrl"), "signatures", "signature_drawn")

    attachments = {
        "document_front": document_front_url,
        "document_back": document_back_url,
        "signature_upload": signature_upload_url,
        "signature_drawn": signature_drawn_url,
    }

    if linked_lead:
        linked_lead.name = applicant_name or linked_lead.name
        linked_lead.email = applicant_email or linked_lead.email
        linked_lead.phone = applicant_phone or linked_lead.phone
        linked_lead.message = _build_public_credit_lead_message(payload)
        linked_lead.updated_at = datetime.datetime.now()

        process_detail = db.query(models.LeadProcessDetail).filter(
            models.LeadProcessDetail.lead_id == linked_lead.id
        ).first()
        if not process_detail:
            process_detail = models.LeadProcessDetail(
                lead_id=linked_lead.id,
                has_vehicle=False,
                desired_vehicle=desired_vehicle,
            )
            db.add(process_detail)
        else:
            process_detail.desired_vehicle = desired_vehicle or process_detail.desired_vehicle

        summary_lines = [
            f"Formulario de crédito completado por {applicant_name}.",
            f"Vehículo de interés: {desired_vehicle}.",
            f"Documento: {personal.get('documentType') or 'N/A'} {document_number}.",
            f"Valor vehículo: ${int(vehicle.get('vehicleValue') or 0):,}".replace(",", "."),
            f"Monto solicitado: ${int(vehicle.get('requestedAmount') or 0):,}".replace(",", "."),
            f"Ingresos totales: ${int(income.get('totalIncome') or 0):,}".replace(",", "."),
            f"Actividad económica: {employment.get('activity') or 'N/A'}.",
        ]
        if personal.get("city"):
            summary_lines.append(f"Ciudad: {personal.get('city')}.")
        if personal.get("address"):
            summary_lines.append(f"Dirección: {personal.get('address')}.")
        if references:
            summary_lines.append("Referencias registradas en formulario de crédito.")

        db.add(models.LeadHistory(
            lead_id=linked_lead.id,
            user_id=None,
            previous_status=linked_lead.status,
            new_status=linked_lead.status,
            comment="Formulario de crédito validado y firmado por el cliente desde el acceso enviado por el asesor.",
        ))

        upsert_credit_application_from_lead(db, linked_lead, " ".join(summary_lines))
        related_credit = next(
            (
                record for record in db.query(models.CreditApplication).filter(
                    models.CreditApplication.lead_id == linked_lead.id
                ).order_by(
                    models.CreditApplication.updated_at.desc(),
                    models.CreditApplication.created_at.desc()
                ).all()
                if not is_purchase_request_record(record)
            ),
            None
        )
        if related_credit:
            related_credit.client_name = applicant_name
            related_credit.phone = applicant_phone
            related_credit.email = applicant_email
            related_credit.desired_vehicle = desired_vehicle
            related_credit.monthly_income = int(income.get("salaryIncome") or 0)
            related_credit.other_income = int(income.get("otherIncome") or 0)
            related_credit.down_payment = int(consent.get("declaredDownPayment") or 0)
            related_credit.occupation = str(employment.get("activity") or "employee")
            related_credit.application_mode = "individual"
            related_credit.notes = "\n".join(summary_lines)

        if document_front_url:
            db.add(models.LeadFile(
                lead_id=linked_lead.id,
                user_id=None,
                file_name="cedula_frontal",
                file_path=document_front_url,
                file_type=getattr(document_front, "content_type", None) if document_front else getattr(front_capture, "file_type", None),
            ))
        if document_back_url:
            db.add(models.LeadFile(
                lead_id=linked_lead.id,
                user_id=None,
                file_name="cedula_posterior",
                file_path=document_back_url,
                file_type=getattr(document_back, "content_type", None) if document_back else getattr(back_capture, "file_type", None),
            ))
        if signature_upload_url or signature_drawn_url:
            db.add(models.LeadFile(
                lead_id=linked_lead.id,
                user_id=None,
                file_name="firma_cliente",
                file_path=signature_upload_url or signature_drawn_url,
                file_type=getattr(signature_file, "content_type", None) or getattr(signature_capture, "file_type", None) or "image/png",
            ))

        submission = db.query(models.PublicCreditSubmission).filter(
            models.PublicCreditSubmission.lead_id == linked_lead.id
        ).order_by(
            models.PublicCreditSubmission.updated_at.desc(),
            models.PublicCreditSubmission.id.desc(),
        ).first()
        if not submission:
            submission = models.PublicCreditSubmission(
                company_id=company.id,
                lead_id=linked_lead.id,
                email=applicant_email,
                applicant_name=applicant_name,
                status="submitted",
                attachments={},
                consent_text=PUBLIC_CREDIT_POLICY_TEXT,
            )
            db.add(submission)
        submission.email = applicant_email
        submission.applicant_name = applicant_name
        submission.document_number = document_number
        submission.phone = applicant_phone
        submission.desired_vehicle = desired_vehicle
        submission.status = "submitted"
        submission.verification_code = linked_access.code if requires_email_validation else None
        submission.verification_verified_at = linked_access.verified_at
        submission.form_payload = payload
        current_attachments = submission.attachments or {}
        current_attachments.update({key: value for key, value in attachments.items() if value})
        submission.attachments = current_attachments
        submission.consent_text = PUBLIC_CREDIT_POLICY_TEXT
        submission.updated_at = datetime.datetime.now()
        linked_access.used_at = datetime.datetime.now()

        db.commit()
        db.refresh(submission)

        response_message = "Formulario de crédito guardado y asociado al lead correctamente."
        try:
            pdf_bytes = _build_public_credit_submission_pdf(company, submission)
            _send_public_credit_submission_email(db, company, submission, pdf_bytes)
        except Exception as exc:
            print(
                f"Warning: linked public credit submission {submission.id} was saved but email delivery failed: {exc}",
                flush=True,
            )
            response_message = (
                "Formulario guardado correctamente, pero no fue posible enviar el correo con el PDF. "
                "La empresa puede descargarlo desde Solicitudes Públicas."
            )

        return schemas.PublicCreditSubmissionResponse(
            status="ok",
            message=response_message,
            lead_id=linked_lead.id,
            submission_id=submission.id,
        )

    assigned_user_id = None
    auto_assigned_user = choose_auto_assign_user(db, company.id)
    if auto_assigned_user:
        assigned_user_id = auto_assigned_user.id

    effective_status = normalize_company_lead_status(
        models.LeadStatus.CREDIT_STUDY.value,
        company,
        models.LeadStatus.NEW.value,
    )
    assigned_user_id, supervisor_ids, credit_coordinator = maybe_assign_credit_coordinator(
        db=db,
        lead_company_id=company.id,
        previous_status=None,
        target_status=effective_status,
        assigned_to_id=assigned_user_id,
        supervisor_ids=[],
        current_user_id=None,
        previous_assigned_to_id=None,
    )

    lead_message = _build_public_credit_lead_message(payload)

    new_lead = models.Lead(
        source=models.LeadSource.WEB.value,
        name=applicant_name,
        email=applicant_email,
        phone=applicant_phone,
        message=lead_message,
        status=effective_status,
        company_id=company.id,
        assigned_to_id=assigned_user_id,
        created_by_id=None,
        created_at=datetime.datetime.now(),
    )
    db.add(new_lead)
    db.flush()

    sync_lead_supervisors(db, new_lead, supervisor_ids, None)

    process_detail = models.LeadProcessDetail(
        lead_id=new_lead.id,
        has_vehicle=False,
        desired_vehicle=desired_vehicle,
    )
    db.add(process_detail)

    history_note = "Formulario público de crédito enviado"
    if credit_coordinator:
        history_note += f" (Coordinador de crédito añadido como supervisor: {credit_coordinator.full_name or credit_coordinator.email})"

    db.add(models.LeadHistory(
        lead_id=new_lead.id,
        user_id=None,
        previous_status=None,
        new_status=new_lead.status,
        comment=history_note,
    ))

    summary_lines = [
        f"Solicitud pública de crédito enviada por {applicant_name}.",
        f"Vehículo de interés: {desired_vehicle}.",
        f"Documento: {personal.get('documentType') or 'N/A'} {document_number}.",
        f"Valor vehículo: ${int(vehicle.get('vehicleValue') or 0):,}".replace(",", "."),
        f"Monto solicitado: ${int(vehicle.get('requestedAmount') or 0):,}".replace(",", "."),
        f"Ingresos totales: ${int(income.get('totalIncome') or 0):,}".replace(",", "."),
        f"Actividad económica: {employment.get('activity') or 'N/A'}.",
    ]
    if personal.get("city"):
        summary_lines.append(f"Ciudad: {personal.get('city')}.")
    if personal.get("address"):
        summary_lines.append(f"Dirección: {personal.get('address')}.")
    if references:
        summary_lines.append("Referencias registradas en formulario público.")

    db.flush()
    upsert_credit_application_from_lead(db, new_lead, " ".join(summary_lines))

    related_credit = next(
        (
            record for record in db.query(models.CreditApplication).filter(
                models.CreditApplication.lead_id == new_lead.id
            ).order_by(
                models.CreditApplication.updated_at.desc(),
                models.CreditApplication.created_at.desc()
            ).all()
            if not is_purchase_request_record(record)
        ),
        None
    )

    if related_credit:
        related_credit.client_name = applicant_name
        related_credit.phone = applicant_phone
        related_credit.email = applicant_email
        related_credit.desired_vehicle = desired_vehicle
        related_credit.monthly_income = int(income.get("salaryIncome") or 0)
        related_credit.other_income = int(income.get("otherIncome") or 0)
        related_credit.down_payment = int(consent.get("declaredDownPayment") or 0)
        related_credit.occupation = str(employment.get("activity") or "employee")
        related_credit.application_mode = "individual"
        related_credit.notes = "\n".join(summary_lines)

    if document_front_url:
        db.add(models.LeadFile(
            lead_id=new_lead.id,
            user_id=None,
            file_name="cedula_frontal",
            file_path=document_front_url,
            file_type=getattr(document_front, "content_type", None) if document_front else getattr(front_capture, "file_type", None),
        ))
    if document_back_url:
        db.add(models.LeadFile(
            lead_id=new_lead.id,
            user_id=None,
            file_name="cedula_posterior",
            file_path=document_back_url,
            file_type=getattr(document_back, "content_type", None) if document_back else getattr(back_capture, "file_type", None),
        ))
    if signature_upload_url or signature_drawn_url:
        db.add(models.LeadFile(
            lead_id=new_lead.id,
            user_id=None,
            file_name="firma_cliente",
            file_path=signature_upload_url or signature_drawn_url,
            file_type=getattr(signature_file, "content_type", None) or getattr(signature_capture, "file_type", None) or "image/png",
        ))

    submission = models.PublicCreditSubmission(
        company_id=company.id,
        lead_id=new_lead.id,
        email=applicant_email,
        applicant_name=applicant_name,
        document_number=document_number,
        phone=applicant_phone,
        desired_vehicle=desired_vehicle,
        status="submitted",
        verification_code=verification.code if verification else None,
        verification_verified_at=verification.verified_at if verification else auto_verified_at,
        form_payload=payload,
        attachments=attachments,
        consent_text=PUBLIC_CREDIT_POLICY_TEXT,
    )
    db.add(submission)
    db.commit()
    db.refresh(submission)

    response_message = "Solicitud de crédito enviada correctamente."
    try:
        pdf_bytes = _build_public_credit_submission_pdf(company, submission)
        _send_public_credit_submission_email(db, company, submission, pdf_bytes)
    except Exception as exc:
        print(
            f"Warning: public credit submission {submission.id} was saved but email delivery failed: {exc}",
            flush=True,
        )
        response_message = (
            "Solicitud guardada correctamente, pero no fue posible enviar el correo con el PDF. "
            "La empresa puede descargarlo desde Solicitudes Públicas."
        )

    return schemas.PublicCreditSubmissionResponse(
        status="ok",
        message=response_message,
        lead_id=new_lead.id,
        submission_id=submission.id,
    )


@app.get("/public-credit-submissions", response_model=schemas.PublicCreditSubmissionList)
def read_public_credit_submissions(
    q: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    ensure_public_credit_review_permissions(current_user)

    query = db.query(models.PublicCreditSubmission).options(
        joinedload(models.PublicCreditSubmission.lead)
    )

    if current_user.company_id:
        query = query.filter(models.PublicCreditSubmission.company_id == current_user.company_id)

    normalized_status = str(status or "").strip().lower()
    if normalized_status:
        query = query.filter(func.lower(models.PublicCreditSubmission.status) == normalized_status)

    normalized_q = str(q or "").strip().lower()
    if normalized_q:
        like_value = f"%{normalized_q}%"
        query = query.filter(
            or_(
                func.lower(models.PublicCreditSubmission.applicant_name).like(like_value),
                func.lower(models.PublicCreditSubmission.email).like(like_value),
                func.lower(func.coalesce(models.PublicCreditSubmission.document_number, "")).like(like_value),
                func.lower(func.coalesce(models.PublicCreditSubmission.phone, "")).like(like_value),
                func.lower(func.coalesce(models.PublicCreditSubmission.desired_vehicle, "")).like(like_value),
            )
        )

    total = query.count()
    items = query.order_by(
        models.PublicCreditSubmission.created_at.desc(),
        models.PublicCreditSubmission.id.desc(),
    ).offset(skip).limit(limit).all()

    return schemas.PublicCreditSubmissionList(
        items=[schemas.PublicCreditSubmissionItem(**serialize_public_credit_submission(item)) for item in items],
        total=total,
    )


@app.get("/public-credit-submissions/{submission_id}", response_model=schemas.PublicCreditSubmissionDetail)
def read_public_credit_submission_detail(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    ensure_public_credit_review_permissions(current_user)

    query = db.query(models.PublicCreditSubmission).options(
        joinedload(models.PublicCreditSubmission.lead)
    ).filter(models.PublicCreditSubmission.id == submission_id)

    if current_user.company_id:
        query = query.filter(models.PublicCreditSubmission.company_id == current_user.company_id)

    submission = query.first()
    if not submission:
        raise HTTPException(status_code=404, detail="Solicitud pública no encontrada")

    return schemas.PublicCreditSubmissionDetail(**serialize_public_credit_submission(submission))


@app.get("/public-credit-submissions/{submission_id}/pdf")
def download_public_credit_submission_pdf(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    ensure_public_credit_review_permissions(current_user)
    query = db.query(models.PublicCreditSubmission).filter(
        models.PublicCreditSubmission.id == submission_id
    )
    if current_user.company_id:
        query = query.filter(models.PublicCreditSubmission.company_id == current_user.company_id)
    submission = query.first()
    if not submission:
        raise HTTPException(status_code=404, detail="Solicitud pública no encontrada")

    company = db.query(models.Company).filter(models.Company.id == submission.company_id).first()
    pdf_bytes = _build_public_credit_submission_pdf(company, submission)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", submission.applicant_name or "solicitante").strip("_")
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                f'attachment; filename="formulario_credito_{safe_name or submission.id}.pdf"'
            )
        },
    )


@app.put("/public-credit-submissions/{submission_id}", response_model=schemas.PublicCreditSubmissionDetail)
def update_public_credit_submission(
    submission_id: int,
    payload: schemas.PublicCreditSubmissionUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    ensure_public_credit_review_permissions(current_user)

    query = db.query(models.PublicCreditSubmission).options(
        joinedload(models.PublicCreditSubmission.lead)
    ).filter(models.PublicCreditSubmission.id == submission_id)

    if current_user.company_id:
        query = query.filter(models.PublicCreditSubmission.company_id == current_user.company_id)

    submission = query.first()
    if not submission:
        raise HTTPException(status_code=404, detail="Solicitud pública no encontrada")

    normalized_status = str(payload.status or "").strip().lower()
    allowed_statuses = {"submitted", "reviewing", "approved", "rejected"}
    if normalized_status not in allowed_statuses:
        raise HTTPException(status_code=400, detail="Estado de solicitud pública no válido")

    submission.status = normalized_status
    submission.updated_at = datetime.datetime.now()
    db.commit()
    db.refresh(submission)

    return schemas.PublicCreditSubmissionDetail(**serialize_public_credit_submission(submission))


@app.delete("/public-credit-submissions/{submission_id}", status_code=204)
def delete_public_credit_submission(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    ensure_public_credit_delete_permissions(current_user)

    query = db.query(models.PublicCreditSubmission).filter(
        models.PublicCreditSubmission.id == submission_id
    )
    if current_user.company_id:
        query = query.filter(models.PublicCreditSubmission.company_id == current_user.company_id)

    submission = query.first()
    if not submission:
        raise HTTPException(status_code=404, detail="Solicitud pública no encontrada")

    db.delete(submission)
    db.commit()
    return Response(status_code=204)


def _get_accessible_lead_for_credit_form(
    db: Session,
    current_user: models.User,
    lead_id: int,
) -> models.Lead:
    query = apply_lead_access_filters(db.query(models.Lead), db, current_user)
    lead = query.filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado o sin acceso")
    return lead


def _default_internal_credit_form_payload(lead: models.Lead) -> Dict[str, Any]:
    process_detail = getattr(lead, "process_detail", None)
    return {
        "vehicle": {
            "vehicleValue": "",
            "requestedAmount": "",
            "requestDate": bogota_today().isoformat(),
            "make": "",
            "model": "",
            "vehicleType": "Automóvil",
            "label": getattr(process_detail, "desired_vehicle", None) or "",
        },
        "personal": {
            "firstName": lead.name or "",
            "lastName": "",
            "documentType": "C.C",
            "documentNumber": "",
            "issuePlace": "",
            "birthDate": "",
            "gender": "",
            "profession": "",
            "birthPlace": "",
            "maritalStatus": "",
            "childrenCount": "",
            "educationLevel": "",
            "livesWith": "",
            "housingType": "",
            "mobilePhone": lead.phone or "",
            "city": "",
            "address": "",
            "email": lead.email or "",
        },
        "employment": {
            "activity": "",
            "companyName": "",
            "companyCity": "",
            "companyAddress": "",
            "jobTitle": "",
            "companyEmail": "",
            "startDate": "",
            "salary": "",
            "contractType": "",
            "previousCompanyName": "",
            "previousCompanyActivity": "",
            "previousCompanyRole": "",
            "previousEmploymentTime": "",
        },
        "income": {
            "salaryIncome": "",
            "commissionsIncome": "",
            "otherIncome": "",
            "otherIncomeDetail": "",
            "totalIncome": "",
        },
        "references": {
            "commercial": {"names": "", "lastNames": "", "phone": "", "city": ""},
            "personal1": {"names": "", "lastNames": "", "phone": "", "city": ""},
            "personal2": {"names": "", "lastNames": "", "phone": "", "city": ""},
        },
        "consent": {
            "accepted": False,
            "signatureMode": "",
            "signatureName": "",
        },
    }


@app.get("/leads/{lead_id}/credit-form")
def read_lead_credit_form(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    lead = _get_accessible_lead_for_credit_form(db, current_user, lead_id)
    submission = db.query(models.PublicCreditSubmission).filter(
        models.PublicCreditSubmission.lead_id == lead.id
    ).order_by(
        models.PublicCreditSubmission.updated_at.desc(),
        models.PublicCreditSubmission.id.desc(),
    ).first()
    if submission:
        return {
            "exists": True,
            "origin": "public" if submission.verification_verified_at else "internal",
            "submission": serialize_public_credit_submission(submission),
            "form_payload": submission.form_payload or _default_internal_credit_form_payload(lead),
        }
    return {
        "exists": False,
        "origin": "internal",
        "submission": None,
        "form_payload": _default_internal_credit_form_payload(lead),
    }


@app.put("/leads/{lead_id}/credit-form", response_model=schemas.PublicCreditSubmissionDetail)
def save_lead_credit_form(
    lead_id: int,
    payload: schemas.LeadCreditFormUpsert,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    lead = db.query(models.Lead).options(joinedload(models.Lead.supervisors)).filter(
        models.Lead.id == lead_id
    ).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    ensure_can_modify_lead(current_user, lead)

    form_payload = payload.form_payload if isinstance(payload.form_payload, dict) else {}
    personal = form_payload.get("personal", {}) or {}
    vehicle = form_payload.get("vehicle", {}) or {}
    applicant_name = " ".join(filter(None, [
        str(personal.get("firstName") or "").strip(),
        str(personal.get("lastName") or "").strip(),
    ])).strip() or lead.name or "Cliente"
    applicant_email = str(personal.get("email") or lead.email or "").strip().lower()
    applicant_phone = str(personal.get("mobilePhone") or lead.phone or "").strip()
    desired_vehicle = str(vehicle.get("label") or " ".join(filter(None, [
        vehicle.get("make"), vehicle.get("model"), vehicle.get("vehicleType")
    ]))).strip()

    submission = db.query(models.PublicCreditSubmission).filter(
        models.PublicCreditSubmission.lead_id == lead.id
    ).order_by(
        models.PublicCreditSubmission.updated_at.desc(),
        models.PublicCreditSubmission.id.desc(),
    ).first()
    created = submission is None
    if created:
        submission = models.PublicCreditSubmission(
            company_id=lead.company_id,
            lead_id=lead.id,
            email=applicant_email,
            applicant_name=applicant_name,
            status="reviewing",
            attachments={},
            consent_text=PUBLIC_CREDIT_POLICY_TEXT,
        )
        db.add(submission)
    else:
        previous_consent = (submission.form_payload or {}).get("consent", {}) or {}
        if submission.verification_verified_at:
            form_payload["consent"] = previous_consent

    submission.email = applicant_email
    submission.applicant_name = applicant_name
    submission.document_number = str(personal.get("documentNumber") or "").strip() or None
    submission.phone = applicant_phone or None
    submission.desired_vehicle = desired_vehicle or None
    submission.form_payload = form_payload
    submission.updated_at = datetime.datetime.now()

    db.add(models.LeadHistory(
        lead_id=lead.id,
        user_id=current_user.id,
        previous_status=lead.status,
        new_status=lead.status,
        comment=(
            "Formulario de crédito interno creado por el asesor"
            if created else "Formulario de crédito actualizado desde el detalle del lead"
        ),
    ))
    db.commit()
    db.refresh(submission)
    return schemas.PublicCreditSubmissionDetail(**serialize_public_credit_submission(submission))


@app.post("/leads/{lead_id}/credit-form/capture-session", response_model=schemas.PublicCreditCaptureSessionResponse)
def create_lead_credit_capture_session(
    lead_id: int,
    payload: schemas.PublicCreditCaptureSessionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    lead = _get_accessible_lead_for_credit_form(db, current_user, lead_id)
    ensure_can_modify_lead(current_user, lead)

    side = str(payload.side or "").strip().lower()
    if side not in {"front", "back", "signature"}:
        raise HTTPException(status_code=400, detail="El tipo de captura no es válido.")

    session = models.PublicCreditCaptureSession(
        company_id=lead.company_id,
        token=secrets.token_urlsafe(32),
        side=side,
        expires_at=datetime.datetime.now() + datetime.timedelta(minutes=30),
    )
    db.add(session)
    db.commit()
    db.refresh(session)
    return _serialize_public_credit_capture_session(session)


def _upsert_lead_credit_form_submission(
    db: Session,
    lead: models.Lead,
    current_user: models.User,
    form_payload: Dict[str, Any],
    attachments: Optional[Dict[str, Any]] = None,
) -> models.PublicCreditSubmission:
    personal = form_payload.get("personal", {}) or {}
    vehicle = form_payload.get("vehicle", {}) or {}
    applicant_name = " ".join(filter(None, [
        str(personal.get("firstName") or "").strip(),
        str(personal.get("lastName") or "").strip(),
    ])).strip() or lead.name or "Cliente"
    applicant_email = str(personal.get("email") or lead.email or "").strip().lower()
    applicant_phone = str(personal.get("mobilePhone") or lead.phone or "").strip()
    desired_vehicle = str(vehicle.get("label") or " ".join(filter(None, [
        vehicle.get("make"), vehicle.get("model"), vehicle.get("vehicleType")
    ]))).strip()

    submission = db.query(models.PublicCreditSubmission).filter(
        models.PublicCreditSubmission.lead_id == lead.id
    ).order_by(
        models.PublicCreditSubmission.updated_at.desc(),
        models.PublicCreditSubmission.id.desc(),
    ).first()
    created = submission is None
    if created:
        submission = models.PublicCreditSubmission(
            company_id=lead.company_id,
            lead_id=lead.id,
            email=applicant_email or f"lead-{lead.id}@sin-correo.local",
            applicant_name=applicant_name,
            status="reviewing",
            attachments={},
            consent_text=PUBLIC_CREDIT_POLICY_TEXT,
        )
        db.add(submission)
    else:
        previous_consent = (submission.form_payload or {}).get("consent", {}) or {}
        if submission.verification_verified_at:
            form_payload["consent"] = previous_consent

    submission.email = applicant_email or submission.email or f"lead-{lead.id}@sin-correo.local"
    submission.applicant_name = applicant_name
    submission.document_number = str(personal.get("documentNumber") or "").strip() or None
    submission.phone = applicant_phone or None
    submission.desired_vehicle = desired_vehicle or None
    submission.form_payload = form_payload
    if attachments:
        current_attachments = submission.attachments or {}
        current_attachments.update({key: value for key, value in attachments.items() if value})
        submission.attachments = current_attachments
    submission.updated_at = datetime.datetime.now()

    db.add(models.LeadHistory(
        lead_id=lead.id,
        user_id=current_user.id,
        previous_status=lead.status,
        new_status=lead.status,
        comment=(
            "Formulario de crédito interno creado por el asesor"
            if created else "Formulario de crédito actualizado desde el detalle del lead"
        ),
    ))
    return submission


@app.put("/leads/{lead_id}/credit-form/files", response_model=schemas.PublicCreditSubmissionDetail)
async def save_lead_credit_form_with_files(
    lead_id: int,
    payload_json: str = Form(...),
    document_front: Optional[UploadFile] = File(None),
    document_back: Optional[UploadFile] = File(None),
    signature_file: Optional[UploadFile] = File(None),
    document_front_capture_token: Optional[str] = Form(None),
    document_back_capture_token: Optional[str] = Form(None),
    signature_capture_token: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    lead = db.query(models.Lead).options(joinedload(models.Lead.supervisors)).filter(
        models.Lead.id == lead_id
    ).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="No autorizado")
    ensure_can_modify_lead(current_user, lead)

    try:
        form_payload = json.loads(payload_json)
    except Exception:
        raise HTTPException(status_code=400, detail="La estructura del formulario no es válida.")
    if not isinstance(form_payload, dict):
        raise HTTPException(status_code=400, detail="La estructura del formulario no es válida.")

    front_capture = None
    back_capture = None
    signature_capture = None
    if document_front_capture_token:
        front_capture = _get_public_credit_capture_session(
            db,
            str(document_front_capture_token).strip(),
            company_id=lead.company_id,
            require_upload=True,
        )
        if front_capture.side != "front":
            raise HTTPException(status_code=400, detail="La captura frontal no corresponde a la cara frontal.")
    if document_back_capture_token:
        back_capture = _get_public_credit_capture_session(
            db,
            str(document_back_capture_token).strip(),
            company_id=lead.company_id,
            require_upload=True,
        )
        if back_capture.side != "back":
            raise HTTPException(status_code=400, detail="La captura posterior no corresponde a la cara posterior.")
    if signature_capture_token:
        signature_capture = _get_public_credit_capture_session(
            db,
            str(signature_capture_token).strip(),
            company_id=lead.company_id,
            require_upload=True,
        )
        if signature_capture.side != "signature":
            raise HTTPException(status_code=400, detail="La captura de firma no corresponde a una firma.")

    consent = form_payload.get("consent", {}) or {}
    document_front_url = _save_public_credit_upload(document_front, "documents", "document_front") if document_front else (front_capture.file_path if front_capture else None)
    document_back_url = _save_public_credit_upload(document_back, "documents", "document_back") if document_back else (back_capture.file_path if back_capture else None)
    signature_upload_url = _save_public_credit_upload(signature_file, "signatures", "signature_upload") if signature_file else (signature_capture.file_path if signature_capture else None)
    signature_drawn_url = _save_public_credit_data_url(consent.get("signatureDrawnDataUrl"), "signatures", "signature_drawn")

    attachments = {
        "document_front": document_front_url,
        "document_back": document_back_url,
        "signature_upload": signature_upload_url,
        "signature_drawn": signature_drawn_url,
    }
    submission = _upsert_lead_credit_form_submission(db, lead, current_user, form_payload, attachments)

    if document_front_url:
        db.add(models.LeadFile(
            lead_id=lead.id,
            user_id=current_user.id,
            file_name="cedula_frontal",
            file_path=document_front_url,
            file_type=getattr(document_front, "content_type", None) if document_front else getattr(front_capture, "file_type", None),
        ))
    if document_back_url:
        db.add(models.LeadFile(
            lead_id=lead.id,
            user_id=current_user.id,
            file_name="cedula_posterior",
            file_path=document_back_url,
            file_type=getattr(document_back, "content_type", None) if document_back else getattr(back_capture, "file_type", None),
        ))
    if signature_upload_url or signature_drawn_url:
        db.add(models.LeadFile(
            lead_id=lead.id,
            user_id=current_user.id,
            file_name="firma_cliente",
            file_path=signature_upload_url or signature_drawn_url,
            file_type=getattr(signature_file, "content_type", None) or getattr(signature_capture, "file_type", None) or "image/png",
        ))

    db.commit()
    db.refresh(submission)
    return schemas.PublicCreditSubmissionDetail(**serialize_public_credit_submission(submission))


def _build_company_public_credit_form_url(company: Optional[models.Company], request: Request) -> str:
    scheme = request.headers.get("x-forwarded-proto") or request.url.scheme or "https"
    domain = None
    if company:
        public_domains = normalize_public_domains(getattr(company, "public_domains", []), company.public_domain)
        domain = public_domains[0] if public_domains else None
    host = domain or request.headers.get("host") or request.url.netloc
    return f"{scheme}://{host}/crm/credito"


def _create_lead_credit_form_access(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> Dict[str, Any]:
    lead = _get_accessible_lead_for_credit_form(db, current_user, lead_id)
    ensure_can_modify_lead(current_user, lead)
    target_email = str(lead.email or "").strip().lower()
    if not target_email:
        raise HTTPException(status_code=400, detail="El lead no tiene correo para generar el acceso.")

    company = db.query(models.Company).filter(models.Company.id == lead.company_id).first()
    requires_email_validation = company_requires_public_credit_email_validation(company)
    verification_code = f"{random.randint(100000, 999999)}"
    access = models.PublicCreditLeadAccess(
        company_id=lead.company_id,
        lead_id=lead.id,
        created_by_id=current_user.id,
        email=target_email,
        token=secrets.token_urlsafe(32),
        code=verification_code,
        expires_at=datetime.datetime.now() + datetime.timedelta(days=7),
    )
    db.add(access)
    db.flush()

    form_url = f"{_build_company_public_credit_form_url(company, request)}?access={access.token}"
    return {
        "lead": lead,
        "company": company,
        "access": access,
        "form_url": form_url,
        "requires_email_validation": requires_email_validation,
        "verification_code": verification_code,
    }


@app.post("/leads/{lead_id}/credit-form/create-access")
def create_lead_credit_form_access(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    access_context = _create_lead_credit_form_access(lead_id, request, db, current_user)
    lead = access_context["lead"]
    db.add(models.LeadHistory(
        lead_id=lead.id,
        user_id=current_user.id,
        previous_status=lead.status,
        new_status=lead.status,
        comment="Enlace de formulario de crédito generado para compartir con el cliente",
    ))
    db.commit()
    return {
        "status": "ok",
        "form_url": access_context["form_url"],
        "requires_email_validation": access_context["requires_email_validation"],
        "verification_code": (
            access_context["verification_code"]
            if access_context["requires_email_validation"]
            else None
        ),
        "expires_at": access_context["access"].expires_at,
        "email": access_context["access"].email,
    }


@app.post("/leads/{lead_id}/credit-form/send-access")
def send_lead_credit_form_access(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    access_context = _create_lead_credit_form_access(lead_id, request, db, current_user)
    lead = access_context["lead"]
    company = access_context["company"]
    form_url = access_context["form_url"]
    requires_email_validation = access_context["requires_email_validation"]
    verification_code = access_context["verification_code"]
    smtp_settings = _get_public_credit_smtp_settings(db, company)
    company_name = getattr(company, "name", None) or "AutosQP"
    code_text_lines = []
    if requires_email_validation:
        code_text_lines = [
            f"Código de validación: {verification_code}",
            "",
        ]
    code_html = ""
    if requires_email_validation:
        code_html = f"""
                <div style="margin:18px 0;padding:18px;border-radius:16px;background:#f8fafc;border:1px solid #e2e8f0;text-align:center;">
                  <div style="font-size:13px;color:#64748b;">Código de validación</div>
                  <div style="font-size:32px;font-weight:800;letter-spacing:6px;color:#0f172a;">{escape(verification_code)}</div>
                </div>
        """.strip()

    message = EmailMessage()
    message["Subject"] = (
        f"Código y firma del formulario de crédito - {company_name}"
        if requires_email_validation
        else f"Firma del formulario de crédito - {company_name}"
    )
    message["From"] = smtp_settings["sender"]
    message["To"] = target_email
    message.set_content(
        "\n".join([
            f"Hola {lead.name or ''}.".strip(),
            "",
            "Tu asesor está diligenciando tu formulario de crédito.",
            *code_text_lines,
            "Ingresa al siguiente enlace para revisar, adjuntar documentos y firmar:",
            form_url,
            "",
            "Este acceso vence en 7 días.",
        ]),
        charset="utf-8",
    )
    message.add_alternative(
        f"""
        <html><body style="margin:0;background:#f8fafc;font-family:Arial,sans-serif;color:#0f172a;">
          <div style="max-width:640px;margin:0 auto;padding:28px 18px;">
            <div style="overflow:hidden;border:1px solid #e2e8f0;border-radius:22px;background:#fff;">
              <div style="padding:26px;background:linear-gradient(135deg,{escape(getattr(company, 'secondary_color', None) or '#0f172a')},{escape(getattr(company, 'primary_color', None) or '#2563eb')});color:#fff;">
                <div style="font-size:13px;opacity:.85;">{escape(company_name)}</div>
                <h1 style="margin:8px 0 0;font-size:26px;">Formulario de crédito</h1>
              </div>
              <div style="padding:28px;">
                <p style="margin-top:0;color:#475569;">Hola {escape(lead.name or '')}, tu asesor está diligenciando tu formulario de crédito. Revisa la información, adjunta los documentos pendientes y firma desde este acceso.</p>
                {code_html}
                <a href="{escape(form_url)}" style="display:inline-block;margin-top:14px;border-radius:14px;background:{escape(getattr(company, 'primary_color', None) or '#2563eb')};color:#fff;text-decoration:none;font-weight:700;padding:14px 20px;">Abrir formulario</a>
                <p style="margin-top:24px;color:#64748b;font-size:13px;">Si el botón no abre, copia este enlace: {escape(form_url)}</p>
                <p style="margin-top:8px;color:#64748b;font-size:13px;">Este acceso vence en 7 días.</p>
              </div>
            </div>
          </div>
        </body></html>
        """.strip(),
        subtype="html",
    )
    _deliver_smtp_message(message, smtp_settings)
    db.add(models.LeadHistory(
        lead_id=lead.id,
        user_id=current_user.id,
        previous_status=lead.status,
        new_status=lead.status,
        comment=(
            "Código y acceso de firma del formulario de crédito enviados al cliente por correo"
            if requires_email_validation
            else "Acceso de firma del formulario de crédito enviado al cliente por correo"
        ),
    ))
    db.commit()
    return {
        "status": "ok",
        "message": (
            "Código y acceso enviados al correo del cliente."
            if requires_email_validation
            else "Acceso enviado al correo del cliente."
        ),
    }


@app.post("/public/credit-request", response_model=schemas.PublicCreditRequestResponse)
def create_public_credit_request(
    payload: schemas.PublicCreditRequestCreate,
    request: Request,
    db: Session = Depends(get_db)
):
    raise HTTPException(
        status_code=410,
        detail="Este endpoint quedó obsoleto. Usa /public/credit-request/send-code, /verify-code y /submit.",
    )

def serialize_integration_settings(settings: models.IntegrationSettings) -> Dict[str, Any]:
    data = schemas.IntegrationSettings.model_validate(settings).model_dump()
    data["smtp_password_configured"] = bool((getattr(settings, "smtp_password", None) or "").strip())
    data["whatsapp_calling_provider_token_configured"] = bool((getattr(settings, "whatsapp_calling_provider_token", None) or "").strip())
    # Never return stored SMTP secrets to the browser. A blank value means "keep current password".
    data["smtp_password"] = ""
    data["whatsapp_calling_provider_token"] = ""
    return data


@app.get("/companies/{company_id}/integrations", response_model=schemas.IntegrationSettings)
def read_integration_settings(company_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    # Check permissions
    if current_user.company_id and current_user.company_id != company_id:
        raise HTTPException(status_code=403, detail="Not authorized to view these settings")
    
    settings = db.query(models.IntegrationSettings).filter(models.IntegrationSettings.company_id == company_id).first()
    if not settings:
        # Create default empty settings if not exists
        settings = models.IntegrationSettings(company_id=company_id)
        db.add(settings)
        db.commit()
        db.refresh(settings)
        
    return serialize_integration_settings(settings)

@app.get("/roles/views")
def read_role_views(current_user: models.User = Depends(get_current_user)):
    ensure_role_management_permissions(current_user)
    company = getattr(current_user, "company", None)
    items = [
        view for view in SYSTEM_VIEWS
        if view.get("scope") == "global" or view["id"] in apply_company_module_limits([view["id"]], company)
    ]
    return {"items": items}


@app.get("/roles/", response_model=list[schemas.Role])
def read_roles(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    query = db.query(models.Role)
    if current_user.company_id:
        query = query.filter(
            or_(
                models.Role.company_id == current_user.company_id,
                and_(models.Role.company_id.is_(None), models.Role.is_system == True)
            )
        )
    roles = query.order_by(models.Role.is_system.desc(), models.Role.label.asc()).all()
    if current_user.company_id:
        override_names = {role.base_role_name for role in roles if role.company_id == current_user.company_id and role.base_role_name}
        roles = [role for role in roles if not (role.is_system and role.name in override_names)]
        roles = [role for role in roles if is_role_enabled_for_company(role, current_user.company)]
    return [serialize_role(role) for role in roles]


@app.post("/roles/", response_model=schemas.Role)
def create_role(
    role: schemas.RoleCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    ensure_role_management_permissions(current_user)
    if not current_user.company_id and current_user.role and current_user.role.name != "super_admin":
        raise HTTPException(status_code=403, detail="Not authorized")

    sanitized_permissions = sanitize_view_ids(role.permissions, current_user.company_id)
    sanitized_menu_order = sanitize_view_ids(role.menu_order, current_user.company_id)
    sanitized_assignable_role_ids = sanitize_assignable_role_ids(db, role.assignable_role_ids, current_user.company_id)
    advisor_tracking_enabled = bool(role.advisor_tracking_enabled)
    for view_id in sanitized_permissions:
        if view_id not in sanitized_menu_order:
            sanitized_menu_order.append(view_id)

    slug_base = re.sub(r"[^a-z0-9]+", "_", role.label.lower()).strip("_") or "rol"
    generated_name = f"custom_{current_user.company_id or 'global'}_{slug_base}_{int(time.time())}"

    db_role = models.Role(
        name=generated_name,
        label=role.label.strip(),
        company_id=current_user.company_id,
        permissions_json=json.dumps(sanitized_permissions),
        menu_order_json=json.dumps(sanitized_menu_order),
        auto_assign_leads=bool(role.auto_assign_leads),
        assignable_role_ids_json=json.dumps(sanitized_assignable_role_ids),
        advisor_tracking_enabled=advisor_tracking_enabled,
        tracked_advisor_ids_json=json.dumps([]),
        is_system=False,
        base_role_name=None
    )
    db.add(db_role)
    db.commit()
    db.refresh(db_role)
    return serialize_role(db_role)


@app.put("/roles/{role_id}", response_model=schemas.Role)
def update_role(
    role_id: int,
    role_update: schemas.RoleUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    ensure_role_management_permissions(current_user)
    db_role = db.query(models.Role).filter(models.Role.id == role_id).first()
    if not db_role:
        raise HTTPException(status_code=404, detail="Role not found")
    if current_user.company_id and db_role.company_id not in {None, current_user.company_id}:
        raise HTTPException(status_code=403, detail="Not authorized")

    if current_user.company_id and db_role.is_system and db_role.company_id is None:
        sanitized_permissions = sanitize_view_ids(role_update.permissions, current_user.company_id) if role_update.permissions is not None else sanitize_view_ids(DEFAULT_ROLE_VIEW_ACCESS.get(db_role.name, []), current_user.company_id)
        sanitized_menu_order = sanitize_view_ids(role_update.menu_order, current_user.company_id) if role_update.menu_order is not None else sanitize_view_ids(DEFAULT_ROLE_MENU_ORDER.get(db_role.name, sanitized_permissions), current_user.company_id)
        current_assignable_role_ids = get_role_assignable_role_ids(db_role)
        sanitized_assignable_role_ids = sanitize_assignable_role_ids(
            db,
            role_update.assignable_role_ids if role_update.assignable_role_ids is not None else current_assignable_role_ids,
            current_user.company_id
        )
        advisor_tracking_enabled = bool(
            role_update.advisor_tracking_enabled
            if role_update.advisor_tracking_enabled is not None
            else getattr(db_role, "advisor_tracking_enabled", False)
        )
        for view_id in sanitized_permissions:
            if view_id not in sanitized_menu_order:
                sanitized_menu_order.append(view_id)

        override_role = get_company_role_override(db, current_user.company_id, db_role.name)
        if not override_role:
            override_role = models.Role(
                name=f"override_{current_user.company_id}_{db_role.name}",
                label=(role_update.label.strip() if role_update.label is not None else db_role.label),
                company_id=current_user.company_id,
                permissions_json=json.dumps(sanitized_permissions),
                menu_order_json=json.dumps(sanitize_view_ids(sanitized_menu_order, current_user.company_id)),
                auto_assign_leads=bool(role_update.auto_assign_leads if role_update.auto_assign_leads is not None else getattr(db_role, "auto_assign_leads", False)),
                assignable_role_ids_json=json.dumps(sanitized_assignable_role_ids),
                advisor_tracking_enabled=advisor_tracking_enabled,
                tracked_advisor_ids_json=json.dumps([]),
                is_system=False,
                base_role_name=db_role.name
            )
            db.add(override_role)
            db.commit()
            db.refresh(override_role)
        else:
            if role_update.label is not None:
                override_role.label = role_update.label.strip()
            override_role.permissions_json = json.dumps(sanitized_permissions)
            override_role.menu_order_json = json.dumps(sanitize_view_ids(sanitized_menu_order, current_user.company_id))
            if role_update.auto_assign_leads is not None:
                override_role.auto_assign_leads = bool(role_update.auto_assign_leads)
            if role_update.assignable_role_ids is not None:
                override_role.assignable_role_ids_json = json.dumps(sanitized_assignable_role_ids)
            if role_update.advisor_tracking_enabled is not None:
                override_role.advisor_tracking_enabled = advisor_tracking_enabled
                if not advisor_tracking_enabled:
                    override_role.tracked_advisor_ids_json = json.dumps([])
            db.commit()
            db.refresh(override_role)

        db.query(models.User).filter(
            models.User.company_id == current_user.company_id,
            models.User.role_id == db_role.id
        ).update({"role_id": override_role.id}, synchronize_session=False)
        db.commit()
        db.refresh(override_role)
        return serialize_role(override_role)

    if current_user.company_id and db_role.company_id is None:
        raise HTTPException(status_code=403, detail="No puedes modificar roles globales del sistema")

    if role_update.label is not None:
        db_role.label = role_update.label.strip()
    if role_update.auto_assign_leads is not None:
        db_role.auto_assign_leads = bool(role_update.auto_assign_leads)
    if role_update.assignable_role_ids is not None:
        db_role.assignable_role_ids_json = json.dumps(
            sanitize_assignable_role_ids(
                db,
                role_update.assignable_role_ids,
                current_user.company_id or db_role.company_id
            )
        )
    if role_update.advisor_tracking_enabled is not None:
        db_role.advisor_tracking_enabled = bool(role_update.advisor_tracking_enabled)
        if not db_role.advisor_tracking_enabled:
            db_role.tracked_advisor_ids_json = json.dumps([])
    if role_update.permissions is not None:
        sanitized_permissions = sanitize_view_ids(role_update.permissions, current_user.company_id or db_role.company_id)
        db_role.permissions_json = json.dumps(sanitized_permissions)
        current_menu = sanitize_view_ids(role_update.menu_order, current_user.company_id or db_role.company_id) if role_update.menu_order is not None else parse_json_list(db_role.menu_order_json, sanitized_permissions)
        for view_id in sanitized_permissions:
            if view_id not in current_menu:
                current_menu.append(view_id)
        db_role.menu_order_json = json.dumps(sanitize_view_ids(current_menu, current_user.company_id or db_role.company_id))
    elif role_update.menu_order is not None:
        current_permissions = sanitize_view_ids(parse_json_list(db_role.permissions_json, DEFAULT_ROLE_VIEW_ACCESS.get(db_role.name, [])), current_user.company_id or db_role.company_id)
        current_menu = sanitize_view_ids(role_update.menu_order, current_user.company_id or db_role.company_id)
        for view_id in current_permissions:
            if view_id not in current_menu:
                current_menu.append(view_id)
        db_role.menu_order_json = json.dumps(current_menu)

    db.commit()
    db.refresh(db_role)
    return serialize_role(db_role)


@app.delete("/roles/{role_id}", status_code=204)
def delete_role(
    role_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    ensure_role_management_permissions(current_user)
    db_role = db.query(models.Role).filter(models.Role.id == role_id).first()
    if not db_role:
        raise HTTPException(status_code=404, detail="Role not found")
    if db_role.is_system:
        raise HTTPException(status_code=400, detail="System roles cannot be deleted")
    if current_user.company_id and db_role.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    in_use = db.query(models.User).filter(models.User.role_id == role_id).count()
    if in_use:
        raise HTTPException(status_code=400, detail="No se puede eliminar un rol que tiene usuarios asignados")

    db.delete(db_role)
    db.commit()
    return None

@app.put("/companies/{company_id}/integrations", response_model=schemas.IntegrationSettings)
def update_integration_settings(company_id: int, settings_update: schemas.IntegrationSettingsUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    # Check permissions (only admins or super admins)
    role_name = get_user_role_name(current_user)
    if role_name not in ["super_admin", "admin"]:
         raise HTTPException(status_code=403, detail="Not authorized")

    if role_name != "super_admin" and current_user.company_id and current_user.company_id != company_id:
        raise HTTPException(status_code=403, detail="Not authorized to modify these settings")
        
    settings = db.query(models.IntegrationSettings).filter(models.IntegrationSettings.company_id == company_id).first()
    if not settings:
        settings = models.IntegrationSettings(company_id=company_id)
        db.add(settings)
    
    # Update fields. Empty SMTP password must preserve the current saved secret.
    for field, value in settings_update.dict(exclude_unset=True).items():
        if field in {"smtp_password", "whatsapp_calling_provider_token"} and (value is None or str(value).strip() == ""):
            continue
        if field in {"smtp_password_configured", "whatsapp_calling_provider_token_configured"}:
            continue
        setattr(settings, field, value)
        
    try:
        db.commit()
        db.refresh(settings)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
        
    return serialize_integration_settings(settings)


@app.post("/companies/{company_id}/integrations/test-smtp")
def test_company_smtp_settings(
    company_id: int,
    payload: schemas.IntegrationSettingsSmtpTest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    role_name = get_user_role_name(current_user)
    if role_name not in ["super_admin", "admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    if role_name != "super_admin" and current_user.company_id and current_user.company_id != company_id:
        raise HTTPException(status_code=403, detail="Not authorized to test these settings")

    company = db.query(models.Company).filter(models.Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")

    settings = db.query(models.IntegrationSettings).filter(models.IntegrationSettings.company_id == company_id).first()
    saved_password = (getattr(settings, "smtp_password", None) or "").strip() if settings else ""
    password = (payload.smtp_password or "").strip() or saved_password
    host = (payload.smtp_host or "").strip()
    sender = (payload.smtp_from or payload.smtp_username or "").strip()
    username = (payload.smtp_username or "").strip()
    port = int(payload.smtp_port or 587)

    if not payload.smtp_enabled:
        raise HTTPException(status_code=400, detail="Activa SMTP propio para probar esta configuración.")
    if not host or not sender:
        raise HTTPException(status_code=400, detail="Configura servidor SMTP y correo remitente antes de probar.")
    if username and not password:
        raise HTTPException(status_code=400, detail="Falta la contraseña SMTP o una contraseña guardada para este usuario.")

    smtp_settings = {
        "host": host,
        "port": port,
        "username": username,
        "password": password,
        "sender": sender,
        "use_tls": bool(payload.smtp_use_tls),
    }

    message = EmailMessage()
    message["Subject"] = f"Correo de prueba SMTP - {company.name}"
    message["From"] = sender
    message["To"] = str(payload.test_email)
    message.set_content(
        (
            f"Hola.\n\n"
            f"Este es un correo de prueba enviado desde la configuración SMTP de {company.name}.\n"
            f"Si recibiste este mensaje, la configuración de correo está funcionando correctamente.\n"
        ),
        charset="utf-8",
    )
    message.add_alternative(
        f"""
        <html>
          <body style="margin:0;padding:24px;background:#f8fafc;font-family:Arial,sans-serif;color:#0f172a;">
            <div style="max-width:620px;margin:0 auto;background:#ffffff;border:1px solid #e2e8f0;border-radius:18px;padding:24px;">
              <h1 style="margin:0 0 12px;font-size:22px;">Correo de prueba SMTP</h1>
              <p style="font-size:15px;line-height:1.6;">Este mensaje fue enviado desde la configuración SMTP de <strong>{escape(company.name)}</strong>.</p>
              <p style="font-size:14px;line-height:1.6;color:#475569;">Si recibiste este correo, la configuración está funcionando correctamente.</p>
            </div>
          </body>
        </html>
        """.strip(),
        subtype="html",
    )

    try:
        _deliver_smtp_message(message, smtp_settings)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo enviar el correo de prueba: {exc}") from exc

    return {"status": "sent", "message": f"Correo de prueba enviado a {payload.test_email}."}

# --- LEADS ENDPOINTS ---

@app.get("/leads", response_model=schemas.LeadList)
def read_leads(
    source: str = None, 
    status: str = None,
    board_scope: str = None,
    q: str = None,
    skip: int = 0, 
    limit: int = 5000, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    query = apply_lead_access_filters(
        build_lead_summary_query(db),
        db,
        current_user,
        source=source,
        status=status,
        board_scope=board_scope,
        q=q,
    )

    total = query.count()
    leads = query.order_by(models.Lead.id.desc()).offset(skip).limit(limit).all()
    hydrate_lead_summary_fields(db, leads)

    return {"items": leads, "total": total}


@app.get("/leads/board", response_model=schemas.LeadBoardResponse)
def read_leads_board(
    board_scope: str = None,
    q: str = None,
    exact_date: str = None,
    assigned_mode: str = None,
    responsible_user_id: int = None,
    global_status: str = None,
    only_my_leads: bool = False,
    load_all_matching: bool = False,
    status_limits: str = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    company = get_company_by_id(db, current_user.company_id)
    parsed_status_limits: Dict[str, int] = {}
    if status_limits:
        try:
            raw_limits = json.loads(status_limits)
            if isinstance(raw_limits, dict):
                parsed_status_limits = {
                    normalize_lead_status_value(key): max(1, int(value))
                    for key, value in raw_limits.items()
                    if value is not None
                }
        except (TypeError, ValueError, json.JSONDecodeError):
            raise HTTPException(status_code=400, detail="Formato de límites por estado inválido")

    columns: List[schemas.LeadBoardColumn] = []

    for status_key in get_company_allowed_lead_statuses(company):
        normalized_status = normalize_lead_status_value(status_key)
        if global_status and normalize_lead_status_value(global_status) != normalized_status:
            columns.append(
                schemas.LeadBoardColumn(
                    status=normalized_status,
                    label=LEAD_STATUS_LABELS.get(normalized_status, normalized_status),
                    total=0,
                    items=[],
                )
            )
            continue

        base_query = apply_lead_access_filters(
            db.query(models.Lead),
            db,
            current_user,
            board_scope=board_scope,
            q=q,
            exact_date=exact_date,
            assigned_mode=assigned_mode,
            responsible_user_id=responsible_user_id,
            only_my_leads=only_my_leads,
            status=normalized_status,
        )

        total = base_query.count()
        status_limit = total if load_all_matching else parsed_status_limits.get(normalized_status, 10)

        items_query = apply_lead_access_filters(
            build_lead_summary_query(db),
            db,
            current_user,
            board_scope=board_scope,
            q=q,
            exact_date=exact_date,
            assigned_mode=assigned_mode,
            responsible_user_id=responsible_user_id,
            only_my_leads=only_my_leads,
            status=normalized_status,
        )
        recent_activity_field = func.coalesce(models.Lead.status_updated_at, models.Lead.created_at)
        items = items_query.order_by(
            recent_activity_field.desc(),
            models.Lead.created_at.desc(),
            models.Lead.id.desc()
        ).limit(status_limit).all()
        hydrate_lead_summary_fields(db, items)

        columns.append(
            schemas.LeadBoardColumn(
                status=normalized_status,
                label=LEAD_STATUS_LABELS.get(normalized_status, normalized_status),
                total=total,
                items=items,
            )
        )

    return {"columns": columns}


@app.get("/leads/{lead_id:int}", response_model=schemas.Lead)
def get_lead_detail(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    lead = db.query(models.Lead).options(
        joinedload(models.Lead.assigned_to),
        joinedload(models.Lead.created_by),
        joinedload(models.Lead.deleted_by),
        selectinload(models.Lead.supervisors),
        selectinload(models.Lead.history).joinedload(models.LeadHistory.user),
        joinedload(models.Lead.process_detail),
        selectinload(models.Lead.notes).joinedload(models.LeadNote.user),
        selectinload(models.Lead.files).joinedload(models.LeadFile.user),
        selectinload(models.Lead.purchase_options).joinedload(models.PurchaseOption.user),
        selectinload(models.Lead.purchase_options).joinedload(models.PurchaseOption.decision_user),
    ).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    related_records = db.query(models.CreditApplication).filter(
        models.CreditApplication.lead_id == lead.id
    ).order_by(
        models.CreditApplication.updated_at.desc(),
        models.CreditApplication.created_at.desc()
    ).all()
    related_credit = pick_related_credit_record(related_records, lead.status)
    related_purchase = pick_related_purchase_record(related_records)
    lead.status = normalize_lead_status_value(lead.status)
    if related_credit and is_credit_stage_status(lead.status):
        lead.credit_application_id = related_credit.id
        lead.credit_application_status = related_credit.status
        lead.credit_application_updated_at = related_credit.updated_at
    if related_purchase and should_attach_purchase_request(lead, related_purchase):
        lead.purchase_request_id = related_purchase.id
        lead.purchase_request_status = related_purchase.status
        lead.purchase_request_updated_at = related_purchase.updated_at
        lead.purchase_request_notes = related_purchase.notes

    sanitize_lead_assignment_for_response(lead)
    return lead


@app.get("/leads/deleted", response_model=schemas.LeadList)
def read_deleted_leads(
    skip: int = 0,
    limit: int = 500,
    q: str = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver leads eliminados")

    query = db.query(models.Lead).options(
        joinedload(models.Lead.assigned_to),
        joinedload(models.Lead.deleted_by),
        joinedload(models.Lead.supervisors)
    ).filter(models.Lead.deleted_at.is_not(None))

    if current_user.company_id:
        query = query.filter(models.Lead.company_id == current_user.company_id)

    if q:
        search = f"%{q}%"
        query = query.filter(
            or_(
                models.Lead.name.ilike(search),
                models.Lead.email.ilike(search),
                models.Lead.phone.ilike(search),
                models.Lead.deleted_reason.ilike(search)
            )
        )

    total = query.count()
    items = query.order_by(models.Lead.deleted_at.desc(), models.Lead.id.desc()).offset(skip).limit(limit).all()
    return {"items": items, "total": total}


@app.post("/leads/{lead_id}/restore", response_model=schemas.Lead)
def restore_deleted_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Solo administradores pueden restaurar leads eliminados")

    lead = db.query(models.Lead).options(
        joinedload(models.Lead.assigned_to),
        joinedload(models.Lead.created_by),
        joinedload(models.Lead.deleted_by),
        joinedload(models.Lead.supervisors),
        joinedload(models.Lead.history).joinedload(models.LeadHistory.user),
        joinedload(models.Lead.conversation).joinedload(models.Conversation.messages),
        joinedload(models.Lead.process_detail),
        joinedload(models.Lead.notes).joinedload(models.LeadNote.user),
        joinedload(models.Lead.files).joinedload(models.LeadFile.user),
        joinedload(models.Lead.purchase_options).joinedload(models.PurchaseOption.user),
        joinedload(models.Lead.purchase_options).joinedload(models.PurchaseOption.decision_user),
    ).filter(models.Lead.id == lead_id).first()

    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="No tienes permisos para restaurar leads de otra empresa")

    if not lead.deleted_at:
        raise HTTPException(status_code=400, detail="El lead no está eliminado")

    deleted_reason_snapshot = lead.deleted_reason
    lead.deleted_at = None
    lead.deleted_reason = None
    lead.deleted_by_id = None

    db.add(models.LeadHistory(
        lead_id=lead.id,
        user_id=current_user.id,
        previous_status=lead.status,
        new_status=lead.status,
        comment=(
            f"Lead restaurado al tablero."
            f"{f' Motivo original de eliminación: {deleted_reason_snapshot}' if deleted_reason_snapshot else ''}"
        )
    ))

    db.commit()
    db.refresh(lead)
    return lead

@app.put("/leads/bulk-assign", status_code=200)
def bulk_assign_leads(
    payload: schemas.LeadBulkAssign, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Verify assigned_to user exists and is in same company
    target_user = db.query(models.User).join(models.Role, isouter=True).filter(models.User.id == payload.assigned_to_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="Target user not found")
    if not is_active_user(target_user):
        raise HTTPException(status_code=400, detail="No puedes asignar leads a un usuario inactivo")
        
    if current_user.company_id and target_user.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Target user is in a different company")
    if not is_advisor_role(target_user.role):
        raise HTTPException(status_code=400, detail="Solo se pueden asignar leads a usuarios con rol asesor")

    leads_to_update = db.query(models.Lead).options(joinedload(models.Lead.supervisors)).filter(
        models.Lead.id.in_(payload.lead_ids)
    )
    
    if current_user.company_id:
        leads_to_update = leads_to_update.filter(models.Lead.company_id == current_user.company_id)
    
    leads = leads_to_update.all()
    result = 0
    for lead in leads:
        previous_assigned_to_id = lead.assigned_to_id
        lead.assigned_to_id = payload.assigned_to_id
        supervisor_ids = normalize_supervisor_ids(getattr(lead, "supervisor_ids", []))
        if should_keep_assigner_as_supervisor(current_user.id, payload.assigned_to_id):
            supervisor_ids = ensure_user_in_supervisors(supervisor_ids, current_user.id)
        sync_lead_supervisors(db, lead, supervisor_ids, current_user.id)

        if previous_assigned_to_id != payload.assigned_to_id:
            db.add(models.LeadHistory(
                lead_id=lead.id,
                user_id=current_user.id,
                previous_status=lead.status,
                new_status=lead.status,
                comment=f"Lead asignado a {target_user.full_name or target_user.email}; quien asigna queda en supervision"
            ))
            db.add(models.Notification(
                user_id=payload.assigned_to_id,
                title="Lead Asignado",
                message=f"Se te asigno el lead: {lead.name}. Continúa con la gestion.",
                type="info",
                link=build_lead_board_link(target_user, lead.id)
            ))
        result += 1

    db.commit()
    return {"message": f"Successfully assigned {result} leads to user {target_user.email}"}


@app.post("/leads", response_model=schemas.Lead)
def create_lead(lead: schemas.LeadCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    import datetime
    
    # 1. Determine Company
    company_id = lead.company_id
    if not company_id:
        if current_user.company_id:
            company_id = current_user.company_id
        else:
             raise HTTPException(status_code=400, detail="Company ID required for assignment")

    # 2. Assignment Logic
    company = get_company_by_id(db, company_id)
    assigned_user_id = lead.assigned_to_id
    supervisor_ids = normalize_supervisor_ids(getattr(lead, "supervisor_ids", []))
    if supervisor_ids and not is_company_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo un administrador puede agregar o quitar supervisores de un lead.")
    
    # Manual leads created by advisors stay assigned to themselves. Random
    # assignment here is only for super admin-created leads.
    if not assigned_user_id:
        if should_self_assign_manual_lead(current_user):
            assigned_user_id = current_user.id
        elif should_auto_assign_manual_lead(current_user):
            auto_assigned_user = choose_auto_assign_user(db, company_id)
            if auto_assigned_user:
                assigned_user_id = auto_assigned_user.id
            else:
                raise HTTPException(
                    status_code=400,
                    detail="No hay asesores vendedores activos con asignacion automatica habilitada"
                )
    else:
        target_user = db.query(models.User).join(models.Role, isouter=True).filter(models.User.id == assigned_user_id).first()
        if not target_user or target_user.company_id != company_id:
            raise HTTPException(status_code=400, detail="Invalid assigned user (not in company)")
        if not is_active_user(target_user):
            raise HTTPException(status_code=400, detail="No puedes crear un lead asignado a un usuario inactivo")
        if not can_assign_lead_to_user(current_user, target_user):
            raise HTTPException(status_code=400, detail="Tu rol no tiene permitido reasignar leads a este rol")
    
    effective_status = enforce_ally_managed_status(
        db=db,
        current_user=current_user,
        base_status=lead.status,
        assigned_to_id=assigned_user_id
    )
    effective_status = normalize_company_lead_status(
        effective_status,
        company,
        models.LeadStatus.NEW.value
    )
    assigned_user_id, supervisor_ids, credit_coordinator = maybe_assign_credit_coordinator(
        db=db,
        lead_company_id=company_id,
        previous_status=None,
        target_status=effective_status,
        assigned_to_id=assigned_user_id,
        supervisor_ids=supervisor_ids,
        current_user_id=current_user.id
    )
    if should_supervise_manual_lead(current_user.id, assigned_user_id):
        supervisor_ids = ensure_user_in_supervisors(supervisor_ids, current_user.id)

    ensure_company_lead_creation_capacity(db, company_id)

    new_lead = models.Lead(
        created_at=datetime.datetime.now(),
        source=lead.source,
        name=lead.name,
        email=lead.email,
        phone=lead.phone,
        message=lead.message,
        status=effective_status,
        company_id=company_id,
        assigned_to_id=assigned_user_id,
        created_by_id=current_user.id
    )
    db.add(new_lead)
    db.flush()
    sync_lead_supervisors(db, new_lead, supervisor_ids, current_user.id)

    initial_history = models.LeadHistory(
        lead_id=new_lead.id,
        user_id=current_user.id,
        previous_status=None,
        new_status=new_lead.status,
        comment=(
            f"{new_lead.message or 'Lead creado manualmente'} "
            f"(Coordinador de credito añadido como supervisor: {credit_coordinator.full_name or credit_coordinator.email})"
            if credit_coordinator else
            (new_lead.message or "Lead creado manualmente")
        )
    )
    db.add(initial_history)
    db.commit()
    db.refresh(new_lead)
    return new_lead

@app.get("/reports/stats", response_model=schemas.ReportsStats)
def get_reports_stats(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    try:
        # Base query filtered by company
        query = db.query(models.Lead).options(
            joinedload(models.Lead.assigned_to),
            joinedload(models.Lead.supervisors),
            joinedload(models.Lead.purchase_options)
        )
        if current_user.company_id:
            query = query.filter(models.Lead.company_id == current_user.company_id)
        
        leads = query.all()
        total_leads = len(leads)
        ally_user_ids = set(get_company_ally_user_ids(db, current_user.company_id))
        credit_query = db.query(models.CreditApplication)
        purchase_query = db.query(models.CreditApplication).filter(
            models.CreditApplication.lead_id.isnot(None),
            models.CreditApplication.lead.has(
                and_(
                    models.Lead.status == models.LeadStatus.IN_PROCESS.value,
                    models.Lead.process_detail.has(models.LeadProcessDetail.has_vehicle == False),  # noqa: E712
                    models.Lead.process_detail.has(models.LeadProcessDetail.desired_vehicle.isnot(None))
                )
            )
        )
        vehicle_query = db.query(models.Vehicle)
        sales_query = db.query(models.Sale)

        if current_user.company_id:
            credit_query = credit_query.filter(models.CreditApplication.company_id == current_user.company_id)
            purchase_query = purchase_query.filter(models.CreditApplication.company_id == current_user.company_id)
            vehicle_query = vehicle_query.filter(models.Vehicle.company_id == current_user.company_id)
            sales_query = sales_query.filter(models.Sale.company_id == current_user.company_id)

        credit_applications = credit_query.all()
        purchase_requests = purchase_query.all()
        vehicles = vehicle_query.all()
        sales = sales_query.all()
        
        # Calculate Stats
        leads_by_status = {}
        leads_by_source = {}
        leads_by_advisor = {}
        converted_count = 0
        active_pipeline_count = 0
        unread_replies_count = 0
        unread_replies_by_source = {}
        assignment_split = {"assigned": 0, "unassigned": 0}
        ally_status_split = {}
        credit_status_split = {}
        purchase_status_split = {}
        purchase_option_decision_split = {"pending": 0, "accepted": 0, "rejected": 0}
        vehicle_status_split = {}
        sales_status_split = {}

        today = datetime.datetime.utcnow().date()
        recent_dates = [today - datetime.timedelta(days=offset) for offset in range(6, -1, -1)]
        recent_leads_by_day = {day.strftime("%d/%m"): 0 for day in recent_dates}
        ally_board_count = 0
        
        for lead in leads:
            # Status
            status = lead.status or "unknown"
            leads_by_status[status] = leads_by_status.get(status, 0) + 1
            if status == "sold":
                converted_count += 1
            if status not in ["sold", "lost"]:
                active_pipeline_count += 1
                
            # Source
            source = lead.source or "manual"
            leads_by_source[source] = leads_by_source.get(source, 0) + 1
            
            # Advisor
            advisor_name = "Sin Asignar"
            if lead.assigned_to:
                advisor_name = lead.assigned_to.email
                assignment_split["assigned"] += 1
            else:
                assignment_split["unassigned"] += 1
            leads_by_advisor[advisor_name] = leads_by_advisor.get(advisor_name, 0) + 1

            if lead.has_unread_reply:
                unread_replies_count += 1
                unread_replies_by_source[source] = unread_replies_by_source.get(source, 0) + 1

            supervisor_ids = {supervisor.id for supervisor in (lead.supervisors or []) if supervisor and supervisor.id}
            touches_ally_board = bool(
                ally_user_ids and (
                    (lead.assigned_to_id in ally_user_ids) or
                    bool(supervisor_ids & ally_user_ids)
                )
            )
            if touches_ally_board:
                ally_board_count += 1
                ally_status_split[status] = ally_status_split.get(status, 0) + 1

            created_at = lead.created_at.date() if lead.created_at else None
            if created_at in recent_dates:
                recent_leads_by_day[created_at.strftime("%d/%m")] += 1

            for option in lead.purchase_options or []:
                decision_status = (getattr(option, "decision_status", None) or "pending").lower()
                if decision_status not in purchase_option_decision_split:
                    purchase_option_decision_split[decision_status] = 0
                purchase_option_decision_split[decision_status] += 1

        for application in credit_applications:
            status_key = application.status or "pending"
            credit_status_split[status_key] = credit_status_split.get(status_key, 0) + 1

        for purchase in purchase_requests:
            status_key = purchase.status or "pending"
            purchase_status_split[status_key] = purchase_status_split.get(status_key, 0) + 1

        for vehicle in vehicles:
            status_key = vehicle.status or "available"
            vehicle_status_split[status_key] = vehicle_status_split.get(status_key, 0) + 1

        for sale in sales:
            status_key = sale.status or "pending"
            sales_status_split[status_key] = sales_status_split.get(status_key, 0) + 1

        conversion_rate = (converted_count / total_leads * 100) if total_leads > 0 else 0.0

        return {
            "total_leads": total_leads,
            "conversion_rate": round(conversion_rate, 2),
            "active_pipeline_count": active_pipeline_count,
            "unread_replies_count": unread_replies_count,
            "ally_board_count": ally_board_count,
            "credit_applications_count": len(credit_applications),
            "purchase_requests_count": len(purchase_requests),
            "available_inventory_count": vehicle_status_split.get("available", 0),
            "approved_sales_count": sales_status_split.get("approved", 0),
            "pending_sales_count": sales_status_split.get("pending", 0),
            "leads_by_status": leads_by_status,
            "leads_by_source": leads_by_source,
            "leads_by_advisor": leads_by_advisor,
            "recent_leads_by_day": recent_leads_by_day,
            "unread_replies_by_source": unread_replies_by_source,
            "assignment_split": assignment_split,
            "ally_status_split": ally_status_split,
            "credit_status_split": credit_status_split,
            "purchase_status_split": purchase_status_split,
            "purchase_option_decision_split": purchase_option_decision_split,
            "vehicle_status_split": vehicle_status_split,
            "sales_status_split": sales_status_split,
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# --- STATIC FILES & UPLOAD ---
from fastapi.staticfiles import StaticFiles
from fastapi import UploadFile, File, Form
import shutil
import os
import uuid

# Ensure static directory exists
os.makedirs("static", exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.post("/upload/")
async def upload_image(request: Request, file: UploadFile = File(...)):
    # Generate unique filename
    file_extension = file.filename.split(".")[-1]
    unique_filename = f"{uuid.uuid4()}.{file_extension}"
    file_path = f"static/{unique_filename}"
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    relative_url = f"/api/static/{unique_filename}"
    base_url = str(request.base_url).rstrip("/")
    # Behind reverse proxies, base_url can already include /api.
    if base_url.endswith("/api"):
        base_url = base_url[:-4]
    return {
        "url": f"{base_url}{relative_url}",
        "url_relative": relative_url
    }

@app.get("/internal-files/{storage_name}")
def get_internal_file(storage_name: str):
    safe_name = os.path.basename(storage_name)
    base_dir = os.path.abspath("static/internal_chat")
    file_path = os.path.abspath(os.path.join(base_dir, safe_name))

    if not file_path.startswith(base_dir):
        raise HTTPException(status_code=400, detail="Invalid file path")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(path=file_path, filename=safe_name)

# --- BRANDS & MODELS ENDPOINTS ---

@app.get("/brands/", response_model=List[schemas.CarBrand])
def read_brands(db: Session = Depends(get_db)):
    return db.query(models.CarBrand).order_by(models.CarBrand.name).all()

@app.get("/brands/{brand_id}/models/", response_model=List[schemas.CarModel])
def read_models_by_brand(brand_id: int, db: Session = Depends(get_db)):
    return db.query(models.CarModel).filter(models.CarModel.brand_id == brand_id).order_by(models.CarModel.name).all()

def normalize_catalog_value(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = " ".join(str(value).strip().split())
    return cleaned if cleaned else None

def catalog_key(value: Optional[str]) -> str:
    normalized = normalize_catalog_value(value) or ""
    return re.sub(r"[^a-z0-9]", "", normalized.lower())

def upsert_brand_model(db: Session, make: Optional[str], model: Optional[str]):
    make_name = normalize_catalog_value(make)
    model_name = normalize_catalog_value(model)
    if not make_name:
        return

    brand = db.query(models.CarBrand).filter(func.lower(models.CarBrand.name) == make_name.lower()).first()
    if not brand:
        brand = models.CarBrand(name=make_name)
        db.add(brand)
        db.flush()

    if model_name:
        existing_model = db.query(models.CarModel).filter(
            models.CarModel.brand_id == brand.id,
            func.lower(models.CarModel.name) == model_name.lower()
        ).first()
        if not existing_model:
            db.add(models.CarModel(name=model_name, brand_id=brand.id))


def get_dashboard_period_bounds(
    period: Optional[str],
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    oldest_record_at: Optional[datetime.datetime] = None
):
    normalized_period = (period or "month").strip().lower()
    now = datetime.datetime.utcnow()

    if start_date and end_date:
        try:
            start_day = datetime.date.fromisoformat(start_date)
            end_day = datetime.date.fromisoformat(end_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Las fechas del dashboard deben usar formato YYYY-MM-DD")

        if end_day < start_day:
            raise HTTPException(status_code=400, detail="La fecha final no puede ser menor a la inicial")

        start = datetime.datetime.combine(start_day, datetime.time.min)
        end = datetime.datetime.combine(end_day + datetime.timedelta(days=1), datetime.time.min)
        total_days = (end_day - start_day).days + 1

        if total_days <= 31:
            bucket_labels = [
                (start_day + datetime.timedelta(days=offset)).strftime("%d/%m")
                for offset in range(total_days)
            ]

            def bucket_label(dt_value: datetime.datetime):
                return dt_value.strftime("%d/%m")

            return "custom", start, end, bucket_labels, bucket_label

        month_cursor = datetime.date(start_day.year, start_day.month, 1)
        month_end = datetime.date(end_day.year, end_day.month, 1)
        bucket_labels = []
        while month_cursor <= month_end:
            bucket_labels.append(month_cursor.strftime("%m/%Y"))
            if month_cursor.month == 12:
                month_cursor = datetime.date(month_cursor.year + 1, 1, 1)
            else:
                month_cursor = datetime.date(month_cursor.year, month_cursor.month + 1, 1)

        def bucket_label(dt_value: datetime.datetime):
            return dt_value.strftime("%m/%Y")

        return "custom", start, end, bucket_labels, bucket_label

    if normalized_period == "all":
        oldest = oldest_record_at or now
        start = oldest.replace(hour=0, minute=0, second=0, microsecond=0)
        end = (now + datetime.timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        start_month = datetime.date(start.year, start.month, 1)
        end_month = datetime.date(end.year, end.month, 1)
        total_months = max(1, (end_month.year - start_month.year) * 12 + (end_month.month - start_month.month) + 1)

        if total_months <= 24:
            bucket_labels = []
            month_cursor = start_month
            while month_cursor <= end_month:
                bucket_labels.append(month_cursor.strftime("%m/%Y"))
                if month_cursor.month == 12:
                    month_cursor = datetime.date(month_cursor.year + 1, 1, 1)
                else:
                    month_cursor = datetime.date(month_cursor.year, month_cursor.month + 1, 1)

            def bucket_label(dt_value: datetime.datetime):
                return dt_value.strftime("%m/%Y")

            return normalized_period, start, end, bucket_labels, bucket_label

        start_year = start.year
        end_year = end.year
        bucket_labels = [str(year) for year in range(start_year, end_year + 1)]

        def bucket_label(dt_value: datetime.datetime):
            return str(dt_value.year)

        return normalized_period, start, end, bucket_labels, bucket_label

    if normalized_period == "day":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + datetime.timedelta(days=1)
        bucket_labels = [f"{hour:02d}:00" for hour in range(24)]

        def bucket_label(dt_value: datetime.datetime):
            return f"{dt_value.hour:02d}:00"

        return normalized_period, start, end, bucket_labels, bucket_label

    if normalized_period == "year":
        start = datetime.datetime(now.year, 1, 1)
        end = datetime.datetime(now.year + 1, 1, 1)
        bucket_labels = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

        def bucket_label(dt_value: datetime.datetime):
            return bucket_labels[dt_value.month - 1]

        return normalized_period, start, end, bucket_labels, bucket_label

    normalized_period = "month"
    start = datetime.datetime(now.year, now.month, 1)
    if now.month == 12:
        end = datetime.datetime(now.year + 1, 1, 1)
    else:
        end = datetime.datetime(now.year, now.month + 1, 1)
    total_days = (end - start).days
    bucket_labels = [f"{day:02d}" for day in range(1, total_days + 1)]

    def bucket_label(dt_value: datetime.datetime):
        return f"{dt_value.day:02d}"

    return normalized_period, start, end, bucket_labels, bucket_label

@app.get("/stats/advisor", response_model=schemas.RoleDashboardStats)
def read_advisor_stats(
    period: str = Query("month"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="Dashboard disponible solo para usuarios de empresa")

    role_name = get_user_role_name(current_user) or ""
    permissions = set(get_role_permissions(current_user.role))
    company_scope = role_name in {"admin", "super_admin"}
    ally_user_ids = set(get_company_ally_user_ids(db, current_user.company_id))
    tracked_advisor_ids = get_user_tracked_advisor_ids(current_user)

    visible_leads_query = db.query(models.Lead).filter(
        models.Lead.company_id == current_user.company_id
    )
    if not company_scope:
        visibility_conditions = [
            models.Lead.assigned_to_id == current_user.id,
            models.Lead.supervisors.any(models.User.id == current_user.id)
        ]
        if tracked_advisor_ids:
            visibility_conditions.extend([
                models.Lead.assigned_to_id.in_(tracked_advisor_ids),
                models.Lead.supervisors.any(models.User.id.in_(tracked_advisor_ids))
            ])
        visible_leads_query = visible_leads_query.filter(or_(*visibility_conditions))

    oldest_visible_lead = visible_leads_query.order_by(models.Lead.created_at.asc()).first()
    normalized_period, period_start, period_end, trend_labels, get_trend_bucket = get_dashboard_period_bounds(
        period,
        start_date,
        end_date,
        getattr(oldest_visible_lead, "created_at", None)
    )

    leads_query = visible_leads_query.options(
        joinedload(models.Lead.supervisors),
        joinedload(models.Lead.purchase_options)
    )
    all_leads = leads_query.all()

    def is_within_dashboard_range(dt_value: Optional[datetime.datetime]) -> bool:
        return bool(dt_value and period_start <= dt_value < period_end)

    def lead_activity_at(lead: models.Lead) -> Optional[datetime.datetime]:
        activity_candidates = [lead.last_reply_at, lead.status_updated_at, lead.created_at]
        return max((candidate for candidate in activity_candidates if candidate is not None), default=None)

    leads = [lead for lead in all_leads if is_within_dashboard_range(lead_activity_at(lead))]

    status_distribution = {}
    ally_status_distribution = {}
    source_distribution = {}
    ally_source_distribution = {}
    unread_replies_count = 0
    ally_unread_replies_count = 0
    active_pipeline_count = 0
    ally_active_pipeline_count = 0
    purchase_option_decision_distribution = {"pending": 0, "accepted": 0, "rejected": 0}
    recent_leads_by_day = {label: 0 for label in trend_labels}
    ally_recent_leads_by_day = {label: 0 for label in trend_labels}
    tracked_advisor_users = []
    if tracked_advisor_ids:
        tracked_advisor_users = db.query(models.User).options(
            joinedload(models.User.role)
        ).filter(
            models.User.id.in_(tracked_advisor_ids),
            models.User.company_id == current_user.company_id
        ).all()
    supervised_advisor_map: Dict[int, Dict[str, Any]] = {
        advisor.id: {
            "user_id": advisor.id,
            "full_name": advisor.full_name,
            "email": advisor.email,
            "role_label": getattr(getattr(advisor, "role", None), "label", None),
            "total_leads": 0,
            "leads_new": 0,
            "leads_sold": 0,
            "active_pipeline_count": 0,
            "unread_replies_count": 0,
            "new_leads_in_range": 0,
            "status_changes_in_range": 0,
            "status_distribution": {},
        }
        for advisor in tracked_advisor_users
    }
    supervised_advisor_ids_by_lead: Dict[int, List[int]] = {}

    relevant_lead_ids = []
    ally_lead_ids = []
    autos_lead_ids = []
    ally_total = 0
    ally_leads_new = 0
    ally_leads_sold = 0
    ally_new_leads_in_range = 0
    autos_new_leads_in_range = 0
    for lead in leads:
        relevant_lead_ids.append(lead.id)
        status_key = lead.status or "new"
        lead_reference_date = lead_activity_at(lead)

        supervisor_ids = {supervisor.id for supervisor in (lead.supervisors or []) if supervisor and supervisor.id}
        touched_tracked_advisor_ids = [
            advisor_id for advisor_id in tracked_advisor_ids
            if advisor_id in supervised_advisor_map
            and (lead.assigned_to_id == advisor_id or advisor_id in supervisor_ids)
        ]
        if touched_tracked_advisor_ids:
            supervised_advisor_ids_by_lead[lead.id] = touched_tracked_advisor_ids
            for advisor_id in touched_tracked_advisor_ids:
                current_supervised = supervised_advisor_map[advisor_id]
                current_supervised["total_leads"] += 1
                current_supervised["status_distribution"][status_key] = current_supervised["status_distribution"].get(status_key, 0) + 1
                if status_key == "new":
                    current_supervised["leads_new"] += 1
                if status_key == "sold":
                    current_supervised["leads_sold"] += 1
                if status_key not in {"sold", "lost"}:
                    current_supervised["active_pipeline_count"] += 1
                if lead.has_unread_reply:
                    current_supervised["unread_replies_count"] += 1
                if is_within_dashboard_range(getattr(lead, "created_at", None)):
                    current_supervised["new_leads_in_range"] += 1
        touches_ally_board = bool(
            ally_user_ids and (
                (lead.assigned_to_id in ally_user_ids) or
                bool(supervisor_ids & ally_user_ids)
            )
        )
        if touches_ally_board:
            ally_total += 1
            ally_lead_ids.append(lead.id)
            ally_status_distribution[status_key] = ally_status_distribution.get(status_key, 0) + 1
            ally_source_key = (lead.source or "sin_fuente").strip().lower() or "sin_fuente"
            ally_source_distribution[ally_source_key] = ally_source_distribution.get(ally_source_key, 0) + 1
            if status_key == "new":
                ally_leads_new += 1
            if status_key == "sold":
                ally_leads_sold += 1
            if status_key not in {"sold", "lost"}:
                ally_active_pipeline_count += 1
            if lead.has_unread_reply:
                ally_unread_replies_count += 1
            if is_within_dashboard_range(getattr(lead, "created_at", None)):
                ally_new_leads_in_range += 1
            if lead_reference_date and period_start <= lead_reference_date < period_end:
                trend_bucket = get_trend_bucket(lead_reference_date)
                if trend_bucket in ally_recent_leads_by_day:
                    ally_recent_leads_by_day[trend_bucket] += 1
            continue

        autos_lead_ids.append(lead.id)
        status_distribution[status_key] = status_distribution.get(status_key, 0) + 1
        source_key = (lead.source or "sin_fuente").strip().lower() or "sin_fuente"
        source_distribution[source_key] = source_distribution.get(source_key, 0) + 1
        if status_key not in {"sold", "lost"}:
            active_pipeline_count += 1
        if lead.has_unread_reply:
            unread_replies_count += 1
        if is_within_dashboard_range(getattr(lead, "created_at", None)):
            autos_new_leads_in_range += 1
        if lead_reference_date and period_start <= lead_reference_date < period_end:
            trend_bucket = get_trend_bucket(lead_reference_date)
            if trend_bucket in recent_leads_by_day:
                recent_leads_by_day[trend_bucket] += 1
        for option in lead.purchase_options or []:
            decision_status = (getattr(option, "decision_status", None) or "pending").lower()
            if decision_status not in purchase_option_decision_distribution:
                purchase_option_decision_distribution[decision_status] = 0
            purchase_option_decision_distribution[decision_status] += 1

    total_leads = len(autos_lead_ids)
    leads_sold = status_distribution.get("sold", 0)
    leads_new = status_distribution.get("new", 0)
    conversion_rate = (leads_sold / total_leads * 100) if total_leads else 0
    ally_conversion_rate = (ally_leads_sold / ally_total * 100) if ally_total else 0
    visible_lead_ids = [lead.id for lead in all_leads]
    autos_visible_lead_ids = set()
    ally_visible_lead_ids = set()
    for lead in all_leads:
        supervisor_ids = {supervisor.id for supervisor in (lead.supervisors or []) if supervisor and supervisor.id}
        touches_ally_board = bool(
            ally_user_ids and (
                (lead.assigned_to_id in ally_user_ids) or
                bool(supervisor_ids & ally_user_ids)
            )
        )
        if touches_ally_board:
            ally_visible_lead_ids.add(lead.id)
        else:
            autos_visible_lead_ids.add(lead.id)

    history_entries = []
    if visible_lead_ids:
        history_entries = db.query(models.LeadHistory).options(
            joinedload(models.LeadHistory.user).joinedload(models.User.role)
        ).filter(
            models.LeadHistory.lead_id.in_(visible_lead_ids),
            models.LeadHistory.created_at >= period_start,
            models.LeadHistory.created_at < period_end
        ).all()

    def is_automatic_alert_history(entry: models.LeadHistory) -> bool:
        comment = normalize_role_text(getattr(entry, "comment", None))
        if not comment:
            return False
        return (
            "[auto_alert]" in comment
            or "apariciones de alerta" in comment
            or "por regla" in comment and "alerta" in comment
            or "lead reasignado automaticamente" in comment
            or "lead reasignado automáticamente" in comment
        )

    human_history_entries = [
        entry for entry in history_entries
        if not is_automatic_alert_history(entry)
    ]
    status_change_entries = [
        entry for entry in human_history_entries
        if (entry.previous_status or "") != (entry.new_status or "")
    ]
    ally_status_change_entries = [
        entry for entry in status_change_entries
        if entry.lead_id in ally_visible_lead_ids
    ]
    autos_status_change_entries = [
        entry for entry in status_change_entries
        if entry.lead_id in autos_visible_lead_ids
    ]
    autos_history_entries = [
        entry for entry in human_history_entries
        if entry.lead_id in autos_visible_lead_ids
    ]
    ally_history_entries = [
        entry for entry in human_history_entries
        if entry.lead_id in ally_visible_lead_ids
    ]
    status_changes_in_range = len(autos_status_change_entries)
    ally_status_changes_in_range = len(ally_status_change_entries)
    for entry in status_change_entries:
        for advisor_id in supervised_advisor_ids_by_lead.get(entry.lead_id, []):
            if advisor_id in supervised_advisor_map:
                supervised_advisor_map[advisor_id]["status_changes_in_range"] += 1
    manager_activity_map: Dict[int, Dict[str, Any]] = {}
    status_mover_map: Dict[int, Dict[str, Any]] = {}
    ally_manager_activity_map: Dict[int, Dict[str, Any]] = {}
    for entry in autos_history_entries:
        if not entry.user_id:
            continue
        current_item = manager_activity_map.setdefault(
            entry.user_id,
            {
                "user_id": entry.user_id,
                "full_name": getattr(getattr(entry, "user", None), "full_name", None),
                "email": getattr(getattr(entry, "user", None), "email", None),
                "role_name": get_user_role_name(getattr(entry, "user", None)),
                "role_label": getattr(getattr(getattr(entry, "user", None), "role", None), "label", None),
                "count": 0,
            }
        )
        current_item["count"] += 1
        if not current_item.get("full_name"):
            current_item["full_name"] = getattr(getattr(entry, "user", None), "full_name", None)
        if not current_item.get("email"):
            current_item["email"] = getattr(getattr(entry, "user", None), "email", None)
        if not current_item.get("role_name"):
            current_item["role_name"] = get_user_role_name(getattr(entry, "user", None))
        if not current_item.get("role_label"):
            current_item["role_label"] = getattr(getattr(getattr(entry, "user", None), "role", None), "label", None)
    for entry in autos_status_change_entries:
        if not entry.user_id:
            continue
        current_item = status_mover_map.setdefault(
            entry.user_id,
            {
                "user_id": entry.user_id,
                "full_name": getattr(getattr(entry, "user", None), "full_name", None),
                "email": getattr(getattr(entry, "user", None), "email", None),
                "role_name": get_user_role_name(getattr(entry, "user", None)),
                "role_label": getattr(getattr(getattr(entry, "user", None), "role", None), "label", None),
                "count": 0,
            }
        )
        current_item["count"] += 1
        if not current_item.get("full_name"):
            current_item["full_name"] = getattr(getattr(entry, "user", None), "full_name", None)
        if not current_item.get("email"):
            current_item["email"] = getattr(getattr(entry, "user", None), "email", None)
        if not current_item.get("role_name"):
            current_item["role_name"] = get_user_role_name(getattr(entry, "user", None))
        if not current_item.get("role_label"):
            current_item["role_label"] = getattr(getattr(getattr(entry, "user", None), "role", None), "label", None)
    for entry in ally_history_entries:
        if not entry.user_id:
            continue
        current_item = ally_manager_activity_map.setdefault(
            entry.user_id,
            {
                "user_id": entry.user_id,
                "full_name": getattr(getattr(entry, "user", None), "full_name", None),
                "email": getattr(getattr(entry, "user", None), "email", None),
                "role_name": get_user_role_name(getattr(entry, "user", None)),
                "role_label": getattr(getattr(getattr(entry, "user", None), "role", None), "label", None),
                "count": 0,
            }
        )
        current_item["count"] += 1
        if not current_item.get("full_name"):
            current_item["full_name"] = getattr(getattr(entry, "user", None), "full_name", None)
        if not current_item.get("email"):
            current_item["email"] = getattr(getattr(entry, "user", None), "email", None)
        if not current_item.get("role_name"):
            current_item["role_name"] = get_user_role_name(getattr(entry, "user", None))
        if not current_item.get("role_label"):
            current_item["role_label"] = getattr(getattr(getattr(entry, "user", None), "role", None), "label", None)
    top_managers = sorted(
        manager_activity_map.values(),
        key=lambda item: (-item["count"], item.get("full_name") or item.get("email") or "")
    )
    top_status_movers = sorted(
        status_mover_map.values(),
        key=lambda item: (-item["count"], item.get("full_name") or item.get("email") or "")
    )
    ally_top_managers = sorted(
        ally_manager_activity_map.values(),
        key=lambda item: (-item["count"], item.get("full_name") or item.get("email") or "")
    )

    credit_status_distribution = {}
    credit_total = 0
    if "credits" in permissions:
        credits_query = db.query(models.CreditApplication).filter(
            models.CreditApplication.company_id == current_user.company_id
        )
        credits = [
            credit for credit in credits_query.all()
            if is_within_dashboard_range(getattr(credit, "updated_at", None) or getattr(credit, "created_at", None))
        ]
        credit_total = len(credits)
        for credit in credits:
            status_key = credit.status or "pending"
            credit_status_distribution[status_key] = credit_status_distribution.get(status_key, 0) + 1

    purchase_status_distribution = {}
    purchase_total = 0
    if "purchase_board" in permissions:
        purchases_query = db.query(models.CreditApplication).filter(
            models.CreditApplication.company_id == current_user.company_id
        )
        purchases = [
            purchase for purchase in purchases_query.all()
            if is_within_dashboard_range(getattr(purchase, "updated_at", None) or getattr(purchase, "created_at", None))
        ]
        purchase_total = len(purchases)
        for purchase in purchases:
            status_key = purchase.status or "pending"
            purchase_status_distribution[status_key] = purchase_status_distribution.get(status_key, 0) + 1

    sales_status_distribution = {}
    sales_total = 0
    sales_approved = 0
    sales_pending = 0
    if "sales" in permissions or "my_sales" in permissions:
        sales_query = db.query(models.Sale).filter(
            models.Sale.company_id == current_user.company_id
        )
        if "my_sales" in permissions and not company_scope:
            sales_query = sales_query.filter(models.Sale.seller_id == current_user.id)
        sales = [
            sale for sale in sales_query.all()
            if is_within_dashboard_range(getattr(sale, "sale_date", None) or getattr(sale, "created_at", None))
        ]
        sales_total = len(sales)
        for sale in sales:
            status_key = sale.status or "pending"
            sales_status_distribution[status_key] = sales_status_distribution.get(status_key, 0) + 1
        sales_approved = sales_status_distribution.get("approved", 0)
        sales_pending = sales_status_distribution.get("pending", 0)

    inventory_status_distribution = {}
    inventory_total = 0
    if "inventory" in permissions:
        vehicles = db.query(models.Vehicle).filter(
            models.Vehicle.company_id == current_user.company_id
        ).all()
        inventory_total = len(vehicles)
        for vehicle in vehicles:
            status_key = vehicle.status or "available"
            inventory_status_distribution[status_key] = inventory_status_distribution.get(status_key, 0) + 1

    appointments_total = 0
    appointments_today = 0
    appointments_upcoming = 0
    appointments_by_user_map: Dict[int, Dict[str, Any]] = {}
    if "appointments_calendar" in permissions:
        appointments_query = db.query(models.LeadAppointment).options(
            joinedload(models.LeadAppointment.user),
            joinedload(models.LeadAppointment.lead)
        ).join(
            models.Lead,
            models.Lead.id == models.LeadAppointment.lead_id
        ).filter(
            models.Lead.company_id == current_user.company_id
        )

        if not company_scope:
            appointments_query = appointments_query.filter(models.LeadAppointment.user_id == current_user.id)

        appointments = [
            appointment for appointment in appointments_query.all()
            if is_within_dashboard_range(getattr(appointment, "appointment_date", None))
        ]

        appointments_total = len(appointments)
        now_bogota = datetime.datetime.now(BOGOTA_TZ).replace(tzinfo=None)
        today_bogota = now_bogota.date()

        for appointment in appointments:
            appointment_date = getattr(appointment, "appointment_date", None)
            if not appointment_date:
                continue
            if appointment_date.date() == today_bogota:
                appointments_today += 1
            if appointment_date >= now_bogota and (appointment.status or "scheduled") != "cancelled":
                appointments_upcoming += 1

            if not appointment.user_id:
                continue

            current_item = appointments_by_user_map.setdefault(
                appointment.user_id,
                {
                    "user_id": appointment.user_id,
                    "full_name": getattr(getattr(appointment, "user", None), "full_name", None),
                    "email": getattr(getattr(appointment, "user", None), "email", None),
                    "role_name": get_user_role_name(getattr(appointment, "user", None)),
                    "role_label": getattr(getattr(getattr(appointment, "user", None), "role", None), "label", None),
                    "count": 0,
                }
            )
            current_item["count"] += 1
            if not current_item.get("full_name"):
                current_item["full_name"] = getattr(getattr(appointment, "user", None), "full_name", None)
            if not current_item.get("email"):
                current_item["email"] = getattr(getattr(appointment, "user", None), "email", None)
            if not current_item.get("role_name"):
                current_item["role_name"] = get_user_role_name(getattr(appointment, "user", None))
            if not current_item.get("role_label"):
                current_item["role_label"] = getattr(getattr(getattr(appointment, "user", None), "role", None), "label", None)

    appointments_by_user = sorted(
        appointments_by_user_map.values(),
        key=lambda item: (-item["count"], item.get("full_name") or item.get("email") or "")
    )
    supervised_advisors = sorted(
        supervised_advisor_map.values(),
        key=lambda item: (
            -item["total_leads"],
            item.get("full_name") or item.get("email") or ""
        )
    )

    return {
        "total_leads": total_leads,
        "leads_new": leads_new,
        "leads_sold": leads_sold,
        "ally_total": ally_total,
        "ally_leads_new": ally_leads_new,
        "ally_leads_sold": ally_leads_sold,
        "ally_new_leads_in_range": ally_new_leads_in_range,
        "ally_status_changes_in_range": ally_status_changes_in_range,
        "new_leads_in_range": autos_new_leads_in_range,
        "status_changes_in_range": status_changes_in_range,
        "conversion_rate": round(conversion_rate, 2),
        "ally_conversion_rate": round(ally_conversion_rate, 2),
        "response_time_min": 37,
        "active_pipeline_count": active_pipeline_count,
        "ally_active_pipeline_count": ally_active_pipeline_count,
        "unread_replies_count": unread_replies_count,
        "ally_unread_replies_count": ally_unread_replies_count,
        "status_distribution": status_distribution,
        "ally_status_distribution": ally_status_distribution,
        "source_distribution": source_distribution,
        "ally_source_distribution": ally_source_distribution,
        "recent_leads_by_day": recent_leads_by_day,
        "ally_recent_leads_by_day": ally_recent_leads_by_day,
        "credit_total": credit_total,
        "credit_status_distribution": credit_status_distribution,
        "purchase_total": purchase_total,
        "purchase_status_distribution": purchase_status_distribution,
        "purchase_option_decision_distribution": purchase_option_decision_distribution,
        "sales_total": sales_total,
        "sales_approved": sales_approved,
        "sales_pending": sales_pending,
        "sales_status_distribution": sales_status_distribution,
        "inventory_total": inventory_total,
        "inventory_status_distribution": inventory_status_distribution,
        "appointments_total": appointments_total,
        "appointments_today": appointments_today,
        "appointments_upcoming": appointments_upcoming,
        "top_managers": top_managers,
        "top_status_movers": top_status_movers,
        "ally_top_managers": ally_top_managers,
        "appointments_by_user": appointments_by_user,
        "supervised_advisors": supervised_advisors,
    }

@app.post("/seed/brands")
def seed_brands(db: Session = Depends(get_db)):
    # Data for common cars in Colombia
    data = {
        "Chevrolet": ["Spark", "Sail", "Onix", "Tracker", "Captiva", "Colorado", "Tahoe", "Joy", "Beat", "Equinox"],
        "Renault": ["Logan", "Sandero", "Stepway", "Duster", "Kwid", "Captur", "Koleos", "Alaskan", "Oroch", "Twingo"],
        "Mazda": ["Mazda 2", "Mazda 3", "Mazda 6", "CX-30", "CX-5", "CX-50", "CX-9", "CX-90", "MX-5"],
        "Kia": ["Picanto", "Rio", "Cerato", "Sportage", "Sorento", "Niro", "Stonic", "Soluto", "Eko Taxi"],
        "Toyota": ["Hilux", "Fortuner", "Corolla", "Prado", "Yaris", "Rav4", "Land Cruiser", "4Runner", "TXL"],
        "Nissan": ["Versa", "March", "Sentra", "Kicks", "Qashqai", "X-Trail", "Frontier", "Murano", "Patrol"],
        "Ford": ["Fiesta", "Focus", "Ecosport", "Escape", "Edge", "Explorer", "Ranger", "F-150", "Bronco"],
        "Volkswagen": ["Gol", "Voyage", "Polo", "Virtus", "T-Cross", "Nivus", "Amarok", "Jetta", "Tiguan"],
        "Hyundai": ["Grand i10", "Accent", "HB20", "Tucson", "Santa Fe", "Creta", "Elantra", "Venue", "Palisade"],
        "Suzuki": ["Swift", "Alto", "S-Presso", "Vitara", "Jimny", "Baleno", "Ertiga"],
        "Honda": ["Civic", "CR-V", "HR-V", "Pilot", "City", "WR-V"],
        "BMW": ["Serie 1", "Serie 2", "Serie 3", "Serie 4", "Serie 5", "X1", "X3", "X4", "X5"],
        "Mercedes-Benz": ["Clase A", "Clase C", "Clase E", "Clase GLA", "Clase GLC", "Clase GLE"],
        "Audi": ["A3", "A4", "Q2", "Q3", "Q5", "Q7", "Q8"]
    }

    count = 0
    for brand_name, models_list in data.items():
        # Check if brand exists
        brand = db.query(models.CarBrand).filter(models.CarBrand.name == brand_name).first()
        if not brand:
            brand = models.CarBrand(name=brand_name)
            db.add(brand)
            db.commit()
            db.refresh(brand)
        
        for model_name in models_list:
            # Check if model exists
            exists = db.query(models.CarModel).filter(
                models.CarModel.brand_id == brand.id,
                models.CarModel.name == model_name
            ).first()
            if not exists:
                new_model = models.CarModel(name=model_name, brand_id=brand.id)
                db.add(new_model)
                count += 1
    
    db.commit()
    return {"message": f"Seeded {count} new models successfully."}

@app.post("/seed/brands/from-vehicles")
def seed_brands_from_vehicles(db: Session = Depends(get_db)):
    rows = db.query(models.Vehicle.make, models.Vehicle.model).all()
    if not rows:
        return {
            "message": "No hay vehiculos para sincronizar.",
            "brands_created": 0,
            "models_created": 0
        }

    existing_brands = db.query(models.CarBrand).all()
    brands_by_key = {b.name.lower(): b for b in existing_brands if b.name}

    existing_models = db.query(models.CarModel).all()
    models_by_key = {
        (m.brand_id, m.name.lower()): m
        for m in existing_models
        if m.name and m.brand_id
    }

    brands_created = 0
    models_created = 0

    for make_raw, model_raw in rows:
        make_name = normalize_catalog_value(make_raw)
        model_name = normalize_catalog_value(model_raw)

        if not make_name:
            continue

        brand_key = make_name.lower()
        brand = brands_by_key.get(brand_key)
        if not brand:
            brand = models.CarBrand(name=make_name)
            db.add(brand)
            db.flush()
            brands_by_key[brand_key] = brand
            brands_created += 1

        if not model_name:
            continue

        model_key = (brand.id, model_name.lower())
        if model_key not in models_by_key:
            new_model = models.CarModel(name=model_name, brand_id=brand.id)
            db.add(new_model)
            db.flush()
            models_by_key[model_key] = new_model
            models_created += 1

    db.commit()
    return {
        "message": "Sincronizacion completada",
        "brands_created": brands_created,
        "models_created": models_created
    }

@app.post("/seed/brands/external")
def seed_brands_external(
    max_makes: int = Query(80, ge=1, le=250),
    max_models_per_make: int = Query(120, ge=1, le=500),
    include_models: bool = Query(True),
    only_common: bool = Query(True),
    db: Session = Depends(get_db)
):
    """
    Pobla catalogo desde la API publica vPIC de NHTSA.
    Fuente:
    - https://vpic.nhtsa.dot.gov/api/vehicles/GetAllMakes?format=json
    - https://vpic.nhtsa.dot.gov/api/vehicles/GetModelsForMakeId/{make_id}?format=json
    """
    base_url = "https://vpic.nhtsa.dot.gov/api/vehicles"

    try:
        makes_resp = requests.get(f"{base_url}/GetAllMakes?format=json", timeout=30)
        makes_resp.raise_for_status()
        makes_data = makes_resp.json().get("Results", [])
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"No se pudo consultar catalogo externo: {str(exc)}")

    common_brands_colombia = {
        catalog_key("Chevrolet"),
        catalog_key("Renault"),
        catalog_key("Mazda"),
        catalog_key("Kia"),
        catalog_key("Toyota"),
        catalog_key("Nissan"),
        catalog_key("Ford"),
        catalog_key("Volkswagen"),
        catalog_key("Hyundai"),
        catalog_key("Suzuki"),
        catalog_key("Honda"),
        catalog_key("BMW"),
        catalog_key("Mercedes-Benz"),
        catalog_key("Audi"),
        catalog_key("Jeep"),
        catalog_key("Peugeot"),
        catalog_key("Citroen"),
        catalog_key("Seat"),
        catalog_key("Volvo"),
        catalog_key("Subaru"),
        catalog_key("Mitsubishi"),
        catalog_key("Fiat"),
        catalog_key("Chery"),
        catalog_key("Great Wall"),
        catalog_key("JAC"),
        catalog_key("BYD"),
        catalog_key("MG"),
        catalog_key("DFSK"),
    }

    if only_common:
        makes_data = [
            item for item in makes_data
            if catalog_key(item.get("Make_Name")) in common_brands_colombia
        ]

    if not makes_data:
        return {
            "message": "La fuente externa no devolvio marcas.",
            "brands_created": 0,
            "models_created": 0,
            "makes_processed": 0,
            "filter": "common_colombia" if only_common else "all"
        }

    # Marcas existentes por nombre normalizado
    existing_brands = db.query(models.CarBrand).all()
    brands_by_key = {b.name.lower(): b for b in existing_brands if b.name}

    # Modelos existentes por (brand_id, nombre normalizado)
    existing_models = db.query(models.CarModel).all()
    models_by_key = {
        (m.brand_id, m.name.lower()): m
        for m in existing_models
        if m.name and m.brand_id
    }

    brands_created = 0
    models_created = 0
    makes_processed = 0

    for raw_make in makes_data[:max_makes]:
        make_name = normalize_catalog_value(raw_make.get("Make_Name"))
        make_id = raw_make.get("Make_ID")
        if not make_name:
            continue

        make_key = make_name.lower()
        brand = brands_by_key.get(make_key)
        if not brand:
            brand = models.CarBrand(name=make_name)
            db.add(brand)
            db.flush()
            brands_by_key[make_key] = brand
            brands_created += 1

        makes_processed += 1

        if not include_models or not make_id:
            continue

        try:
            models_resp = requests.get(
                f"{base_url}/GetModelsForMakeId/{int(make_id)}?format=json",
                timeout=30
            )
            models_resp.raise_for_status()
            models_data = models_resp.json().get("Results", [])
        except Exception:
            # Si una marca falla, continuamos con las demas.
            continue

        for raw_model in models_data[:max_models_per_make]:
            model_name = normalize_catalog_value(raw_model.get("Model_Name"))
            if not model_name:
                continue

            model_key = (brand.id, model_name.lower())
            if model_key not in models_by_key:
                new_model = models.CarModel(name=model_name, brand_id=brand.id)
                db.add(new_model)
                db.flush()
                models_by_key[model_key] = new_model
                models_created += 1

        # Evita activar control de trafico del proveedor externo.
        time.sleep(0.05)

    db.commit()
    return {
        "message": "Catalogo externo sincronizado correctamente",
        "brands_created": brands_created,
        "models_created": models_created,
        "makes_processed": makes_processed,
        "filter": "common_colombia" if only_common else "all"
    }


# --- PUBLIC SALES CHATBOT ENDPOINTS ---

def get_public_company_id(db: Session, explicit_company_id: Optional[int] = None) -> Optional[int]:
    if explicit_company_id:
        company = db.query(models.Company).filter(models.Company.id == explicit_company_id).first()
        if company:
            return company.id
    env_company_id = os.getenv("PUBLIC_CHAT_COMPANY_ID")
    if env_company_id and env_company_id.isdigit():
        company = db.query(models.Company).filter(models.Company.id == int(env_company_id)).first()
        if company:
            return company.id
    company_with_ai = db.query(models.IntegrationSettings).filter(
        models.IntegrationSettings.openai_api_key.isnot(None)
    ).order_by(models.IntegrationSettings.company_id.asc()).first()
    if company_with_ai and company_with_ai.company_id:
        return company_with_ai.company_id

    first_company = db.query(models.Company).order_by(models.Company.id.asc()).first()
    return first_company.id if first_company else None

def get_company_ai_settings(db: Session, company_id: Optional[int]) -> tuple[Optional[str], str]:
    model_name = "gpt-4o-mini"
    api_key = (os.getenv("OPENAI_API_KEY") or "").strip() or None
    if company_id:
        settings = db.query(models.IntegrationSettings).filter(models.IntegrationSettings.company_id == company_id).first()
        if settings:
            if settings.openai_api_key and settings.openai_api_key.strip():
                api_key = settings.openai_api_key.strip()
            if settings.gw_model:
                model_name = settings.gw_model
    return api_key, model_name

def get_company_chatbot_settings(db: Session, company_id: Optional[int]) -> tuple[str, int, int]:
    bot_name = "Jennifer Quimbayo"
    typing_min_ms = 7000
    typing_max_ms = 18000

    if company_id:
        settings = db.query(models.IntegrationSettings).filter(models.IntegrationSettings.company_id == company_id).first()
        if settings:
            if settings.chatbot_bot_name and settings.chatbot_bot_name.strip():
                bot_name = settings.chatbot_bot_name.strip()
            if settings.chatbot_typing_min_ms is not None:
                typing_min_ms = int(settings.chatbot_typing_min_ms)
            if settings.chatbot_typing_max_ms is not None:
                typing_max_ms = int(settings.chatbot_typing_max_ms)

    typing_min_ms = max(0, typing_min_ms)
    typing_max_ms = max(typing_min_ms, typing_max_ms)
    return bot_name, typing_min_ms, typing_max_ms


def is_company_admin_user(user: models.User) -> bool:
    return get_user_role_name(user) in ["admin", "super_admin"]


def resolve_sale_seller(
    db: Session,
    current_user: models.User,
    seller_type: Optional[str],
    seller_id: Optional[int],
    external_seller_name: Optional[str]
) -> tuple[str, Optional[models.User], Optional[str]]:
    normalized_type = (seller_type or "internal").strip().lower()
    if normalized_type not in {"internal", "external"}:
        raise HTTPException(status_code=400, detail="Tipo de vendedor invalido")

    if normalized_type == "external":
        cleaned_name = (external_seller_name or "").strip()
        if not cleaned_name:
            raise HTTPException(status_code=400, detail="Debes indicar el asesor externo")
        return "external", None, cleaned_name

    target_seller = current_user
    if seller_id and seller_id != current_user.id:
        if not is_company_admin_user(current_user):
            raise HTTPException(status_code=403, detail="No puedes cambiar el asesor interno de la venta")
        target_seller = db.query(models.User).options(joinedload(models.User.role)).filter(
            models.User.id == seller_id
        ).first()
        if not target_seller:
            raise HTTPException(status_code=404, detail="Asesor interno no encontrado")
        if current_user.company_id and target_seller.company_id != current_user.company_id:
            raise HTTPException(status_code=400, detail="El asesor interno pertenece a otra empresa")
    return "internal", target_seller, None


def compute_sale_numbers(sale_price: int, seller: Optional[models.User]) -> tuple[int, int, int]:
    safe_price = int(sale_price or 0)
    commission_pct = int(getattr(seller, "commission_percentage", 0) or 0)
    commission_amount = int(safe_price * commission_pct / 100)
    net_revenue = safe_price - commission_amount
    return commission_pct, commission_amount, net_revenue


def get_sale_display_name(sale: Optional[models.Sale]) -> Optional[str]:
    if not sale:
        return None
    if (sale.seller_type or "internal") == "external":
        return (sale.external_seller_name or "").strip() or "Asesor externo"
    if sale.seller:
        return sale.seller.full_name or sale.seller.email
    return None


def _pick_best_receipt_display_name(candidates: List[str]) -> Optional[str]:
    cleaned_candidates = []
    for candidate in candidates or []:
        normalized_candidate = (candidate or "").strip()
        if normalized_candidate:
            cleaned_candidates.append(normalized_candidate)

    if not cleaned_candidates:
        return None

    def _display_name_score(value: str) -> tuple[int, int, int]:
        has_digits = 1 if any(char.isdigit() for char in value) else 0
        word_count = len(value.split())
        return (has_digits, word_count, len(value))

    return max(cleaned_candidates, key=_display_name_score)


def enrich_receipt_display_names(
    db: Session,
    receipts: List[models.PaymentReceipt],
    company_id: Optional[int] = None
) -> List[models.PaymentReceipt]:
    if not receipts:
        return receipts

    missing_receipt_numbers = {
        (receipt.receipt_number or "").strip()
        for receipt in receipts
        if not (receipt.display_name or "").strip() and (receipt.receipt_number or "").strip()
    }
    missing_sale_ids = {
        int(receipt.sale_id)
        for receipt in receipts
        if not (receipt.display_name or "").strip() and getattr(receipt, "sale_id", None)
    }

    if not missing_receipt_numbers and not missing_sale_ids:
        return receipts

    lookup_filters = []
    if missing_receipt_numbers:
        lookup_filters.append(models.PaymentReceipt.receipt_number.in_(missing_receipt_numbers))
    if missing_sale_ids:
        lookup_filters.append(models.PaymentReceipt.sale_id.in_(missing_sale_ids))

    sibling_query = db.query(models.PaymentReceipt).filter(
        models.PaymentReceipt.display_name.isnot(None),
        models.PaymentReceipt.display_name != "",
        or_(*lookup_filters)
    )
    if company_id:
        sibling_query = sibling_query.filter(models.PaymentReceipt.company_id == company_id)

    sibling_rows = sibling_query.order_by(
        models.PaymentReceipt.payment_date.desc(),
        models.PaymentReceipt.id.desc()
    ).all()

    display_name_candidates_by_receipt_number: Dict[str, List[str]] = {}
    display_name_candidates_by_sale_id: Dict[int, List[str]] = {}

    for sibling in sibling_rows:
        sibling_name = (sibling.display_name or "").strip()
        if not sibling_name:
            continue
        sibling_receipt_number = (sibling.receipt_number or "").strip()
        if sibling_receipt_number:
            display_name_candidates_by_receipt_number.setdefault(sibling_receipt_number, []).append(sibling_name)
        sibling_sale_id = getattr(sibling, "sale_id", None)
        if sibling_sale_id:
            display_name_candidates_by_sale_id.setdefault(int(sibling_sale_id), []).append(sibling_name)

    for receipt in receipts:
        current_name = (receipt.display_name or "").strip()
        if not current_name:
            continue
        current_receipt_number = (receipt.receipt_number or "").strip()
        if current_receipt_number:
            display_name_candidates_by_receipt_number.setdefault(current_receipt_number, []).append(current_name)
        current_sale_id = getattr(receipt, "sale_id", None)
        if current_sale_id:
            display_name_candidates_by_sale_id.setdefault(int(current_sale_id), []).append(current_name)

    display_name_by_receipt_number = {
        receipt_number: _pick_best_receipt_display_name(candidate_names)
        for receipt_number, candidate_names in display_name_candidates_by_receipt_number.items()
    }
    display_name_by_sale_id = {
        sale_id: _pick_best_receipt_display_name(candidate_names)
        for sale_id, candidate_names in display_name_candidates_by_sale_id.items()
    }

    for receipt in receipts:
        receipt_number = (receipt.receipt_number or "").strip()
        sale_id = int(receipt.sale_id) if getattr(receipt, "sale_id", None) else None
        fallback_name = (
            display_name_by_receipt_number.get(receipt_number)
            or (display_name_by_sale_id.get(sale_id) if sale_id else None)
        )
        if fallback_name:
            receipt.display_name = fallback_name

    return receipts


def apply_receipt_group_search_filter(
    query,
    db: Session,
    search_term: Optional[str],
    company_id: Optional[int] = None
):
    normalized_search = (search_term or "").strip()
    if not normalized_search:
        return query

    search = f"%{normalized_search}%"
    matched_rows_query = query.join(
        models.Sale, models.PaymentReceipt.sale_id == models.Sale.id, isouter=True
    ).join(
        models.Vehicle, models.Sale.vehicle_id == models.Vehicle.id, isouter=True
    ).join(
        models.Lead, models.Sale.lead_id == models.Lead.id, isouter=True
    ).join(
        models.CreditApplication, models.CreditApplication.lead_id == models.Lead.id, isouter=True
    ).join(
        models.PublicCreditSubmission, models.PublicCreditSubmission.lead_id == models.Lead.id, isouter=True
    ).filter(
        or_(
            models.PaymentReceipt.display_name.ilike(search),
            models.PaymentReceipt.customer_name.ilike(search),
            models.PaymentReceipt.customer_document.ilike(search),
            models.PaymentReceipt.concept.ilike(search),
            models.PaymentReceipt.receipt_number.ilike(search),
            models.PaymentReceipt.notes.ilike(search),
            models.PaymentReceipt.payment_method.ilike(search),
            models.PaymentReceipt.bank.ilike(search),
            models.Vehicle.make.ilike(search),
            models.Vehicle.model.ilike(search),
            models.Vehicle.plate.ilike(search),
            models.Sale.external_seller_name.ilike(search),
            models.Sale.tax_seller_name.ilike(search),
            models.Sale.tax_seller_document.ilike(search),
            models.Sale.tax_buyer_name.ilike(search),
            models.Sale.tax_buyer_document.ilike(search),
            models.Lead.name.ilike(search),
            models.Lead.email.ilike(search),
            models.Lead.phone.ilike(search),
            models.CreditApplication.client_name.ilike(search),
            models.CreditApplication.email.ilike(search),
            models.CreditApplication.phone.ilike(search),
            models.PublicCreditSubmission.applicant_name.ilike(search),
            models.PublicCreditSubmission.document_number.ilike(search),
            models.PublicCreditSubmission.email.ilike(search),
            models.PublicCreditSubmission.phone.ilike(search)
        )
    )

    matched_rows = matched_rows_query.with_entities(
        models.PaymentReceipt.id,
        models.PaymentReceipt.sale_id,
        models.PaymentReceipt.receipt_number
    ).all()

    matched_receipt_ids = {int(row.id) for row in matched_rows if getattr(row, "id", None)}
    matched_sale_ids = {int(row.sale_id) for row in matched_rows if getattr(row, "sale_id", None)}
    matched_receipt_numbers = {
        (row.receipt_number or "").strip()
        for row in matched_rows
        if (row.receipt_number or "").strip()
    }

    if not matched_receipt_ids and not matched_sale_ids and not matched_receipt_numbers:
        return query.filter(false())

    group_filters = []
    if matched_receipt_ids:
        group_filters.append(models.PaymentReceipt.id.in_(matched_receipt_ids))
    if matched_sale_ids:
        group_filters.append(models.PaymentReceipt.sale_id.in_(matched_sale_ids))
    if matched_receipt_numbers:
        group_filters.append(models.PaymentReceipt.receipt_number.in_(matched_receipt_numbers))

    group_scope_query = db.query(models.PaymentReceipt.id)
    if company_id:
        group_scope_query = group_scope_query.filter(models.PaymentReceipt.company_id == company_id)
    group_scope_query = group_scope_query.filter(or_(*group_filters))

    grouped_receipt_ids = [row.id for row in group_scope_query.all()]
    if not grouped_receipt_ids:
        return query.filter(false())

    return query.filter(models.PaymentReceipt.id.in_(grouped_receipt_ids))


def attach_sale_metadata_to_vehicle(vehicle: models.Vehicle, sale: Optional[models.Sale]) -> models.Vehicle:
    setattr(vehicle, "sold_price", getattr(sale, "sale_price", None))
    setattr(vehicle, "sold_by_type", getattr(sale, "seller_type", None))
    setattr(vehicle, "sold_by_name", get_sale_display_name(sale))
    setattr(vehicle, "sold_by_user", getattr(sale, "seller", None))
    return vehicle

def call_openai_chat(api_key: str, model_name: str, messages: List[Dict[str, str]]) -> str:
    url = "https://api.openai.com/v1/chat/completions"
    payload = {
        "model": model_name,
        "messages": messages,
        "temperature": 0.5
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    response = requests.post(url, headers=headers, json=payload, timeout=45)
    response.raise_for_status()
    data = response.json()
    return data["choices"][0]["message"]["content"].strip()

def call_openai_chat_with_fallback(
    primary_key: str,
    model_name: str,
    messages: List[Dict[str, str]],
    fallback_key: Optional[str]
) -> str:
    try:
        return call_openai_chat(primary_key, model_name, messages)
    except requests.HTTPError as exc:
        status_code = exc.response.status_code if exc.response is not None else None
        if status_code == 401 and fallback_key and fallback_key.strip() and fallback_key.strip() != primary_key.strip():
            return call_openai_chat(fallback_key.strip(), model_name, messages)
        raise

def extract_prospect_data_with_ai(api_key: str, model_name: str, full_conversation: str) -> Dict[str, Any]:
    extraction_messages = [
        {
            "role": "system",
            "content": (
                "Extrae datos de prospecto de conversación de compra de autos. "
                "Responde SOLO JSON con estas claves exactas: "
                "name, phone, email, interested_vehicle, payment_type, down_payment_amount, has_credit_report, "
                "report_entity, has_payment_agreement, occupation_type, residence_city, monthly_income, is_ready_to_create_lead. "
                "Si falta un dato usa null. "
                "is_ready_to_create_lead=true si ya hay datos base suficientes para crear el lead: "
                "name, phone e interested_vehicle. "
                "El email, payment_type, down_payment_amount, occupation_type, residence_city, monthly_income, "
                "report_entity y has_payment_agreement son deseables pero no obligatorios para crear el lead."
            )
        },
        {
            "role": "user",
            "content": full_conversation
        }
    ]
    url = "https://api.openai.com/v1/chat/completions"
    payload = {
        "model": model_name,
        "messages": extraction_messages,
        "temperature": 0,
        "response_format": {"type": "json_object"}
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    response = requests.post(url, headers=headers, json=payload, timeout=45)
    response.raise_for_status()
    data = response.json()
    content = data["choices"][0]["message"]["content"]
    return json.loads(content)

def normalize_phone(phone: Optional[str]) -> Optional[str]:
    if not phone:
        return None
    digits = re.sub(r"\D", "", phone)
    if not digits:
        return None
    if len(digits) == 10:
        return f"+57{digits}"
    if len(digits) == 12 and digits.startswith("57"):
        return f"+{digits}"
    if len(digits) > 10:
        return f"+57{digits[-10:]}"
    return None

def phone_variants_for_lookup(phone: Optional[str]) -> List[str]:
    normalized = normalize_phone(phone)
    if not normalized:
        return []
    digits = re.sub(r"\D", "", normalized)
    variants = {normalized, digits}
    if digits.startswith("57") and len(digits) == 12:
        variants.add(digits[2:])
        variants.add(f"+{digits[2:]}")
    return list(variants)


def find_existing_active_lead(
    db: Session,
    company_id: int,
    phone: Optional[str],
    email: Optional[str]
) -> Optional[models.Lead]:
    phone_variants = phone_variants_for_lookup(phone)
    normalized_email = (email or "").strip().lower() or None

    filters = []
    if phone_variants:
        filters.append(models.Lead.phone.in_(phone_variants))
    if normalized_email:
        filters.append(func.lower(models.Lead.email) == normalized_email)

    if not filters:
        return None

    return db.query(models.Lead).options(
        joinedload(models.Lead.assigned_to),
        selectinload(models.Lead.supervisors)
    ).filter(
        models.Lead.company_id == company_id,
        models.Lead.deleted_at.is_(None),
        or_(*filters)
    ).order_by(models.Lead.created_at.desc(), models.Lead.id.desc()).first()

def upsert_lead_process_detail(db: Session, lead_id: int, interested_vehicle: str):
    detail = db.query(models.LeadProcessDetail).filter(models.LeadProcessDetail.lead_id == lead_id).first()
    if detail:
        if interested_vehicle and detail.desired_vehicle != interested_vehicle:
            detail.has_vehicle = False
            detail.desired_vehicle = interested_vehicle
    else:
        db.add(models.LeadProcessDetail(
            lead_id=lead_id,
            has_vehicle=False,
            desired_vehicle=interested_vehicle
        ))


def sync_vehicle_reservation_for_lead(
    db: Session,
    lead_id: int,
    previous_vehicle_id: Optional[int],
    next_vehicle_id: Optional[int]
):
    if previous_vehicle_id == next_vehicle_id:
        return

    if previous_vehicle_id:
        other_lead_holds = db.query(models.LeadProcessDetail).filter(
            models.LeadProcessDetail.vehicle_id == previous_vehicle_id,
            models.LeadProcessDetail.lead_id != lead_id
        ).count()
        pending_sale = db.query(models.Sale).filter(
            models.Sale.vehicle_id == previous_vehicle_id,
            models.Sale.status == models.SaleStatus.PENDING.value
        ).first()
        previous_vehicle = db.query(models.Vehicle).filter(models.Vehicle.id == previous_vehicle_id).first()
        if previous_vehicle and previous_vehicle.status == models.VehicleStatus.RESERVED.value and not other_lead_holds and not pending_sale:
            previous_vehicle.status = models.VehicleStatus.AVAILABLE.value

    if next_vehicle_id:
        next_vehicle = db.query(models.Vehicle).filter(models.Vehicle.id == next_vehicle_id).first()
        if not next_vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        if next_vehicle.status == models.VehicleStatus.SOLD.value:
            raise HTTPException(status_code=400, detail="El vehículo ya fue vendido")
        next_vehicle.status = models.VehicleStatus.RESERVED.value

def sync_public_chat_to_lead_conversation(db: Session, session: models.PublicChatSession):
    if not session.lead_id:
        return

    lead = db.query(models.Lead).filter(models.Lead.id == session.lead_id).first()
    if not lead:
        return

    conversation = db.query(models.Conversation).filter(models.Conversation.lead_id == lead.id).first()
    if not conversation:
        conversation = models.Conversation(
            lead_id=lead.id,
            company_id=lead.company_id,
            last_message_at=datetime.datetime.utcnow()
        )
        db.add(conversation)
        db.commit()
        db.refresh(conversation)

    public_messages = db.query(models.PublicChatMessage).filter(
        models.PublicChatMessage.session_id == session.id
    ).order_by(models.PublicChatMessage.created_at.asc()).all()

    for pm in public_messages:
        marker = f"publicchat:{session.session_token}:{pm.id}"
        exists = db.query(models.Message).filter(models.Message.whatsapp_message_id == marker).first()
        if exists:
            continue

        sender_type = "lead" if pm.role == "user" else "user"
        db.add(models.Message(
            conversation_id=conversation.id,
            sender_type=sender_type,
            content=pm.content,
            message_type="text",
            status="delivered",
            whatsapp_message_id=marker,
            created_at=pm.created_at
        ))
        conversation.last_message_at = pm.created_at or datetime.datetime.utcnow()

    db.commit()

def is_inventory_request(message: str) -> bool:
    text = (message or "").strip().lower()
    triggers = [
        "que carros tenemos",
        "qué carros tenemos",
        "que autos tienen",
        "qué autos tienen",
        "que vehiculos tienen",
        "qué vehículos tienen",
        "carros disponibles",
        "autos disponibles",
        "vehiculos disponibles",
        "inventario"
    ]
    return any(t in text for t in triggers)

def build_inventory_response(db: Session, company_id: Optional[int]) -> str:
    company = db.query(models.Company).filter(models.Company.id == company_id).first() if company_id else None
    inventory_url = f"https://{company.public_domain}/autos" if company and company.public_domain else "https://autosqp.com/autos"
    query = db.query(models.Vehicle).filter(models.Vehicle.status == "available")
    if company_id:
        query = query.filter(models.Vehicle.company_id == company_id)
    vehicles = query.order_by(models.Vehicle.id.desc()).limit(5).all()

    if not vehicles:
        return (
            "En este momento no tengo vehículos disponibles para mostrarte en el chat. "
            f"Puedes revisar el inventario completo aquí: {inventory_url}"
        )

    lines = ["Claro. Estos son 5 carros disponibles en este momento:"]
    for idx, v in enumerate(vehicles, start=1):
        price = f"{v.price:,}".replace(",", ".") if v.price is not None else "N/A"
        lines.append(f"{idx}. {v.make} {v.model or ''} {v.year} - COP {price}")
    lines.append(f"Puedes ver más opciones aquí: {inventory_url}")
    return "\n".join(lines)

def maybe_create_public_chat_lead(
    db: Session,
    session: models.PublicChatSession,
    extracted_data: Dict[str, Any]
) -> Optional[int]:
    if session.lead_id:
        return session.lead_id

    name = (extracted_data.get("name") or "").strip()
    first_name = name.split(" ")[0] if name else ""
    phone = normalize_phone(extracted_data.get("phone"))
    email = (extracted_data.get("email") or "").strip() or None
    interested_vehicle = (extracted_data.get("interested_vehicle") or "").strip()
    payment_type = (extracted_data.get("payment_type") or "").strip().lower()
    down_payment_amount = extracted_data.get("down_payment_amount")
    has_credit_report = extracted_data.get("has_credit_report")
    report_entity = (extracted_data.get("report_entity") or "").strip()
    has_payment_agreement = extracted_data.get("has_payment_agreement")
    occupation_type = (extracted_data.get("occupation_type") or "").strip()
    residence_city = (extracted_data.get("residence_city") or "").strip()
    monthly_income = extracted_data.get("monthly_income")
    is_ready = bool(extracted_data.get("is_ready_to_create_lead"))

    has_minimum_lead_data = all([
        name,
        first_name,
        phone,
        interested_vehicle
    ])
    should_create_lead = has_minimum_lead_data
    if not should_create_lead:
        return None

    lookup_phones = phone_variants_for_lookup(phone)
    duplicate_window_days = int(os.getenv("PUBLIC_CHAT_DUPLICATE_WINDOW_DAYS", "30") or "30")
    recent_threshold = datetime.datetime.utcnow() - datetime.timedelta(days=duplicate_window_days)

    existing_lead = None
    if lookup_phones:
        existing_lead = db.query(models.Lead).filter(
            models.Lead.company_id == session.company_id,
            models.Lead.phone.in_(lookup_phones),
            models.Lead.created_at >= recent_threshold
        ).order_by(models.Lead.created_at.desc()).first()

    normalized_source_page = (session.source_page or "").lower()
    is_tiktok_landing = "tiktok" in normalized_source_page
    lead_source = models.LeadSource.TIKTOK.value if is_tiktok_landing else models.LeadSource.WEB.value
    lead_channel_label = "chatbot TikTok" if is_tiktok_landing else "chatbot web"

    lead_message = (
        f"Interes detectado por {lead_channel_label}: {interested_vehicle} | "
        f"Pago: {payment_type} | Cuota inicial: {down_payment_amount} | "
        f"Ingresos: {monthly_income} | Ocupacion: {occupation_type} | "
        f"Residencia: {residence_city} | Reportado: {has_credit_report} | "
        f"Entidad reporte: {report_entity or 'N/A'} | Acuerdo/Paz y salvo: {has_payment_agreement}"
    )

    if existing_lead:
        existing_lead.name = name or existing_lead.name
        existing_lead.email = email or existing_lead.email
        existing_lead.message = lead_message
        upsert_lead_process_detail(db, existing_lead.id, interested_vehicle)
        session.lead_id = existing_lead.id
        db.commit()
        sync_public_chat_to_lead_conversation(db, session)
        return existing_lead.id

    assigned_user_id = None
    if session.company_id:
        auto_assigned_user = choose_auto_assign_user(db, session.company_id)
        if auto_assigned_user:
            assigned_user_id = auto_assigned_user.id

    ensure_company_lead_creation_capacity(db, session.company_id)

    new_lead = models.Lead(
        source=lead_source,
        name=name,
        phone=phone,
        email=email,
        message=lead_message,
        status=models.LeadStatus.NEW.value,
        company_id=session.company_id,
        assigned_to_id=assigned_user_id
    )
    db.add(new_lead)
    db.commit()
    db.refresh(new_lead)

    upsert_lead_process_detail(db, new_lead.id, interested_vehicle)
    session.lead_id = new_lead.id
    db.commit()
    sync_public_chat_to_lead_conversation(db, session)
    return new_lead.id

@app.post("/public-chat/session", response_model=schemas.PublicChatSession)
def create_public_chat_session(payload: schemas.PublicChatSessionCreate, db: Session = Depends(get_db)):
    company_id = get_public_company_id(db, payload.company_id)
    ensure_public_chat_enabled_for_company(db, company_id=company_id)
    session_token = str(uuid.uuid4())
    session = models.PublicChatSession(
        session_token=session_token,
        company_id=company_id,
        source_page=payload.source_page or "/autos",
        status="active"
    )
    db.add(session)
    db.commit()
    db.refresh(session)
    return session

@app.post("/public-chat/config")
def get_public_chat_config(payload: schemas.PublicChatInactiveCheck, db: Session = Depends(get_db)):
    session = db.query(models.PublicChatSession).filter(
        models.PublicChatSession.session_token == payload.session_token
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    ensure_public_chat_enabled_for_company(db, company_id=session.company_id)

    bot_name, typing_min_ms, typing_max_ms = get_company_chatbot_settings(db, session.company_id)
    return {
        "bot_name": bot_name,
        "typing_min_ms": typing_min_ms,
        "typing_max_ms": typing_max_ms
    }

@app.get("/public-chat/{session_token}/messages", response_model=List[schemas.PublicChatMessageItem])
def get_public_chat_messages(session_token: str, db: Session = Depends(get_db)):
    session = db.query(models.PublicChatSession).filter(models.PublicChatSession.session_token == session_token).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    ensure_public_chat_enabled_for_company(db, company_id=session.company_id)
    messages = db.query(models.PublicChatMessage).filter(
        models.PublicChatMessage.session_id == session.id
    ).order_by(models.PublicChatMessage.created_at.asc()).all()
    return [{"role": m.role, "content": m.content, "created_at": m.created_at} for m in messages]

@app.post("/public-chat/check-inactive")
def public_chat_check_inactive(payload: schemas.PublicChatInactiveCheck, db: Session = Depends(get_db)):
    """
    If assistant asked a question and user has not replied for 5+ minutes,
    send a follow-up check-in message once.
    """
    token = (payload.session_token or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="Session token is required")

    session = db.query(models.PublicChatSession).filter(models.PublicChatSession.session_token == token).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    ensure_public_chat_enabled_for_company(db, company_id=session.company_id)

    last_message = db.query(models.PublicChatMessage).filter(
        models.PublicChatMessage.session_id == session.id
    ).order_by(models.PublicChatMessage.created_at.desc()).first()

    if not last_message or last_message.role != "assistant":
        return {"nudged": False}

    assistant_text = (last_message.content or "").strip()
    asked_question = ("?" in assistant_text) or ("¿" in assistant_text)
    if not asked_question:
        return {"nudged": False}

    now_utc = datetime.datetime.utcnow()
    if (now_utc - last_message.created_at).total_seconds() < 300:
        return {"nudged": False}

    has_user_reply = db.query(models.PublicChatMessage).filter(
        models.PublicChatMessage.session_id == session.id,
        models.PublicChatMessage.role == "user",
        models.PublicChatMessage.created_at > last_message.created_at
    ).first()
    if has_user_reply:
        return {"nudged": False}

    existing_checkin = db.query(models.PublicChatMessage).filter(
        models.PublicChatMessage.session_id == session.id,
        models.PublicChatMessage.role == "assistant",
        models.PublicChatMessage.content == "¿Sigues en línea? Si quieres, te ayudo a continuar."
    ).order_by(models.PublicChatMessage.created_at.desc()).first()
    if existing_checkin and existing_checkin.created_at > last_message.created_at:
        return {"nudged": False}

    nudge_text = "¿Sigues en línea? Si quieres, te ayudo a continuar."
    db.add(models.PublicChatMessage(session_id=session.id, role="assistant", content=nudge_text))
    db.commit()
    return {"nudged": True, "message": nudge_text}

@app.post("/public-chat/message", response_model=schemas.PublicChatMessageResponse)
def public_chat_message(payload: schemas.PublicChatMessageCreate, db: Session = Depends(get_db)):
    user_message = (payload.message or "").strip()
    if not user_message:
        raise HTTPException(status_code=400, detail="Message is required")

    session = db.query(models.PublicChatSession).filter(
        models.PublicChatSession.session_token == payload.session_token
    ).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    ensure_public_chat_enabled_for_company(db, company_id=session.company_id)

    if payload.source_page:
        session.source_page = payload.source_page

    api_key, model_name = get_company_ai_settings(db, session.company_id)
    if not api_key:
        raise HTTPException(status_code=400, detail="OpenAI API key no configurada")
    env_fallback_key = (os.getenv("OPENAI_API_KEY") or "").strip() or None

    db.add(models.PublicChatMessage(session_id=session.id, role="user", content=user_message))
    db.commit()

    history = db.query(models.PublicChatMessage).filter(
        models.PublicChatMessage.session_id == session.id
    ).order_by(models.PublicChatMessage.created_at.asc()).all()

    vehicle_hint = ""
    if payload.vehicle_id:
        vehicle = db.query(models.Vehicle).filter(models.Vehicle.id == payload.vehicle_id).first()
        if vehicle:
            vehicle_hint = (
                f"Vehículo que está viendo el cliente: {vehicle.make} {vehicle.model} {vehicle.year}, "
                f"precio {vehicle.price} COP."
            )

    bot_name, _, _ = get_company_chatbot_settings(db, session.company_id)
    system_prompt = (
        f"Eres {bot_name}, asesora comercial de AutosQP en Colombia. "
        "Habla en tono amigable, cercano y comercial. "
        "En tu primera respuesta de cada sesion debes presentarte explicitamente como: "
        f"\"Hola, soy {bot_name}, asesora comercial de Autos QP\" y luego continuar con la asesoria. "
        "Si el cliente da nombre completo, dirigete solo por su primer nombre. "
        "Nunca ofrezcas opciones de financiacion, planes de credito ni preguntes si deseas compartir opciones de financiacion. "
        "Tampoco uses frases como '¿Te gustaria que te compartiera opciones de financiacion?'. "
        "Si necesitas perfilar la forma de pago, pregunta solo si la compra seria de contado o por credito, sin ofrecer productos financieros. "
        "Si el cliente responde que la compra es de contado, no preguntes nada sobre centrales de riesgo, reportes, entidades, acuerdos de pago, paz y salvo, ingresos ni perfil crediticio. "
        "Si es de contado, despues de confirmar vehiculo, nombre, telefono y correo, cierra indicando que la informacion sera analizada y que nos contactaremos tan pronto tengamos informacion. "
        "Nunca pidas documentos, soportes, cedula, certificaciones, extractos ni facturas. "
        "Cuando ya tengas todos los datos base obligatorios, no solicites nada adicional: responde unicamente que la informacion sera analizada "
        "y que tan pronto tengamos informacion nos contactaremos. "
        "Debes perfilar al cliente con preguntas claras, una por turno, en este orden: "
        "1) vehiculo de interes, 2) nombre, 3) telefono, 4) correo (OBLIGATORIO), "
        "5) pago de contado o con credito, 6) cuota inicial y monto, "
        "7) si tiene reportes en centrales de riesgo, "
        "8) si esta reportado: entidad y si tiene acuerdo de pago o paz y salvo, "
        "9) ocupacion (empleado, independiente u otro), "
        "10) lugar de residencia, 11) ingresos mensuales. "
        "Ese orden aplica solo para escenarios de credito. Si es contado, detente despues del paso 5 y no hagas las preguntas 6 a 11. "
        "No inventes informacion ni cierres perfilado hasta tener todo lo obligatorio. "
        f"{vehicle_hint}"
    )

    if is_inventory_request(user_message):
        assistant_reply = build_inventory_response(db, session.company_id)
    else:
        chat_messages = [{"role": "system", "content": system_prompt}]
        for msg in history[-20:]:
            if msg.role in ["user", "assistant"]:
                chat_messages.append({"role": msg.role, "content": msg.content})

        try:
            assistant_reply = call_openai_chat_with_fallback(api_key, model_name, chat_messages, env_fallback_key)
        except requests.HTTPError as exc:
            status_code = exc.response.status_code if exc.response is not None else 500
            if status_code == 401:
                raise HTTPException(status_code=400, detail="OpenAI API key inválida o expirada para el chatbot público")
            raise HTTPException(status_code=500, detail=f"Error consultando OpenAI: {str(exc)}")
    db.add(models.PublicChatMessage(session_id=session.id, role="assistant", content=assistant_reply))
    db.commit()

    full_text = "\n".join([f"{m.role}: {m.content}" for m in history[-30:]] + [f"assistant: {assistant_reply}"])
    lead_created = False
    lead_id = session.lead_id
    try:
        extraction_key = api_key
        if env_fallback_key and env_fallback_key.strip():
            extraction_key = env_fallback_key if api_key != env_fallback_key else api_key
        extracted = extract_prospect_data_with_ai(extraction_key, model_name, full_text)
        lead_id = maybe_create_public_chat_lead(db, session, extracted)
        lead_created = bool(lead_id and not session.lead_id is None)
    except Exception:
        # Extraction failure should not break chat response.
        lead_created = False

    # Keep lead conversation in sync after each turn once lead exists.
    if session.lead_id:
        sync_public_chat_to_lead_conversation(db, session)

    latest_messages = db.query(models.PublicChatMessage).filter(
        models.PublicChatMessage.session_id == session.id
    ).order_by(models.PublicChatMessage.created_at.asc()).all()

    return {
        "session_token": session.session_token,
        "reply": assistant_reply,
        "lead_created": bool(session.lead_id),
        "lead_id": session.lead_id,
        "messages": [{"role": m.role, "content": m.content, "created_at": m.created_at} for m in latest_messages]
    }


# --- LEADS ENDPOINTS ---


# Remove duplicate read_leads


def _legacy_create_lead_unused(
    lead: schemas.LeadCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Assign to current user's company if not provided
    company_id = lead.company_id or current_user.company_id
    
    if not company_id:
        raise HTTPException(status_code=400, detail="Company ID required for assignment")

    existing_lead = find_existing_active_lead(db, company_id, lead.phone, lead.email)
    if existing_lead:
        if lead.name and not existing_lead.name:
            existing_lead.name = lead.name
        if lead.email and not existing_lead.email:
            existing_lead.email = lead.email
        if lead.phone and not existing_lead.phone:
            existing_lead.phone = lead.phone
        if lead.message:
            existing_lead.message = lead.message
        db.commit()
        db.refresh(existing_lead)
        return existing_lead
    
    # Exclude company_id because we pass it manually
    lead_data = lead.dict(exclude={'company_id'})
    supervisor_ids = normalize_supervisor_ids(getattr(lead, "supervisor_ids", []))
    if supervisor_ids and not is_company_admin(current_user):
        raise HTTPException(status_code=403, detail="Solo un administrador puede agregar o quitar supervisores de un lead.")
    
    # Manual Assignment by Aliado/Admin
    assigned_to_id = lead.assigned_to_id
    
    # Manual leads created by advisors stay assigned to themselves. Random
    # assignment here is only for super admin-created leads.
    if not assigned_to_id:
        if should_self_assign_manual_lead(current_user):
            assigned_to_id = current_user.id
        elif should_auto_assign_manual_lead(current_user):
            auto_assigned_user = choose_auto_assign_user(db, company_id)
            if auto_assigned_user:
                assigned_to_id = auto_assigned_user.id
            else:
                raise HTTPException(
                    status_code=400,
                    detail="No hay asesores vendedores activos con asignacion automatica habilitada"
                )
    
    # Verify the manually assigned user belongs to the company
    elif assigned_to_id:
        target_user = db.query(models.User).join(models.Role, isouter=True).filter(models.User.id == assigned_to_id).first()
        if not target_user or target_user.company_id != company_id:
             raise HTTPException(status_code=400, detail="Invalid assigned user (not in company)")
        if not can_manually_assign_to_any_role(current_user) and not is_advisor_role(target_user.role):
             raise HTTPException(status_code=400, detail="Solo se pueden asignar leads a usuarios con rol asesor")

    lead_data["status"] = enforce_ally_managed_status(
        db=db,
        current_user=current_user,
        base_status=lead_data.get("status"),
        assigned_to_id=assigned_to_id
    )

    db_lead = models.Lead(
        **lead_data, 
        company_id=company_id, 
        assigned_to_id=assigned_to_id,
        created_by_id=current_user.id # Track who created it
    )
    db.add(db_lead)
    db.commit()
    db.refresh(db_lead)

    initial_history = models.LeadHistory(
        lead_id=db_lead.id,
        user_id=current_user.id,
        previous_status=None,
        new_status=db_lead.status,
        comment=(db_lead.message or "Lead creado manualmente")
    )
    db.add(initial_history)
    db.commit()
    db.refresh(db_lead)
    
    log_action_to_db(db, current_user.id, "CREATE", "Lead", db_lead.id, f"Lead creado: {db_lead.name} ({db_lead.email or db_lead.phone})")
    
    
    # Create Notification for assigned user
    if db_lead.assigned_to_id:
        target_user = db.query(models.User).join(models.Role, isouter=True).filter(models.User.id == db_lead.assigned_to_id).first()
        notification = models.Notification(
            user_id=db_lead.assigned_to_id,
            title="Nuevo Lead Asignado",
            message=f"Se te ha asignado un nuevo lead: {db_lead.name}. Continúa con la gestion.",
            type="info",
            link=build_lead_board_link(target_user, db_lead.id)
        )
        db.add(notification)
        db.commit()
        
    return db_lead


@app.delete("/leads/{lead_id}", status_code=200)
def delete_lead(
    lead_id: int,
    payload: schemas.LeadDeleteRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    reason = (payload.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Debes indicar el motivo de eliminación")

    lead = db.query(models.Lead).options(
        joinedload(models.Lead.assigned_to),
        joinedload(models.Lead.supervisors)
    ).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    ensure_can_modify_lead(current_user, lead)

    if lead.deleted_at:
        raise HTTPException(status_code=400, detail="El lead ya fue eliminado")

    lead.deleted_at = datetime.datetime.utcnow()
    lead.deleted_reason = reason
    lead.deleted_by_id = current_user.id

    db.add(models.LeadHistory(
        lead_id=lead.id,
        user_id=current_user.id,
        previous_status=lead.status,
        new_status=lead.status,
        comment=f"Lead eliminado del tablero. Motivo: {reason}"
    ))

    log_action_to_db(
        db,
        current_user.id,
        "DELETE",
        "Lead",
        lead.id,
        f"Lead ocultado de tableros. Motivo: {reason}"
    )

    db.commit()
    return {"message": "Lead eliminado correctamente"}

@app.put("/leads/{lead_id}", response_model=schemas.Lead)
def update_lead(
    lead_id: int, 
    lead_update: schemas.LeadUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
        
    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    update_data = lead_update.dict(exclude_unset=True)
    requested_fields = set(update_data.keys())
    contact_edit_fields = {"name", "email", "phone"}
    is_contact_only_update = bool(requested_fields) and requested_fields.issubset(contact_edit_fields)

    if not is_contact_only_update:
        ensure_can_modify_lead(current_user, lead)
    if lead_update.supervisor_ids is not None:
        ensure_can_manage_lead_supervision(current_user, lead)

    payload_update = lead_update.dict(exclude={'comment', 'process_detail', 'supervisor_ids'}, exclude_unset=True)
    assignment_context = get_effective_lead_assignment_for_update(
        db=db,
        lead=lead,
        lead_update=lead_update,
        current_user=current_user
    )
    previous_assigned_user = assignment_context["previous_assigned_user"]
    target_assigned_user = assignment_context["target_assigned_user"]
    target_assigned_to_id = assignment_context["target_assigned_to_id"]
    target_supervisor_ids = assignment_context["target_supervisor_ids"]
    effective_status = assignment_context["effective_status"]
    board_transfer = assignment_context["board_transfer"]
    credit_coordinator = assignment_context["credit_coordinator"]
    target_role_name = assignment_context["target_role_name"]

    if payload_update.get('status') is not None or payload_update.get('assigned_to_id') is not None:
        payload_update['status'] = effective_status
    if lead_update.assigned_to_id is not None:
        payload_update['assigned_to_id'] = target_assigned_to_id

    # Check for status change OR comment to log history
    # If status changes, or if there's a comment (even without status change), we log it.
    has_status_change = effective_status != lead.status
    auto_transfer_comment = None
    if board_transfer:
        auto_transfer_comment = (
            "Lead enviado al tablero de aliados"
            if target_role_name == "aliado"
            else "Lead devuelto al tablero general"
        )
    auto_credit_comment = None
    if credit_coordinator and credit_coordinator.id != getattr(previous_assigned_user, "id", None):
        auto_credit_comment = f"Coordinador de credito añadido como supervisor: {credit_coordinator.full_name or credit_coordinator.email}"
    process_detail_data = lead_update.process_detail
    previous_process_detail = db.query(models.LeadProcessDetail).filter(models.LeadProcessDetail.lead_id == lead.id).first()
    previous_vehicle_id = previous_process_detail.vehicle_id if previous_process_detail else None
    process_detail_data = normalize_interested_process_detail(
        effective_status,
        process_detail_data,
        previous_process_detail,
        lead
    )
    if not is_contact_only_update:
        validate_interested_process_detail(effective_status, process_detail_data, previous_process_detail)

    history_comment_parts = []
    if lead_update.comment and lead_update.comment.strip():
        history_comment_parts.append(lead_update.comment.strip())
    elif auto_credit_comment or auto_transfer_comment:
        history_comment_parts.append(auto_credit_comment or auto_transfer_comment)

    if normalize_lead_status_value(effective_status) == models.LeadStatus.CREDIT_STUDY.value:
        context_detail = process_detail_data or previous_process_detail
        vehicle_context = get_lead_credit_vehicle_context(db, lead, context_detail)
        vehicle_phrase = f"{vehicle_context['mode'].capitalize()}: {vehicle_context['vehicle']}"
        if not any(vehicle_phrase.lower() in part.lower() for part in history_comment_parts):
            history_comment_parts.append(vehicle_phrase)

    if normalize_lead_status_value(effective_status) == models.LeadStatus.RESERVED.value:
        context_detail = process_detail_data or previous_process_detail
        reservation_amount = getattr(context_detail, "reservation_amount", None)
        reservation_payment_method = (getattr(context_detail, "reservation_payment_method", None) or "").strip().lower()
        if reservation_amount:
            reservation_phrase = f"Separación registrada: ${int(reservation_amount):,}".replace(",", ".")
            if not any(reservation_phrase.lower() in part.lower() for part in history_comment_parts):
                history_comment_parts.append(reservation_phrase)
        if reservation_payment_method:
            payment_phrase = f"Medio de separación: {reservation_payment_method.capitalize()}"
            if not any(payment_phrase.lower() in part.lower() for part in history_comment_parts):
                history_comment_parts.append(payment_phrase)

    history_comment = " | ".join(part for part in history_comment_parts if part).strip() or None
    has_comment = bool(history_comment)

    if has_status_change or has_comment:
        if has_status_change:
            lead.status_updated_at = datetime.datetime.utcnow()
            
        new_history = models.LeadHistory(
            lead_id=lead.id,
            user_id=current_user.id,
            previous_status=lead.status,
            new_status=effective_status,
            comment=history_comment
        )
        db.add(new_history)
        
    allowed_lead_update_fields = {
        "name",
        "email",
        "phone",
        "status",
        "message",
        "assigned_to_id",
    }
    for field, value in payload_update.items():
        if field not in allowed_lead_update_fields:
            continue
        setattr(lead, field, value)

    sync_lead_supervisors(db, lead, target_supervisor_ids, current_user.id)
        
    # Handle Process Detail Upsert
    if process_detail_data:
        process_detail_payload = process_detail_data.dict(exclude_unset=True)
        existing_detail = db.query(models.LeadProcessDetail).filter(models.LeadProcessDetail.lead_id == lead.id).first()
        if existing_detail:
            for k, v in process_detail_payload.items():
                setattr(existing_detail, k, v)
        else:
            new_detail = models.LeadProcessDetail(
                lead_id=lead.id,
                **process_detail_payload
            )
            db.add(new_detail)

        sync_vehicle_reservation_for_lead(
            db=db,
            lead_id=lead.id,
            previous_vehicle_id=previous_vehicle_id,
            next_vehicle_id=process_detail_data.vehicle_id if process_detail_data.has_vehicle else None
        )

    # Auto-create/update credit request queue when lead enters credit stage
    upsert_credit_application_from_lead(db, lead, lead_update.comment)
    upsert_credit_approval_data_for_lead(db, lead, lead_update)

    # Auto-create/update purchase request queue for role "compras"
    upsert_purchase_request_from_lead(db, lead)

    related_credit_records = db.query(models.CreditApplication).filter(
        models.CreditApplication.lead_id == lead.id
    ).order_by(
        models.CreditApplication.updated_at.desc(),
        models.CreditApplication.created_at.desc()
    ).all()
    related_credit = next(
        (record for record in related_credit_records if not is_purchase_request_record(record)),
        None
    )
    if related_credit and normalize_lead_status_value(lead.status) == models.LeadStatus.SOLD.value:
        if related_credit.status != models.CreditStatus.COMPLETED.value:
            related_credit.status = models.CreditStatus.COMPLETED.value
            related_credit.notes = (
                f"{(related_credit.notes or '').rstrip()}\n"
                "Marcado automáticamente como vendido desde el tablero de leads."
            ).strip()

    db.commit()
    db.refresh(lead)

    if credit_coordinator and credit_coordinator.id != getattr(previous_assigned_user, "id", None):
        notification = models.Notification(
            user_id=credit_coordinator.id,
            title="Lead en solicitud de crédito",
            message=f"Ahora haces parte de la gestión del lead: {lead.name}.",
            type="info",
            link=build_lead_board_link(credit_coordinator, lead.id)
        )
        db.add(notification)
        db.commit()
        db.refresh(lead)

    if target_assigned_to_id and target_assigned_to_id != getattr(previous_assigned_user, "id", None):
        notification = models.Notification(
            user_id=target_assigned_to_id,
            title="Lead Reasignado" if not credit_coordinator else "Lead de crédito asignado",
            message=(
                f"Se te ha asignado el lead: {lead.name}. Continua con la gestion."
                if not credit_coordinator
                else f"Se te ha asignado el lead en solicitud de crédito: {lead.name}"
            ),
            type="info",
            link=build_lead_board_link(target_assigned_user, lead.id)
        )
        db.add(notification)
        db.commit()
        db.refresh(lead)
    
    log_action_to_db(db, current_user.id, "UPDATE", "Lead", lead.id, f"Lead modificado: {lead.name} a estado {lead.status}")
    
    return lead

# --- LEAD NOTES AND FILES ENDPOINTS ---

@app.post("/leads/{lead_id}/notes", response_model=schemas.LeadNote)
def create_lead_note(
    lead_id: int, 
    note: schemas.LeadNoteCreate, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
        
    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    ensure_can_modify_lead(current_user, lead)
        
    db_note = models.LeadNote(
        lead_id=lead_id,
        user_id=current_user.id,
        content=note.content
    )
    db_history = models.LeadHistory(
        lead_id=lead_id,
        user_id=current_user.id,
        previous_status=lead.status,
        new_status=lead.status,
        comment=f"Nota agregada: {note.content}"
    )
    db.add(db_note)
    db.add(db_history)
    db.commit()
    db.refresh(db_note)
    db_note = db.query(models.LeadNote).options(
        joinedload(models.LeadNote.user)
    ).filter(
        models.LeadNote.id == db_note.id
    ).first()
    
    log_action_to_db(db, current_user.id, "CREATE", "LeadNote", db_note.id, f"Nota agregada al lead {lead.name}")
    
    return db_note

@app.get("/leads/{lead_id}/notes", response_model=List[schemas.LeadNote])
def get_lead_notes(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    return db.query(models.LeadNote).options(
        joinedload(models.LeadNote.user)
    ).filter(
        models.LeadNote.lead_id == lead_id
    ).order_by(models.LeadNote.created_at.desc()).all()

@app.post("/leads/{lead_id}/files", response_model=schemas.LeadFile)
async def upload_lead_file(
    lead_id: int, 
    file: UploadFile = File(...), 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    import os
    import uuid
    import shutil
    
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
        
    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    ensure_can_modify_lead(current_user, lead)
        
    os.makedirs("static/leads", exist_ok=True)
    file_extension = file.filename.split(".")[-1]
    unique_filename = f"{uuid.uuid4()}.{file_extension}"
    file_path = f"static/leads/{unique_filename}"
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    db_file = models.LeadFile(
        lead_id=lead_id,
        user_id=current_user.id,
        file_name=file.filename,
        file_path=f"/{file_path}",
        file_type=file.content_type
    )
    db.add(db_file)
    db.commit()
    db.refresh(db_file)
    
    log_action_to_db(db, current_user.id, "CREATE", "LeadFile", db_file.id, f"Archivo adjuntado al lead {lead.name}")
    
    return db_file

@app.get("/leads/{lead_id}/files", response_model=List[schemas.LeadFile])
def get_lead_files(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    return db.query(models.LeadFile).options(
        joinedload(models.LeadFile.user)
    ).filter(
        models.LeadFile.lead_id == lead_id
    ).order_by(models.LeadFile.created_at.desc()).all()

@app.delete("/leads/{lead_id}/files/{file_id}")
def delete_lead_file(
    lead_id: int,
    file_id: int,
    payload: schemas.LeadFileDeleteRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    ensure_can_modify_lead(current_user, lead)

    reason = (payload.reason or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="Deletion reason is required")

    lead_file = db.query(models.LeadFile).filter(
        models.LeadFile.id == file_id,
        models.LeadFile.lead_id == lead_id
    ).first()
    if not lead_file:
        raise HTTPException(status_code=404, detail="Lead file not found")

    stored_path = (lead_file.file_path or "").lstrip("/")
    if stored_path and os.path.exists(stored_path):
        try:
            os.remove(stored_path)
        except OSError:
            pass

    history_entry = models.LeadHistory(
        lead_id=lead.id,
        user_id=current_user.id,
        previous_status=lead.status,
        new_status=lead.status,
        comment=f"Documento eliminado: {lead_file.file_name}. Motivo: {reason}"
    )
    db.add(history_entry)

    db_note = models.LeadNote(
        lead_id=lead.id,
        user_id=current_user.id,
        content=f"Se eliminó el documento '{lead_file.file_name}'. Motivo: {reason}"
    )
    db.add(db_note)

    file_name = lead_file.file_name
    db.delete(lead_file)
    db.commit()

    log_action_to_db(
        db,
        current_user.id,
        "DELETE",
        "LeadFile",
        file_id,
        f"Archivo {file_name} eliminado del lead {lead.name}. Motivo: {reason}"
    )

    return {"message": "Lead file deleted successfully"}

@app.get("/leads/{lead_id}/messages")
def get_lead_messages(
    lead_id: int, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    # Retrieve the lead first to check company scope
    lead = db.query(models.Lead).filter(models.Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
        
    if current_user.company_id and lead.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    if lead.has_unread_reply:
        lead.has_unread_reply = 0
        db.commit()
        
    conversation_ids = {
        row.id
        for row in db.query(models.Conversation.id).filter(models.Conversation.lead_id == lead_id).all()
    }

    if lead.phone:
        lookup_phones = phone_variants_for_lookup(lead.phone)
        channel_query = db.query(models.ChannelChatSession).filter(
            models.ChannelChatSession.company_id == lead.company_id,
            models.ChannelChatSession.conversation_id.isnot(None),
        )
        channel_query = channel_query.filter(
            or_(
                models.ChannelChatSession.lead_id == lead.id,
                models.ChannelChatSession.external_user_id.in_(lookup_phones or [lead.phone]),
            )
        )
        for channel_session in channel_query.all():
            if channel_session.conversation_id:
                conversation_ids.add(channel_session.conversation_id)
                if channel_session.lead_id != lead.id:
                    channel_session.lead_id = lead.id
                    db.commit()

    # Backward-compatible linking for older channel conversations.
    if not conversation_ids and lead.source in {"facebook", "instagram", "whatsapp"} and lead.phone:
        channel_session = db.query(models.ChannelChatSession).filter(
            models.ChannelChatSession.company_id == lead.company_id,
            models.ChannelChatSession.source == lead.source,
            models.ChannelChatSession.external_user_id == lead.phone,
            models.ChannelChatSession.conversation_id.isnot(None),
        ).first()
        if channel_session and channel_session.conversation_id:
            conversation = db.query(models.Conversation).filter(
                models.Conversation.id == channel_session.conversation_id
            ).first()
            if conversation and conversation.lead_id != lead.id:
                conversation.lead_id = lead.id
                if channel_session.lead_id != lead.id:
                    channel_session.lead_id = lead.id
                db.commit()
                conversation_ids.add(conversation.id)
    if not conversation_ids:
        return [] # No messages yet
        
    messages = db.query(models.Message).filter(
        models.Message.conversation_id.in_(conversation_ids)
    ).order_by(models.Message.created_at.desc()).all()
    
    return messages

@app.post("/webhooks/leads/{source}")
def receive_webhook_lead(
    source: str, 
    data: dict
):
    """
    Mock endpoint to receive leads from external sources (Webhooks).
    Payload expected: { "name": "...", "phone": "...", "message": "..." }
    In a real scenario, you'd map Facebook/WhatsApp payloads here.
    """
    # Simple mapping for demonstration
    name = data.get("name", "Unknown Lead")
    phone = data.get("phone") or data.get("from") # FB often uses 'from'
    email = data.get("email")
    message = data.get("message") or data.get("text")
    
    # Default to first company found if no ID logic (For demo purposes)
    # In prod, you'd map the webhook token/ID to a specific company
    company = db.query(models.Company).first()
    company_id = company.id if company else None

    # Create Lead
    ensure_company_lead_creation_capacity(db, company_id)
    new_lead = models.Lead(
        name=name,
        phone=phone,
        email=email,
        message=message,
        source=source, # facebook, whatsapp, etc
        company_id=company_id,
        status="new"
    )
    db.add(new_lead)
    db.commit()
    return {"status": "received", "lead_id": new_lead.id}

# --- VEHICLES ENDPOINTS ---

@app.get("/vehicles/makes")
def get_public_makes(request: Request, db: Session = Depends(get_db)):
    # Returns distinct makes from available vehicles
    company = resolve_public_company(db, request)
    query = db.query(models.Vehicle.make).filter(models.Vehicle.status == 'available')
    if company:
        query = query.filter(models.Vehicle.company_id == company.id)
    makes = query.distinct().all()
    return [m[0] for m in makes if m[0]]

@app.get("/vehicles/public")
def get_public_vehicles(
    request: Request,
    q: Optional[str] = None,
    make: Optional[str] = None,
    model: Optional[str] = None,
    year_from: Optional[int] = None,
    year_to: Optional[int] = None,
    price_min: Optional[int] = None,
    price_max: Optional[int] = None,
    mileage_min: Optional[int] = None,
    mileage_max: Optional[int] = None,
    color: Optional[str] = None,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    query = db.query(models.Vehicle).filter(models.Vehicle.status == 'available')
    company = resolve_public_company(db, request)
    if company:
        query = query.filter(models.Vehicle.company_id == company.id)
    
    if q:
        search = f"%{q}%"
        query = query.filter(
            (models.Vehicle.make.ilike(search)) |
            (models.Vehicle.model.ilike(search)) |
            (models.Vehicle.plate.ilike(search))
        )
    if make:
        query = query.filter(models.Vehicle.make.ilike(f"%{make}%"))
    if model:
        query = query.filter(models.Vehicle.model.ilike(f"%{model}%"))
    if color:
        query = query.filter(models.Vehicle.color.ilike(f"%{color}%"))
    if year_from is not None:
        query = query.filter(models.Vehicle.year >= year_from)
    if year_to is not None:
        query = query.filter(models.Vehicle.year <= year_to)
    if price_min is not None:
        query = query.filter(models.Vehicle.price >= price_min)
    if price_max is not None:
        query = query.filter(models.Vehicle.price <= price_max)
    if mileage_min is not None:
        query = query.filter(models.Vehicle.mileage >= mileage_min)
    if mileage_max is not None:
        query = query.filter(models.Vehicle.mileage <= mileage_max)
        
    vehicles = query.order_by(models.Vehicle.id.desc()).limit(limit).all()
    # Serialize to dict to match expected frontend structure if needed, or rely on FastAPI's response_model
    return vehicles

@app.get("/vehicles/", response_model=schemas.VehicleList)
def read_vehicles(
    q: str = None,
    status: str = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    query = db.query(models.Vehicle)
    
    # Filter by Company
    if current_user.company_id:
        query = query.filter(models.Vehicle.company_id == current_user.company_id)
    query = query.filter(models.Vehicle.status != "inactive")
        
    if q:
        search = f"%{q}%"
        query = query.filter(
            (models.Vehicle.make.ilike(search)) |
            (models.Vehicle.model.ilike(search)) |
            (models.Vehicle.plate.ilike(search))
        )

    effective_status = status if is_inventory_editor(current_user) else "available"
    if effective_status:
        query = query.filter(models.Vehicle.status == effective_status)
        
    total = query.count()
    vehicles = query.order_by(models.Vehicle.id.desc()).offset(skip).limit(limit).all()
    if vehicles:
        vehicle_ids = [vehicle.id for vehicle in vehicles]
        sales_query = db.query(models.Sale).options(joinedload(models.Sale.seller)).filter(
            models.Sale.vehicle_id.in_(vehicle_ids)
        )
        if current_user.company_id:
            sales_query = sales_query.filter(models.Sale.company_id == current_user.company_id)
        sales = sales_query.all()
        sales_by_vehicle_id = {sale.vehicle_id: sale for sale in sales}
        for vehicle in vehicles:
            attach_sale_metadata_to_vehicle(vehicle, sales_by_vehicle_id.get(vehicle.id))
    return {"items": vehicles, "total": total}

@app.post("/vehicles/", response_model=schemas.Vehicle)
def create_vehicle(
    vehicle: schemas.VehicleCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    ensure_inventory_editor(current_user)

    if current_user.company_id and vehicle.company_id and vehicle.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Cannot create vehicle for another company")
        
    # Default to user's company if not provided and user has one
    if not vehicle.company_id and current_user.company_id:
        vehicle.company_id = current_user.company_id

    upsert_brand_model(db, vehicle.make, vehicle.model)
        
    db_vehicle = models.Vehicle(**vehicle.dict())
    db.add(db_vehicle)
    db.commit()
    db.refresh(db_vehicle)
    return db_vehicle

@app.put("/vehicles/{vehicle_id}", response_model=schemas.Vehicle)
def update_vehicle(
    vehicle_id: int, 
    vehicle_update: schemas.VehicleUpdate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    ensure_inventory_editor(current_user)

    vehicle = db.query(models.Vehicle).filter(models.Vehicle.id == vehicle_id).first()
    if not vehicle or vehicle.status == "inactive":
        raise HTTPException(status_code=404, detail="Vehicle not found")
        
    if current_user.company_id and vehicle.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_payload = vehicle_update.dict(exclude_unset=True)
    manual_sold_sale = None
    if update_payload.get("status") == "sold":
        seller_type = update_payload.get("sold_by_type")
        internal_user_id = update_payload.get("sold_by_internal_user_id")
        external_name = update_payload.get("sold_by_external_name")
        sold_price = int(update_payload.get("sold_price") or update_payload.get("price") or vehicle.price or 0)
        if sold_price <= 0:
            raise HTTPException(status_code=400, detail="Debes indicar un valor de venta valido")

        normalized_seller_type, seller_user, external_seller_name = resolve_sale_seller(
            db,
            current_user,
            seller_type,
            internal_user_id,
            external_name
        )
        commission_pct, commission_amount, net_revenue = compute_sale_numbers(sold_price, seller_user)
        manual_sold_sale = db.query(models.Sale).filter(models.Sale.vehicle_id == vehicle.id).first()
        if manual_sold_sale:
            manual_sold_sale.seller_type = normalized_seller_type
            manual_sold_sale.external_seller_name = external_seller_name
            manual_sold_sale.seller_id = seller_user.id if seller_user else None
            manual_sold_sale.company_id = vehicle.company_id
            manual_sold_sale.sale_price = sold_price
            manual_sold_sale.commission_percentage = commission_pct
            manual_sold_sale.commission_amount = commission_amount
            manual_sold_sale.net_revenue = net_revenue
            manual_sold_sale.status = models.SaleStatus.APPROVED.value
            manual_sold_sale.sale_date = datetime.datetime.utcnow()
        else:
            manual_sold_sale = models.Sale(
                vehicle_id=vehicle.id,
                seller_id=seller_user.id if seller_user else None,
                seller_type=normalized_seller_type,
                external_seller_name=external_seller_name,
                company_id=vehicle.company_id,
                sale_price=sold_price,
                status=models.SaleStatus.APPROVED.value,
                commission_percentage=commission_pct,
                commission_amount=commission_amount,
                net_revenue=net_revenue,
                sale_date=datetime.datetime.utcnow()
            )
            db.add(manual_sold_sale)
        update_payload["price"] = sold_price

    for field, value in update_payload.items():
        if field in {"sold_price", "sold_by_type", "sold_by_internal_user_id", "sold_by_external_name"}:
            continue
        setattr(vehicle, field, value)

    target_make = vehicle_update.make if vehicle_update.make is not None else vehicle.make
    target_model = vehicle_update.model if vehicle_update.model is not None else vehicle.model
    upsert_brand_model(db, target_make, target_model)
        
    db.commit()
    db.refresh(vehicle)
    attach_sale_metadata_to_vehicle(vehicle, manual_sold_sale)
    return vehicle

@app.delete("/vehicles/{vehicle_id}", response_model=dict)
def delete_vehicle(
    vehicle_id: int, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    ensure_inventory_admin_only(current_user)

    vehicle = db.query(models.Vehicle).filter(models.Vehicle.id == vehicle_id).first()
    if not vehicle or vehicle.status == "inactive":
        raise HTTPException(status_code=404, detail="Vehicle not found")
        
    if current_user.company_id and vehicle.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    vehicle.status = "inactive"
    db.commit()
    return {"status": "success", "message": "Vehicle deactivated"}

# --- SALES & COMMISSION ENDPOINTS ---

@app.post("/sales/", response_model=schemas.Sale)
def create_sale(
    sale: schemas.SaleCreate, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(get_current_user)
):
    # 1. Verify Vehicle
    vehicle = db.query(models.Vehicle).filter(models.Vehicle.id == sale.vehicle_id).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
        
    if vehicle.status == "sold":
        raise HTTPException(status_code=400, detail="Vehicle is already sold")

    # If purchases already created a pending sale for this vehicle, reuse it instead of inserting a duplicate.
    existing_sale = db.query(models.Sale).filter(models.Sale.vehicle_id == sale.vehicle_id).first()
    if existing_sale and existing_sale.status in {"approved"}:
        raise HTTPException(status_code=400, detail="Vehicle is already sold")
    
    # 2. Resolve who sold it
    seller_type, seller, external_seller_name = resolve_sale_seller(
        db,
        current_user,
        sale.seller_type,
        sale.seller_id,
        sale.external_seller_name
    )
    
    # 3. Calculate Commission
    commission_pct, commission_amount, net_revenue = compute_sale_numbers(sale.sale_price, seller)
    
    # 4. Create or Update Sale Record
    if existing_sale:
        existing_sale.lead_id = sale.lead_id
        existing_sale.seller_id = seller.id if seller else None
        existing_sale.seller_type = seller_type
        existing_sale.external_seller_name = external_seller_name
        existing_sale.company_id = current_user.company_id or vehicle.company_id
        existing_sale.sale_price = sale.sale_price
        existing_sale.commission_percentage = commission_pct
        existing_sale.commission_amount = commission_amount
        existing_sale.net_revenue = net_revenue
        existing_sale.status = "pending"
        new_sale = existing_sale
    else:
        new_sale = models.Sale(
            vehicle_id=sale.vehicle_id,
            lead_id=sale.lead_id,
            seller_id=seller.id if seller else None,
            seller_type=seller_type,
            external_seller_name=external_seller_name,
            company_id=current_user.company_id or vehicle.company_id,
            sale_price=sale.sale_price,
            commission_percentage=commission_pct,
            commission_amount=commission_amount,
            net_revenue=net_revenue,
            status="pending"
        )
        db.add(new_sale)
    
    # 5. Lock Vehicle (Reserve it until approved)
    vehicle.status = "reserved"
    
    # 6. Update Lead Status if provided
    if sale.lead_id:
        lead = db.query(models.Lead).filter(models.Lead.id == sale.lead_id).first()
        if lead:
            lead.status = "sold"  # Mark as sold on frontend triggers this, so sync it back
            
    db.commit()
    db.refresh(new_sale)
    return new_sale

@app.get("/sales/", response_model=schemas.SaleList)
def read_sales(
    status: str = None, # Optional status
    month: int = None,
    year: int = None,
    q: str = None, # Search query
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    query = db.query(models.Sale).options(
        joinedload(models.Sale.vehicle),
        joinedload(models.Sale.lead),
        joinedload(models.Sale.seller),
        joinedload(models.Sale.payment_receipts)
    )
    
    # Permission Logic
    if is_company_admin_user(current_user):
        # Admin sees all company sales
        if current_user.company_id:
            query = query.filter(models.Sale.company_id == current_user.company_id)
    else:
        # Advisors/Others sees only THEIR sales
        query = query.filter(models.Sale.seller_id == current_user.id)
        
    # Filters
    if status:
        query = query.filter(models.Sale.status == status)

    if start_date and end_date:
        try:
            range_start = datetime.datetime.combine(datetime.date.fromisoformat(start_date), datetime.time.min)
            range_end = datetime.datetime.combine(datetime.date.fromisoformat(end_date) + datetime.timedelta(days=1), datetime.time.min)
        except ValueError:
            raise HTTPException(status_code=400, detail="Las fechas deben usar formato YYYY-MM-DD")
        if range_end <= range_start:
            raise HTTPException(status_code=400, detail="La fecha final debe ser mayor o igual a la inicial")
        query = query.filter(models.Sale.sale_date >= range_start, models.Sale.sale_date < range_end)
        
    if month and year:
        # Filter by specific month/year
        # Note: SQLite/Postgres specific might vary, using Generic extract or range
        # Easier to construct a date range for the month
        import calendar
        _, last_day = calendar.monthrange(year, month)
        start_date = datetime.datetime(year, month, 1)
        end_date = datetime.datetime(year, month, last_day, 23, 59, 59)
        query = query.filter(models.Sale.sale_date >= start_date, models.Sale.sale_date <= end_date)
        
    if q:
        search = f"%{q}%"
        query = query.join(models.Vehicle).filter(
            (models.Vehicle.make.ilike(search)) |
            (models.Vehicle.model.ilike(search)) |
            (models.Vehicle.plate.ilike(search))
        )
        
    total = query.count()
    sales = query.order_by(models.Sale.sale_date.desc()).offset(skip).limit(limit).all()
    return {"items": sales, "total": total}

@app.put("/sales/{sale_id}", response_model=schemas.Sale)
def update_sale(
    sale_id: int,
    sale_update: schemas.SaleUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    sale = db.query(models.Sale).options(
        joinedload(models.Sale.vehicle),
        joinedload(models.Sale.seller)
    ).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    if current_user.company_id and sale.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    can_edit = is_company_admin_user(current_user) or (
        sale.status == models.SaleStatus.PENDING.value and sale.seller_id == current_user.id
    )
    if not can_edit:
        raise HTTPException(status_code=403, detail="No tienes permisos para editar esta venta")

    if int(sale_update.sale_price or 0) <= 0:
        raise HTTPException(status_code=400, detail="Debes indicar un valor de venta valido")

    commission_pct, commission_amount, net_revenue = compute_sale_numbers(sale_update.sale_price, sale.seller)
    sale.sale_price = int(sale_update.sale_price)
    sale.commission_percentage = commission_pct
    sale.commission_amount = commission_amount
    sale.net_revenue = net_revenue
    if sale.vehicle:
        sale.vehicle.price = int(sale_update.sale_price)

    db.commit()
    db.refresh(sale)
    return sale


@app.delete("/sales/{sale_id}")
def delete_sale(
    sale_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if not is_company_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Solo un administrador puede eliminar ventas")

    sale = db.query(models.Sale).options(
        joinedload(models.Sale.vehicle)
    ).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    if current_user.company_id and sale.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    vehicle = sale.vehicle
    if vehicle and vehicle.status in {"reserved", "sold"}:
        vehicle.status = "available"

    db.delete(sale)
    db.commit()
    return {"message": "Venta eliminada correctamente"}

TAX_MONTH_LABELS = {
    1: "ENE", 2: "FEB", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
    7: "JULIO", 8: "AGOSTO", 9: "SEPT", 10: "OCT", 11: "NOV", 12: "DIC",
}


def _tax_clean(value: Optional[str]) -> Optional[str]:
    cleaned = (value or "").strip()
    return cleaned or None


def _get_sale_tax_query(db: Session, current_user: models.User):
    query = db.query(models.Sale).options(
        joinedload(models.Sale.vehicle),
        joinedload(models.Sale.lead),
        joinedload(models.Sale.seller),
    )
    if current_user.company_id:
        query = query.filter(models.Sale.company_id == current_user.company_id)
    return query


def _sale_tax_row(sale: models.Sale) -> dict:
    sale_date = sale.sale_date or datetime.datetime.utcnow()
    vehicle = sale.vehicle
    lead = sale.lead
    purchase_price = int(getattr(vehicle, "purchase_price", None) or 0)
    commission_base = int(round(purchase_price * 0.03))
    tax_iva = int(round(commission_base * 0.19))

    return {
        "source": "sale",
        "sale_id": sale.id,
        "manual_entry_id": None,
        "month": TAX_MONTH_LABELS.get(sale_date.month),
        "year": sale_date.year,
        "make": getattr(vehicle, "make", None),
        "reference": getattr(vehicle, "model", None),
        "plate": getattr(vehicle, "plate", None),
        "model_year": getattr(vehicle, "year", None),
        "purchase_price": purchase_price,
        "commission_base": commission_base,
        "tax_iva": tax_iva,
        "sale_price": int(sale.sale_price or 0),
        "iva_base": commission_base,
        "transaction_type": _tax_clean(getattr(sale, "tax_transaction_type", None)) or "INTERMEDIACION",
        "transfer_to_cars": _tax_clean(getattr(sale, "tax_transfer_to_cars", None)),
        "seller_name": _tax_clean(getattr(sale, "tax_seller_name", None)) or _tax_clean(getattr(sale, "external_seller_name", None)),
        "seller_document": _tax_clean(getattr(sale, "tax_seller_document", None)),
        "seller_email": _tax_clean(getattr(sale, "tax_seller_email", None)),
        "seller_address": _tax_clean(getattr(sale, "tax_seller_address", None)),
        "seller_phone": _tax_clean(getattr(sale, "tax_seller_phone", None)),
        "seller_payment_method": _tax_clean(getattr(sale, "tax_seller_payment_method", None)),
        "buyer_name": _tax_clean(getattr(sale, "tax_buyer_name", None)) or _tax_clean(getattr(lead, "name", None)),
        "buyer_document": _tax_clean(getattr(sale, "tax_buyer_document", None)),
        "buyer_email": _tax_clean(getattr(sale, "tax_buyer_email", None)) or _tax_clean(getattr(lead, "email", None)),
        "buyer_address": _tax_clean(getattr(sale, "tax_buyer_address", None)),
        "buyer_phone": _tax_clean(getattr(sale, "tax_buyer_phone", None)) or _tax_clean(getattr(lead, "phone", None)),
        "buyer_payment_method": _tax_clean(getattr(sale, "tax_buyer_payment_method", None)),
        "buyer_financing_entity": _tax_clean(getattr(sale, "tax_buyer_financing_entity", None)),
    }


def _manual_tax_row(entry: models.TaxReportEntry) -> dict:
    purchase_price = int(entry.purchase_price or 0)
    commission_base = int(round(purchase_price * 0.03))
    tax_iva = int(round(commission_base * 0.19))
    return {
        "source": "manual",
        "sale_id": None,
        "manual_entry_id": entry.id,
        "month": _tax_clean(entry.month),
        "year": entry.year,
        "make": _tax_clean(entry.make),
        "reference": _tax_clean(entry.reference),
        "plate": _tax_clean(entry.plate),
        "model_year": entry.model_year,
        "purchase_price": purchase_price,
        "commission_base": commission_base,
        "tax_iva": tax_iva,
        "sale_price": int(entry.sale_price or 0),
        "iva_base": commission_base,
        "transaction_type": _tax_clean(entry.transaction_type) or "INTERMEDIACION",
        "transfer_to_cars": _tax_clean(entry.transfer_to_cars),
        "seller_name": _tax_clean(entry.seller_name),
        "seller_document": _tax_clean(entry.seller_document),
        "seller_email": _tax_clean(entry.seller_email),
        "seller_address": _tax_clean(entry.seller_address),
        "seller_phone": _tax_clean(entry.seller_phone),
        "seller_payment_method": _tax_clean(entry.seller_payment_method),
        "buyer_name": _tax_clean(entry.buyer_name),
        "buyer_document": _tax_clean(entry.buyer_document),
        "buyer_email": _tax_clean(entry.buyer_email),
        "buyer_address": _tax_clean(entry.buyer_address),
        "buyer_phone": _tax_clean(entry.buyer_phone),
        "buyer_payment_method": _tax_clean(entry.buyer_payment_method),
        "buyer_financing_entity": _tax_clean(entry.buyer_financing_entity),
    }


def _manual_tax_query(db: Session, current_user: models.User):
    query = db.query(models.TaxReportEntry)
    if current_user.company_id:
        query = query.filter(models.TaxReportEntry.company_id == current_user.company_id)
    return query


def _tax_row_matches(row: dict, q: Optional[str]) -> bool:
    if not q:
        return True
    needle = q.strip().lower()
    return any(needle in str(row.get(field) or "").lower() for field in [
        "make", "reference", "plate", "seller_name", "buyer_name", "seller_document", "buyer_document"
    ])


def _tax_row_sort_key(row: dict):
    month_index = {label: number for number, label in TAX_MONTH_LABELS.items()}.get((row.get("month") or "").upper(), 13)
    return (row.get("year") or 9999, month_index, row.get("sale_id") or row.get("manual_entry_id") or 0)


@app.get("/finance/tax-report", response_model=schemas.TaxReportList)
def read_tax_report(
    year: Optional[int] = None,
    q: Optional[str] = None,
    skip: int = 0,
    limit: int = 300,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    query = _get_sale_tax_query(db, current_user).filter(models.Sale.status == models.SaleStatus.APPROVED.value)
    if year:
        start = datetime.datetime(year, 1, 1)
        end = datetime.datetime(year + 1, 1, 1)
        query = query.filter(models.Sale.sale_date >= start, models.Sale.sale_date < end)
    rows = [_sale_tax_row(sale) for sale in query.all()]

    manual_query = _manual_tax_query(db, current_user)
    if year:
        manual_query = manual_query.filter(models.TaxReportEntry.year == year)
    rows.extend(_manual_tax_row(entry) for entry in manual_query.all())
    rows = [row for row in rows if _tax_row_matches(row, q)]
    rows.sort(key=_tax_row_sort_key)
    return {"items": rows[skip:skip + limit], "total": len(rows)}


@app.put("/sales/{sale_id}/tax-info", response_model=schemas.TaxReportRow)
def update_sale_tax_info(
    sale_id: int,
    payload: schemas.SaleTaxInfoUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Solo un administrador puede editar tributacion")

    sale = _get_sale_tax_query(db, current_user).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(sale, field, _tax_clean(value))

    db.commit()
    db.refresh(sale)
    return _sale_tax_row(sale)


def _apply_manual_tax_payload(entry: models.TaxReportEntry, payload: schemas.TaxReportManualEntryCreate):
    numeric_fields = {"year", "model_year", "purchase_price", "sale_price"}
    for field, value in payload.model_dump(exclude_unset=True).items():
        if field in numeric_fields:
            setattr(entry, field, int(value or 0) if value is not None else None)
        else:
            setattr(entry, field, _tax_clean(value))


@app.post("/finance/tax-report/manual", response_model=schemas.TaxReportRow)
def create_manual_tax_report_entry(
    payload: schemas.TaxReportManualEntryCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Solo un administrador puede crear tributacion manual")
    company_id = current_user.company_id
    if not company_id:
        raise HTTPException(status_code=400, detail="Debes estar asociado a una empresa")
    entry = models.TaxReportEntry(company_id=company_id)
    _apply_manual_tax_payload(entry, payload)
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return _manual_tax_row(entry)


@app.put("/finance/tax-report/manual/{entry_id}", response_model=schemas.TaxReportRow)
def update_manual_tax_report_entry(
    entry_id: int,
    payload: schemas.TaxReportManualEntryCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Solo un administrador puede editar tributacion manual")
    entry = _manual_tax_query(db, current_user).filter(models.TaxReportEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Registro manual no encontrado")
    _apply_manual_tax_payload(entry, payload)
    db.commit()
    db.refresh(entry)
    return _manual_tax_row(entry)


@app.delete("/finance/tax-report/manual/{entry_id}")
def delete_manual_tax_report_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Solo un administrador puede eliminar tributacion manual")
    entry = _manual_tax_query(db, current_user).filter(models.TaxReportEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Registro manual no encontrado")
    db.delete(entry)
    db.commit()
    return {"status": "success"}


@app.get("/finance/tax-report.xlsx")
def download_tax_report_xlsx(
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_user_from_anywhere)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    query = _get_sale_tax_query(db, current_user).filter(models.Sale.status == models.SaleStatus.APPROVED.value)
    if year:
        start = datetime.datetime(year, 1, 1)
        end = datetime.datetime(year + 1, 1, 1)
        query = query.filter(models.Sale.sale_date >= start, models.Sale.sale_date < end)
    rows = [_sale_tax_row(sale) for sale in query.all()]
    manual_query = _manual_tax_query(db, current_user)
    if year:
        manual_query = manual_query.filter(models.TaxReportEntry.year == year)
    rows.extend(_manual_tax_row(entry) for entry in manual_query.all())
    rows.sort(key=_tax_row_sort_key)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="La libreria openpyxl no esta instalada en el servidor")

    wb = Workbook()
    ws = wb.active
    ws.title = f"ANUAL {year or datetime.datetime.utcnow().year}"
    ws.merge_cells("A1:N2")
    ws.merge_cells("O1:T1")
    ws.merge_cells("O2:T2")
    ws.merge_cells("U1:AA1")
    ws.merge_cells("U2:AA2")
    ws["A1"] = "QUIEN ME VENDIO EL CARRO"
    ws["O1"] = "A QUIEN LE COMPRE EL CARRO"
    ws["O2"] = "DATOS CLIENTE"
    ws["U1"] = "A QUIEN LE VENDI EL CARRO"
    ws["U2"] = "DATOS CLIENTE"
    headers = [
        "MES", "AÑO", "MARCA", "REFERENCIA", "PLACA", "MODELO",
        "PRECIO COMPRA / CONSIGNACIÓN", "COMISION", "IVA", "PRECIO VENTA",
        "BASE DEL IVA", "IVA", "INTERMEDIACION O VENTA COMPLETA", "TRASPASO A CARS SI /NO",
        "NOMBRES O RAZON SOCIAL", "DOCUMENTO / NIT", "EMAIL", "DIRECCIÓN", "TELÉFONO",
        "FORMA DE PAGO A VENDEDOR", "NOMBRES", "DOCUMENTO", "EMAIL", "DIRECCIÓN",
        "TELÉFONO", "FORMA DE PAGO DEL COMPRADOR", "ENTIDAD / OBSERVACION",
    ]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1E293B")
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill

    for index, row in enumerate(rows, start=4):
        ws.append([
            row["month"], row["year"], row["make"], row["reference"], row["plate"], row["model_year"],
            row["purchase_price"], f"=G{index}*3%", f"=H{index}*0.19", row["sale_price"],
            f"=H{index}", None, row["transaction_type"], row["transfer_to_cars"],
            row["seller_name"], row["seller_document"], row["seller_email"], row["seller_address"],
            row["seller_phone"], row["seller_payment_method"], row["buyer_name"], row["buyer_document"],
            row["buyer_email"], row["buyer_address"], row["buyer_phone"], row["buyer_payment_method"],
            row["buyer_financing_entity"],
        ])

    widths = [10, 8, 16, 22, 12, 10, 18, 14, 14, 16, 14, 12, 24, 18, 26, 18, 28, 34, 16, 24, 26, 18, 28, 34, 16, 28, 28]
    for idx, width_value in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(3, idx).column_letter].width = width_value
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in {7, 8, 9, 10, 11, 12}:
                cell.number_format = '"$"#,##0'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename_year = year or datetime.datetime.utcnow().year
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="cuadro_tributacion_{filename_year}.xlsx"'}
    )

@app.put("/sales/{sale_id}/approve", response_model=schemas.Sale)
def approve_sale(
    sale_id: int, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Only Admin or Super Admin
    if not is_company_admin_user(current_user):
         raise HTTPException(status_code=403, detail="Only admins can approve sales")
         
    sale = db.query(models.Sale).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
        
    if sale.status == "approved":
        raise HTTPException(status_code=400, detail="Sale already approved")
        
    # Finalize
    sale.status = "approved"
    # Create Automatic Receipt for the Sale
    try:
        count = db.query(models.PaymentReceipt).filter(models.PaymentReceipt.company_id == sale.company_id).count()
        receipt_number = f"REC-{count + 1}"
        automatic_receipt = models.PaymentReceipt(
            company_id=sale.company_id,
            sale_id=sale.id,
            user_id=current_user.id,
            concept=f"Venta de vehiculo #{sale.id}",
            movement_type="income",
            receipt_number=receipt_number,
            payment_date=datetime.datetime.utcnow(),
            amount=sale.sale_price,
            category="ingreso_venta",
            notes=f"Recibo generado automaticamente al aprobar la venta #{sale.id}"
        )
        db.add(automatic_receipt)
    except Exception as e:
        print(f"Error creating automatic receipt: {e}")
    sale.approved_by_id = current_user.id
    sale.sale_date = datetime.datetime.utcnow()
    
    # Mark Vehicle as Sold
    vehicle = db.query(models.Vehicle).filter(models.Vehicle.id == sale.vehicle_id).first()
    if vehicle:
        vehicle.status = "sold"
        vehicle.price = sale.sale_price
        
    db.commit()
    db.refresh(sale)
    return sale


@app.put("/sales/{sale_id}/reject", response_model=schemas.Sale)
def reject_sale(
    sale_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if not is_company_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Only admins can reject sales")

    sale = db.query(models.Sale).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    if sale.status == models.SaleStatus.REJECTED.value:
        raise HTTPException(status_code=400, detail="Sale already rejected")
    if sale.status == models.SaleStatus.APPROVED.value:
        raise HTTPException(status_code=400, detail="Approved sales cannot be rejected")

    sale.status = models.SaleStatus.REJECTED.value
    sale.approved_by_id = current_user.id

    vehicle = db.query(models.Vehicle).filter(models.Vehicle.id == sale.vehicle_id).first()
    if vehicle and vehicle.status != models.VehicleStatus.SOLD.value:
        vehicle.status = models.VehicleStatus.AVAILABLE.value

    db.commit()
    db.refresh(sale)
    return sale

@app.get("/finance/stats")
def get_finance_stats(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Only Admin/Super Admin
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
         raise HTTPException(status_code=403, detail="Not authorized")

    range_start = None
    range_end = None
    if start_date and end_date:
        try:
            range_start = datetime.datetime.combine(datetime.date.fromisoformat(start_date), datetime.time.min)
            range_end = datetime.datetime.combine(datetime.date.fromisoformat(end_date) + datetime.timedelta(days=1), datetime.time.min)
        except ValueError:
            raise HTTPException(status_code=400, detail="Las fechas deben usar formato YYYY-MM-DD")
        if range_end <= range_start:
            raise HTTPException(status_code=400, detail="La fecha final debe ser mayor o igual a la inicial")
         
    # 1. Total Stats (All time)
    query = db.query(models.Sale).filter(models.Sale.status == "approved")
    if current_user.company_id:
        query = query.filter(models.Sale.company_id == current_user.company_id)
    if range_start and range_end:
        query = query.filter(models.Sale.sale_date >= range_start, models.Sale.sale_date < range_end)
    sales = query.all()
    
    total_revenue = sum(s.net_revenue for s in sales)
    total_commissions = sum(s.commission_amount for s in sales)
    total_sales_count = len(sales)
    
    # 2. Pending Count
    pending_query = db.query(models.Sale).filter(models.Sale.status == "pending")
    if current_user.company_id:
        pending_query = pending_query.filter(models.Sale.company_id == current_user.company_id)
    if range_start and range_end:
        pending_query = pending_query.filter(models.Sale.sale_date >= range_start, models.Sale.sale_date < range_end)
    pending_count = pending_query.count()

    # 3. Monthly Stats (Current Month)
    now = datetime.datetime.utcnow()
    current_month_sales = sales if (range_start and range_end) else [s for s in sales if s.sale_date and s.sale_date.month == now.month and s.sale_date.year == now.year]
    
    monthly_revenue = sum(s.net_revenue for s in current_month_sales)
    monthly_commissions = sum(s.commission_amount for s in current_month_sales)

    # 4. Payroll Expenses (Sum of base_salary of all users)
    # Filter by company if needed
    user_query = db.query(models.User)
    if current_user.company_id:
        user_query = user_query.filter(models.User.company_id == current_user.company_id)
    
    users = user_query.all()
    payroll_expenses = sum(u.base_salary or 0 for u in users)

    receipts_query = db.query(models.PaymentReceipt)
    if current_user.company_id:
        receipts_query = receipts_query.filter(models.PaymentReceipt.company_id == current_user.company_id)
    if range_start and range_end:
        receipt_date_field = func.coalesce(models.PaymentReceipt.payment_date, models.PaymentReceipt.created_at)
        receipts_query = receipts_query.filter(receipt_date_field >= range_start, receipt_date_field < range_end)
    receipts = receipts_query.all()
    current_month_receipts = receipts if (range_start and range_end) else [
        receipt for receipt in receipts
        if receipt.payment_date and receipt.payment_date.month == now.month and receipt.payment_date.year == now.year
    ]
    total_income = sum((receipt.amount or 0) for receipt in receipts if (receipt.movement_type or "income") == "income")
    total_expense = sum((receipt.amount or 0) for receipt in receipts if (receipt.movement_type or "income") == "expense")
    monthly_income = sum((receipt.amount or 0) for receipt in current_month_receipts if (receipt.movement_type or "income") == "income")
    monthly_expense = sum((receipt.amount or 0) for receipt in current_month_receipts if (receipt.movement_type or "income") == "expense")
    
    return {
        "total_revenue": total_revenue,
        "total_commissions": total_commissions,
        "total_sales_count": total_sales_count,
        "pending_sales_count": pending_count,
        "monthly_revenue": monthly_revenue, 
        "monthly_commissions": monthly_commissions,
        "payroll_expenses": payroll_expenses,
        "receipts_total_count": len(receipts),
        "receipts_total_amount": sum(receipt.amount or 0 for receipt in receipts),
        "receipts_monthly_amount": sum(receipt.amount or 0 for receipt in current_month_receipts),
        "accounting_income_total": total_income,
        "accounting_expense_total": total_expense,
        "accounting_balance_total": total_income - total_expense,
        "accounting_income_monthly": monthly_income,
        "accounting_expense_monthly": monthly_expense,
        "accounting_balance_monthly": monthly_income - monthly_expense
    }


def _projection_money(value) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _projection_date(value):
    if not value:
        return None
    if isinstance(value, datetime.datetime):
        return value
    return None


def _projection_month_label(value):
    date_value = _projection_date(value) or datetime.datetime.utcnow()
    return TAX_MONTH_LABELS.get(date_value.month, "")


def _receipt_sum(receipts, movement_type=None, categories=None):
    total = 0
    category_set = set(categories or [])
    for receipt in receipts or []:
        if movement_type and (receipt.movement_type or "income") != movement_type:
            continue
        if category_set and receipt.category not in category_set:
            continue
        total += _projection_money(receipt.amount)
    return total


def _first_receipt(receipts, movement_type=None, categories=None):
    category_set = set(categories or [])
    ordered = sorted(receipts or [], key=lambda item: item.payment_date or item.created_at or datetime.datetime.min)
    for receipt in ordered:
        if movement_type and (receipt.movement_type or "income") != movement_type:
            continue
        if category_set and receipt.category not in category_set:
            continue
        return receipt
    return None


def _accounting_category_label(category: Optional[str]) -> str:
    labels = {
        "ingreso_venta": "Ingresos por venta",
        "sale_payment": "Ingresos por venta",
        "other_income": "Otros ingresos",
        "costo_vehiculo": "Costo de vehículo",
        "vehicle_purchase": "Compra de vehículo",
        "purchase_payment": "Abono compra",
        "vehicle_separation": "Separación del vehículo",
        "gasto_tramites": "Gastos de trámites y alistamiento",
        "vehicle_expense": "Gastos del vehículo",
        "expense": "Gastos generales",
        "gasto_operativo": "Gastos operativos y administrativos",
        "comisiones": "Pago de comisiones",
        "otros": "Otros movimientos",
    }
    if not category:
        return "Sin cuenta"
    return labels.get(category, category.replace("_", " ").capitalize())


def _accounting_movement_label(movement_type: Optional[str]) -> str:
    return "Egreso" if (movement_type or "income") == "expense" else "Ingreso"


def _projection_sales_query(db: Session, current_user: models.User):
    query = db.query(models.Sale).options(
        joinedload(models.Sale.vehicle),
        joinedload(models.Sale.lead),
        joinedload(models.Sale.seller),
        joinedload(models.Sale.payment_receipts),
    )
    if current_user.company_id:
        query = query.filter(models.Sale.company_id == current_user.company_id)
    return query


@app.get("/finance/projection.xlsx")
def download_finance_projection_xlsx(
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_user_from_anywhere)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    report_year = year or datetime.datetime.utcnow().year
    start = datetime.datetime(report_year, 1, 1)
    end = datetime.datetime(report_year + 1, 1, 1)
    sales = _projection_sales_query(db, current_user).filter(
        models.Sale.sale_date >= start,
        models.Sale.sale_date < end
    ).order_by(models.Sale.sale_date.asc(), models.Sale.id.asc()).all()

    receipts_query = db.query(models.PaymentReceipt).options(
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.vehicle),
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.lead),
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.seller),
    )
    if current_user.company_id:
        receipts_query = receipts_query.filter(models.PaymentReceipt.company_id == current_user.company_id)
    receipt_date = func.coalesce(models.PaymentReceipt.payment_date, models.PaymentReceipt.created_at)
    receipts = receipts_query.filter(receipt_date >= start, receipt_date < end).all()

    credits_query = db.query(models.CreditApplication).options(joinedload(models.CreditApplication.lead))
    if current_user.company_id:
        credits_query = credits_query.filter(models.CreditApplication.company_id == current_user.company_id)
    credits = credits_query.filter(models.CreditApplication.created_at >= start, models.CreditApplication.created_at < end).all()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="La libreria openpyxl no esta instalada en el servidor")

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1E293B")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def autosize(ws, max_width=34):
        for column_cells in ws.columns:
            length = 10
            for cell in column_cells:
                length = max(length, min(len(str(cell.value or "")) + 2, max_width))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.column_dimensions[column_cells[0].column_letter].width = length

    costs = wb.active
    costs.title = "COSTOS X VH"
    costs.append([
        "MES", "ITEM", "ASESOR VENTA", "EQUIPO", "ESTADO", "CONTINUA CON EL NEGOCIO",
        "NOMBRE DEL CLIENTE", "REFERENCIA", "PLACA", "TIPO DE TRANSACCION", "MODELO",
        "COSTO VEHICULO", "GASTOS VEHICULO", "COSTOS TOTALES", "VALOR DE VENTA",
        "SEPARACION / CUOTA INICIAL", "SALDO", "DESEMBOLSO", "COMISION CREDITO",
        "UTILIDAD NETA"
    ])
    style_header(costs)
    for index, sale in enumerate(sales, start=1):
        vehicle = sale.vehicle
        lead = sale.lead
        sale_receipts = sale.payment_receipts or []
        purchase_cost = _projection_money(getattr(vehicle, "purchase_price", None))
        vehicle_expenses = _receipt_sum(sale_receipts, "expense", ["vehicle_expense", "gasto_tramites", "gasto_operativo"])
        purchase_receipts = _receipt_sum(sale_receipts, "expense", ["vehicle_purchase", "costo_vehiculo"])
        sale_income = _receipt_sum(sale_receipts, "income")
        down_payment = _projection_money(getattr(getattr(lead, "process_detail", None), "reservation_amount", None))
        credit_amount = _projection_money(getattr(getattr(lead, "process_detail", None), "credit_used_amount", None))
        sale_value = _projection_money(sale.sale_price)
        total_cost = purchase_cost + purchase_receipts + vehicle_expenses
        costs.append([
            _projection_month_label(sale.sale_date), index, get_sale_display_name(sale), getattr(sale.company, "name", None),
            sale.status, "CONTINUA" if sale.status != models.SaleStatus.REJECTED.value else "DESISTE",
            getattr(lead, "name", None), " ".join(filter(None, [getattr(vehicle, "make", None), getattr(vehicle, "model", None)])),
            getattr(vehicle, "plate", None), getattr(sale, "tax_transaction_type", None) or "COMPRA",
            getattr(vehicle, "year", None), purchase_cost, vehicle_expenses, total_cost, sale_value,
            down_payment, max(sale_value - down_payment - credit_amount, 0), credit_amount,
            _projection_money(sale.commission_amount), sale_value - total_cost - _projection_money(sale.commission_amount)
        ])
    autosize(costs)

    expenses = wb.create_sheet("GASTOS")
    expenses.append(["FECHA", "CONCEPTO", "CUENTA", "TIPO", "VALOR", "EFEC", "TRANS", "BANCO", "VENTA", "PLACA", "NOTA"])
    style_header(expenses)
    for receipt in receipts:
        if (receipt.movement_type or "income") != "expense":
            continue
        expenses.append([
            receipt.payment_date or receipt.created_at,
            receipt.concept,
            _accounting_category_label(receipt.category),
            _accounting_movement_label(receipt.movement_type),
            _projection_money(receipt.amount),
            "X" if (receipt.payment_method or "").lower() == "efectivo" else "",
            "X" if (receipt.payment_method or "").lower() == "transferencia" else "",
            receipt.bank,
            getattr(receipt.sale, "id", None),
            getattr(getattr(receipt.sale, "vehicle", None), "plate", None), receipt.notes
        ])
    autosize(expenses)

    process = wb.create_sheet("VENTAS EN PROCESO")
    process.append([
        "ITEM", "FECHA INGRESO", "ESTADO", "NOMBRE DEL EJECUTIVO", "NOMBRE DEL CLIENTE",
        "TELEFONO", "FECHA DE RADICACION", "BANCO", "RTA BANCO", "CARRO", "PLACA",
        "% DE FINANCIACION", "VALOR DEL VH", "CUANTO LE PRESTAN", "VALOR CUOTA INICIAL", "OBSERVACION"
    ])
    style_header(process)
    for index, credit in enumerate(credits, start=1):
        lead = credit.lead
        process_detail = getattr(lead, "process_detail", None)
        vehicle_value = _projection_money(credit.purchase_sale_price or credit.approved_amount)
        approved_amount = _projection_money(credit.approved_amount)
        financing_pct = round(approved_amount / vehicle_value, 2) if vehicle_value else None
        process.append([
            index, credit.created_at, credit.status, getattr(getattr(lead, "assigned_to", None), "full_name", None),
            credit.client_name or getattr(lead, "name", None), credit.phone or getattr(lead, "phone", None),
            credit.updated_at, getattr(process_detail, "reservation_payment_method", None), credit.notes,
            credit.purchase_vehicle_name or credit.desired_vehicle, credit.purchase_vehicle_plate,
            financing_pct, vehicle_value, approved_amount,
            _projection_money(getattr(process_detail, "reservation_amount", None) or credit.down_payment or credit.approved_down_payment),
            credit.notes
        ])
    autosize(process)

    payments = wb.create_sheet("RELACION DE PAGOS X CARRO")
    payments.append([
        "FECHA", "VEHICULO", "PLACA", "MODELO", "VENDEDOR", "CEDULA", "CORREO", "DIRECCION",
        "CELULAR", "OBSERVACIONES", "VALOR COMPRA", "SEPARACION", "VALOR", "BANCO",
        "ABONOS / PAGOS", "GASTOS", "PRECIO FINAL", "CLIENTE", "VALOR VENTA", "TOTAL INGRESOS",
        "TOTAL GASTOS", "DIFERENCIA", "BASE IMPUESTO", "IVA"
    ])
    style_header(payments)
    for sale in sales:
        vehicle = sale.vehicle
        lead = sale.lead
        sale_receipts = sorted(sale.payment_receipts or [], key=lambda item: item.payment_date or item.created_at or datetime.datetime.min)
        first_income = _first_receipt(sale_receipts, "income")
        total_income = _receipt_sum(sale_receipts, "income")
        total_expense = _receipt_sum(sale_receipts, "expense")
        difference = total_income - total_expense
        base_tax = int(difference / 1.19) if difference else 0
        iva = int(base_tax * 0.19)
        payments.append([
            _projection_month_label(sale.sale_date),
            " ".join(filter(None, [getattr(vehicle, "make", None), getattr(vehicle, "model", None)])),
            getattr(vehicle, "plate", None), getattr(vehicle, "year", None),
            getattr(sale, "tax_seller_name", None) or get_sale_display_name(sale),
            getattr(sale, "tax_seller_document", None), getattr(sale, "tax_seller_email", None),
            getattr(sale, "tax_seller_address", None), getattr(sale, "tax_seller_phone", None),
            getattr(sale, "tax_seller_payment_method", None), _projection_money(getattr(vehicle, "purchase_price", None)),
            getattr(first_income, "payment_date", None), _projection_money(getattr(first_income, "amount", None)),
            getattr(first_income, "bank", None) or getattr(first_income, "receipt_number", None),
            " | ".join(f"{r.payment_date:%d/%m/%Y}: {r.concept or r.category} ${_projection_money(r.amount):,}" for r in sale_receipts if (r.movement_type or "income") == "income"),
            total_expense, _projection_money(sale.sale_price), getattr(lead, "name", None),
            _projection_money(sale.sale_price), total_income, total_expense, difference, base_tax, iva
        ])
    autosize(payments)

    status_ws = wb.create_sheet("ESTADO VENTAS")
    status_ws.append(["FECHA", "ASESOR", "CLIENTE", "ACTIVIDAD PENDIENTE", "RESPONSABLE", "FECHA DE COMPROMISO", "ESTADO"])
    style_header(status_ws)
    for credit in credits:
        lead = credit.lead
        status_ws.append([
            credit.updated_at, getattr(getattr(lead, "assigned_to", None), "full_name", None),
            credit.client_name or getattr(lead, "name", None), credit.notes or credit.status,
            getattr(getattr(credit, "assigned_to", None), "full_name", None), credit.updated_at, credit.status
        ])
    autosize(status_ws)

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.column >= 10:
                    cell.number_format = '"$"#,##0'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="proyeccion_contable_{report_year}.xlsx"'}
    )


@app.get("/finance/receipts", response_model=schemas.PaymentReceiptList)
def read_payment_receipts(
    sale_id: Optional[int] = None,
    receipt_number: Optional[str] = None,
    category: Optional[str] = None,
    movement_type: Optional[str] = None,
    q: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    query = db.query(models.PaymentReceipt).options(
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.vehicle),
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.seller),
        joinedload(models.PaymentReceipt.user)
    )
    if current_user.company_id:
        query = query.filter(models.PaymentReceipt.company_id == current_user.company_id)
    if sale_id:
        query = query.filter(models.PaymentReceipt.sale_id == sale_id)
    if receipt_number:
        query = query.filter(models.PaymentReceipt.receipt_number == receipt_number.strip())
    if category:
        query = query.filter(models.PaymentReceipt.category == category)
    if movement_type:
        query = query.filter(models.PaymentReceipt.movement_type == movement_type)
    if start_date and end_date:
        try:
            range_start = datetime.datetime.combine(datetime.date.fromisoformat(start_date), datetime.time.min)
            range_end = datetime.datetime.combine(datetime.date.fromisoformat(end_date) + datetime.timedelta(days=1), datetime.time.min)
        except ValueError:
            raise HTTPException(status_code=400, detail="Las fechas deben usar formato YYYY-MM-DD")
        if range_end <= range_start:
            raise HTTPException(status_code=400, detail="La fecha final debe ser mayor o igual a la inicial")
        receipt_date_field = func.coalesce(models.PaymentReceipt.payment_date, models.PaymentReceipt.created_at)
        query = query.filter(receipt_date_field >= range_start, receipt_date_field < range_end)
    if q:
        query = apply_receipt_group_search_filter(query, db, q, current_user.company_id)

    total = query.count()
    items = query.order_by(models.PaymentReceipt.payment_date.desc(), models.PaymentReceipt.id.desc()).offset(skip).limit(limit).all()
    items = enrich_receipt_display_names(db, items, current_user.company_id)
    return {"items": items, "total": total}


@app.get("/finance/receipts.xlsx")
def download_payment_receipts_xlsx(
    category: Optional[str] = None,
    movement_type: Optional[str] = None,
    q: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_user_from_anywhere)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="La libreria openpyxl no esta instalada en el servidor")

    query = db.query(models.PaymentReceipt).options(
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.vehicle),
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.seller),
        joinedload(models.PaymentReceipt.user)
    )
    if current_user.company_id:
        query = query.filter(models.PaymentReceipt.company_id == current_user.company_id)
    if category:
        query = query.filter(models.PaymentReceipt.category == category)
    if movement_type:
        query = query.filter(models.PaymentReceipt.movement_type == movement_type)
    if start_date and end_date:
        try:
            range_start = datetime.datetime.combine(datetime.date.fromisoformat(start_date), datetime.time.min)
            range_end = datetime.datetime.combine(datetime.date.fromisoformat(end_date) + datetime.timedelta(days=1), datetime.time.min)
        except ValueError:
            raise HTTPException(status_code=400, detail="Las fechas deben usar formato YYYY-MM-DD")
        receipt_date_field = func.coalesce(models.PaymentReceipt.payment_date, models.PaymentReceipt.created_at)
        query = query.filter(receipt_date_field >= range_start, receipt_date_field < range_end)
    if q:
        query = apply_receipt_group_search_filter(query, db, q, current_user.company_id)

    receipts = query.order_by(models.PaymentReceipt.payment_date.desc(), models.PaymentReceipt.id.desc()).all()
    receipts = enrich_receipt_display_names(db, receipts, current_user.company_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Contabilidad"
    headers = [
        "FECHA", "VENTA", "VEHICULO", "PLACA", "CONCEPTO", "RECIBO", "TIPO", "CUENTA",
        "VALOR", "EFECTIVO", "TRANSFERENCIA", "BANCO", "NOTA", "SOPORTE", "CREADO POR"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")

    for receipt in receipts:
        vehicle = getattr(receipt.sale, "vehicle", None)
        vehicle_label = " ".join(filter(None, [getattr(vehicle, "make", None), getattr(vehicle, "model", None)])).strip()
        payment_method = (receipt.payment_method or "").lower()
        ws.append([
            receipt.payment_date or receipt.created_at,
            getattr(receipt.sale, "id", None),
            vehicle_label,
            getattr(vehicle, "plate", None),
            receipt.concept,
            receipt.receipt_number,
            _accounting_movement_label(receipt.movement_type),
            _accounting_category_label(receipt.category),
            int(receipt.amount or 0),
            "X" if payment_method == "efectivo" else "",
            "X" if payment_method == "transferencia" else "",
            receipt.bank,
            receipt.notes,
            receipt.file_name or "",
            getattr(receipt.user, "full_name", None) or getattr(receipt.user, "email", None)
        ])

    for column_cells in ws.columns:
        width = 12
        for cell in column_cells:
            width = max(width, min(len(str(cell.value or "")) + 2, 45))
        ws.column_dimensions[column_cells[0].column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="contabilidad_recibos.xlsx"'}
    )


def _get_receipt_with_access(
    db: Session,
    receipt_id: int,
    current_user: models.User
) -> models.PaymentReceipt:
    receipt = db.query(models.PaymentReceipt).options(
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.vehicle),
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.seller),
        joinedload(models.PaymentReceipt.user)
    ).filter(models.PaymentReceipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="Receipt not found")
    if current_user.company_id and receipt.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    return receipt


def _get_sale_with_finance_access(
    db: Session,
    sale_id: int,
    current_user: models.User,
    include_attachments: bool = False
) -> models.Sale:
    options = [
        joinedload(models.Sale.vehicle),
        joinedload(models.Sale.seller),
        joinedload(models.Sale.lead)
    ]
    if include_attachments:
        options.append(joinedload(models.Sale.attachments).joinedload(models.SaleAttachment.user))
    sale = db.query(models.Sale).options(*options).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    if current_user.company_id and sale.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    return sale


def _normalize_receipt_payment_method(payment_method: Optional[str]) -> Optional[str]:
    payment_method_value = (payment_method or "").strip().lower() or None
    if payment_method_value and payment_method_value not in {"efectivo", "transferencia"}:
        raise HTTPException(status_code=400, detail="La forma de pago debe ser efectivo o transferencia")
    return payment_method_value


def _generate_accounting_receipt_number(db: Session, company_id: Optional[int]) -> str:
    query = db.query(func.count(models.PaymentReceipt.id)).filter(
        models.PaymentReceipt.sale_id.is_(None)
    )
    if company_id:
        query = query.filter(models.PaymentReceipt.company_id == company_id)
    count = query.scalar() or 0
    return f"CONT-{count + 1:05d}"


@app.put("/finance/receipts/{receipt_id}", response_model=schemas.PaymentReceipt)
def update_payment_receipt(
    receipt_id: int,
    receipt_update: schemas.PaymentReceiptUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Solo un administrador puede editar recibos")

    receipt = _get_receipt_with_access(db, receipt_id, current_user)
    movement_type_value = (receipt_update.movement_type or "income").strip().lower()
    if movement_type_value not in {"income", "expense"}:
        raise HTTPException(status_code=400, detail="movement_type must be income or expense")
    payment_method_value = _normalize_receipt_payment_method(receipt_update.payment_method)
    if int(receipt_update.amount or 0) <= 0:
        raise HTTPException(status_code=400, detail="Debes indicar un valor valido")

    sale = None
    company_id = current_user.company_id or receipt.company_id
    if receipt_update.sale_id:
        sale = db.query(models.Sale).options(
            joinedload(models.Sale.vehicle),
            joinedload(models.Sale.seller)
        ).filter(models.Sale.id == receipt_update.sale_id).first()
        if not sale:
            raise HTTPException(status_code=404, detail="Sale not found")
        if current_user.company_id and sale.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="Not authorized")
        company_id = sale.company_id

    concept_value = (receipt_update.concept or "").strip() or None
    if not sale and not concept_value:
        raise HTTPException(status_code=400, detail="Debes indicar una venta o un concepto contable")

    receipt.sale_id = sale.id if sale else None
    receipt.company_id = company_id
    receipt.display_name = (receipt_update.display_name or "").strip() or None
    receipt.customer_name = (receipt_update.customer_name or "").strip() or None
    receipt.customer_document = (receipt_update.customer_document or "").strip() or None
    receipt.concept = concept_value
    receipt.movement_type = movement_type_value
    receipt.receipt_number = (receipt_update.receipt_number or "").strip() or None
    receipt.payment_date = receipt_update.payment_date or receipt.payment_date
    receipt.amount = int(receipt_update.amount)
    receipt.category = (receipt_update.category or "sale_payment").strip() or "sale_payment"
    receipt.notes = (receipt_update.notes or "").strip() or None
    receipt.payment_method = payment_method_value
    receipt.bank = (receipt_update.bank or "").strip() or None

    db.commit()
    db.refresh(receipt)
    return db.query(models.PaymentReceipt).options(
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.vehicle),
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.seller),
        joinedload(models.PaymentReceipt.user)
    ).filter(models.PaymentReceipt.id == receipt.id).first()


@app.post("/finance/receipts/{receipt_id}/upload", response_model=schemas.PaymentReceipt)
async def upload_payment_receipt_file(
    receipt_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    receipt = _get_receipt_with_access(db, receipt_id, current_user)

    stored_file_name = None
    stored_file_path = None
    stored_file_type = None
    if file and file.filename:
        os.makedirs("static/accounting", exist_ok=True)
        safe_name = os.path.basename(file.filename)
        extension = safe_name.split(".")[-1] if "." in safe_name else "bin"
        unique_filename = f"{uuid.uuid4()}.{extension}"
        file_path = f"static/accounting/{unique_filename}"
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        stored_file_name = safe_name
        stored_file_path = f"/{file_path.replace('\\', '/')}"
        stored_file_type = file.content_type

        receipt.file_name = stored_file_name
        receipt.file_path = stored_file_path
        receipt.file_type = stored_file_type

        db.commit()
        db.refresh(receipt)

    return db.query(models.PaymentReceipt).options(
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.vehicle),
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.seller),
        joinedload(models.PaymentReceipt.user)
    ).filter(models.PaymentReceipt.id == receipt.id).first()


@app.post("/finance/receipts", response_model=schemas.PaymentReceipt)
async def create_payment_receipt(
    request: Request,
    sale_id: Optional[int] = Form(None),
    concept: Optional[str] = Form(None),
    display_name: Optional[str] = Form(None),
    customer_name: Optional[str] = Form(None),
    customer_document: Optional[str] = Form(None),
    amount: int = Form(...),
    movement_type: Optional[str] = Form("income"),
    payment_date: Optional[str] = Form(None),
    receipt_number: Optional[str] = Form(None),
    category: Optional[str] = Form("sale_payment"),
    notes: Optional[str] = Form(None),
    payment_method: Optional[str] = Form(None),
    bank: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    concept_value = (concept or "").strip() or None
    movement_type_value = (movement_type or "income").strip().lower()
    if movement_type_value not in {"income", "expense"}:
        raise HTTPException(status_code=400, detail="movement_type must be income or expense")
    payment_method_value = _normalize_receipt_payment_method(payment_method)
    sale = None
    company_id = current_user.company_id
    if sale_id:
        sale = db.query(models.Sale).options(
            joinedload(models.Sale.vehicle),
            joinedload(models.Sale.seller)
        ).filter(models.Sale.id == sale_id).first()
        if not sale:
            raise HTTPException(status_code=404, detail="Sale not found")
        if current_user.company_id and sale.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="Not authorized")
        company_id = sale.company_id
    elif not concept_value:
        raise HTTPException(status_code=400, detail="Debes indicar una venta o un concepto contable")

    receipt_number_value = (receipt_number or "").strip() or None
    if not sale and not receipt_number_value:
        receipt_number_value = _generate_accounting_receipt_number(db, company_id)

    parsed_payment_date = datetime.datetime.utcnow()
    if payment_date:
        payment_date = payment_date.strip()
        try:
            if len(payment_date) == 10:
                parsed_payment_date = datetime.datetime.strptime(payment_date, "%Y-%m-%d")
            else:
                parsed_payment_date = datetime.datetime.fromisoformat(payment_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid payment date")

    stored_file_name = None
    stored_file_path = None
    stored_file_type = None
    if file and file.filename:
        os.makedirs("static/accounting", exist_ok=True)
        safe_name = os.path.basename(file.filename)
        extension = safe_name.split(".")[-1] if "." in safe_name else "bin"
        unique_filename = f"{uuid.uuid4()}.{extension}"
        file_path = f"static/accounting/{unique_filename}"
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        stored_file_name = safe_name
        stored_file_path = f"/{file_path.replace('\\', '/')}"
        stored_file_type = file.content_type

    receipt = models.PaymentReceipt(
        company_id=company_id,
        sale_id=sale.id if sale else None,
        user_id=current_user.id,
        display_name=(display_name or "").strip() or None,
        customer_name=(customer_name or "").strip() or None,
        customer_document=(customer_document or "").strip() or None,
        concept=concept_value,
        movement_type=movement_type_value,
        receipt_number=receipt_number_value,
        payment_date=parsed_payment_date,
        amount=amount,
        category=(category or "sale_payment").strip() or "sale_payment",
        notes=(notes or "").strip() or None,
        payment_method=payment_method_value,
        bank=(bank or "").strip() or None,
        file_name=stored_file_name,
        file_path=stored_file_path,
        file_type=stored_file_type
    )
    db.add(receipt)
    db.commit()
    db.refresh(receipt)
    return db.query(models.PaymentReceipt).options(
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.vehicle),
        joinedload(models.PaymentReceipt.sale).joinedload(models.Sale.seller),
        joinedload(models.PaymentReceipt.user)
    ).filter(models.PaymentReceipt.id == receipt.id).first()


@app.get("/finance/sales/{sale_id}/attachments", response_model=schemas.SaleAttachmentList)
def read_sale_attachments(
    sale_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    _get_sale_with_finance_access(db, sale_id, current_user)
    query = db.query(models.SaleAttachment).options(joinedload(models.SaleAttachment.user)).filter(
        models.SaleAttachment.sale_id == sale_id
    )
    if current_user.company_id:
        query = query.filter(models.SaleAttachment.company_id == current_user.company_id)
    items = query.order_by(models.SaleAttachment.created_at.desc(), models.SaleAttachment.id.desc()).all()
    return {"items": items, "total": len(items)}


@app.post("/finance/sales/{sale_id}/attachments", response_model=schemas.SaleAttachment)
async def upload_sale_attachment(
    sale_id: int,
    note: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    sale = _get_sale_with_finance_access(db, sale_id, current_user)
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="Debes adjuntar un archivo")

    os.makedirs("static/sale-attachments", exist_ok=True)
    safe_name = os.path.basename(file.filename)
    extension = safe_name.split(".")[-1] if "." in safe_name else "bin"
    unique_filename = f"{uuid.uuid4()}.{extension}"
    file_path = f"static/sale-attachments/{unique_filename}"
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    attachment = models.SaleAttachment(
        company_id=sale.company_id,
        sale_id=sale.id,
        user_id=current_user.id,
        file_name=safe_name,
        file_path=f"/{file_path.replace('\\', '/')}",
        file_type=file.content_type,
        note=(note or "").strip() or None
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    return db.query(models.SaleAttachment).options(
        joinedload(models.SaleAttachment.user)
    ).filter(models.SaleAttachment.id == attachment.id).first()


@app.delete("/finance/sales/{sale_id}/attachments/{attachment_id}")
def delete_sale_attachment(
    sale_id: int,
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    _get_sale_with_finance_access(db, sale_id, current_user)
    attachment = db.query(models.SaleAttachment).filter(
        models.SaleAttachment.id == attachment_id,
        models.SaleAttachment.sale_id == sale_id
    ).first()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
    if current_user.company_id and attachment.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    local_path = attachment.file_path.lstrip("/").replace("/", os.sep)
    db.delete(attachment)
    db.commit()
    if local_path and os.path.exists(local_path):
        try:
            os.remove(local_path)
        except OSError:
            pass
    return {"message": "Comprobante eliminado"}


@app.get("/finance/sales/{sale_id}/invoice.pdf")
def download_sale_invoice_pdf(
    sale_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_user_from_anywhere)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    sale = db.query(models.Sale).options(
        joinedload(models.Sale.vehicle),
        joinedload(models.Sale.seller),
        joinedload(models.Sale.lead),
        joinedload(models.Sale.payment_receipts)
    ).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    if current_user.company_id and sale.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized")

    receipts = sorted(
        sale.payment_receipts or [],
        key=lambda item: (item.payment_date or item.created_at or datetime.datetime.min, item.id or 0)
    )

    try:
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError:
        raise HTTPException(status_code=500, detail="La libreria reportlab no esta instalada en el servidor")

    def money(value: int) -> str:
        return f"COP ${int(value or 0):,}".replace(",", ".")

    def write_row(pdf_doc, y_pos, values, fills=None):
        x_positions = [50, 115, 290, 390, 470]
        widths = [58, 165, 90, 70, 85]
        for index, value in enumerate(values):
            pdf_doc.setFillColor(fills[index] if fills and fills[index] else HexColor("#1e293b"))
            pdf_doc.drawString(x_positions[index], y_pos, str(value)[: int(widths[index] / 4.8)])

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    pdf.setTitle(f"factura_venta_{sale.id}.pdf")

    def draw_header():
        pdf.setFillColor(HexColor("#1e293b"))
        pdf.rect(0, height - 100, width, 100, fill=1, stroke=0)
        pdf.setFillColor(HexColor("#ffffff"))
        pdf.setFont("Helvetica-Bold", 22)
        pdf.drawString(50, height - 58, "AUTOS QP - FACTURA CONSOLIDADA")
        pdf.setFont("Helvetica", 10)
        pdf.drawString(50, height - 78, "Resumen contable de venta y movimientos relacionados")
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawRightString(width - 50, height - 58, f"Venta #{sale.id}")
        pdf.setFont("Helvetica", 10)
        pdf.drawRightString(width - 50, height - 78, f"Fecha: {datetime.datetime.utcnow().strftime('%d/%m/%Y')}")

    draw_header()
    y = height - 135

    vehicle_label = " ".join(filter(None, [
        getattr(sale.vehicle, "make", None),
        getattr(sale.vehicle, "model", None),
        getattr(sale.vehicle, "plate", None)
    ])).strip() or "Sin datos de vehiculo"
    seller_label = (
        getattr(sale.seller, "full_name", None)
        or getattr(sale.seller, "email", None)
        or sale.external_seller_name
        or "Sin vendedor"
    )
    lead_label = getattr(sale.lead, "name", None) or "Sin cliente asociado"

    pdf.setFont("Helvetica-Bold", 14)
    pdf.setFillColor(HexColor("#0f172a"))
    pdf.drawString(50, y, "Datos de la venta")
    y -= 22
    for label, value in [
        ("Vehiculo", vehicle_label),
        ("Cliente", lead_label),
        ("Vendedor", seller_label),
        ("Valor de venta", money(sale.sale_price)),
        ("Estado", (sale.status or "pending").upper())
    ]:
        pdf.setFont("Helvetica-Bold", 10)
        pdf.setFillColor(HexColor("#64748b"))
        pdf.drawString(50, y, f"{label}:")
        pdf.setFont("Helvetica", 10)
        pdf.setFillColor(HexColor("#1e293b"))
        pdf.drawString(150, y, str(value)[:80])
        y -= 18

    y -= 10
    pdf.setFont("Helvetica-Bold", 14)
    pdf.setFillColor(HexColor("#0f172a"))
    pdf.drawString(50, y, "Movimientos relacionados")
    y -= 24

    pdf.setFillColor(HexColor("#f1f5f9"))
    pdf.rect(45, y - 6, width - 90, 22, fill=1, stroke=0)
    pdf.setFont("Helvetica-Bold", 9)
    write_row(pdf, y, ["Fecha", "Concepto", "Cuenta", "Tipo", "Valor"])
    y -= 22

    income_total = 0
    expense_total = 0
    pdf.setFont("Helvetica", 9)
    for receipt in receipts:
        if y < 95:
            pdf.showPage()
            draw_header()
            y = height - 135
            pdf.setFillColor(HexColor("#f1f5f9"))
            pdf.rect(45, y - 6, width - 90, 22, fill=1, stroke=0)
            pdf.setFont("Helvetica-Bold", 9)
            write_row(pdf, y, ["Fecha", "Concepto", "Cuenta", "Tipo", "Valor"])
            y -= 22
            pdf.setFont("Helvetica", 9)

        is_expense = (receipt.movement_type or "income") == "expense"
        if is_expense:
            expense_total += int(receipt.amount or 0)
        else:
            income_total += int(receipt.amount or 0)

        pdf.setStrokeColor(HexColor("#e2e8f0"))
        pdf.line(45, y - 6, width - 45, y - 6)
        write_row(
            pdf,
            y,
            [
                receipt.payment_date.strftime("%d/%m/%Y") if receipt.payment_date else "-",
                receipt.concept or receipt.notes or "Movimiento contable",
                _accounting_category_label(receipt.category),
                _accounting_movement_label(receipt.movement_type),
                money(receipt.amount)
            ],
            [None, None, None, HexColor("#dc2626") if is_expense else HexColor("#059669"), HexColor("#dc2626") if is_expense else HexColor("#059669")]
        )
        y -= 20

    if not receipts:
        pdf.setFont("Helvetica-Oblique", 10)
        pdf.setFillColor(HexColor("#64748b"))
        pdf.drawString(50, y, "Esta venta no tiene movimientos contables asociados.")
        y -= 24

    y -= 16
    if y < 150:
        pdf.showPage()
        draw_header()
        y = height - 150

    balance_total = income_total - expense_total
    pdf.setFillColor(HexColor("#f8fafc"))
    pdf.rect(330, y - 70, width - 380, 88, fill=1, stroke=1)
    pdf.setFont("Helvetica-Bold", 10)
    pdf.setFillColor(HexColor("#059669"))
    pdf.drawString(350, y, f"Ingresos: {money(income_total)}")
    y -= 20
    pdf.setFillColor(HexColor("#dc2626"))
    pdf.drawString(350, y, f"Egresos: {money(expense_total)}")
    y -= 20
    pdf.setFillColor(HexColor("#1d4ed8") if balance_total >= 0 else HexColor("#dc2626"))
    pdf.drawString(350, y, f"Neto: {money(balance_total)}")

    pdf.setStrokeColor(HexColor("#cbd5e1"))
    pdf.line(50, 80, width - 50, 80)
    pdf.setFont("Helvetica-Oblique", 8)
    pdf.setFillColor(HexColor("#94a3b8"))
    pdf.drawCentredString(width / 2, 63, "Este documento consolida los movimientos contables relacionados con la venta.")
    pdf.drawCentredString(width / 2, 51, f"Generado digitalmente el {datetime.datetime.utcnow().strftime('%d/%m/%Y %H:%M')}")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="factura_venta_{sale.id}.pdf"'}
    )


@app.get("/finance/receipts/{receipt_id}/pdf")
def download_payment_receipt_pdf(
    receipt_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_user_from_anywhere)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    receipt = _get_receipt_with_access(db, receipt_id, current_user)

    try:
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError:
        raise HTTPException(status_code=500, detail="La libreria reportlab no esta instalada en el servidor")

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    y = height - 60
    pdf.setTitle(f"recibo_{receipt.id}.pdf")
    
    # Header with background
    pdf.setFillColor(HexColor("#1e293b"))
    pdf.rect(0, height - 100, width, 100, fill=1, stroke=0)
    
    pdf.setFillColor(HexColor("#ffffff"))
    pdf.setFont("Helvetica-Bold", 24)
    pdf.drawString(50, height - 60, "AUTOS QP - COMPRA Y VENTA")
    
    pdf.setFont("Helvetica", 10)
    pdf.drawString(50, height - 80, "Soporte de Contabilidad | Gestion Financiera")
    
    pdf.setFillColor(HexColor("#ffffff"))
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawRightString(width - 50, height - 60, f"No. {receipt.receipt_number or 'PROV'}")
    pdf.setFont("Helvetica", 10)
    pdf.drawRightString(width - 50, height - 80, f"Fecha: {receipt.payment_date.strftime('%d/%m/%Y') if receipt.payment_date else 'N/A'}")

    y = height - 140
    
    # Main Content
    pdf.setFillColor(HexColor("#0f172a"))
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, y, "Detalle del Movimiento")
    y -= 10
    
    pdf.setStrokeColor(HexColor("#e2e8f0"))
    pdf.line(50, y, width - 50, y)
    y -= 30

    display_lines = [
        ("Concepto", (receipt.concept or "Sin concepto").upper()),
        ("Tipo de Movimiento", _accounting_movement_label(receipt.movement_type).upper()),
        ("Cuenta Contable", _accounting_category_label(receipt.category).upper()),
        ("Forma de Pago", (receipt.payment_method or "Sin definir").capitalize()),
        ("Banco / Cuenta", (receipt.bank or "Sin definir").upper()),
        ("Valor Total", f"COP ${int(receipt.amount or 0):,}".replace(",", ".")),
    ]

    if receipt.sale:
        vehicle_label = " ".join(filter(None, [
            getattr(receipt.sale.vehicle, "make", None),
            getattr(receipt.sale.vehicle, "model", None),
            getattr(receipt.sale.vehicle, "plate", None)
        ])).strip()
        display_lines.extend([
            ("Referencia Venta", f"#{receipt.sale.id}"),
            ("Vehiculo Asociado", (vehicle_label or "Sin datos de vehiculo").upper()),
            ("Vendedor", (getattr(receipt.sale.seller, "full_name", None) or getattr(receipt.sale.seller, "email", "") or "N/A").upper())
        ])

    for label, value in display_lines:
        pdf.setFont("Helvetica-Bold", 11)
        pdf.setFillColor(HexColor("#64748b"))
        pdf.drawString(50, y, f"{label}:")
        
        pdf.setFont("Helvetica", 11)
        pdf.setFillColor(HexColor("#1e293b"))
        pdf.drawString(180, y, str(value))
        
        y -= 6
        pdf.setStrokeColor(HexColor("#f1f5f9"))
        pdf.line(50, y, width - 50, y)
        y -= 20

    if receipt.notes:
        y -= 20
        pdf.setFillColor(HexColor("#f8fafc"))
        pdf.rect(50, y - 60, width - 100, 70, fill=1, stroke=1)
        
        pdf.setFillColor(HexColor("#475569"))
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(60, y, "NOTAS Y OBSERVACIONES:")
        y -= 15
        pdf.setFont("Helvetica", 9)
        pdf.setFillColor(HexColor("#1e293b"))
        text_object = pdf.beginText(60, y)
        text_object.setLeading(12)
        for paragraph in str(receipt.notes).splitlines() or [""]:
            text_object.textLine(paragraph[:100])
        pdf.drawText(text_object)
        y -= 60

    # Footer
    pdf.setStrokeColor(HexColor("#cbd5e1"))
    pdf.line(50, 100, width - 50, 100)
    
    pdf.setFont("Helvetica-Oblique", 8)
    pdf.setFillColor(HexColor("#94a3b8"))
    pdf.drawCentredString(width/2, 85, "Este documento es un soporte interno de contabilidad de AUTOS QP.")
    pdf.drawCentredString(width/2, 75, f"Generado digitalmente el {datetime.datetime.utcnow().strftime('%d/%m/%Y %H:%M')}")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="recibo_{receipt.id}.pdf"'}
    )


@app.delete("/finance/receipts/{receipt_id}")
def delete_payment_receipt(
    receipt_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if get_user_role_name(current_user) not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="Solo un administrador puede eliminar recibos")

    receipt = _get_receipt_with_access(db, receipt_id, current_user)
    if receipt.file_path:
        try:
            relative_path = receipt.file_path.lstrip("/").replace("/", os.sep)
            absolute_path = os.path.abspath(relative_path)
            if os.path.exists(absolute_path):
                os.remove(absolute_path)
        except Exception:
            pass

    db.delete(receipt)
    db.commit()
    return {"message": "Recibo eliminado correctamente"}

# --- INTERNAL CHAT ENDPOINTS ---

@app.get("/internal-messages", response_model=List[schemas.InternalMessage])
def get_internal_messages(
    date: str = Query(None), # YYYY-MM-DD
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Filter by company
    query = db.query(models.InternalMessage).options(
        joinedload(models.InternalMessage.sender),
        joinedload(models.InternalMessage.recipient)
    ).filter(models.InternalMessage.company_id == current_user.company_id)
    
    # Filter Permissions: 
    # Show message IF:
    # 1. recipient_id is NULL (Broadcast)
    # 2. recipient_id == current_user.id (Received DM)
    # 3. sender_id == current_user.id (Sent DM)
    from sqlalchemy import or_
    query = query.filter(
        or_(
            models.InternalMessage.recipient_id == None,
            models.InternalMessage.recipient_id == current_user.id,
            models.InternalMessage.sender_id == current_user.id
        )
    )
    
    if date:
        try:
            date_obj = datetime.datetime.strptime(date, "%Y-%m-%d").date()
            # Filter by date range (start of day to end of day)
            start_of_day = datetime.datetime.combine(date_obj, datetime.time.min)
            end_of_day = datetime.datetime.combine(date_obj, datetime.time.max)
            query = query.filter(models.InternalMessage.created_at >= start_of_day, models.InternalMessage.created_at <= end_of_day)
        except ValueError:
             raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        # Default to today
        today = datetime.datetime.utcnow().date()
        start_of_day = datetime.datetime.combine(today, datetime.time.min)
        query = query.filter(models.InternalMessage.created_at >= start_of_day)
        
    return query.order_by(models.InternalMessage.created_at.asc()).all()

@app.post("/internal-messages", response_model=schemas.InternalMessage)
def create_internal_message(
    message: schemas.InternalMessageCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    # Verify recipient strictly if provided
    recipient_id = message.recipient_id
    if recipient_id:
        recipient = db.query(models.User).filter(models.User.id == recipient_id).first()
        if not recipient:
            raise HTTPException(status_code=404, detail="Recipient not found")
        if recipient.company_id != current_user.company_id:
             raise HTTPException(status_code=403, detail="Cannot message users from other companies")

    db_message = models.InternalMessage(
        content=message.content,
        company_id=current_user.company_id,
        sender_id=current_user.id,
        recipient_id=recipient_id
    )
    db.add(db_message)
    db.commit()
    return db.query(models.InternalMessage).options(
        joinedload(models.InternalMessage.sender),
        joinedload(models.InternalMessage.recipient)
    ).filter(models.InternalMessage.id == db_message.id).first()

@app.post("/internal-messages/upload", response_model=schemas.InternalMessage)
async def upload_internal_message_file(
    request: Request,
    file: UploadFile = File(...),
    recipient_id: Optional[int] = Form(None),
    content: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if recipient_id:
        recipient = db.query(models.User).filter(models.User.id == recipient_id).first()
        if not recipient:
            raise HTTPException(status_code=404, detail="Recipient not found")
        if recipient.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="Cannot message users from other companies")

    os.makedirs("static/internal_chat", exist_ok=True)

    safe_name = os.path.basename(file.filename or "archivo")
    ext = safe_name.split(".")[-1] if "." in safe_name else "bin"
    unique_name = f"{uuid.uuid4()}.{ext}"
    rel_path = f"static/internal_chat/{unique_name}"

    with open(rel_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    rel_path_web = rel_path.replace("\\", "/")
    relative_url = f"/api/{rel_path_web}"
    file_url = f"{str(request.base_url).rstrip('/')}{relative_url}"
    payload = {
        "type": "file",
        "file_name": safe_name,
        "storage_name": unique_name,
        "file_url": file_url,
        "file_url_relative": relative_url,
        "file_path": rel_path_web,
        "file_type": file.content_type or "application/octet-stream",
        "file_size": os.path.getsize(rel_path),
        "text": content.strip() if content else ""
    }

    db_message = models.InternalMessage(
        content=f"__FILE__{json.dumps(payload, ensure_ascii=False)}",
        company_id=current_user.company_id,
        sender_id=current_user.id,
        recipient_id=recipient_id
    )
    db.add(db_message)
    db.commit()
    return db.query(models.InternalMessage).options(
        joinedload(models.InternalMessage.sender),
        joinedload(models.InternalMessage.recipient)
    ).filter(models.InternalMessage.id == db_message.id).first()


